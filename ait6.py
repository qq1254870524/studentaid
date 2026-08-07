"""
StudentAid 浏览器自动化工具 v6。

2026-08-08 第十一步稳定版更新：
1. GUI 新增 browser-use / playwright 后端下拉选项；两种模式都为每个 worker
   创建独立浏览器进程和独立临时配置目录，默认使用 2 个处理线程。
2. browser-use 模式直接采用 BrowserProfile 的 Chrome 参数、隔离目录和 CDP
   启动策略，再执行现有确定性表单流程，避免默认共享 daemon 在并发时互相抢焦点。
3. Limit Reached: Try Again in 24 Hours 改为正常明确结果：实时追加累计 CSV，
   最终状态原样写入第 6 列，随后删除对应输入行并继续处理下一条。
4. 停止或批次结束时，每个 worker 都清除 cookie、cache、storage、service worker，
   回到空白页并结束本脚本创建的浏览器进程；browser-use 临时配置目录同步删除。
5. 恢复并发任务队列和单独 SQLite writer，累计输出与输入删行继续使用进程内锁和
   原子文件替换，两个浏览器线程不会同时破坏 CSV/XLSX。
6. 键盘逐字输入延迟由 60ms 调整为 20ms，字段完成后的稳定等待由 500ms 调整为
   200ms；仍保留真实键盘事件、失焦和 Continue 启用校验。
7. 累计 CSV 的 DOB 对 5 列无表头输入保留原始显示格式，与实际输出参考一致。
8. 根据假资料和实际输出参考补齐 ``September 07, 1980`` 这类英文月份 DOB
   导入格式；此类有效资料不再作为格式错误留在输入文件。

2026-08-08 第十步稳定版更新：
1. Account Not Found 明确结果实时保存后，清除 StudentAid 专用浏览器全部站点数据，
   回到 about:blank；处理下一条时重新打开找回页面。
2. Retrieve Your Log-in Information 明确结果实时保存后，点击可见 Cancel，
   等待返回空白输入表单，再填写下一条资料。
3. 采用严格顺序处理，避免多个共享 CDP 页面同时清理 cookie/session 导致提交互相干扰。
4. 累计结果实时追加到用户选择的 CSV，不覆盖旧数据；程序启动时读取输出第一列，
   输入第一列已存在于输出的资料会直接从输入文件原子删除。
5. 每条明确结果按“先写 SQLite、再追加累计 CSV、最后删除输入行”的顺序落盘，
   中断重启时可依靠输出第一列去重并继续。
6. 定位 Continue 转圈根因：Playwright 直接启动 Chrome 时 navigator.webdriver=true；
   新版复用 browser-use 的 AutomationControlled 启动参数，实测启动后该标志为 false。
7. Continue 改用可见按钮坐标鼠标点击；持续 Loading 自动清理、重开、重填并重试。
8. 识别官方 Limit Reached: Try Again in 24 Hours，立即停止批次并保留当前及后续输入。
9. browser-use 实测 Account Not Found、全数据清理、about:blank 重开及官方九次限制；
   Retrieve 结果后的 Cancel 使用真实按钮结构并由 Playwright DOM 回归覆盖。
10. 新增统一一键安装启动 CMD；依赖存在就跳过，缺少时才安装。

2026-08-07 第九步源码校验更新：
1. 使用 browser-use 独立核验 StudentAid 真实页面字段、Continue 按钮和提交状态时间线。
2. 删除“姓名输入框消失即判定可找回”的宽泛条件；只有明确结果标志才允许完成，
   防止页面跳转或短暂重绘被误判为成功。
3. 页面结果等待改为可停止的短轮询；保留 Loading 等待，并准确识别站点错误。
4. CDP 模式复用 Chrome 默认浏览器上下文并新开页面，保持与 browser-use 实测一致；
   输入改用真实键盘事件、字段失焦和提交前稳定等待，修复过快 fill 导致站点返回 unknown error。
5. 检测到因页面停留过久产生的 unknown error 时自动新开页面、重新填表并提交一次。
6. 每条记录实时输出“打开页面、填写资料、已提交、等待结果、已识别结果”等安全阶段日志；
   长时间等待时每 5 秒输出心跳，最终状态实时写 SQLite 并刷新 GUI，不记录原始资料。
7. Playwright 仍是正式运行根基，browser-use 仅用于独立校验。

2026-08-07 第八步实测更新：
1. 保留第七步的 GUI、SQLite WAL、并发处理和 CSV 导出能力。
2. 新增 Chrome CDP 连接模式：优先复用已通过站点网络校验的本机 Chrome，
   解决新启动 Playwright 浏览器访问 StudentAid 时的 HTTP/2 协议错误。
3. 支持 STUDENTAID_CDP_URL 显式指定 CDP 地址；未设置时自动探测
   http://127.0.0.1:9223，设置为 off/none/disabled 可关闭自动探测。
4. 新增页面导航重试、共享浏览器安全断开和浏览器模式日志。
5. 修复输入页标题被误判为找回成功的问题；现在会等待 Loading 结束，
   并把站点未知错误准确记录为失败。
6. 修复纯文本 CSV 中全数字 SSN 被 float 转换吞掉前导 0 的问题。
7. 第八步使用测试资料完成 GUI、浏览器、SQLite 和导出链路实测。

2026-08-07 第七步更新（原 ait2.py 日期记录受本机时钟影响）：
1. 新增 Tkinter GUI，可选择 CSV/SCV/TXT/XLSX 输入文件和结果目录。
2. 导入资料先进入 SQLite；数据库固定启用 WAL。
3. 使用多个独立 Playwright 处理线程和一个专用 SQLite 写入线程。
4. 支持开始、停止、线程数设置和实时进度；停止后不再领取新任务。
5. 每个批次完成或停止后自动导出 UTF-8 BOM CSV。
6. 兼容无表头的 ``SSN,月,日,年,姓,名,地址``、旧版
   ``SSN,DOB,First Name,Last Name,Address``，以及带常见英文表头的资料。

旧版 ait.py 保持不变。本文件不把 SSN 或整条原始资料打印到日志。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import queue
import re
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
from typing import Any, Callable, Iterable, Mapping, Sequence
import urllib.request
import uuid

sys.dont_write_bytecode = True

try:
    from playwright.sync_api import Browser, Page, Playwright, sync_playwright
except ImportError:  # GUI 仍可启动，并在开始处理时给出明确错误。
    Browser = Page = Playwright = Any  # type: ignore[assignment,misc]
    sync_playwright = None


APP_TITLE = "StudentAid 批量处理工具 - 第十一步稳定版"
DATABASE_FILENAME = "studentaid.sqlite3"
CUMULATIVE_OUTPUT_FILENAME = "StudentAid累计结果.csv"
LIMIT_REACHED_HEADING = "Limit Reached: Try Again in 24 Hours"
BROWSER_BACKENDS = ("browser-use", "playwright")
RETRIEVE_ACCOUNT_DETAILS_URL = (
    "https://studentaid.gov/fsa-id/sign-in/retrieve-account-details"
)

FIRST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameFirstName"
LAST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameLastName"
BIRTH_MONTH_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthMonth"
BIRTH_DAY_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthDay"
BIRTH_YEAR_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthYear"
SSN_SELECTOR = "#fsa_Input_ForgotUsernameSsnInput"

RESULT_COLUMNS = (
    "Result_Heading",
    "Masked_Phone",
    "Masked_Email",
    "Recovery_Method",
    "Processing_Status",
    "Error",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
)


@dataclass(frozen=True)
class AccountDetails:
    """StudentAid 账户找回页面所需字段。"""

    first_name: str
    last_name: str
    birth_month: str
    birth_day: str
    birth_year: str
    ssn: str
    original_ssn: str


@dataclass(frozen=True)
class RecoveryResult:
    """页面返回的可公开保存结果；联系方式保持页面原有脱敏形式。"""

    result_code: str
    heading: str
    masked_phone: str
    masked_email: str
    recovery_method: str


@dataclass(frozen=True)
class ImportedRecord:
    """一条已读取的输入资料；校验失败的行也会进入数据库。"""

    source_file: str
    source_sheet: str
    source_row: int
    original_fields: tuple[str, ...]
    details: AccountDetails | None
    address: str
    import_error: str = ""


@dataclass(frozen=True)
class WorkItem:
    """SQLite、浏览器、累计输出和输入删行之间传递的任务。"""

    record_id: int
    details: AccountDetails
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    original_fields: tuple[str, ...] = ()
    address: str = ""

    @property
    def record_key(self) -> str:
        first_value = self.original_fields[0] if self.original_fields else self.details.original_ssn
        return _normalise_record_key(first_value)


class StopRequested(RuntimeError):
    """在网页处理步骤之间收到停止请求。"""


class PageSessionExpired(RuntimeError):
    """StudentAid 页面停留过久后返回 unknown error，需要重新打开页面。"""


class PageSubmissionStalled(RuntimeError):
    """Continue 已提交但站点长时间停在 Loading，需要重建浏览器会话。"""


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _clean_error(exc: BaseException | str, limit: int = 1000) -> str:
    """生成适合写库和显示的单行错误，不包含输入资料。"""
    text = str(exc).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit] or type(exc).__name__


def _normalise_month(value: str) -> str:
    aliases = {
        "january": "01", "jan": "01", "february": "02", "feb": "02",
        "march": "03", "mar": "03", "april": "04", "apr": "04",
        "may": "05", "june": "06", "jun": "06", "july": "07",
        "jul": "07", "august": "08", "aug": "08", "september": "09",
        "sep": "09", "sept": "09", "october": "10", "oct": "10",
        "november": "11", "nov": "11", "december": "12", "dec": "12",
    }
    cleaned = value.strip().lower()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    month = aliases.get(cleaned) or (cleaned.zfill(2) if cleaned.isdigit() else "")
    if month not in {f"{number:02d}" for number in range(1, 13)}:
        raise ValueError("出生月份必须是 1-12、01-12 或英文月份名称")
    return month


def _normalise_day(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if not digits or not 1 <= int(digits) <= 31:
        raise ValueError("出生日期（日）必须是 1-31")
    return digits.zfill(2)


def _normalise_year(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) != 4 or not 1900 <= int(digits) <= date.today().year:
        raise ValueError("出生年份必须是四位数字，且不晚于当前年份")
    return digits


def _normalise_ssn(value: str) -> str:
    cleaned = value.strip()
    # 只移除 Excel 数字单元格常见的 .0 后缀；纯数字文本必须保留前导 0。
    if re.fullmatch(r"\d+\.0+", cleaned):
        cleaned = cleaned.split(".", 1)[0]
    digits = re.sub(r"\D", "", cleaned)
    # Excel 把首位 0 当数字格式丢掉时，允许 8 位数字恢复为 9 位。
    if len(digits) == 8:
        digits = digits.zfill(9)
    if len(digits) != 9:
        raise ValueError("Social Security Number 必须包含 9 位数字")
    return digits


def _validate_details(
    ssn_value: str,
    month_value: str,
    day_value: str,
    year_value: str,
    first_name: str,
    last_name: str,
) -> AccountDetails:
    ssn = _normalise_ssn(ssn_value)
    month = _normalise_month(month_value)
    day = _normalise_day(day_value)
    year = _normalise_year(year_value)
    first_name = first_name.strip()
    last_name = last_name.strip()

    if not first_name or len(first_name) > 35:
        raise ValueError("First Name 不能为空且长度不能超过 35 个字符")
    if not last_name or len(last_name) > 35:
        raise ValueError("Last Name 不能为空且长度不能超过 35 个字符")
    try:
        date(int(year), int(month), int(day))
    except ValueError as exc:
        raise ValueError("出生日期不是有效日期") from exc

    return AccountDetails(
        first_name=first_name,
        last_name=last_name,
        birth_month=month,
        birth_day=day,
        birth_year=year,
        ssn=ssn,
        original_ssn=ssn_value,
    )


def _parse_dob(value: str) -> tuple[str, str, str]:
    value = value.strip()
    for date_format in (
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%y", "%m-%d-%y", "%B %d, %Y", "%b %d, %Y",
        "%B %d %Y", "%b %d %Y",
    ):
        try:
            parsed = datetime.strptime(value, date_format).date()
            return f"{parsed.month:02d}", f"{parsed.day:02d}", f"{parsed.year:04d}"
        except ValueError:
            continue
    raise ValueError(
        "DOB 格式无效，应为 MM/DD/YYYY、MM-DD-YYYY、YYYY-MM-DD "
        "或英文月份日期（如 September 07, 1980）"
    )


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalise_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.strip().casefold())


HEADER_ALIASES: Mapping[str, set[str]] = {
    "ssn": {"ssn", "socialsecuritynumber", "socialsecurityno", "socialsecurity"},
    "dob": {"dob", "dateofbirth", "birthdate"},
    "month": {"birthmonth", "dobmonth", "month", "mm"},
    "day": {"birthday", "dobday", "day", "dd"},
    "year": {"birthyear", "dobyear", "year", "yyyy"},
    "first_name": {"firstname", "givenname", "fname", "first"},
    "last_name": {"lastname", "surname", "familyname", "lname", "last"},
    "address": {"address", "streetaddress", "mailingaddress", "homeaddress"},
}


def _header_mapping(row: Sequence[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        cleaned = _normalise_header(value)
        for canonical, aliases in HEADER_ALIASES.items():
            if cleaned in aliases and canonical not in mapping:
                mapping[canonical] = index
                break
    has_date = "dob" in mapping or {"month", "day", "year"}.issubset(mapping)
    if "ssn" in mapping and has_date and {"first_name", "last_name"}.issubset(mapping):
        return mapping
    return None


def _mapped_value(row: Sequence[str], mapping: Mapping[str, int], key: str) -> str:
    index = mapping.get(key)
    return row[index].strip() if index is not None and index < len(row) else ""


def _row_looks_like_split_date(row: Sequence[str]) -> bool:
    if len(row) < 6:
        return False
    try:
        _normalise_month(row[1])
        _normalise_day(row[2])
        _normalise_year(row[3])
        return True
    except ValueError:
        return False


def _parse_input_row(
    fields: Sequence[str],
    header: Mapping[str, int] | None,
) -> tuple[AccountDetails, str]:
    row = [str(value) for value in fields]
    if header is not None:
        ssn = _mapped_value(row, header, "ssn")
        first_name = _mapped_value(row, header, "first_name")
        last_name = _mapped_value(row, header, "last_name")
        address = _mapped_value(row, header, "address")
        dob = _mapped_value(row, header, "dob")
        if dob:
            month, day, year = _parse_dob(dob)
        else:
            month = _mapped_value(row, header, "month")
            day = _mapped_value(row, header, "day")
            year = _mapped_value(row, header, "year")
    elif _row_looks_like_split_date(row):
        # 当前测试资料：SSN, 月, 日, 年, Last Name, First Name, Address。
        ssn, month, day, year = row[:4]
        last_name, first_name = row[4:6]
        address = ",".join(row[6:]).strip()
    else:
        if len(row) < 4:
            raise ValueError(
                "至少需要 4 列：SSN、DOB、First Name、Last Name"
            )
        # 旧版输入：SSN, DOB, First Name, Last Name, Address。
        ssn, dob, first_name, last_name = row[:4]
        month, day, year = _parse_dob(dob)
        address = ",".join(row[4:]).strip()

    return _validate_details(
        ssn, month, day, year, first_name, last_name
    ), address


def _records_from_rows(
    rows: Iterable[Sequence[Any]],
    input_path: Path,
    sheet_name: str = "",
) -> list[ImportedRecord]:
    materialised: list[tuple[int, list[str]]] = []
    for row_number, raw_row in enumerate(rows, start=1):
        row = [_cell_to_text(value) for value in raw_row]
        while row and row[-1] == "":
            row.pop()
        if row and any(value.strip() for value in row):
            materialised.append((row_number, row))

    if not materialised:
        return []

    mapping = _header_mapping(materialised[0][1])
    data_rows = materialised[1:] if mapping is not None else materialised
    imported: list[ImportedRecord] = []
    for row_number, row in data_rows:
        details: AccountDetails | None = None
        address = ""
        import_error = ""
        try:
            details, address = _parse_input_row(row, mapping)
        except Exception as exc:
            import_error = _clean_error(exc)
        imported.append(
            ImportedRecord(
                source_file=str(input_path.resolve()),
                source_sheet=sheet_name,
                source_row=row_number,
                original_fields=tuple(row),
                details=details,
                address=address,
                import_error=import_error,
            )
        )
    return imported


def _read_text_rows(input_path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = input_path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8") from last_error

    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            delimiter = "\t"
    return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def load_input_records(input_path: Path) -> list[ImportedRecord]:
    """读取 CSV/SCV/TXT/XLSX，返回包括校验失败行在内的导入记录。"""
    input_path = input_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    suffix = input_path.suffix.casefold()
    records: list[ImportedRecord] = []
    if suffix in {".csv", ".scv", ".txt"}:
        records.extend(_records_from_rows(_read_text_rows(input_path), input_path))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                records.extend(
                    _records_from_rows(
                        worksheet.iter_rows(values_only=True),
                        input_path,
                        worksheet.title,
                    )
                )
        finally:
            workbook.close()
    else:
        raise ValueError("只支持 .csv、.scv、.txt 和 .xlsx 输入文件")

    if not records:
        raise ValueError("输入文件中没有可导入的资料行")
    return records


_SOURCE_FILE_LOCK = threading.RLock()


def _normalise_record_key(value: Any) -> str:
    """输出第一列/输入第一列比较键；SSN 忽略连字符、空格和 Excel .0。"""
    text = _cell_to_text(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    digits = re.sub(r"\D", "", text)
    if 8 <= len(digits) <= 9:
        return digits.zfill(9)
    return text.casefold()


def _detect_text_format(path: Path) -> tuple[str, str, list[list[str]]]:
    raw = path.read_bytes()
    encoding = "utf-8-sig"
    text = ""
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8")
    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            delimiter = "\t"
    rows = [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]
    return encoding, delimiter, rows


def read_output_first_column_keys(output_path: Path) -> set[str]:
    """读取累计输出第一列；空文件或不存在时返回空集合。"""
    output_path = output_path.resolve()
    if not output_path.is_file() or output_path.stat().st_size == 0:
        return set()
    _encoding, _delimiter, rows = _detect_text_format(output_path)
    return {
        key for row in rows if row and (key := _normalise_record_key(row[0]))
        and key not in {"ssn", "socialsecuritynumber", "inputcolumn1"}
    }


def _atomic_write_text_rows(
    path: Path, rows: Sequence[Sequence[Any]], encoding: str, delimiter: str
) -> None:
    temporary = path.with_name(f".{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding=encoding, newline="") as stream:
            writer = csv.writer(stream, delimiter=delimiter, lineterminator="\n")
            writer.writerows(rows)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def remove_input_rows_by_keys(input_path: Path, keys: set[str]) -> int:
    """按第一列原子删除输入行；XLSX 保留样式和其他工作表。"""
    keys = {key for value in keys if (key := _normalise_record_key(value))}
    if not keys:
        return 0
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            kept: list[list[str]] = []
            removed = 0
            mapping = _header_mapping(rows[0]) if rows else None
            for index, row in enumerate(rows):
                if index == 0 and mapping is not None:
                    kept.append(row)
                    continue
                key = _normalise_record_key(row[0]) if row else ""
                if key and key in keys:
                    removed += 1
                else:
                    kept.append(row)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path)
            removed = 0
            temporary = input_path.with_name(
                f".{input_path.stem}.{os.getpid()}.{uuid.uuid4().hex}.tmp.xlsx"
            )
            try:
                for worksheet in workbook.worksheets:
                    header_values = [_cell_to_text(cell.value) for cell in worksheet[1]]
                    mapping = _header_mapping(header_values)
                    first_data_row = 2 if mapping is not None else 1
                    rows_to_delete = []
                    for row_number in range(first_data_row, worksheet.max_row + 1):
                        key = _normalise_record_key(worksheet.cell(row_number, 1).value)
                        if key and key in keys:
                            rows_to_delete.append(row_number)
                    for row_number in reversed(rows_to_delete):
                        worksheet.delete_rows(row_number, 1)
                    removed += len(rows_to_delete)
                if removed:
                    workbook.save(temporary)
                    workbook.close()
                    os.replace(temporary, input_path)
                else:
                    workbook.close()
            finally:
                try:
                    workbook.close()
                except Exception:
                    pass
                if temporary.exists():
                    temporary.unlink()
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def _format_output_dob(details: AccountDetails) -> str:
    return f"{int(details.birth_month):02d}/{int(details.birth_day):02d}/{details.birth_year}"


def build_cumulative_output_row(item: WorkItem, result: RecoveryResult) -> list[str]:
    """生成与实际输出参考一致的 5 个资料列 + 4 个结果列。"""
    first_column = (
        item.original_fields[0] if item.original_fields else item.details.original_ssn
    )
    original_dob = ""
    if len(item.original_fields) >= 5 and not _row_looks_like_split_date(item.original_fields):
        original_dob = item.original_fields[1].strip()
    return [
        first_column,
        original_dob or _format_output_dob(item.details),
        item.details.first_name,
        item.details.last_name,
        item.address,
        result.heading,
        result.masked_phone,
        result.masked_email,
        result.recovery_method,
    ]


def append_cumulative_result(
    output_path: Path, item: WorkItem, result: RecoveryResult
) -> bool:
    """实时追加累计结果；第一列已存在时不重复追加。"""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with _SOURCE_FILE_LOCK:
        existing_keys = read_output_first_column_keys(output_path)
        if item.record_key and item.record_key in existing_keys:
            return False
        with output_path.open("a", encoding="utf-8-sig", newline="") as stream:
            csv.writer(stream, lineterminator="\n").writerow(
                build_cumulative_output_row(item, result)
            )
            stream.flush()
            os.fsync(stream.fileno())
    return True


def resolve_output_target(output_target: Path) -> tuple[Path, Path]:
    """兼容旧版目录参数；GUI 第十一步直接选择累计 CSV。"""
    output_target = output_target.resolve()
    if output_target.suffix.casefold() in {".csv", ".scv", ".txt"}:
        return output_target.parent, output_target
    return output_target, output_target / CUMULATIVE_OUTPUT_FILENAME


@dataclass
class BrowserLaunch:
    """浏览器启动结果；默认浏览器完全由当前工具拥有。"""

    browser: Browser
    external_browser: bool
    mode: str
    owns_browser: bool = False
    dedicated_profile: bool = False


class BrowserUseBrowserHost:
    """使用 browser-use BrowserProfile 启动独立 Chrome 和 CDP。

    browser-use CLI 的默认本机 daemon 是共享浏览器，不能安全供两个 worker 并发。
    本类复用 BrowserProfile 的完整 Chrome 参数，为每个 worker 创建独立进程、CDP
    端口和临时 user-data-dir；避免 BrowserSession 看门狗与 Playwright 双控制器在
    多线程结束时竞争标签页。表单由确定性 Playwright 定位器通过该 CDP 会话执行。
    """

    def __init__(self, worker_number: int) -> None:
        self.worker_number = worker_number
        self._cdp_url = ""
        self._process: subprocess.Popen[Any] | None = None
        self._browser_pid: int | None = None
        self._profile_dir = Path(
            tempfile.mkdtemp(prefix=f"browser-use-user-data-dir-studentaid-{worker_number}-")
        )

    @property
    def cdp_url(self) -> str:
        return self._cdp_url

    @staticmethod
    def _find_free_port() -> int:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
            probe.bind(("127.0.0.1", 0))
            return int(probe.getsockname()[1])

    def _wait_until_ready(self, timeout: float) -> None:
        deadline = time.monotonic() + timeout
        last_error: BaseException | None = None
        while time.monotonic() < deadline:
            if self._process is not None and self._process.poll() is not None:
                raise RuntimeError(
                    f"browser-use Chrome 提前退出，退出码 {self._process.returncode}"
                )
            try:
                with urllib.request.urlopen(
                    f"{self._cdp_url}/json/version", timeout=1
                ) as response:
                    payload = json.load(response)
                if payload.get("webSocketDebuggerUrl"):
                    return
            except Exception as exc:
                last_error = exc
            time.sleep(0.1)
        raise RuntimeError("browser-use Chrome CDP 启动超时") from last_error

    def start(self, timeout: float = 60.0) -> None:
        try:
            try:
                from browser_use.browser.chrome import find_chrome_executable
                from browser_use.browser.profile import BrowserProfile
            except ImportError as exc:
                raise RuntimeError(
                    "未安装 browser-use，请双击一键启动 CMD，"
                    "或执行：pip install -r requirements.txt"
                ) from exc

            chrome_path = find_chrome_executable()
            if not chrome_path:
                raise RuntimeError("browser-use 未找到 Google Chrome 可执行文件")
            profile = BrowserProfile(
                executable_path=chrome_path,
                user_data_dir=str(self._profile_dir),
                headless=False,
                keep_alive=False,
                args=[
                    "--disable-blink-features=AutomationControlled",
                    "--window-size=1440,900",
                ],
            )
            debug_port = self._find_free_port()
            launch_args = [*profile.get_args(), f"--remote-debugging-port={debug_port}"]
            self._cdp_url = f"http://127.0.0.1:{debug_port}"
            creationflags = (
                getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                if os.name == "nt"
                else 0
            )
            self._process = subprocess.Popen(
                [str(chrome_path), *launch_args],
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=creationflags,
            )
            self._browser_pid = self._process.pid
            self._wait_until_ready(timeout)
        except Exception:
            try:
                self.close()
            except Exception:
                pass
            raise

    def _terminate_process_tree(self) -> None:
        if not self._browser_pid:
            return
        try:
            import psutil

            parent = psutil.Process(self._browser_pid)
            processes = parent.children(recursive=True) + [parent]
            for process in reversed(processes):
                try:
                    process.terminate()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            _gone, alive = psutil.wait_procs(processes, timeout=4)
            for process in alive:
                try:
                    process.kill()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except Exception:
            pass

    def _remove_profile_dir(self) -> None:
        last_error: OSError | None = None
        for _attempt in range(5):
            try:
                if self._profile_dir.exists():
                    shutil.rmtree(self._profile_dir)
                return
            except OSError as exc:
                last_error = exc
                time.sleep(0.2)
        if self._profile_dir.exists():
            raise RuntimeError(
                f"browser-use 临时缓存目录删除失败：{self._profile_dir}"
            ) from last_error

    def close(self) -> None:
        self._terminate_process_tree()
        if self._process is not None:
            try:
                self._process.wait(timeout=5)
            except Exception:
                try:
                    self._process.kill()
                    self._process.wait(timeout=5)
                except Exception:
                    pass
            self._process = None
        self._remove_profile_dir()


def launch_browser(playwright: Playwright) -> BrowserLaunch:
    """启动独立、隐藏 AutomationControlled 的真实 Chrome。

    实测旧版 ``playwright.chromium.launch(channel="chrome")`` 会令
    ``navigator.webdriver`` 为 true，StudentAid 接口随后长期停在 Loading。
    browser-use 使用的同名 Blink 开关可消除这个差异，同时仍由 Playwright
    管理独立 BrowserContext，便于彻底清除本批次浏览器数据。
    """
    try:
        browser = playwright.chromium.launch(
            channel="chrome",
            headless=False,
            args=["--disable-blink-features=AutomationControlled"],
        )
    except Exception as exc:
        raise RuntimeError(
            "无法启动 Google Chrome。请双击一键启动 CMD 自动检查并安装 Chrome。"
        ) from exc
    return BrowserLaunch(
        browser,
        False,
        "独立 Google Chrome（AutomationControlled 已关闭）",
        owns_browser=True,
        dedicated_profile=True,
    )


def _check_stop(stop_event: threading.Event) -> None:
    if stop_event.is_set():
        raise StopRequested("用户已请求停止")


def step_1_open_retrieve_account_details(
    page: Page, stop_event: threading.Event
) -> None:
    last_error: BaseException | None = None
    for attempt in range(1, 4):
        _check_stop(stop_event)
        try:
            page.goto(
                RETRIEVE_ACCOUNT_DETAILS_URL,
                wait_until="load",
                timeout=60_000,
            )
            _check_stop(stop_event)
            page.locator(FIRST_NAME_SELECTOR).wait_for(
                state="visible", timeout=60_000
            )
            page.wait_for_function(
                """
                () => {
                    const text = document.body?.innerText || "";
                    return !text.includes("Loading...")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameLastName")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameDateOfBirthMonth")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameSsnInput");
                }
                """,
                timeout=30_000,
            )
            return
        except StopRequested:
            raise
        except Exception as exc:
            last_error = exc
            if attempt >= 3:
                break
            page.wait_for_timeout(1_500 * attempt)
    raise RuntimeError(
        "StudentAid 页面连续 3 次无法打开；请先在 Chrome 中确认页面可访问，"
        "或切换 GUI 中的浏览器后端后重试。"
    ) from last_error


def _type_account_field(page: Page, selector: str, value: str) -> None:
    """用真实键盘事件输入并失焦，确保 StudentAid 前端状态完整更新。"""
    field = page.locator(selector)
    field.click(timeout=30_000)
    field.fill("")
    field.press_sequentially(value, delay=20)
    field.press("Tab")


def step_2_fill_account_details(
    page: Page, details: AccountDetails, stop_event: threading.Event
) -> None:
    _check_stop(stop_event)
    _type_account_field(page, FIRST_NAME_SELECTOR, details.first_name)
    _type_account_field(page, LAST_NAME_SELECTOR, details.last_name)
    page.locator(BIRTH_MONTH_SELECTOR).select_option(details.birth_month)
    page.locator(BIRTH_MONTH_SELECTOR).press("Tab")
    _type_account_field(page, BIRTH_DAY_SELECTOR, str(int(details.birth_day)))
    _type_account_field(page, BIRTH_YEAR_SELECTOR, details.birth_year)
    _type_account_field(page, SSN_SELECTOR, details.ssn)
    _check_stop(stop_event)
    page.wait_for_function(
        """
        () => {
            const selectors = [
                "#fsa_Input_ForgotUsernameFirstName",
                "#fsa_Input_ForgotUsernameLastName",
                "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                "#fsa_Input_ForgotUsernameDateOfBirthDay",
                "#fsa_Input_ForgotUsernameDateOfBirthYear",
                "#fsa_Input_ForgotUsernameSsnInput",
            ];
            const button = [...document.querySelectorAll("button")].find(
                element => (element.innerText || "").trim() === "Continue"
            );
            return selectors.every(selector => {
                const field = document.querySelector(selector);
                return field && String(field.value || "").length > 0;
            }) && button && !button.disabled;
        }
        """,
        timeout=30_000,
    )
    page.wait_for_timeout(200)


def step_3_click_continue(page: Page, stop_event: threading.Event) -> None:
    """使用与 browser-use 一致的真实坐标鼠标点击并确认提交已启动。"""
    _check_stop(stop_event)
    continue_element = page.get_by_role("button", name="Continue", exact=True).first
    continue_element.wait_for(state="visible", timeout=30_000)
    if not continue_element.is_enabled():
        raise RuntimeError("Continue 按钮未启用，请检查字段输入状态")
    box = continue_element.bounding_box()
    if not box:
        raise RuntimeError("无法取得 Continue 按钮可见坐标")
    page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
    try:
        page.wait_for_function(
            """
            () => {
                const text = document.body?.innerText || "";
                return text.includes("Loading...")
                    || text.includes("Account Not Found")
                    || text.includes("Recover my account with a photo ID")
                    || text.includes("Limit Reached: Try Again in 24 Hours")
                    || text.includes("An unknown error has occurred")
                    || location.pathname.endsWith("/username");
            }
            """,
            timeout=5_000,
        )
    except Exception as exc:
        raise PageSubmissionStalled(
            "点击 Continue 后页面没有进入 Loading 或结果状态"
        ) from exc


def _report_stage(
    progress_callback: Callable[[str], None] | None, message: str
) -> None:
    """发送不包含输入资料的阶段消息；UI 回调失败不能中断网页处理。"""
    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def step_4_judge_password_recovery(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
    *,
    timeout_ms: int = 45_000,
    poll_interval_ms: int = 125,
    heartbeat_seconds: float = 5.0,
    stalled_loading_seconds: float = 20.0,
) -> str:
    """等待明确页面结果；持续 Loading 会触发重建会话，而不是无限转圈。"""
    started = time.monotonic()
    deadline = started + max(1, timeout_ms) / 1000
    next_heartbeat = max(0.05, heartbeat_seconds)
    loading_started: float | None = None

    while True:
        _check_stop(stop_event)
        body_text = page.evaluate("document.body?.innerText || ''")
        loading = "Loading..." in body_text
        now = time.monotonic()
        if loading:
            loading_started = loading_started or now
            if now - loading_started >= stalled_loading_seconds:
                raise PageSubmissionStalled(
                    f"点击 Continue 后 Loading 已持续 {int(now - loading_started)} 秒"
                )
        else:
            loading_started = None
            if "An unknown error has occurred" in body_text:
                raise PageSessionExpired(
                    "StudentAid 页面会话已失效，需要重新打开页面"
                )
            if "Something went wrong" in body_text:
                raise RuntimeError("StudentAid 返回页面错误，请稍后重试")
            if LIMIT_REACHED_HEADING in body_text:
                _report_stage(progress_callback, f"已识别结果：{LIMIT_REACHED_HEADING}")
                return "limit_reached"
            if re.search(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                body_text,
                re.IGNORECASE | re.MULTILINE,
            ):
                _report_stage(progress_callback, "已识别结果：Account Not Found")
                return "account_not_found"
            if (
                "Retrieve Your Log-in Information" in body_text
                and "Recover my account with a photo ID" in body_text
            ):
                _report_stage(progress_callback, "已识别结果：Retrieve Your Log-in Information")
                return "can_recover"

        elapsed = now - started
        if now >= deadline:
            raise RuntimeError(
                f"等待 StudentAid 明确结果超时（{timeout_ms / 1000:g} 秒）"
            )
        if elapsed >= next_heartbeat:
            _report_stage(
                progress_callback,
                f"官方页面仍在处理中，已等待 {int(elapsed)} 秒",
            )
            while next_heartbeat <= elapsed:
                next_heartbeat += max(0.05, heartbeat_seconds)
        page.wait_for_timeout(
            min(poll_interval_ms, max(1, int((deadline - now) * 1000)))
        )


def _visible_text(page: Page, text: str) -> str:
    locator = page.get_by_text(text, exact=True).first
    return locator.inner_text().strip() if locator.is_visible() else ""


def collect_recovery_result(page: Page, recovery_status: str) -> RecoveryResult:
    if recovery_status == "limit_reached":
        heading = _visible_text(page, LIMIT_REACHED_HEADING) or LIMIT_REACHED_HEADING
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_not_found":
        locator = page.get_by_text(
            re.compile(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                re.IGNORECASE,
            )
        ).first
        heading = locator.inner_text().strip() if locator.is_visible() else "Account Not Found"
        return RecoveryResult(recovery_status, heading, "", "", "")

    heading = _visible_text(page, "Retrieve Your Log-in Information")
    masked_phone = ""
    masked_email = ""
    for value in (line.strip() for line in page.locator("body").inner_text().splitlines()):
        if not value:
            continue
        if "@" in value and not masked_email:
            masked_email = value
        elif ("*" in value or "•" in value) and re.search(r"\d{4}\s*$", value):
            masked_phone = masked_phone or value
    recovery_method = _visible_text(page, "Recover my account with a photo ID")
    return RecoveryResult(
        recovery_status,
        heading,
        masked_phone,
        masked_email,
        recovery_method,
    )


class BrowserRecoverySession:
    """第十一步 worker 会话；每个线程拥有独立浏览器和缓存目录。"""

    def __init__(self, worker_number: int, backend: str = "playwright") -> None:
        backend = backend.strip().casefold()
        if backend not in BROWSER_BACKENDS:
            raise ValueError(f"浏览器后端必须是：{', '.join(BROWSER_BACKENDS)}")
        self.worker_number = worker_number
        self.backend = backend
        self._playwright: Playwright | None = None
        self._browser: Browser | None = None
        self._browser_use_host: BrowserUseBrowserHost | None = None
        self._context: Any = None
        self._page: Page | None = None
        self._owns_context = False
        self._external_browser = False
        self._owns_browser = False
        self._dedicated_profile = False
        self.browser_mode = "未启动"

    def start(self) -> None:
        if sync_playwright is None:
            raise RuntimeError(
                "未安装 playwright，请双击一键启动 CMD，或执行：pip install -r requirements.txt"
            )
        self._playwright = sync_playwright().start()
        try:
            if self.backend == "browser-use":
                self._browser_use_host = BrowserUseBrowserHost(self.worker_number)
                self._browser_use_host.start()
                try:
                    browser = self._playwright.chromium.connect_over_cdp(
                        self._browser_use_host.cdp_url, timeout=30_000
                    )
                except Exception as exc:
                    raise RuntimeError("Playwright 无法接管 browser-use 独立 CDP 会话") from exc
                launch = BrowserLaunch(
                    browser=browser,
                    external_browser=True,
                    mode=(
                        "browser-use 独立 Chrome "
                        f"(worker {self.worker_number}, AutomationControlled 已关闭)"
                    ),
                    owns_browser=False,
                    dedicated_profile=True,
                )
            else:
                launch = launch_browser(self._playwright)
            self._browser = launch.browser
            self._external_browser = launch.external_browser
            self._owns_browser = launch.owns_browser
            self._dedicated_profile = launch.dedicated_profile
            self.browser_mode = launch.mode
            browser_contexts = list(getattr(self._browser, "contexts", []))
            if self._external_browser:
                if not browser_contexts:
                    raise RuntimeError("Chrome CDP 没有可用的默认浏览器上下文")
                self._context = browser_contexts[0]
            else:
                self._context = self._browser.new_context(
                    viewport={"width": 1440, "height": 900}, locale="en-US"
                )
                self._owns_context = True
            pages = list(getattr(self._context, "pages", []))
            self._page = pages[-1] if pages else self._context.new_page()
            try:
                self._page.set_viewport_size({"width": 1440, "height": 900})
            except Exception:
                pass
            if self._page.evaluate("navigator.webdriver === true"):
                raise RuntimeError(
                    "检测到浏览器处于 Playwright 自动化启动模式；"
                    "该模式会让 StudentAid 在 Continue 后持续 Loading"
                )
            if self._dedicated_profile:
                # 上次进程若被强制关闭，先消除可能残留的 cookie、cache 和 storage。
                self._clear_browser_data_and_blank()
        except Exception:
            self.close()
            raise

    def _close_page(self) -> None:
        if self._page is not None:
            try:
                if not self._page.is_closed():
                    self._page.close()
            except Exception:
                pass
            self._page = None

    def _new_blank_page(self) -> Page:
        if self._context is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        self._close_page()
        self._page = self._context.new_page()
        return self._page

    def _ensure_form_page(
        self,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None,
    ) -> Page:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        form_ready = False
        try:
            form_ready = bool(
                page.locator(FIRST_NAME_SELECTOR).is_visible(timeout=1_000)
                and "Loading..." not in page.locator("body").inner_text(timeout=1_000)
            )
        except Exception:
            form_ready = False
        if not form_ready:
            _report_stage(progress_callback, "正在打开 StudentAid 页面")
            step_1_open_retrieve_account_details(page, stop_event)
        return page

    def _clear_browser_data_and_blank(
        self,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        _report_stage(progress_callback, "正在清除 StudentAid 专用浏览器数据")
        origins = {"https://studentaid.gov"}
        try:
            discovered = page.evaluate(
                """
                () => [...new Set([
                    location.origin,
                    ...performance.getEntriesByType("resource").map(entry => {
                        try { return new URL(entry.name).origin; }
                        catch (_) { return ""; }
                    })
                ])]
                """
            )
            origins.update(
                origin for origin in discovered
                if isinstance(origin, str)
                and re.fullmatch(r"https://(?:[a-z0-9-]+\.)*studentaid\.gov", origin)
            )
        except Exception:
            pass
        try:
            session = self._context.new_cdp_session(page)
            try:
                session.send("Network.clearBrowserCache")
                if self._dedicated_profile:
                    session.send("Network.clearBrowserCookies")
                try:
                    session.send("ServiceWorker.stopAllWorkers")
                except Exception:
                    pass
                for origin in origins:
                    session.send(
                        "Storage.clearDataForOrigin",
                        {"origin": origin, "storageTypes": "all"},
                    )
            finally:
                try:
                    session.detach()
                except Exception:
                    pass
        except Exception:
            pass
        if self._dedicated_profile:
            try:
                self._context.clear_cookies()
            except Exception:
                pass
        self._new_blank_page()
        _report_stage(progress_callback, "浏览器数据已清除，已回到空白页")

    def process(
        self,
        item: WorkItem,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> RecoveryResult:
        if self._browser is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        last_error: BaseException | None = None
        for attempt in range(1, 4):
            try:
                page = self._ensure_form_page(stop_event, progress_callback)
                if attempt > 1:
                    _report_stage(
                        progress_callback,
                        f"已重建页面会话，正在重新填写资料（自动重试 {attempt - 1}/2）",
                    )
                else:
                    _report_stage(progress_callback, "页面已打开，正在填写资料")
                step_2_fill_account_details(page, item.details, stop_event)
                _report_stage(progress_callback, "资料已填写，正在点击 Continue")
                step_3_click_continue(page, stop_event)
                _report_stage(progress_callback, "已提交，正在等待官方结果")
                status = step_4_judge_password_recovery(
                    page, stop_event, progress_callback
                )
                return collect_recovery_result(page, status)
            except StopRequested:
                raise
            except (PageSessionExpired, PageSubmissionStalled) as exc:
                last_error = exc
                if attempt >= 3:
                    break
                _report_stage(
                    progress_callback,
                    "页面会话失效或提交持续转圈，正在自动清理并重建页面",
                )
                self._clear_browser_data_and_blank(progress_callback)
        raise RuntimeError(
            "StudentAid 页面清理并重填后仍未返回明确结果"
        ) from last_error

    def prepare_for_next(
        self,
        result_code: str,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        """必须在结果写入 SQLite/累计 CSV 后调用。"""
        _check_stop(stop_event)
        page = self._page
        if result_code in {"account_not_found", "limit_reached"}:
            self._clear_browser_data_and_blank(progress_callback)
            return
        if result_code != "can_recover":
            self._close_page()
            return
        if page is None or page.is_closed():
            raise RuntimeError("记录账户找回结果后页面已关闭，无法点击 Cancel")
        _report_stage(progress_callback, "结果已保存，正在点击 Cancel")
        cancel = page.locator("#fsa_Button_ForgotUsernameCancel").first
        try:
            cancel.wait_for(state="visible", timeout=5_000)
        except Exception:
            cancel = page.get_by_role("button", name="Cancel", exact=True).first
            cancel.wait_for(state="visible", timeout=10_000)
        box = cancel.bounding_box()
        if not box:
            raise RuntimeError("无法取得 Cancel 按钮可见坐标")
        page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
        page.locator(FIRST_NAME_SELECTOR).wait_for(state="visible", timeout=30_000)
        page.wait_for_function(
            """
            () => {
                const ids = [
                    "#fsa_Input_ForgotUsernameFirstName",
                    "#fsa_Input_ForgotUsernameLastName",
                    "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                    "#fsa_Input_ForgotUsernameDateOfBirthDay",
                    "#fsa_Input_ForgotUsernameDateOfBirthYear",
                    "#fsa_Input_ForgotUsernameSsnInput",
                ];
                return ids.every(id => !(document.querySelector(id)?.value || ""));
            }
            """,
            timeout=15_000,
        )
        _report_stage(progress_callback, "Cancel 已完成，空白表单已就绪")

    def recover_after_cleanup_error(self) -> None:
        self._close_page()

    def close(self) -> None:
        if self._context is not None and self._browser is not None:
            try:
                self._clear_browser_data_and_blank()
            except Exception:
                pass
        if self.backend == "browser-use":
            # 保留最后一个 about:blank，避免 browser-use 看门狗在结束前自动补建标签页。
            self._page = None
        else:
            self._close_page()
        if self._context is not None and self._owns_context:
            try:
                self._context.close()
            except Exception:
                pass
        self._context = None
        self._owns_context = False
        if self._browser is not None and self._owns_browser:
            try:
                self._browser.close()
            except Exception:
                pass
        self._browser = None
        self._external_browser = False
        self._owns_browser = False
        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None
        if self._browser_use_host is not None:
            try:
                self._browser_use_host.close()
            finally:
                self._browser_use_host = None
        self._dedicated_profile = False


@dataclass
class _DbCommand:
    action: str
    payload: dict[str, Any]
    response: queue.Queue[tuple[bool, Any]]


class DatabaseWriter:
    """SQLite 的唯一写入者；工作线程不得直接执行写 SQL。"""

    def __init__(self, database_path: Path) -> None:
        self.database_path = database_path.resolve()
        self._commands: queue.Queue[_DbCommand] = queue.Queue()
        self._ready = threading.Event()
        self._thread = threading.Thread(
            target=self._run,
            name="studentaid-db-writer",
            daemon=False,
        )
        self._startup_error: BaseException | None = None

    def start(self) -> None:
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._thread.start()
        if not self._ready.wait(30):
            raise TimeoutError("SQLite 写入线程启动超时")
        if self._startup_error is not None:
            raise RuntimeError("SQLite 初始化失败") from self._startup_error

    def request(self, action: str, **payload: Any) -> Any:
        if not self._thread.is_alive():
            raise RuntimeError("SQLite 写入线程未运行")
        response: queue.Queue[tuple[bool, Any]] = queue.Queue(maxsize=1)
        self._commands.put(_DbCommand(action, payload, response))
        try:
            ok, value = response.get(timeout=300)
        except queue.Empty as exc:
            raise TimeoutError(f"SQLite 写入命令超时：{action}") from exc
        if ok:
            return value
        raise RuntimeError(f"SQLite 写入失败：{action}") from value

    def close(self) -> None:
        if self._thread.is_alive():
            try:
                self.request("shutdown")
            finally:
                self._thread.join(timeout=30)

    def _run(self) -> None:
        connection: sqlite3.Connection | None = None
        try:
            connection = sqlite3.connect(self.database_path, timeout=30)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA busy_timeout=30000")
            journal_mode = connection.execute("PRAGMA journal_mode=WAL").fetchone()[0]
            if str(journal_mode).casefold() != "wal":
                raise RuntimeError(f"无法启用 SQLite WAL，当前模式：{journal_mode}")
            connection.execute("PRAGMA synchronous=NORMAL")
            connection.execute("PRAGMA wal_autocheckpoint=1000")
            self._create_schema(connection)
            connection.commit()
        except BaseException as exc:
            self._startup_error = exc
            self._ready.set()
            if connection is not None:
                connection.close()
            return

        self._ready.set()
        assert connection is not None
        try:
            while True:
                command = self._commands.get()
                try:
                    value = self._handle(connection, command.action, command.payload)
                    connection.commit()
                    command.response.put((True, value))
                    if command.action == "shutdown":
                        break
                except BaseException as exc:
                    connection.rollback()
                    command.response.put((False, exc))
                finally:
                    self._commands.task_done()
        finally:
            connection.close()

    @staticmethod
    def _create_schema(connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS batches (
                batch_id TEXT PRIMARY KEY,
                input_path TEXT NOT NULL,
                output_directory TEXT NOT NULL,
                thread_count INTEGER NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                finished_at TEXT,
                export_path TEXT,
                total_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id TEXT NOT NULL REFERENCES batches(batch_id),
                source_file TEXT NOT NULL,
                source_sheet TEXT NOT NULL DEFAULT '',
                source_row INTEGER NOT NULL,
                original_fields_json TEXT NOT NULL,
                ssn TEXT,
                birth_month TEXT,
                birth_day TEXT,
                birth_year TEXT,
                first_name TEXT,
                last_name TEXT,
                address TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL,
                attempt_count INTEGER NOT NULL DEFAULT 0,
                result_code TEXT NOT NULL DEFAULT '',
                result_heading TEXT NOT NULL DEFAULT '',
                masked_phone TEXT NOT NULL DEFAULT '',
                masked_email TEXT NOT NULL DEFAULT '',
                recovery_method TEXT NOT NULL DEFAULT '',
                error TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_records_batch_status
                ON records(batch_id, status, id);
            """
        )

    @staticmethod
    def _handle(
        connection: sqlite3.Connection, action: str, payload: Mapping[str, Any]
    ) -> Any:
        if action == "create_batch":
            connection.execute(
                """
                INSERT INTO batches(
                    batch_id, input_path, output_directory, thread_count,
                    status, created_at
                ) VALUES (?, ?, ?, ?, 'importing', ?)
                """,
                (
                    payload["batch_id"], payload["input_path"],
                    payload["output_directory"], payload["thread_count"], _now_iso(),
                ),
            )
            return None

        if action == "insert_records":
            batch_id = str(payload["batch_id"])
            work_items: list[WorkItem] = []
            records: Sequence[ImportedRecord] = payload["records"]
            for record in records:
                details = record.details
                status = "failed" if record.import_error else "pending"
                cursor = connection.execute(
                    """
                    INSERT INTO records(
                        batch_id, source_file, source_sheet, source_row,
                        original_fields_json, ssn, birth_month, birth_day,
                        birth_year, first_name, last_name, address, status,
                        error, created_at, finished_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id, record.source_file, record.source_sheet,
                        record.source_row,
                        json.dumps(record.original_fields, ensure_ascii=False),
                        details.ssn if details else None,
                        details.birth_month if details else None,
                        details.birth_day if details else None,
                        details.birth_year if details else None,
                        details.first_name if details else None,
                        details.last_name if details else None,
                        record.address, status, record.import_error, _now_iso(),
                        _now_iso() if record.import_error else None,
                    ),
                )
                if details is not None and not record.import_error:
                    work_items.append(
                        WorkItem(
                            int(cursor.lastrowid), details, record.source_file,
                            record.source_sheet, record.source_row,
                            record.original_fields, record.address,
                        )
                    )
            connection.execute(
                """
                UPDATE batches SET status='running', total_count=? WHERE batch_id=?
                """,
                (len(records), batch_id),
            )
            return work_items

        if action == "mark_processing":
            connection.execute(
                """
                UPDATE records
                SET status='processing', started_at=?, attempt_count=attempt_count+1,
                    error=''
                WHERE id=? AND status='pending'
                """,
                (_now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_completed":
            result: RecoveryResult = payload["result"]
            connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE id=?
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["record_id"],
                ),
            )
            return None

        if action in {"mark_failed", "mark_stopped"}:
            status = "failed" if action == "mark_failed" else "stopped"
            connection.execute(
                "UPDATE records SET status=?, error=?, finished_at=? WHERE id=?",
                (status, payload.get("error", ""), _now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_remaining":
            status = str(payload["status"])
            connection.execute(
                """
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND status IN ('pending', 'processing')
                """,
                (
                    status, payload.get("error", ""), _now_iso(), payload["batch_id"],
                ),
            )
            return None

        if action == "status_counts":
            rows = connection.execute(
                """
                SELECT status, COUNT(*) AS count FROM records
                WHERE batch_id=? GROUP BY status
                """,
                (payload["batch_id"],),
            ).fetchall()
            return {str(row["status"]): int(row["count"]) for row in rows}

        if action == "finish_batch":
            connection.execute(
                """
                UPDATE batches SET status=?, finished_at=?, export_path=?
                WHERE batch_id=?
                """,
                (
                    payload["status"], _now_iso(), payload.get("export_path", ""),
                    payload["batch_id"],
                ),
            )
            return None

        if action == "checkpoint":
            return connection.execute("PRAGMA wal_checkpoint(PASSIVE)").fetchone()

        if action == "shutdown":
            return None

        raise ValueError(f"未知 SQLite 写入命令：{action}")


def export_batch_csv(
    database_path: Path, batch_id: str, output_directory: Path
) -> Path:
    """只读查询当前批次并原子生成带 BOM 的最终 CSV。"""
    output_directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_directory / f"studentaid_results_{timestamp}_{batch_id[:8]}.csv"
    temporary_path = output_path.with_name(f".{output_path.name}.{os.getpid()}.tmp")

    connection = sqlite3.connect(database_path, timeout=30)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            "SELECT * FROM records WHERE batch_id=? ORDER BY id",
            (batch_id,),
        ).fetchall()
    finally:
        connection.close()

    original_rows: list[list[str]] = []
    maximum_columns = 0
    for row in rows:
        fields = [str(value) for value in json.loads(row["original_fields_json"])]
        original_rows.append(fields)
        maximum_columns = max(maximum_columns, len(fields))

    try:
        with temporary_path.open("w", encoding="utf-8-sig", newline="") as output_file:
            writer = csv.writer(output_file, lineterminator="\n")
            writer.writerow(
                [*(f"Input_Column_{index}" for index in range(1, maximum_columns + 1)),
                 *RESULT_COLUMNS]
            )
            for row, original_fields in zip(rows, original_rows):
                padded = [*original_fields, *("" for _ in range(maximum_columns - len(original_fields)))]
                writer.writerow(
                    [
                        *padded,
                        row["result_heading"], row["masked_phone"],
                        row["masked_email"], row["recovery_method"], row["status"],
                        row["error"], Path(row["source_file"]).name,
                        row["source_sheet"], row["source_row"],
                    ]
                )
            output_file.flush()
            os.fsync(output_file.fileno())
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path


EventCallback = Callable[[str, dict[str, Any]], None]
SessionFactory = Callable[..., Any]


class BatchEngine:
    """第十一步并发处理：每个 worker 独立浏览器，写库由专用线程串行化。"""

    def __init__(
        self,
        event_callback: EventCallback | None = None,
        session_factory: SessionFactory = BrowserRecoverySession,
    ) -> None:
        self._callback = event_callback or (lambda _kind, _payload: None)
        self._session_factory = session_factory
        self._stop_event = threading.Event()
        self._state_lock = threading.Lock()
        self._running = False
        self._coordinator: threading.Thread | None = None

    @property
    def is_running(self) -> bool:
        with self._state_lock:
            return self._running

    def start(
        self,
        input_path: Path,
        output_target: Path,
        thread_count: int = 2,
        backend: str = "playwright",
    ) -> None:
        input_path = input_path.resolve()
        output_directory, output_path = resolve_output_target(output_target)
        backend = backend.strip().casefold()
        if not input_path.is_file():
            raise FileNotFoundError(f"输入文件不存在：{input_path}")
        if input_path == output_path:
            raise ValueError("输入文件和累计输出文件不能是同一个文件")
        if not 1 <= thread_count <= 8:
            raise ValueError("处理线程数必须是 1-8")
        if backend not in BROWSER_BACKENDS:
            raise ValueError(f"浏览器后端必须是：{', '.join(BROWSER_BACKENDS)}")
        output_directory.mkdir(parents=True, exist_ok=True)
        with self._state_lock:
            if self._running:
                raise RuntimeError("已有批次正在运行")
            self._running = True
        self._stop_event.clear()
        self._coordinator = threading.Thread(
            target=self._run,
            args=(input_path, output_path, thread_count, backend),
            name="studentaid-coordinator",
            daemon=False,
        )
        self._coordinator.start()

    def stop(self) -> None:
        if self.is_running:
            self._stop_event.set()
            self._emit(
                "log",
                message="已收到停止请求；正在停止领取新任务、清除缓存并结束浏览器进程。",
            )

    def wait(self, timeout: float | None = None) -> bool:
        coordinator = self._coordinator
        if coordinator is not None:
            coordinator.join(timeout)
        return not self.is_running

    def _emit(self, kind: str, **payload: Any) -> None:
        try:
            self._callback(kind, payload)
        except Exception:
            pass

    def _publish_progress(
        self, writer: DatabaseWriter, batch_id: str, total: int
    ) -> dict[str, int]:
        counts = writer.request("status_counts", batch_id=batch_id)
        payload = {
            "total": total,
            "pending": int(counts.get("pending", 0)),
            "processing": int(counts.get("processing", 0)),
            "completed": int(counts.get("completed", 0)),
            "failed": int(counts.get("failed", 0)),
            "stopped": int(counts.get("stopped", 0)),
        }
        self._emit("progress", **payload)
        return payload

    def _new_session(self, worker_number: int, backend: str) -> Any:
        if self._session_factory is BrowserRecoverySession:
            return self._session_factory(worker_number, backend)
        return self._session_factory(worker_number)

    def _worker_loop(
        self,
        worker_number: int,
        backend: str,
        tasks: queue.Queue[WorkItem | None],
        writer: DatabaseWriter,
        batch_id: str,
        total: int,
        input_path: Path,
        output_path: Path,
    ) -> None:
        session = self._new_session(worker_number, backend)
        try:
            session.start()
            self._emit(
                "log",
                message=(
                    f"处理线程 {worker_number} 已启动："
                    f"{getattr(session, 'browser_mode', '模式未知')}。"
                ),
            )
        except Exception as exc:
            self._emit(
                "log",
                message=f"处理线程 {worker_number} 启动失败：{_clean_error(exc)}",
            )
            try:
                session.close()
            except Exception:
                pass
            return

        try:
            while not self._stop_event.is_set():
                try:
                    item = tasks.get(timeout=0.2)
                except queue.Empty:
                    continue
                if item is None:
                    tasks.task_done()
                    break
                if self._stop_event.is_set():
                    tasks.task_done()
                    break

                writer.request("mark_processing", record_id=item.record_id)
                self._publish_progress(writer, batch_id, total)
                progress = lambda stage, record_id=item.record_id: self._emit(
                    "log", message=f"记录 #{record_id}：{stage}"
                )
                try:
                    result = session.process(item, self._stop_event, progress)
                    writer.request(
                        "mark_completed", record_id=item.record_id, result=result
                    )
                    self._emit(
                        "log",
                        message=f"记录 #{item.record_id}：明确结果已实时写入 SQLite。",
                    )
                    appended = append_cumulative_result(output_path, item, result)
                    if appended:
                        self._emit(
                            "log",
                            message=f"记录 #{item.record_id}：已实时追加到累计输出 CSV。",
                        )
                    else:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：累计输出第一列已存在，"
                                "跳过重复追加。"
                            ),
                        )
                    deleted = remove_input_rows_by_keys(input_path, {item.record_key})
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：明确结果已落盘，"
                            f"输入文件已删除 {deleted} 条匹配资料。"
                        ),
                    )
                    try:
                        prepare = getattr(session, "prepare_for_next", None)
                        if callable(prepare):
                            prepare(result.result_code, self._stop_event, progress)
                    except StopRequested:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "收到停止请求，跳过下一页准备。"
                            ),
                        )
                    except Exception as cleanup_exc:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "页面清理失败，将为下一条重建页面："
                                f"{_clean_error(cleanup_exc)}"
                            ),
                        )
                        recover = getattr(session, "recover_after_cleanup_error", None)
                        if callable(recover):
                            recover()
                except StopRequested as exc:
                    writer.request(
                        "mark_stopped", record_id=item.record_id, error=_clean_error(exc)
                    )
                except Exception as exc:
                    writer.request(
                        "mark_failed", record_id=item.record_id, error=_clean_error(exc)
                    )
                    self._emit(
                        "log",
                        message=f"记录 #{item.record_id} 处理失败：{_clean_error(exc)}",
                    )
                    recover = getattr(session, "recover_after_cleanup_error", None)
                    if callable(recover):
                        try:
                            recover()
                        except Exception:
                            pass
                finally:
                    tasks.task_done()
                    self._publish_progress(writer, batch_id, total)
        finally:
            try:
                session.close()
                self._emit(
                    "log",
                    message=(
                        f"处理线程 {worker_number} 已清除浏览器缓存并结束浏览器进程。"
                    ),
                )
            except Exception as exc:
                self._emit(
                    "log",
                    message=f"处理线程 {worker_number} 关闭浏览器失败：{_clean_error(exc)}",
                )

    def _run(
        self,
        input_path: Path,
        output_path: Path,
        thread_count: int,
        backend: str,
    ) -> None:
        batch_id = uuid.uuid4().hex
        output_directory = output_path.parent
        database_path = output_directory / DATABASE_FILENAME
        writer: DatabaseWriter | None = None
        total = 0
        try:
            self._emit("log", message="正在读取累计输出第一列并同步输入文件……")
            existing_keys = read_output_first_column_keys(output_path)
            removed_before_start = remove_input_rows_by_keys(input_path, existing_keys)
            if removed_before_start:
                self._emit(
                    "log",
                    message=f"输入文件已删除 {removed_before_start} 条输出中已存在的资料。",
                )
            try:
                records = load_input_records(input_path)
            except ValueError as exc:
                if "没有可导入的资料行" not in str(exc):
                    raise
                records = []
            total = len(records)
            if self._stop_event.is_set():
                raise StopRequested("导入阶段已停止")

            writer = DatabaseWriter(database_path)
            writer.start()
            writer.request(
                "create_batch",
                batch_id=batch_id,
                input_path=str(input_path),
                output_directory=str(output_directory),
                thread_count=thread_count,
            )
            work_items: list[WorkItem] = writer.request(
                "insert_records", batch_id=batch_id, records=records
            )
            invalid_count = len(records) - len(work_items)
            self._emit(
                "log",
                message=(
                    f"输入剩余 {len(records)} 条；可处理 {len(work_items)} 条，"
                    f"格式错误 {invalid_count} 条。后端 {backend}，"
                    f"处理线程 {thread_count}。"
                ),
            )
            self._publish_progress(writer, batch_id, total)

            tasks: queue.Queue[WorkItem | None] = queue.Queue()
            for item in work_items:
                tasks.put(item)
            worker_count = min(thread_count, len(work_items))
            for _ in range(worker_count):
                tasks.put(None)
            workers = [
                threading.Thread(
                    target=self._worker_loop,
                    args=(
                        number,
                        backend,
                        tasks,
                        writer,
                        batch_id,
                        total,
                        input_path,
                        output_path,
                    ),
                    name=f"studentaid-worker-{number}",
                    daemon=False,
                )
                for number in range(1, worker_count + 1)
            ]
            for worker in workers:
                worker.start()
            for worker in workers:
                worker.join()

            if self._stop_event.is_set():
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="stopped",
                    error="用户停止，任务尚未处理",
                )
                batch_status = "stopped"
            else:
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="failed",
                    error="任务未被浏览器处理",
                )
                batch_status = "completed"
            final_counts = self._publish_progress(writer, batch_id, total)
            writer.request(
                "finish_batch", batch_id=batch_id, status=batch_status,
                export_path=str(output_path),
            )
            writer.request("checkpoint")
            self._emit(
                "finished",
                batch_id=batch_id,
                database_path=str(database_path),
                export_path=str(output_path),
                status=batch_status,
                counts=final_counts,
            )
        except StopRequested as exc:
            self._emit("fatal", message=_clean_error(exc), stopped=True)
        except Exception as exc:
            self._emit(
                "fatal",
                message=_clean_error(exc),
                details=traceback.format_exc(limit=8),
                stopped=False,
            )
        finally:
            if writer is not None:
                try:
                    writer.close()
                except Exception as exc:
                    self._emit("log", message=f"关闭数据库时出错：{_clean_error(exc)}")
            with self._state_lock:
                self._running = False
            self._emit("idle", export_path=str(output_path))


class StudentAidApp:
    """第十一步稳定版桌面 GUI。"""

    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("920x650")
        self.root.minsize(820, 570)

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar(value=str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME))
        self.backend_var = tk.StringVar(value="browser-use")
        self.thread_var = tk.StringVar(value="2")
        self.status_var = tk.StringVar(value="就绪")
        self.progress_text_var = tk.StringVar(value="总数 0 | 完成 0 | 失败 0 | 停止 0")
        self._ui_events: queue.Queue[tuple[str, dict[str, Any]]] = queue.Queue()
        self.engine = BatchEngine(self._backend_event)
        self._closing = False

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._poll_events)

    def _build_ui(self) -> None:
        ttk = self.ttk
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(7, weight=1)

        ttk.Label(main, text="导入文件：").grid(row=0, column=0, sticky="w", pady=6)
        self.input_entry = ttk.Entry(main, textvariable=self.input_var)
        self.input_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        self.input_button = ttk.Button(main, text="选择文件", command=self._choose_input)
        self.input_button.grid(row=0, column=2, pady=6)

        ttk.Label(main, text="累计输出 CSV：").grid(row=1, column=0, sticky="w", pady=6)
        self.output_entry = ttk.Entry(main, textvariable=self.output_var)
        self.output_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        self.output_button = ttk.Button(main, text="选择文件", command=self._choose_output)
        self.output_button.grid(row=1, column=2, pady=6)

        ttk.Label(main, text="浏览器后端：").grid(row=2, column=0, sticky="w", pady=6)
        self.backend_combo = ttk.Combobox(
            main,
            textvariable=self.backend_var,
            values=BROWSER_BACKENDS,
            width=18,
            state="readonly",
        )
        self.backend_combo.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="browser-use 和 playwright 都使用每线程独立浏览器。")\
            .grid(row=2, column=1, sticky="w", padx=(180, 0), pady=6)

        ttk.Label(main, text="处理线程数：").grid(row=3, column=0, sticky="w", pady=6)
        self.thread_spin = ttk.Spinbox(
            main, from_=1, to=8, textvariable=self.thread_var, width=10
        )
        self.thread_spin.grid(row=3, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="默认 2 线程；每个线程独立缓存、页面和浏览器进程。")\
            .grid(row=3, column=1, sticky="w", padx=(100, 0), pady=6)

        button_bar = ttk.Frame(main)
        button_bar.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(10, 8))
        self.start_button = ttk.Button(button_bar, text="开始", command=self._start)
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(
            button_bar, text="停止", command=self._stop, state="disabled"
        )
        self.stop_button.pack(side="left", padx=8)
        ttk.Label(button_bar, textvariable=self.status_var).pack(side="right")

        self.progress = ttk.Progressbar(main, mode="determinate", maximum=1, value=0)
        self.progress.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(2, 4))
        ttk.Label(main, textvariable=self.progress_text_var).grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        log_frame = ttk.LabelFrame(main, text="运行日志（不显示 SSN 和原始资料）", padding=8)
        log_frame.grid(row=7, column=0, columnspan=3, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.log_text = self.tk.Text(log_frame, wrap="word", state="disabled")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

    def _choose_input(self) -> None:
        path = self.filedialog.askopenfilename(
            title="选择导入文件",
            filetypes=[
                ("支持的资料文件", "*.csv *.scv *.txt *.xlsx"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
                ("纯文本文件", "*.txt *.scv"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            self.input_var.set(path)
            current = self.output_var.get().strip()
            if not current or current == str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME):
                self.output_var.set(str(Path(path).parent / CUMULATIVE_OUTPUT_FILENAME))

    def _choose_output(self) -> None:
        current = Path(self.output_var.get().strip() or CUMULATIVE_OUTPUT_FILENAME)
        path = self.filedialog.asksaveasfilename(
            title="选择累计输出 CSV（已有文件会继续叠加）",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if path:
            self.output_var.set(path)

    def _set_running(self, running: bool) -> None:
        normal_or_disabled = "disabled" if running else "normal"
        self.input_entry.configure(state=normal_or_disabled)
        self.output_entry.configure(state=normal_or_disabled)
        self.backend_combo.configure(state="disabled" if running else "readonly")
        self.thread_spin.configure(state=normal_or_disabled)
        self.input_button.configure(state=normal_or_disabled)
        self.output_button.configure(state=normal_or_disabled)
        self.start_button.configure(state=normal_or_disabled)
        self.stop_button.configure(state="normal" if running else "disabled")

    def _start(self) -> None:
        try:
            input_path = Path(self.input_var.get().strip())
            output_target = Path(self.output_var.get().strip())
            thread_count = int(self.thread_var.get().strip())
            backend = self.backend_var.get().strip()
            self.engine.start(input_path, output_target, thread_count, backend)
        except Exception as exc:
            self.messagebox.showerror(APP_TITLE, _clean_error(exc))
            return
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self.status_var.set("运行中")
        self._set_running(True)
        self._append_log(
            f"批次已启动：后端 {self.backend_var.get()}，线程 {self.thread_var.get()}。"
        )

    def _stop(self) -> None:
        self.stop_button.configure(state="disabled")
        self.status_var.set("停止中")
        self.engine.stop()

    def _backend_event(self, kind: str, payload: dict[str, Any]) -> None:
        self._ui_events.put((kind, payload))

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _poll_events(self) -> None:
        try:
            while True:
                kind, payload = self._ui_events.get_nowait()
                if kind == "log":
                    self._append_log(str(payload.get("message", "")))
                elif kind == "progress":
                    total = int(payload.get("total", 0))
                    completed = int(payload.get("completed", 0))
                    failed = int(payload.get("failed", 0))
                    stopped = int(payload.get("stopped", 0))
                    processing = int(payload.get("processing", 0))
                    terminal = completed + failed + stopped
                    self.progress.configure(maximum=max(1, total), value=terminal)
                    self.progress_text_var.set(
                        f"总数 {total} | 完成 {completed} | 失败 {failed} | "
                        f"停止 {stopped} | 处理中 {processing}"
                    )
                elif kind == "finished":
                    output_path = str(payload.get("export_path", ""))
                    database_path = str(payload.get("database_path", ""))
                    counts = payload.get("counts", {})
                    failed = int(counts.get("failed", 0)) if isinstance(counts, dict) else 0
                    if payload.get("status") == "stopped":
                        display_status = "已停止"
                    elif failed:
                        display_status = "完成（有失败）"
                    else:
                        display_status = "已完成"
                    self.status_var.set(display_status)
                    self._append_log(f"累计 CSV：{output_path}")
                    self._append_log(f"SQLite 数据库：{database_path}")
                    self.messagebox.showinfo(
                        APP_TITLE,
                        f"批次处理结束。\n\n累计 CSV：{output_path}\n数据库：{database_path}",
                    )
                elif kind == "fatal":
                    message = str(payload.get("message", "未知错误"))
                    self.status_var.set("已停止" if payload.get("stopped") else "失败")
                    self._append_log(message)
                    if not payload.get("stopped"):
                        self.messagebox.showerror(APP_TITLE, message)
                elif kind == "idle":
                    self._set_running(False)
                self._ui_events.task_done()
        except queue.Empty:
            pass

        if self._closing and not self.engine.is_running:
            self.root.destroy()
            return
        self.root.after(100, self._poll_events)

    def _on_close(self) -> None:
        if self.engine.is_running:
            if not self.messagebox.askyesno(APP_TITLE, "任务仍在运行，是否停止并退出？"):
                return
            self._closing = True
            self._set_running(True)
            self.stop_button.configure(state="disabled")
            self.status_var.set("停止中")
            self.engine.stop()
        else:
            self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    StudentAidApp().run()


if __name__ == "__main__":
    main()
