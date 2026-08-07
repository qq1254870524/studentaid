"""
StudentAid 浏览器自动化工具 v4。

2026-08-07 第九步源码校验更新：
1. 使用 browser-use 独立核验 StudentAid 真实页面字段、Continue 按钮和提交状态时间线。
2. 删除“姓名输入框消失即判定可找回”的宽泛条件；只有明确结果标志才允许完成，
   防止页面跳转或短暂重绘被误判为成功。
3. 页面结果等待改为可停止的短轮询；保留 Loading 等待，并准确识别站点错误。
4. CDP 模式复用 Chrome 默认浏览器上下文并新开页面，保持与 browser-use 实测一致；
   输入改用真实键盘事件、字段失焦和提交前稳定等待，修复过快 fill 导致站点返回 unknown error。
5. 检测到因页面停留过久产生的 unknown error 时自动新开页面、重新填表并提交一次。
6. 每条记录实时输出“打开页面、填写资料、已提交、等待结果、已识别结果”等安全阶段日志；
   长时间等待时每 5 秒输出心跳，最终状态实时写 SQLite 并刷新 GUI，不记录原始资料。
7. Playwright 仍是正式运行根基，browser-use 仅用于独立校验。

2026-08-07 第八步实测更新：
1. 保留第七步的 GUI、SQLite WAL、并发处理和 CSV 导出能力。
2. 新增 Chrome CDP 连接模式：优先复用已通过站点网络校验的本机 Chrome，
   解决新启动 Playwright 浏览器访问 StudentAid 时的 HTTP/2 协议错误。
3. 支持 STUDENTAID_CDP_URL 显式指定 CDP 地址；未设置时自动探测
   http://127.0.0.1:9223，设置为 off/none/disabled 可关闭自动探测。
4. 新增页面导航重试、共享浏览器安全断开和浏览器模式日志。
5. 修复输入页标题被误判为找回成功的问题；现在会等待 Loading 结束，
   并把站点未知错误准确记录为失败。
6. 修复纯文本 CSV 中全数字 SSN 被 float 转换吞掉前导 0 的问题。
7. 第八步使用测试资料完成 GUI、浏览器、SQLite 和导出链路实测。

2026-08-07 第七步更新（原 ait2.py 日期记录受本机时钟影响）：
1. 新增 Tkinter GUI，可选择 CSV/SCV/TXT/XLSX 输入文件和结果目录。
2. 导入资料先进入 SQLite；数据库固定启用 WAL。
3. 使用多个独立 Playwright 处理线程和一个专用 SQLite 写入线程。
4. 支持开始、停止、线程数设置和实时进度；停止后不再领取新任务。
5. 每个批次完成或停止后自动导出 UTF-8 BOM CSV。
6. 兼容无表头的 ``SSN,月,日,年,姓,名,地址``、旧版
   ``SSN,DOB,First Name,Last Name,Address``，以及带常见英文表头的资料。

旧版 ait.py 保持不变。本文件不把 SSN 或整条原始资料打印到日志。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import queue
import re
import sqlite3
import sys
import threading
import time
import traceback
import urllib.request
from typing import Any, Callable, Iterable, Mapping, Sequence
import uuid

sys.dont_write_bytecode = True

try:
    from playwright.sync_api import Browser, Page, Playwright, sync_playwright
except ImportError:  # GUI 仍可启动，并在开始处理时给出明确错误。
    Browser = Page = Playwright = Any  # type: ignore[assignment,misc]
    sync_playwright = None


APP_TITLE = "StudentAid 批量处理工具 - 第九步源码校验"
DATABASE_FILENAME = "studentaid.sqlite3"
RETRIEVE_ACCOUNT_DETAILS_URL = (
    "https://studentaid.gov/fsa-id/sign-in/retrieve-account-details"
)

FIRST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameFirstName"
LAST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameLastName"
BIRTH_MONTH_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthMonth"
BIRTH_DAY_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthDay"
BIRTH_YEAR_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthYear"
SSN_SELECTOR = "#fsa_Input_ForgotUsernameSsnInput"

RESULT_COLUMNS = (
    "Result_Heading",
    "Masked_Phone",
    "Masked_Email",
    "Recovery_Method",
    "Processing_Status",
    "Error",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
)


@dataclass(frozen=True)
class AccountDetails:
    """StudentAid 账户找回页面所需字段。"""

    first_name: str
    last_name: str
    birth_month: str
    birth_day: str
    birth_year: str
    ssn: str
    original_ssn: str


@dataclass(frozen=True)
class RecoveryResult:
    """页面返回的可公开保存结果；联系方式保持页面原有脱敏形式。"""

    result_code: str
    heading: str
    masked_phone: str
    masked_email: str
    recovery_method: str


@dataclass(frozen=True)
class ImportedRecord:
    """一条已读取的输入资料；校验失败的行也会进入数据库。"""

    source_file: str
    source_sheet: str
    source_row: int
    original_fields: tuple[str, ...]
    details: AccountDetails | None
    address: str
    import_error: str = ""


@dataclass(frozen=True)
class WorkItem:
    """数据库记录和浏览器处理线程之间传递的最小任务。"""

    record_id: int
    details: AccountDetails


class StopRequested(RuntimeError):
    """在网页处理步骤之间收到停止请求。"""


class PageSessionExpired(RuntimeError):
    """StudentAid 页面停留过久后返回 unknown error，需要重新打开页面。"""


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _clean_error(exc: BaseException | str, limit: int = 1000) -> str:
    """生成适合写库和显示的单行错误，不包含输入资料。"""
    text = str(exc).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit] or type(exc).__name__


def _normalise_month(value: str) -> str:
    aliases = {
        "january": "01", "jan": "01", "february": "02", "feb": "02",
        "march": "03", "mar": "03", "april": "04", "apr": "04",
        "may": "05", "june": "06", "jun": "06", "july": "07",
        "jul": "07", "august": "08", "aug": "08", "september": "09",
        "sep": "09", "sept": "09", "october": "10", "oct": "10",
        "november": "11", "nov": "11", "december": "12", "dec": "12",
    }
    cleaned = value.strip().lower()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    month = aliases.get(cleaned) or (cleaned.zfill(2) if cleaned.isdigit() else "")
    if month not in {f"{number:02d}" for number in range(1, 13)}:
        raise ValueError("出生月份必须是 1-12、01-12 或英文月份名称")
    return month


def _normalise_day(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if not digits or not 1 <= int(digits) <= 31:
        raise ValueError("出生日期（日）必须是 1-31")
    return digits.zfill(2)


def _normalise_year(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) != 4 or not 1900 <= int(digits) <= date.today().year:
        raise ValueError("出生年份必须是四位数字，且不晚于当前年份")
    return digits


def _normalise_ssn(value: str) -> str:
    cleaned = value.strip()
    # 只移除 Excel 数字单元格常见的 .0 后缀；纯数字文本必须保留前导 0。
    if re.fullmatch(r"\d+\.0+", cleaned):
        cleaned = cleaned.split(".", 1)[0]
    digits = re.sub(r"\D", "", cleaned)
    # Excel 把首位 0 当数字格式丢掉时，允许 8 位数字恢复为 9 位。
    if len(digits) == 8:
        digits = digits.zfill(9)
    if len(digits) != 9:
        raise ValueError("Social Security Number 必须包含 9 位数字")
    return digits


def _validate_details(
    ssn_value: str,
    month_value: str,
    day_value: str,
    year_value: str,
    first_name: str,
    last_name: str,
) -> AccountDetails:
    ssn = _normalise_ssn(ssn_value)
    month = _normalise_month(month_value)
    day = _normalise_day(day_value)
    year = _normalise_year(year_value)
    first_name = first_name.strip()
    last_name = last_name.strip()

    if not first_name or len(first_name) > 35:
        raise ValueError("First Name 不能为空且长度不能超过 35 个字符")
    if not last_name or len(last_name) > 35:
        raise ValueError("Last Name 不能为空且长度不能超过 35 个字符")
    try:
        date(int(year), int(month), int(day))
    except ValueError as exc:
        raise ValueError("出生日期不是有效日期") from exc

    return AccountDetails(
        first_name=first_name,
        last_name=last_name,
        birth_month=month,
        birth_day=day,
        birth_year=year,
        ssn=ssn,
        original_ssn=ssn_value,
    )


def _parse_dob(value: str) -> tuple[str, str, str]:
    value = value.strip()
    for date_format in (
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%y", "%m-%d-%y",
    ):
        try:
            parsed = datetime.strptime(value, date_format).date()
            return f"{parsed.month:02d}", f"{parsed.day:02d}", f"{parsed.year:04d}"
        except ValueError:
            continue
    raise ValueError("DOB 格式无效，应为 MM/DD/YYYY、MM-DD-YYYY 或 YYYY-MM-DD")


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalise_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.strip().casefold())


HEADER_ALIASES: Mapping[str, set[str]] = {
    "ssn": {"ssn", "socialsecuritynumber", "socialsecurityno", "socialsecurity"},
    "dob": {"dob", "dateofbirth", "birthdate"},
    "month": {"birthmonth", "dobmonth", "month", "mm"},
    "day": {"birthday", "dobday", "day", "dd"},
    "year": {"birthyear", "dobyear", "year", "yyyy"},
    "first_name": {"firstname", "givenname", "fname", "first"},
    "last_name": {"lastname", "surname", "familyname", "lname", "last"},
    "address": {"address", "streetaddress", "mailingaddress", "homeaddress"},
}


def _header_mapping(row: Sequence[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        cleaned = _normalise_header(value)
        for canonical, aliases in HEADER_ALIASES.items():
            if cleaned in aliases and canonical not in mapping:
                mapping[canonical] = index
                break
    has_date = "dob" in mapping or {"month", "day", "year"}.issubset(mapping)
    if "ssn" in mapping and has_date and {"first_name", "last_name"}.issubset(mapping):
        return mapping
    return None


def _mapped_value(row: Sequence[str], mapping: Mapping[str, int], key: str) -> str:
    index = mapping.get(key)
    return row[index].strip() if index is not None and index < len(row) else ""


def _row_looks_like_split_date(row: Sequence[str]) -> bool:
    if len(row) < 6:
        return False
    try:
        _normalise_month(row[1])
        _normalise_day(row[2])
        _normalise_year(row[3])
        return True
    except ValueError:
        return False


def _parse_input_row(
    fields: Sequence[str],
    header: Mapping[str, int] | None,
) -> tuple[AccountDetails, str]:
    row = [str(value) for value in fields]
    if header is not None:
        ssn = _mapped_value(row, header, "ssn")
        first_name = _mapped_value(row, header, "first_name")
        last_name = _mapped_value(row, header, "last_name")
        address = _mapped_value(row, header, "address")
        dob = _mapped_value(row, header, "dob")
        if dob:
            month, day, year = _parse_dob(dob)
        else:
            month = _mapped_value(row, header, "month")
            day = _mapped_value(row, header, "day")
            year = _mapped_value(row, header, "year")
    elif _row_looks_like_split_date(row):
        # 当前测试资料：SSN, 月, 日, 年, Last Name, First Name, Address。
        ssn, month, day, year = row[:4]
        last_name, first_name = row[4:6]
        address = ",".join(row[6:]).strip()
    else:
        if len(row) < 4:
            raise ValueError(
                "至少需要 4 列：SSN、DOB、First Name、Last Name"
            )
        # 旧版输入：SSN, DOB, First Name, Last Name, Address。
        ssn, dob, first_name, last_name = row[:4]
        month, day, year = _parse_dob(dob)
        address = ",".join(row[4:]).strip()

    return _validate_details(
        ssn, month, day, year, first_name, last_name
    ), address


def _records_from_rows(
    rows: Iterable[Sequence[Any]],
    input_path: Path,
    sheet_name: str = "",
) -> list[ImportedRecord]:
    materialised: list[tuple[int, list[str]]] = []
    for row_number, raw_row in enumerate(rows, start=1):
        row = [_cell_to_text(value) for value in raw_row]
        while row and row[-1] == "":
            row.pop()
        if row and any(value.strip() for value in row):
            materialised.append((row_number, row))

    if not materialised:
        return []

    mapping = _header_mapping(materialised[0][1])
    data_rows = materialised[1:] if mapping is not None else materialised
    imported: list[ImportedRecord] = []
    for row_number, row in data_rows:
        details: AccountDetails | None = None
        address = ""
        import_error = ""
        try:
            details, address = _parse_input_row(row, mapping)
        except Exception as exc:
            import_error = _clean_error(exc)
        imported.append(
            ImportedRecord(
                source_file=str(input_path.resolve()),
                source_sheet=sheet_name,
                source_row=row_number,
                original_fields=tuple(row),
                details=details,
                address=address,
                import_error=import_error,
            )
        )
    return imported


def _read_text_rows(input_path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = input_path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8") from last_error

    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            delimiter = "\t"
    return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def load_input_records(input_path: Path) -> list[ImportedRecord]:
    """读取 CSV/SCV/TXT/XLSX，返回包括校验失败行在内的导入记录。"""
    input_path = input_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    suffix = input_path.suffix.casefold()
    records: list[ImportedRecord] = []
    if suffix in {".csv", ".scv", ".txt"}:
        records.extend(_records_from_rows(_read_text_rows(input_path), input_path))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                records.extend(
                    _records_from_rows(
                        worksheet.iter_rows(values_only=True),
                        input_path,
                        worksheet.title,
                    )
                )
        finally:
            workbook.close()
    else:
        raise ValueError("只支持 .csv、.scv、.txt 和 .xlsx 输入文件")

    if not records:
        raise ValueError("输入文件中没有可导入的资料行")
    return records


def _candidate_cdp_urls() -> list[str]:
    """返回可尝试的 Chrome DevTools 地址，不启动或关闭用户浏览器。"""
    configured = os.getenv("STUDENTAID_CDP_URL", "").strip()
    if configured.casefold() in {"off", "none", "disabled", "0", "false"}:
        return []
    if configured:
        return [configured.rstrip("/")]
    return ["http://127.0.0.1:9223"]


def _cdp_endpoint_available(url: str, timeout: float = 1.0) -> bool:
    try:
        with urllib.request.urlopen(f"{url}/json/version", timeout=timeout) as response:
            payload = json.load(response)
        return bool(payload.get("webSocketDebuggerUrl"))
    except Exception:
        return False


def launch_browser(playwright: Playwright) -> tuple[Browser, bool, str]:
    """连接已运行 Chrome；不可用时回退到 Playwright 管理的浏览器。"""
    for cdp_url in _candidate_cdp_urls():
        if not _cdp_endpoint_available(cdp_url):
            continue
        try:
            browser = playwright.chromium.connect_over_cdp(cdp_url, timeout=10_000)
            return browser, True, f"共享 Chrome CDP ({cdp_url})"
        except Exception:
            continue

    headless = os.getenv("STUDENTAID_HEADLESS", "").strip().casefold() in {
        "1", "true", "yes", "on"
    }
    try:
        browser = playwright.chromium.launch(channel="chrome", headless=headless)
        return browser, False, "独立 Google Chrome"
    except Exception:
        browser = playwright.chromium.launch(headless=headless)
        return browser, False, "独立 Playwright Chromium"


def _check_stop(stop_event: threading.Event) -> None:
    if stop_event.is_set():
        raise StopRequested("用户已请求停止")


def step_1_open_retrieve_account_details(
    page: Page, stop_event: threading.Event
) -> None:
    last_error: BaseException | None = None
    for attempt in range(1, 4):
        _check_stop(stop_event)
        try:
            page.goto(
                RETRIEVE_ACCOUNT_DETAILS_URL,
                wait_until="load",
                timeout=60_000,
            )
            _check_stop(stop_event)
            page.locator(FIRST_NAME_SELECTOR).wait_for(
                state="visible", timeout=60_000
            )
            page.wait_for_function(
                """
                () => {
                    const text = document.body?.innerText || "";
                    return !text.includes("Loading...")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameLastName")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameDateOfBirthMonth")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameSsnInput");
                }
                """,
                timeout=30_000,
            )
            return
        except StopRequested:
            raise
        except Exception as exc:
            last_error = exc
            if attempt >= 3:
                break
            page.wait_for_timeout(1_500 * attempt)
    raise RuntimeError(
        "StudentAid 页面连续 3 次无法打开；请先在 Chrome 中确认页面可访问，"
        "或设置 STUDENTAID_CDP_URL 复用已打开的 Chrome。"
    ) from last_error


def _type_account_field(page: Page, selector: str, value: str) -> None:
    """用真实键盘事件输入并失焦，确保 StudentAid 前端状态完整更新。"""
    field = page.locator(selector)
    field.click(timeout=30_000)
    field.fill("")
    field.press_sequentially(value, delay=60)
    field.press("Tab")


def step_2_fill_account_details(
    page: Page, details: AccountDetails, stop_event: threading.Event
) -> None:
    _check_stop(stop_event)
    _type_account_field(page, FIRST_NAME_SELECTOR, details.first_name)
    _type_account_field(page, LAST_NAME_SELECTOR, details.last_name)
    page.locator(BIRTH_MONTH_SELECTOR).select_option(details.birth_month)
    page.locator(BIRTH_MONTH_SELECTOR).press("Tab")
    _type_account_field(page, BIRTH_DAY_SELECTOR, str(int(details.birth_day)))
    _type_account_field(page, BIRTH_YEAR_SELECTOR, details.birth_year)
    _type_account_field(page, SSN_SELECTOR, details.ssn)
    _check_stop(stop_event)
    page.wait_for_function(
        """
        () => {
            const selectors = [
                "#fsa_Input_ForgotUsernameFirstName",
                "#fsa_Input_ForgotUsernameLastName",
                "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                "#fsa_Input_ForgotUsernameDateOfBirthDay",
                "#fsa_Input_ForgotUsernameDateOfBirthYear",
                "#fsa_Input_ForgotUsernameSsnInput",
            ];
            const button = [...document.querySelectorAll("button")].find(
                element => (element.innerText || "").trim() === "Continue"
            );
            return selectors.every(selector => {
                const field = document.querySelector(selector);
                return field && String(field.value || "").length > 0;
            }) && button && !button.disabled;
        }
        """,
        timeout=30_000,
    )
    page.wait_for_timeout(500)


def step_3_click_continue(page: Page, stop_event: threading.Event) -> None:
    _check_stop(stop_event)
    continue_element = page.get_by_role("button", name="Continue", exact=True).first
    continue_element.wait_for(state="visible", timeout=30_000)
    if not continue_element.is_enabled():
        raise RuntimeError("Continue 按钮未启用，请检查字段输入状态")
    continue_element.click(timeout=30_000)


def _report_stage(
    progress_callback: Callable[[str], None] | None, message: str
) -> None:
    """发送不包含输入资料的阶段消息；UI 回调失败不能中断网页处理。"""
    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def step_4_judge_password_recovery(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
    *,
    timeout_ms: int = 60_000,
    poll_interval_ms: int = 250,
    heartbeat_seconds: float = 5.0,
) -> str:
    """等待明确页面结果；DOM 消失或页面重绘本身绝不作为成功依据。"""
    started = time.monotonic()
    deadline = started + max(1, timeout_ms) / 1000
    next_heartbeat = max(0.05, heartbeat_seconds)

    while True:
        _check_stop(stop_event)
        body_text = page.evaluate("document.body?.innerText || ''")
        loading = "Loading..." in body_text

        if not loading:
            if "An unknown error has occurred" in body_text:
                raise PageSessionExpired(
                    "StudentAid 页面会话已失效，需要重新打开页面"
                )
            if "Something went wrong" in body_text:
                raise RuntimeError("StudentAid 返回页面错误，请稍后重试")

            if re.search(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                body_text,
                re.IGNORECASE | re.MULTILINE,
            ):
                _report_stage(progress_callback, "已识别结果：Account Not Found")
                return "account_not_found"
            if "Recover my account with a photo ID" in body_text:
                _report_stage(progress_callback, "已识别明确的账户找回结果")
                return "can_recover"

        now = time.monotonic()
        elapsed = now - started
        if now >= deadline:
            raise RuntimeError(
                f"等待 StudentAid 明确结果超时（{timeout_ms / 1000:g} 秒）"
            )
        if elapsed >= next_heartbeat:
            _report_stage(
                progress_callback,
                f"官方页面仍在处理中，已等待 {int(elapsed)} 秒",
            )
            while next_heartbeat <= elapsed:
                next_heartbeat += max(0.05, heartbeat_seconds)

        page.wait_for_timeout(min(poll_interval_ms, max(1, int((deadline - now) * 1000))))


def _visible_text(page: Page, text: str) -> str:
    locator = page.get_by_text(text, exact=True).first
    return locator.inner_text().strip() if locator.is_visible() else ""


def collect_recovery_result(page: Page, recovery_status: str) -> RecoveryResult:
    if recovery_status == "account_not_found":
        locator = page.get_by_text(
            re.compile(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                re.IGNORECASE,
            )
        ).first
        heading = locator.inner_text().strip() if locator.is_visible() else "Account Not Found"
        return RecoveryResult(recovery_status, heading, "", "", "")

    heading = _visible_text(page, "Retrieve Your Log-in Information")
    masked_phone = ""
    masked_email = ""
    for value in (line.strip() for line in page.locator("body").inner_text().splitlines()):
        if not value:
            continue
        if "@" in value and not masked_email:
            masked_email = value
        elif ("*" in value or "•" in value) and re.search(r"\d{4}\s*$", value):
            masked_phone = masked_phone or value
    recovery_method = _visible_text(page, "Recover my account with a photo ID")
    return RecoveryResult(
        recovery_status,
        heading,
        masked_phone,
        masked_email,
        recovery_method,
    )


class BrowserRecoverySession:
    """一个处理线程独占一个 Playwright 和 Browser，避免跨线程共享。"""

    def __init__(self, worker_number: int) -> None:
        self.worker_number = worker_number
        self._playwright: Playwright | None = None
        self._browser: Browser | None = None
        self._external_browser = False
        self.browser_mode = "未启动"

    def start(self) -> None:
        if sync_playwright is None:
            raise RuntimeError(
                "未安装 playwright，请执行：pip install playwright && playwright install chromium"
            )
        self._playwright = sync_playwright().start()
        self._browser, self._external_browser, self.browser_mode = launch_browser(
            self._playwright
        )

    def process(
        self,
        item: WorkItem,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> RecoveryResult:
        if self._browser is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        browser_contexts = list(getattr(self._browser, "contexts", []))
        owns_context = not (self._external_browser and browser_contexts)
        if owns_context:
            context = self._browser.new_context(
                viewport={"width": 1440, "height": 900},
                locale="en-US",
            )
        else:
            # connect_over_cdp 的默认 Chrome 上下文与 browser-use 实测环境一致。
            # 只关闭本任务新开的 Page，绝不关闭用户的 Context 或 Chrome。
            context = browser_contexts[0]
        try:
            for attempt in range(1, 3):
                page = context.new_page()
                try:
                    if attempt == 1:
                        _report_stage(progress_callback, "正在打开 StudentAid 页面")
                    else:
                        _report_stage(
                            progress_callback,
                            "页面会话失效，正在新开页面并重新填写资料（自动重试 1/1）",
                        )
                    step_1_open_retrieve_account_details(page, stop_event)
                    _report_stage(progress_callback, "页面已打开，正在填写资料")
                    step_2_fill_account_details(page, item.details, stop_event)
                    _report_stage(progress_callback, "资料已填写，正在点击 Continue")
                    step_3_click_continue(page, stop_event)
                    _report_stage(progress_callback, "已提交，正在等待官方结果")
                    status = step_4_judge_password_recovery(
                        page, stop_event, progress_callback
                    )
                    return collect_recovery_result(page, status)
                except PageSessionExpired as exc:
                    if attempt >= 2:
                        raise RuntimeError(
                            "StudentAid 页面重新打开并重填后仍提示会话失效"
                        ) from exc
                finally:
                    page.close()
            raise RuntimeError("StudentAid 页面自动重试未返回结果")
        finally:
            if owns_context:
                context.close()

    def close(self) -> None:
        if self._browser is not None:
            if not self._external_browser:
                try:
                    self._browser.close()
                except Exception:
                    pass
            self._browser = None
            self._external_browser = False
        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None


@dataclass
class _DbCommand:
    action: str
    payload: dict[str, Any]
    response: queue.Queue[tuple[bool, Any]]


class DatabaseWriter:
    """SQLite 的唯一写入者；工作线程不得直接执行写 SQL。"""

    def __init__(self, database_path: Path) -> None:
        self.database_path = database_path.resolve()
        self._commands: queue.Queue[_DbCommand] = queue.Queue()
        self._ready = threading.Event()
        self._thread = threading.Thread(
            target=self._run,
            name="studentaid-db-writer",
            daemon=False,
        )
        self._startup_error: BaseException | None = None

    def start(self) -> None:
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._thread.start()
        if not self._ready.wait(30):
            raise TimeoutError("SQLite 写入线程启动超时")
        if self._startup_error is not None:
            raise RuntimeError("SQLite 初始化失败") from self._startup_error

    def request(self, action: str, **payload: Any) -> Any:
        if not self._thread.is_alive():
            raise RuntimeError("SQLite 写入线程未运行")
        response: queue.Queue[tuple[bool, Any]] = queue.Queue(maxsize=1)
        self._commands.put(_DbCommand(action, payload, response))
        try:
            ok, value = response.get(timeout=300)
        except queue.Empty as exc:
            raise TimeoutError(f"SQLite 写入命令超时：{action}") from exc
        if ok:
            return value
        raise RuntimeError(f"SQLite 写入失败：{action}") from value

    def close(self) -> None:
        if self._thread.is_alive():
            try:
                self.request("shutdown")
            finally:
                self._thread.join(timeout=30)

    def _run(self) -> None:
        connection: sqlite3.Connection | None = None
        try:
            connection = sqlite3.connect(self.database_path, timeout=30)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA busy_timeout=30000")
            journal_mode = connection.execute("PRAGMA journal_mode=WAL").fetchone()[0]
            if str(journal_mode).casefold() != "wal":
                raise RuntimeError(f"无法启用 SQLite WAL，当前模式：{journal_mode}")
            connection.execute("PRAGMA synchronous=NORMAL")
            connection.execute("PRAGMA wal_autocheckpoint=1000")
            self._create_schema(connection)
            connection.commit()
        except BaseException as exc:
            self._startup_error = exc
            self._ready.set()
            if connection is not None:
                connection.close()
            return

        self._ready.set()
        assert connection is not None
        try:
            while True:
                command = self._commands.get()
                try:
                    value = self._handle(connection, command.action, command.payload)
                    connection.commit()
                    command.response.put((True, value))
                    if command.action == "shutdown":
                        break
                except BaseException as exc:
                    connection.rollback()
                    command.response.put((False, exc))
                finally:
                    self._commands.task_done()
        finally:
            connection.close()

    @staticmethod
    def _create_schema(connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS batches (
                batch_id TEXT PRIMARY KEY,
                input_path TEXT NOT NULL,
                output_directory TEXT NOT NULL,
                thread_count INTEGER NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                finished_at TEXT,
                export_path TEXT,
                total_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id TEXT NOT NULL REFERENCES batches(batch_id),
                source_file TEXT NOT NULL,
                source_sheet TEXT NOT NULL DEFAULT '',
                source_row INTEGER NOT NULL,
                original_fields_json TEXT NOT NULL,
                ssn TEXT,
                birth_month TEXT,
                birth_day TEXT,
                birth_year TEXT,
                first_name TEXT,
                last_name TEXT,
                address TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL,
                attempt_count INTEGER NOT NULL DEFAULT 0,
                result_code TEXT NOT NULL DEFAULT '',
                result_heading TEXT NOT NULL DEFAULT '',
                masked_phone TEXT NOT NULL DEFAULT '',
                masked_email TEXT NOT NULL DEFAULT '',
                recovery_method TEXT NOT NULL DEFAULT '',
                error TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_records_batch_status
                ON records(batch_id, status, id);
            """
        )

    @staticmethod
    def _handle(
        connection: sqlite3.Connection, action: str, payload: Mapping[str, Any]
    ) -> Any:
        if action == "create_batch":
            connection.execute(
                """
                INSERT INTO batches(
                    batch_id, input_path, output_directory, thread_count,
                    status, created_at
                ) VALUES (?, ?, ?, ?, 'importing', ?)
                """,
                (
                    payload["batch_id"], payload["input_path"],
                    payload["output_directory"], payload["thread_count"], _now_iso(),
                ),
            )
            return None

        if action == "insert_records":
            batch_id = str(payload["batch_id"])
            work_items: list[WorkItem] = []
            records: Sequence[ImportedRecord] = payload["records"]
            for record in records:
                details = record.details
                status = "failed" if record.import_error else "pending"
                cursor = connection.execute(
                    """
                    INSERT INTO records(
                        batch_id, source_file, source_sheet, source_row,
                        original_fields_json, ssn, birth_month, birth_day,
                        birth_year, first_name, last_name, address, status,
                        error, created_at, finished_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id, record.source_file, record.source_sheet,
                        record.source_row,
                        json.dumps(record.original_fields, ensure_ascii=False),
                        details.ssn if details else None,
                        details.birth_month if details else None,
                        details.birth_day if details else None,
                        details.birth_year if details else None,
                        details.first_name if details else None,
                        details.last_name if details else None,
                        record.address, status, record.import_error, _now_iso(),
                        _now_iso() if record.import_error else None,
                    ),
                )
                if details is not None and not record.import_error:
                    work_items.append(WorkItem(int(cursor.lastrowid), details))
            connection.execute(
                """
                UPDATE batches SET status='running', total_count=? WHERE batch_id=?
                """,
                (len(records), batch_id),
            )
            return work_items

        if action == "mark_processing":
            connection.execute(
                """
                UPDATE records
                SET status='processing', started_at=?, attempt_count=attempt_count+1,
                    error=''
                WHERE id=? AND status='pending'
                """,
                (_now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_completed":
            result: RecoveryResult = payload["result"]
            connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE id=?
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["record_id"],
                ),
            )
            return None

        if action in {"mark_failed", "mark_stopped"}:
            status = "failed" if action == "mark_failed" else "stopped"
            connection.execute(
                "UPDATE records SET status=?, error=?, finished_at=? WHERE id=?",
                (status, payload.get("error", ""), _now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_remaining":
            status = str(payload["status"])
            connection.execute(
                """
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND status IN ('pending', 'processing')
                """,
                (
                    status, payload.get("error", ""), _now_iso(), payload["batch_id"],
                ),
            )
            return None

        if action == "status_counts":
            rows = connection.execute(
                """
                SELECT status, COUNT(*) AS count FROM records
                WHERE batch_id=? GROUP BY status
                """,
                (payload["batch_id"],),
            ).fetchall()
            return {str(row["status"]): int(row["count"]) for row in rows}

        if action == "finish_batch":
            connection.execute(
                """
                UPDATE batches SET status=?, finished_at=?, export_path=?
                WHERE batch_id=?
                """,
                (
                    payload["status"], _now_iso(), payload.get("export_path", ""),
                    payload["batch_id"],
                ),
            )
            return None

        if action == "checkpoint":
            return connection.execute("PRAGMA wal_checkpoint(PASSIVE)").fetchone()

        if action == "shutdown":
            return None

        raise ValueError(f"未知 SQLite 写入命令：{action}")


def export_batch_csv(
    database_path: Path, batch_id: str, output_directory: Path
) -> Path:
    """只读查询当前批次并原子生成带 BOM 的最终 CSV。"""
    output_directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_directory / f"studentaid_results_{timestamp}_{batch_id[:8]}.csv"
    temporary_path = output_path.with_name(f".{output_path.name}.{os.getpid()}.tmp")

    connection = sqlite3.connect(database_path, timeout=30)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            "SELECT * FROM records WHERE batch_id=? ORDER BY id",
            (batch_id,),
        ).fetchall()
    finally:
        connection.close()

    original_rows: list[list[str]] = []
    maximum_columns = 0
    for row in rows:
        fields = [str(value) for value in json.loads(row["original_fields_json"])]
        original_rows.append(fields)
        maximum_columns = max(maximum_columns, len(fields))

    try:
        with temporary_path.open("w", encoding="utf-8-sig", newline="") as output_file:
            writer = csv.writer(output_file, lineterminator="\n")
            writer.writerow(
                [*(f"Input_Column_{index}" for index in range(1, maximum_columns + 1)),
                 *RESULT_COLUMNS]
            )
            for row, original_fields in zip(rows, original_rows):
                padded = [*original_fields, *("" for _ in range(maximum_columns - len(original_fields)))]
                writer.writerow(
                    [
                        *padded,
                        row["result_heading"], row["masked_phone"],
                        row["masked_email"], row["recovery_method"], row["status"],
                        row["error"], Path(row["source_file"]).name,
                        row["source_sheet"], row["source_row"],
                    ]
                )
            output_file.flush()
            os.fsync(output_file.fileno())
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path


EventCallback = Callable[[str, dict[str, Any]], None]
SessionFactory = Callable[[int], Any]


class BatchEngine:
    """负责导入、排队、并发处理、单线程写库和最终导出。"""

    def __init__(
        self,
        event_callback: EventCallback | None = None,
        session_factory: SessionFactory = BrowserRecoverySession,
    ) -> None:
        self._callback = event_callback or (lambda _kind, _payload: None)
        self._session_factory = session_factory
        self._stop_event = threading.Event()
        self._state_lock = threading.Lock()
        self._running = False
        self._coordinator: threading.Thread | None = None

    @property
    def is_running(self) -> bool:
        with self._state_lock:
            return self._running

    def start(self, input_path: Path, output_directory: Path, thread_count: int) -> None:
        input_path = input_path.resolve()
        output_directory = output_directory.resolve()
        if not input_path.is_file():
            raise FileNotFoundError(f"输入文件不存在：{input_path}")
        if not 1 <= thread_count <= 64:
            raise ValueError("线程数必须是 1-64")
        output_directory.mkdir(parents=True, exist_ok=True)

        with self._state_lock:
            if self._running:
                raise RuntimeError("已有批次正在运行")
            self._running = True
        self._stop_event.clear()
        self._coordinator = threading.Thread(
            target=self._run,
            args=(input_path, output_directory, thread_count),
            name="studentaid-coordinator",
            daemon=False,
        )
        self._coordinator.start()

    def stop(self) -> None:
        if self.is_running:
            self._stop_event.set()
            self._emit("log", message="已收到停止请求；正在结束当前步骤并停止领取新任务。")

    def wait(self, timeout: float | None = None) -> bool:
        coordinator = self._coordinator
        if coordinator is not None:
            coordinator.join(timeout)
        return not self.is_running

    def _emit(self, kind: str, **payload: Any) -> None:
        try:
            self._callback(kind, payload)
        except Exception:
            pass

    def _publish_progress(
        self, writer: DatabaseWriter, batch_id: str, total: int
    ) -> dict[str, int]:
        counts = writer.request("status_counts", batch_id=batch_id)
        payload = {
            "total": total,
            "pending": int(counts.get("pending", 0)),
            "processing": int(counts.get("processing", 0)),
            "completed": int(counts.get("completed", 0)),
            "failed": int(counts.get("failed", 0)),
            "stopped": int(counts.get("stopped", 0)),
        }
        self._emit("progress", **payload)
        return payload

    def _worker_loop(
        self,
        worker_number: int,
        tasks: queue.Queue[WorkItem | None],
        writer: DatabaseWriter,
        batch_id: str,
        total: int,
    ) -> None:
        session = self._session_factory(worker_number)
        try:
            session.start()
            browser_mode = getattr(session, "browser_mode", "浏览器模式未知")
            self._emit(
                "log",
                message=f"处理线程 {worker_number} 已启动：{browser_mode}。",
            )
        except Exception as exc:
            self._emit(
                "log",
                message=f"处理线程 {worker_number} 启动失败：{_clean_error(exc)}",
            )
            try:
                session.close()
            except Exception:
                pass
            return

        try:
            while True:
                try:
                    item = tasks.get(timeout=0.2)
                except queue.Empty:
                    if self._stop_event.is_set():
                        break
                    continue
                if item is None:
                    tasks.task_done()
                    break
                if self._stop_event.is_set():
                    tasks.task_done()
                    break

                writer.request("mark_processing", record_id=item.record_id)
                self._publish_progress(writer, batch_id, total)
                try:
                    result = session.process(
                        item,
                        self._stop_event,
                        lambda stage, record_id=item.record_id: self._emit(
                            "log", message=f"记录 #{record_id}：{stage}"
                        ),
                    )
                    writer.request(
                        "mark_completed", record_id=item.record_id, result=result
                    )
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：结果 {result.result_code} "
                            "已实时写入 SQLite"
                        ),
                    )
                except StopRequested as exc:
                    writer.request(
                        "mark_stopped", record_id=item.record_id, error=_clean_error(exc)
                    )
                except Exception as exc:
                    writer.request(
                        "mark_failed", record_id=item.record_id, error=_clean_error(exc)
                    )
                    self._emit(
                        "log",
                        message=f"记录 #{item.record_id} 处理失败：{_clean_error(exc)}",
                    )
                finally:
                    tasks.task_done()
                    self._publish_progress(writer, batch_id, total)
                if self._stop_event.is_set():
                    break
        finally:
            session.close()
            self._emit("log", message=f"处理线程 {worker_number} 已结束。")

    def _run(
        self, input_path: Path, output_directory: Path, thread_count: int
    ) -> None:
        batch_id = uuid.uuid4().hex
        database_path = output_directory / DATABASE_FILENAME
        writer: DatabaseWriter | None = None
        export_path: Path | None = None
        try:
            self._emit("log", message="正在读取并校验输入文件……")
            records = load_input_records(input_path)
            if self._stop_event.is_set():
                raise StopRequested("导入阶段已停止")

            writer = DatabaseWriter(database_path)
            writer.start()
            writer.request(
                "create_batch",
                batch_id=batch_id,
                input_path=str(input_path),
                output_directory=str(output_directory),
                thread_count=thread_count,
            )
            work_items: list[WorkItem] = writer.request(
                "insert_records", batch_id=batch_id, records=records
            )
            invalid_count = len(records) - len(work_items)
            self._emit(
                "log",
                message=(
                    f"已导入 {len(records)} 条资料到 SQLite；"
                    f"可处理 {len(work_items)} 条，格式错误 {invalid_count} 条。"
                ),
            )
            self._publish_progress(writer, batch_id, len(records))

            tasks: queue.Queue[WorkItem | None] = queue.Queue()
            for item in work_items:
                tasks.put(item)
            for _ in range(thread_count):
                tasks.put(None)

            workers = [
                threading.Thread(
                    target=self._worker_loop,
                    args=(number, tasks, writer, batch_id, len(records)),
                    name=f"studentaid-worker-{number}",
                    daemon=False,
                )
                for number in range(1, thread_count + 1)
            ] if work_items else []
            for worker in workers:
                worker.start()
            for worker in workers:
                worker.join()

            if self._stop_event.is_set():
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="stopped",
                    error="用户停止，任务尚未处理",
                )
                batch_status = "stopped"
            else:
                # 浏览器线程全部启动失败时，不能让 pending 永久悬空。
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="failed",
                    error="没有可用的处理线程完成此任务",
                )
                batch_status = "completed"

            final_counts = self._publish_progress(writer, batch_id, len(records))
            export_path = export_batch_csv(database_path, batch_id, output_directory)
            writer.request(
                "finish_batch", batch_id=batch_id, status=batch_status,
                export_path=str(export_path),
            )
            writer.request("checkpoint")
            self._emit(
                "finished",
                batch_id=batch_id,
                database_path=str(database_path),
                export_path=str(export_path),
                status=batch_status,
                counts=final_counts,
            )
        except StopRequested as exc:
            self._emit("fatal", message=_clean_error(exc), stopped=True)
        except Exception as exc:
            self._emit(
                "fatal",
                message=_clean_error(exc),
                details=traceback.format_exc(limit=6),
                stopped=False,
            )
        finally:
            if writer is not None:
                try:
                    writer.close()
                except Exception as exc:
                    self._emit("log", message=f"关闭数据库时出错：{_clean_error(exc)}")
            with self._state_lock:
                self._running = False
            self._emit("idle", export_path=str(export_path or ""))


class StudentAidApp:
    """第九步源码校验桌面 GUI。"""

    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("920x650")
        self.root.minsize(820, 570)

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar(value=str(Path.cwd()))
        self.thread_var = tk.StringVar(value=str(min(4, max(1, os.cpu_count() or 1))))
        self.status_var = tk.StringVar(value="就绪")
        self.progress_text_var = tk.StringVar(value="总数 0 | 完成 0 | 失败 0 | 停止 0")
        self._ui_events: queue.Queue[tuple[str, dict[str, Any]]] = queue.Queue()
        self.engine = BatchEngine(self._backend_event)
        self._closing = False

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._poll_events)

    def _build_ui(self) -> None:
        ttk = self.ttk
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(6, weight=1)

        ttk.Label(main, text="导入文件：").grid(row=0, column=0, sticky="w", pady=6)
        self.input_entry = ttk.Entry(main, textvariable=self.input_var)
        self.input_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        self.input_button = ttk.Button(main, text="选择文件", command=self._choose_input)
        self.input_button.grid(row=0, column=2, pady=6)

        ttk.Label(main, text="导出目录：").grid(row=1, column=0, sticky="w", pady=6)
        self.output_entry = ttk.Entry(main, textvariable=self.output_var)
        self.output_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        self.output_button = ttk.Button(main, text="选择目录", command=self._choose_output)
        self.output_button.grid(row=1, column=2, pady=6)

        ttk.Label(main, text="处理线程数：").grid(row=2, column=0, sticky="w", pady=6)
        self.thread_spin = ttk.Spinbox(
            main, from_=1, to=64, textvariable=self.thread_var, width=10
        )
        self.thread_spin.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="每个线程使用独立浏览器；建议先用 2-4。")\
            .grid(row=2, column=1, sticky="w", padx=(100, 0), pady=6)

        button_bar = ttk.Frame(main)
        button_bar.grid(row=3, column=0, columnspan=3, sticky="ew", pady=(10, 8))
        self.start_button = ttk.Button(button_bar, text="开始", command=self._start)
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(
            button_bar, text="停止", command=self._stop, state="disabled"
        )
        self.stop_button.pack(side="left", padx=8)
        ttk.Label(button_bar, textvariable=self.status_var).pack(side="right")

        self.progress = ttk.Progressbar(main, mode="determinate", maximum=1, value=0)
        self.progress.grid(row=4, column=0, columnspan=3, sticky="ew", pady=(2, 4))
        ttk.Label(main, textvariable=self.progress_text_var).grid(
            row=5, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        log_frame = ttk.LabelFrame(main, text="运行日志（不显示 SSN 和原始资料）", padding=8)
        log_frame.grid(row=6, column=0, columnspan=3, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.log_text = self.tk.Text(log_frame, wrap="word", state="disabled")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

    def _choose_input(self) -> None:
        path = self.filedialog.askopenfilename(
            title="选择导入文件",
            filetypes=[
                ("支持的资料文件", "*.csv *.scv *.txt *.xlsx"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
                ("纯文本文件", "*.txt *.scv"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            self.input_var.set(path)
            if not self.output_var.get().strip():
                self.output_var.set(str(Path(path).parent))

    def _choose_output(self) -> None:
        path = self.filedialog.askdirectory(title="选择导出目录")
        if path:
            self.output_var.set(path)

    def _set_running(self, running: bool) -> None:
        normal_or_disabled = "disabled" if running else "normal"
        self.input_entry.configure(state=normal_or_disabled)
        self.output_entry.configure(state=normal_or_disabled)
        self.thread_spin.configure(state=normal_or_disabled)
        self.input_button.configure(state=normal_or_disabled)
        self.output_button.configure(state=normal_or_disabled)
        self.start_button.configure(state=normal_or_disabled)
        self.stop_button.configure(state="normal" if running else "disabled")

    def _start(self) -> None:
        try:
            input_path = Path(self.input_var.get().strip())
            output_directory = Path(self.output_var.get().strip())
            thread_count = int(self.thread_var.get().strip())
            self.engine.start(input_path, output_directory, thread_count)
        except Exception as exc:
            self.messagebox.showerror(APP_TITLE, _clean_error(exc))
            return
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self.status_var.set("运行中")
        self._set_running(True)
        self._append_log("批次已启动。")

    def _stop(self) -> None:
        self.stop_button.configure(state="disabled")
        self.status_var.set("停止中")
        self.engine.stop()

    def _backend_event(self, kind: str, payload: dict[str, Any]) -> None:
        self._ui_events.put((kind, payload))

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _poll_events(self) -> None:
        try:
            while True:
                kind, payload = self._ui_events.get_nowait()
                if kind == "log":
                    self._append_log(str(payload.get("message", "")))
                elif kind == "progress":
                    total = int(payload.get("total", 0))
                    completed = int(payload.get("completed", 0))
                    failed = int(payload.get("failed", 0))
                    stopped = int(payload.get("stopped", 0))
                    processing = int(payload.get("processing", 0))
                    terminal = completed + failed + stopped
                    self.progress.configure(maximum=max(1, total), value=terminal)
                    self.progress_text_var.set(
                        f"总数 {total} | 完成 {completed} | 失败 {failed} | "
                        f"停止 {stopped} | 处理中 {processing}"
                    )
                elif kind == "finished":
                    output_path = str(payload.get("export_path", ""))
                    database_path = str(payload.get("database_path", ""))
                    counts = payload.get("counts", {})
                    failed = int(counts.get("failed", 0)) if isinstance(counts, dict) else 0
                    if payload.get("status") == "stopped":
                        display_status = "已停止"
                    elif failed:
                        display_status = "完成（有失败）"
                    else:
                        display_status = "已完成"
                    self.status_var.set(display_status)
                    self._append_log(f"CSV 已导出：{output_path}")
                    self._append_log(f"SQLite 数据库：{database_path}")
                    self.messagebox.showinfo(
                        APP_TITLE,
                        f"批次处理结束。\n\nCSV：{output_path}\n数据库：{database_path}",
                    )
                elif kind == "fatal":
                    message = str(payload.get("message", "未知错误"))
                    self.status_var.set("已停止" if payload.get("stopped") else "失败")
                    self._append_log(message)
                    if not payload.get("stopped"):
                        self.messagebox.showerror(APP_TITLE, message)
                elif kind == "idle":
                    self._set_running(False)
                self._ui_events.task_done()
        except queue.Empty:
            pass

        if self._closing and not self.engine.is_running:
            self.root.destroy()
            return
        self.root.after(100, self._poll_events)

    def _on_close(self) -> None:
        if self.engine.is_running:
            if not self.messagebox.askyesno(APP_TITLE, "任务仍在运行，是否停止并退出？"):
                return
            self._closing = True
            self._set_running(True)
            self.stop_button.configure(state="disabled")
            self.status_var.set("停止中")
            self.engine.stop()
        else:
            self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    StudentAidApp().run()


if __name__ == "__main__":
    main()
