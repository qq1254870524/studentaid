"""
StudentAid 批量搜索工具 v32。

2026-08-18 第三十七步 TXT 分隔符与无表头身份列修复版更新：
1. TXT/CSV 同时含 Tab 和字段内逗号时，按稳定列宽优先识别真正的 Tab 分隔，避免年龄描述中的
   单个逗号让 12 列资料被误读成 2 列；导入和结果落盘后的输入删行共用同一检测规则。
2. 无表头宽表的姓名推断优先选择连续姓名列组，并排除大写美国州缩写；手机号、标记、年龄、
   邮箱列表、街道、城市、州、邮编、SSN、DOB 这类布局可稳定定位 First/Last Name。
3. 现场 ``导入资料/8.18 强.txt`` 的 4 行、12 列全部导入成功，SSN/DOB/姓名/街道映射正确；
   全部 12 个原始输入字段仍按原顺序保留，累计输出继续只在末尾追加四个结果字段。
4. 仅修复文本分隔和无表头身份列推断；Windows 输出权限、累计结果末列、SQLite、HTTP/浏览器、
   完整行去重和其他已稳定流程保持不变。

2026-08-18 第三十六步 Windows 输出权限自动修复版更新：
1. 开始批次和清空数据库前，自动为当前输出目录补齐当前 Windows 账户的可继承完全控制；
   新建累计 CSV、SQLite ``-wal``/``-shm`` 不再因提升权限运行后只归 Administrators 而变只读。
2. 对已经存在的累计 CSV、数据库及 WAL/SHM sidecar 同步补写当前账户权限；后台清库不再报
   ``attempt to write a readonly database``，累计结果文件退出程序后仍可由 Excel/文本编辑器修改。
3. 权限修复仅作用于用户选择的输出目录和明确的 StudentAid 输出文件，不改输入资料内容、
   数据库表结构、累计结果行或网页处理流程；外部程序真实占用文件时仍保留原有重试与报错。

2026-08-17 第三十五步累计结果状态末列统一版更新：
1. 累计 CSV 统一按全文件最大原始输入宽度补齐空单元格，所有行固定同一列数；
   Result Heading 不再随 6/20/21/22 列输入漂移，固定为唯一最后一列。
2. 保留 Masked Phone、Masked Email、Recovery Method 三项结果详情，并固定放在状态列前；
   原始输入每个非空单元格及顺序保持不变，只增加对齐空格和调整四个结果字段顺序。
3. 启动时自动迁移旧累计文件的 ``Heading/Phone/Email/Method`` 尾部布局；后续遇到
   更宽输入时先原子扩展既有行，再实时追加，不产生前后错列。
4. 累计输出完整行去重会忽略程序增加的尾部对齐空格，SQLite、HTTP/浏览器、输入删行、
   大型 XLSX 批末一次写回和其他已稳定功能保持不变。

2026-08-17 第三十四步大批量 XLSX 持久化不卡队列版更新：
1. 修复 4 万行级 XLSX 每累计 20 个结果就在结果持久化线程内重写输入文件，导致该
   线程数分钟不能消费 CSV 写入命令、30 个 HTTP worker 最终报“累计输出写入等待超时”。
2. XLSX 处理期间只实时写 SQLite 和累计 CSV，完整行删除键在内存合并；所有 worker
   结束后只重写一次输入 XLSX。CSV/TXT 原有定时合并逻辑保持不变。
3. 启动时已有累计结果不再先重写大型 XLSX，而是在只读导入后按完整行键直接跳过，
   并把匹配键并入批末一次性删除，避免恢复任务启动阶段再次长时间停滞。
4. XLSX 删除由逐行调用 ``delete_rows`` 改为 ZIP 内 worksheet XML 单遍过滤并重编号，
   样式和其他工作表部件原样保留，消除 O(删除数×总行数) 单元格搬移。
5. GUI 异步清库、完整行去重、HTTP/浏览器、输出列和数据库结构均保持不变。

2026-08-17 第三十三步 GUI 异步清空数据库版更新：
1. “清空数据库”确认后的 DELETE、WAL checkpoint 和 VACUUM 全部移到独立后台线程，
   Tk 主线程只负责刷新状态和接收完成事件，数据库较大或短暂锁定时窗口仍可拖动、刷新。
2. 清理期间显示“正在清空数据库”，并临时禁用开始、路径、后端和清空按钮，防止清理
   与新批次同时操作同一个 SQLite；完成或失败后由 GUI 事件队列统一恢复控件。
3. 关闭窗口时若数据库仍在清理，可选择等待后台清理结束后自动退出，不会中断事务或
   在后台线程直接调用 Tk；原有清理范围、表结构保留和文件保护规则不变。

2026-08-17 第三十二步 GUI 清空数据库版更新：
1. GUI 开始/停止按钮旁新增“清空数据库”按钮，清理对象固定为当前累计输出目录下的
   ``studentaid.sqlite3``；累计 CSV、输入文件和 GUI 配置均不删除。
2. 点击后先显示数据库绝对路径和影响范围并要求确认；确认后清空 batches/records、
   重置自增序列、截断 WAL 并 VACUUM，数据库结构保留供下次批次直接复用。
3. 批次运行期间按钮禁用，方法层也再次检查运行状态；数据库不存在、被占用或损坏时
   均显示明确提示，不影响其他文件。
4. 仅增加数据库维护入口；完整行去重、自动格式适配、HTTP/浏览器和输出流程不变。

2026-08-17 第三十一步完整行去重全路径统一版更新：
1. 导入、批次队列、SQLite 分组、累计输出判重、启动同步和结果完成后的输入删行，
   全部统一使用原始输入行每一列生成的完整行键，不使用 SSN 或任意单列去重。
2. 无表头宽表的输入删行与启动同步现在复用和导入完全相同的自动列推断，确保结果
   落盘后能按完整行精确找到原记录；同一 SSN 但其他列不同的资料保持独立。
3. 完全相同的整行只搜索一次并删除全部相同副本；只要任意一列内容不同，即使 SSN
   相同，也会分别搜索、分别输出、分别删除。
4. 仅统一完整行键重建路径；字段识别、HTTP/浏览器搜索、SQLite、输出列和 GUI 不变。

2026-08-17 第三十步无表头自动列推断版更新：
1. 当首行不是表头时，自动统计整表各列的 SSN、DOB、姓名和地址特征，不再只依赖
   固定的四列、五列、七列位置；首行作为真实数据保留，不会误删。
2. 新增 ``SSN / First / Middle / Last / DOB / Email / Phone / Address / ...``
   等无表头宽表识别；Middle Name 自动跳过，First/Last Name 与 DOB 精确定位。
3. ``导入资料\\8.17 测试.xlsx`` 的 936 条、10 列资料现在全部可识别，原始 10 列
   逐格保留，累计输出仍只在末尾追加四个结果列。
4. 已有带表头万能适配、旧无表头四/五/七列、HTTP、浏览器、去重、SQLite、
   累计输出和输入删行流程保持不变。

2026-08-17 第二十九步万能表头自动适配版更新：
1. 自动识别任意列顺序中的原始字段、``result_*`` 结果字段及常见前后缀表头；支持
   SSN、DOB/出生日期、First/Last/Full Name 和 Address 的英文、中文及语义别名。
2. 同一行同时存在原始字段与搜索结果字段时，优先使用 ``result_ssn``、
   ``result_birth_date``、``result_first_name``、``result_last_name``；结果字段为空时
   逐字段回退原始姓名，避免丢弃仍可搜索的完整资料。
3. ``8.11 背调手机号 SSN.xlsx`` 现在可自动跳过表头、识别有效资料并进入现有
   HTTP/浏览器搜索；37 个原始输入列仍逐格保留，只在末尾追加既有四个结果列。
4. 只扩展输入识别；HTTP、browser-use、playwright、默认后端、整行去重、SQLite、
   累计输出、输入删行、线程和 GUI 配置流程保持不变。

2026-08-17 第二十八步 HTTP 搜索后端增量更新：
1. 在现有 browser-use、playwright 浏览器搜索之外新增 ``http`` 搜索后端；每个
   worker 使用独立 HTTP 会话，按官网当前会话初始化、CSRF 和 JSON 接口流程查询。
2. HTTP 明确结果映射到现有 Result Heading、Masked Phone、Masked Email、
   Recovery Method 四列；接口返回的联系方式先按官网规则遮挡，再进入既有落盘流程。
3. 会话失效时自动重新初始化一次；限流、服务端错误和未知返回继续进入既有队尾重试，
   不把不明确响应误写为完成结果。
4. 只增加 HTTP 搜索选项；输入字段映射、默认 browser-use 后端、浏览器流程、整行去重、
   SQLite、累计输出、输入删行、线程和 GUI 配置规则保持不变。

2026-08-12 第二十七步原始输入列完整保留版更新：
1. 累计输出严格保持输入资料的列数、列顺序和每个单元格内容，不删除、不替换、
   不前插身份列；只在原始行最后追加 Result Heading、Masked Phone、Masked Email、
   Recovery Method 四个页面结果字段。
2. 完整行去重键同步改为逐字段使用原始输入行，确保输出前缀与输入整行完全一致；
   同一 SSN 但任意其他列不同的资料仍分别处理。
3. MyLife 当前 20 列输入固定输出为原始 20 列加结果 4 列，共 24 列；旧四列、五列、
   七列和任意动态列输入同样只在末尾增加四列。
4. 只修改输入列透传边界；网页识别、Continue、Cancel、浏览器复用、线程、GUI、
   SQLite、明确结果、输入删行及缓存清理逻辑保持不变。

2026-08-12 第二十六步整行去重版更新：
1. 批次内、累计输出和启动同步全部改为按完整输入资料行去重；同一 SSN 但其他任意
   字段不同的资料视为不同记录，分别进入浏览器并分别输出。
2. 完整行比较使用规范化后的逻辑输入列：SSN、标准 DOB、First Name、Last Name
   加全部原始/顺延字段；页面结果四列不参与输入资料是否重复的判断。
3. 完全相同的输入行仍只处理一次；明确结果后只删除与该完整行一致的输入记录，
   不再因为 SSN 相同删除其他不同资料。
4. SQLite 完成、重试、失败和停止改为按完整行键联动完全相同的副本；不同整行之间
   不再共享状态或页面结果，队尾重试仍保持原有三轮规则。
5. 只修改去重/同步边界；MyLife 新格式、旧格式、网页状态、Continue、Cancel、
   浏览器复用、持久化流水线和 GUI 实时效率看板保持不变。

2026-08-12 第二十五步 MyLife 新格式自适应版更新：
1. 自动识别包含 ``primary_phone,ssn,known_birthday,full_name,...,生日`` 的
   MyLife 结果 CSV；SSN 不再要求位于第一列，中文生日优先、known_birthday 备用。
2. full_name 自动拆为 First Name 和 Last Name，current_address 映射为 Address；
   页面仍只填写 StudentAid 所需六项资料，其余原始列按原顺序完整透传到累计输出。
3. 新格式累计输出强制以真实 SSN 为第一列，并将标准 DOB、First Name、Last Name
   放在前四列；原始输入的其他列随后顺延，确保累计去重和启动同步删行都按 SSN。
4. 输入明确完成后的删行、输出第一列已存在时的启动删行均按表头中的 SSN 列执行；
   两条生日为空的资料继续按既有缺少必填字段规则直接删除，不进入浏览器。
5. 只增加输入格式适配；网页状态、Continue、Cancel、浏览器复用、实时效率看板、
   持久化线程和所有既有输入格式保持不变。

2026-08-12 第二十四步 GUI 实时效率看板版更新：
1. GUI 新增实时效率信息：完成百分比、已处理/剩余/等待/处理中数量、最近一分钟
   处理数量、批次平均速度、预计剩余时间和已运行时间。
2. 速度采用批次首个数据库进度快照作为基线，避免把导入阶段已经判定的格式错误
   误算成浏览器吞吐量；最近一分钟采用滚动 60 秒窗口，ETA 随进度实时刷新。
3. 并发进度事件如果发生到达顺序交错，旧快照不会让百分比或已处理数量倒退；
   批次结束后冻结耗时和 ETA，便于保留最终现场信息。
4. 只调整 GUI 进度展示与纯计算辅助类；网页识别、Continue、Cancel、浏览器复用、
   结果状态、输入删除、累计输出和持久化流水线保持不变。

2026-08-12 第二十三步 20 线程持久化流水线优化版更新：
1. 新增单独的结果持久化线程；浏览器 worker 只等待累计 CSV 实时落盘，不再等待
   XLSX 整表重写，页面清理和下一条资料可以立即继续。
2. 输出第一列在批次启动时只读取一次并驻留内存；每次追加后同步更新去重集合，
   不再为每条结果重复扫描整个累计 CSV。
3. 输入文件删除改为最多 20 个结果或 5 秒合并一次，停止、正常结束和队列关闭时
   强制刷新；顺序仍为 SQLite、累计 CSV、输入删除，异常退出后继续由启动同步兜底。
4. 只调整结果文件持久化路径；网页识别、Continue、Cancel、浏览器复用、结果状态、
   动态列、DOB、队尾重试和线程数规则保持不变。

2026-08-12 第二十二步 Account Not Found 完整状态版更新：
1. 点击 Continue 后，页面精确显示 ``Account Not Found: Create a New Account``
   时，将该完整文案作为正常明确结果写入输出，而不是只保留缩短的状态名。
2. 该条仍沿用现有 Account Not Found 的清理、输入删除、累计输出和下一条资料流程；
   GUI 配置持久化及其他浏览器、字段、结果逻辑保持不变。

2026-08-12 第二十一步 GUI 配置持久化版更新：
1. GUI 现在把输入文件、累计输出、浏览器后端、窗口/无头和线程数保存到运行目录
   的 ``studentaid_gui_config.json``；源码启动和独立 EXE 下次打开都会恢复上次值。
2. 配置采用临时文件原子替换，启动时遇到损坏或旧值会回退到默认值，不影响批处理。
3. 只增加配置读写和窗口关闭/开始时保存；网页流程、浏览器复用、SQLite、输入输出
   和既有明确结果逻辑保持不变。

2026-08-09 第二十步账户恢复进行中状态版更新：
1. 点击 Continue 后若页面显示 Account Recovery In Progress，立即将该完整文案
   作为正常明确结果输出，不再等待超时或进入页面重建、队尾重试。
2. 结果落盘和输入删行后沿用现有终态清理流程；Step19 动态列、SSN 校验以及
   其他页面结果、浏览器复用、批处理和输出逻辑保持不变。

2026-08-08 第十九步动态输入列与 SSN 页面校验版更新：
1. 无表头 DOB 输入只要求前四列为 SSN、DOB、First Name、Last Name；第 5 列及
   后续任意列不再合并到 Address，而是按原顺序逐列透传到累计 CSV，最后再追加四个
   页面结果字段。原五列输入仍输出九列，七列输入自然输出十一列。
2. 点击 Continue 后若页面显示 ``Enter a valid Social Security number.``，立即将
   该完整文案作为正常明确结果输出，不再反复清页重填或进入队尾重试。
3. 拆分月/日/年旧输入、生日 ``MM/DD/YYYY``、两个 Step18 状态、Cancel、浏览器
   复用、清缓存、累计去重和任意正整数线程等既有逻辑保持不变。

2026-08-08 第十八步新增明确状态与任意线程版更新：
1. 点击 Continue 后若出现 ``Account Lookup Issue: Get Help``，将该完整文案作为
   正常明确结果写入累计 CSV，随后沿用终态清理流程处理下一条。
2. 点击 Continue 后若出现 ``We are unable to retrieve your log-in information. Access
   your account by recovering your account with a photo ID.``，将该完整文案作为正常
   明确结果写入累计 CSV，不再误入普通 Retrieve 联系方式提取或等待超时。
3. GUI 线程输入改为可直接填写任意正整数，批处理层同步移除 8 线程硬上限；默认值、
   每线程独立浏览器、停止清缓存、结果落盘和输入删除等既有逻辑保持不变。

2026-08-08 第十七步瞬时失败队尾重试版更新：
1. 浏览器任务完成三次页面会话重建后仍未取得明确结果时，不再立即作为本批终态
   失败；任务会回到队尾，先让其余资料继续处理，避免单条慢响应持续占用线程。
2. 每个唯一任务最多执行三轮队列尝试，每轮仍保留原有三次页面清理、重开、重填；
   下一轮会在当前队列后方执行，让官网瞬时故障有充分恢复时间。
3. SQLite 新增同组 retry 状态回退：代表任务与同第一列重复行一起回到 pending，
   attempt_count 继续累计；最终明确结果仍只追加一行并一次删除全部匹配输入行。
4. 协调器改为等待动态队列真正清空后再发送线程结束信号，保证运行中追加到队尾的
   重试任务不会落在结束信号之后而被遗漏；停止按钮仍可立即结束并保留未完成输入。

2026-08-08 第十六步大批量去重稳定版更新：
1. 同一批输入按规范化后的第一列只创建一个浏览器任务；重复资料不再被 8 个线程
   同时领取，避免同一 SSN 重复打开页面、重复点击 Continue 和重复触发官网限流。
2. 唯一任务取得明确结果后，SQLite 会把同一批次、同一第一列的全部重复记录一起
   标记完成；累计 CSV 仍只写一行，输入文件仍一次删除全部匹配行。
3. 唯一任务失败或停止时，同组重复记录一起进入相同终态并保留在输入文件，后续
   批次可以整体重试，不会出现没有队列任务却长期停在 pending 的情况。
4. 累计 CSV 启动规范化现在同时删除全空物理行；即使文件已经是全字段双引号格式，
   也会清除文件开头或中间的空白行，保持固定 9 列数据连续。

2026-08-08 第十五步批量输入兼容版更新：
1. 兼容 ``MM/DD/YYY`` 或 ``MM-DD-YYY`` 的三位年份：仅当补前导 ``1`` 后落在
   1900 到当前年份且能组成真实日期时才接受，避免把其他错误日期静默改写。
2. First Name、Last Name、Month、Day、Year、SSN 任意必填项为空时，直接从输入
   文件删除该源行，不写累计输出，也不启动浏览器处理该行。
3. 输入处理发生在校验阶段；累计输出生日仍严格为单列 ``MM/DD/YYYY``，
   其他无法可靠恢复的格式错误继续保留在输入文件，不会被误删。

2026-08-08 第十四步生日格式修复版更新：
1. 累计输出第 2 列生日强制统一为 ``MM/DD/YYYY``，月份和日期始终补足两位。
2. 不再保留五列输入中的英文月份、短日期或其他原始显示格式；输入可以继续使用
   支持的多种生日写法，但输出只采用校验后的 month/day/year 生成单列标准日期。
3. 七列拆分生日输入仍按月、日、年读取，累计 CSV 不会把生日拆成三列；每行继续
   固定为 5 个资料列加 4 个结果列，共 9 列。
4. 现场已有累计 CSV 的 17 条生日已全部是单列 ``MM/DD/YYYY``，无需重写现有结果。

2026-08-08 第十三步联系方式修复版更新：
1. 修复找回结果页手机号漏记：StudentAid 实际掩码字符为 ``⦁``（U+2981），
   旧版只识别 ``*`` 和 ``•``，导致 ``(⦁⦁⦁) ⦁⦁⦁ 8139`` 未写入累计 CSV。
2. 联系方式改为优先读取结果卡片中可见的 ``p.fsa-color-gray-60``，手机号和
   邮箱分别按掩码格式判断；页面有值就原样记录，没有就严格保持空白。
3. 兜底扫描也只限定在可见 ``p`` 元素，避免把页脚邮箱或页面其他文字误当成
   找回邮箱；兼容 ``⦁``、``•``、``●`` 和 ``*`` 掩码字符。
4. Retrieve 的 Cancel 改为直接点击可见 ``<span>Cancel</span>``，并验证结果页
   实际离开；旧版仅发出按钮坐标点击，现场存在没有点上的情况。点击成功后复用同一个
   worker、浏览器、页面会话和空白表单继续填下一条，不清缓存、不重启 Chrome。
5. Account Not Found 保存后清除全部浏览器数据、回到 about:blank，下一条才重新打开
   指定网址；这条清理规则不用于 Retrieve。
6. 保持累计输出、明确结果后删除输入行，以及启动时按累计 CSV 第一列清理输入行。

2026-08-08 第十二步无头稳定版更新：
1. GUI 新增“浏览器显示”下拉框，选项严格为“窗口”和“无头”；默认“无头”。
2. 窗口模式显示 900x720 的小 Chrome 窗口，页面 viewport 固定为 880x650；
   无头模式完全不创建可见浏览器窗口。
3. browser-use / playwright 与窗口 / 无头可自由组合；每个 worker 继续拥有独立
   Chrome 进程、CDP/Context 和缓存目录，不复用用户已有浏览器。
4. 批处理日志记录本次后端、显示模式和线程数；运行中禁止修改两个下拉选项。
5. 停止或结束时两种显示模式都执行相同的 cookie/cache/storage/service worker
   清理，并结束本批次浏览器；browser-use 临时 profile 同步删除。
6. 新增四组合启动参数、GUI 默认值和浏览器生命周期回归；使用假资料测试4.xlsx
   的隔离副本对四种组合逐一现场验证。
7. 修复小窗口下 Continue 坐标与页面缩放不一致导致鼠标落点未触发：优先使用
   Playwright Locator 的真实按钮点击，未进入提交状态时依次用 Enter 和 DOM click 兜底，
   每种方式都必须确认 Loading、按钮禁用、结果文案或目标路径后才进入结果等待。
8. 实测 Chromium 原生 headless 会被 StudentAid 以 HTTP/2 协议错误拒绝；“无头”改用
   Windows 隐藏的普通 Chrome 网络栈，启动时只隐藏本批次精确 PID 的窗口，用户看不到
   浏览器且不会影响已有 Chrome。导航以 commit 返回，随后仍严格等待表单字段和 Loading。
9. browser-use 结束顺序改为先结束精确 PID 树、删除一次性 profile，再断开 CDP，避免
   失败页面令关闭阶段卡住；四组合 2 线程现场矩阵 8/8 完成且残留 Chrome 为 0。

2026-08-08 第十一步稳定版更新：
1. GUI 新增 browser-use / playwright 后端下拉选项；两种模式都为每个 worker
   创建独立浏览器进程和独立临时配置目录，默认使用 2 个处理线程。
2. browser-use 模式直接采用 BrowserProfile 的 Chrome 参数、隔离目录和 CDP
   启动策略，再执行现有确定性表单流程，避免默认共享 daemon 在并发时互相抢焦点。
3. Limit Reached: Try Again in 24 Hours 改为正常明确结果：实时追加累计 CSV，
   最终状态原样写入第 6 列，随后删除对应输入行并继续处理下一条。
4. 停止或批次结束时，每个 worker 都清除 cookie、cache、storage、service worker，
   回到空白页并结束本脚本创建的浏览器进程；browser-use 临时配置目录同步删除。
5. 恢复并发任务队列和单独 SQLite writer，累计输出与输入删行继续使用进程内锁和
   原子文件替换，两个浏览器线程不会同时破坏 CSV/XLSX。
6. 键盘逐字输入延迟由 60ms 调整为 20ms，字段完成后的稳定等待由 500ms 调整为
   200ms；仍保留真实键盘事件、失焦和 Continue 启用校验。
7. 累计 CSV 的 DOB 对 5 列无表头输入保留原始显示格式，与实际输出参考一致。
8. 根据假资料和实际输出参考补齐 ``September 07, 1980`` 这类英文月份 DOB
   导入格式；此类有效资料不再作为格式错误留在输入文件。

2026-08-08 第十步稳定版更新：
1. Account Not Found 明确结果实时保存后，清除 StudentAid 专用浏览器全部站点数据，
   回到 about:blank；处理下一条时重新打开找回页面。
2. Retrieve Your Log-in Information 明确结果实时保存后，点击可见 Cancel，
   等待返回空白输入表单，再填写下一条资料。
3. 采用严格顺序处理，避免多个共享 CDP 页面同时清理 cookie/session 导致提交互相干扰。
4. 累计结果实时追加到用户选择的 CSV，不覆盖旧数据；程序启动时读取输出第一列，
   输入第一列已存在于输出的资料会直接从输入文件原子删除。
5. 每条明确结果按“先写 SQLite、再追加累计 CSV、最后删除输入行”的顺序落盘，
   中断重启时可依靠输出第一列去重并继续。
6. 定位 Continue 转圈根因：Playwright 直接启动 Chrome 时 navigator.webdriver=true；
   新版复用 browser-use 的 AutomationControlled 启动参数，实测启动后该标志为 false。
7. Continue 改用可见按钮坐标鼠标点击；持续 Loading 自动清理、重开、重填并重试。
8. 识别官方 Limit Reached: Try Again in 24 Hours，立即停止批次并保留当前及后续输入。
9. browser-use 实测 Account Not Found、全数据清理、about:blank 重开及官方九次限制；
   Retrieve 结果后的 Cancel 使用真实按钮结构并由 Playwright DOM 回归覆盖。
10. 新增统一一键安装启动 CMD；依赖存在就跳过，缺少时才安装。

2026-08-07 第九步源码校验更新：
1. 使用 browser-use 独立核验 StudentAid 真实页面字段、Continue 按钮和提交状态时间线。
2. 删除“姓名输入框消失即判定可找回”的宽泛条件；只有明确结果标志才允许完成，
   防止页面跳转或短暂重绘被误判为成功。
3. 页面结果等待改为可停止的短轮询；保留 Loading 等待，并准确识别站点错误。
4. CDP 模式复用 Chrome 默认浏览器上下文并新开页面，保持与 browser-use 实测一致；
   输入改用真实键盘事件、字段失焦和提交前稳定等待，修复过快 fill 导致站点返回 unknown error。
5. 检测到因页面停留过久产生的 unknown error 时自动新开页面、重新填表并提交一次。
6. 每条记录实时输出“打开页面、填写资料、已提交、等待结果、已识别结果”等安全阶段日志；
   长时间等待时每 5 秒输出心跳，最终状态实时写 SQLite 并刷新 GUI，不记录原始资料。
7. Playwright 仍是正式运行根基，browser-use 仅用于独立校验。

2026-08-07 第八步实测更新：
1. 保留第七步的 GUI、SQLite WAL、并发处理和 CSV 导出能力。
2. 新增 Chrome CDP 连接模式：优先复用已通过站点网络校验的本机 Chrome，
   解决新启动 Playwright 浏览器访问 StudentAid 时的 HTTP/2 协议错误。
3. 支持 STUDENTAID_CDP_URL 显式指定 CDP 地址；未设置时自动探测
   http://127.0.0.1:9223，设置为 off/none/disabled 可关闭自动探测。
4. 新增页面导航重试、共享浏览器安全断开和浏览器模式日志。
5. 修复输入页标题被误判为找回成功的问题；现在会等待 Loading 结束，
   并把站点未知错误准确记录为失败。
6. 修复纯文本 CSV 中全数字 SSN 被 float 转换吞掉前导 0 的问题。
7. 第八步使用测试资料完成 GUI、浏览器、SQLite 和导出链路实测。

2026-08-07 第七步更新（原 ait2.py 日期记录受本机时钟影响）：
1. 新增 Tkinter GUI，可选择 CSV/SCV/TXT/XLSX 输入文件和结果目录。
2. 导入资料先进入 SQLite；数据库固定启用 WAL。
3. 使用多个独立 Playwright 处理线程和一个专用 SQLite 写入线程。
4. 支持开始、停止、线程数设置和实时进度；停止后不再领取新任务。
5. 每个批次完成或停止后自动导出 UTF-8 BOM CSV。
6. 兼容无表头的 ``SSN,月,日,年,姓,名,地址``、旧版
   ``SSN,DOB,First Name,Last Name,Address``，以及带常见英文表头的资料。

旧版 ait.py 保持不变。本文件不把 SSN 或整条原始资料打印到日志。
"""

from __future__ import annotations

import csv
from bisect import bisect_left
from collections import deque
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import queue
import re
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
from typing import Any, Callable, Iterable, Mapping, Sequence
import urllib.request
import uuid
import zipfile

sys.dont_write_bytecode = True

try:
    from playwright.sync_api import Browser, Page, Playwright, sync_playwright
except ImportError:  # GUI 仍可启动，并在开始处理时给出明确错误。
    Browser = Page = Playwright = Any  # type: ignore[assignment,misc]
    sync_playwright = None

try:
    import requests
except ImportError:  # HTTP 后端选中时再给出明确依赖错误，GUI 和浏览器后端仍可启动。
    requests = None  # type: ignore[assignment]


APP_TITLE = "StudentAid 批量处理工具 - 第三十七步 TXT 分隔符与无表头身份列修复版"
DATABASE_FILENAME = "studentaid.sqlite3"
CUMULATIVE_OUTPUT_FILENAME = "StudentAid累计结果.csv"
GUI_CONFIG_FILENAME = "studentaid_gui_config.json"
LIMIT_REACHED_HEADING = "Limit Reached: Try Again in 24 Hours"
ACCOUNT_DISABLED_HEADING = "Your Account Is Disabled"
ACCOUNT_LOOKUP_ISSUE_HEADING = "Account Lookup Issue: Get Help"
ACCOUNT_RECOVERY_IN_PROGRESS_HEADING = "Account Recovery In Progress"
PHOTO_ID_RECOVERY_MESSAGE = (
    "We are unable to retrieve your log-in information. "
    "Access your account by recovering your account with a photo ID."
)
INVALID_SSN_MESSAGE = "Enter a valid Social Security number."
ACCOUNT_NOT_FOUND_CREATE_HEADING = "Account Not Found: Create a New Account"
KNOWN_RESULT_HEADINGS = frozenset(
    {
        "Retrieve Your Log-in Information",
        "Account Not Found",
        ACCOUNT_NOT_FOUND_CREATE_HEADING,
        LIMIT_REACHED_HEADING,
        ACCOUNT_DISABLED_HEADING,
        ACCOUNT_LOOKUP_ISSUE_HEADING,
        ACCOUNT_RECOVERY_IN_PROGRESS_HEADING,
        PHOTO_ID_RECOVERY_MESSAGE,
        INVALID_SSN_MESSAGE,
    }
)
CUMULATIVE_RESULT_FIELD_COUNT = 4
CUMULATIVE_RESULT_COLUMNS = (
    "Masked_Phone",
    "Masked_Email",
    "Recovery_Method",
    "Result_Heading",
)
BROWSER_BACKENDS = ("browser-use", "playwright")
PROCESSING_BACKENDS = ("http", *BROWSER_BACKENDS)
DISPLAY_MODES = ("窗口", "无头")
MAX_QUEUE_ATTEMPTS = 3
INPUT_DELETE_BATCH_SIZE = 20
INPUT_DELETE_FLUSH_SECONDS = 5.0
SMALL_WINDOW_SIZE = {"width": 900, "height": 720}
SMALL_VIEWPORT_SIZE = {"width": 880, "height": 650}
RETRIEVE_ACCOUNT_DETAILS_URL = (
    "https://studentaid.gov/fsa-id/sign-in/retrieve-account-details"
)
HTTP_KEEP_SESSION_ALIVE_URL = "https://studentaid.gov/app/keepSessionAlive.action"
HTTP_SESSION_USER_URL = "https://studentaid.gov/app/sessionUser.action"
HTTP_GATEWAY_SESSION_URL = "https://studentaid.gov/app/api/gateway/session"
HTTP_ACCOUNT_LOOKUP_URL = "https://studentaid.gov/app/api/auth/f/a"
HTTP_ORIGIN = "https://studentaid.gov"
HTTP_TIMEOUT = (15, 60)
HTTP_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/140.0.0.0 Safari/537.36"
)


def gui_config_path() -> Path:
    """源码和 EXE 都把 GUI 配置放在实际运行程序同目录。"""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent / GUI_CONFIG_FILENAME
    return Path(__file__).resolve().parent / GUI_CONFIG_FILENAME


def load_gui_config(config_path: Path | None = None) -> dict[str, str]:
    """读取上次 GUI 值；缺失、损坏或不合法字段回退默认值。"""
    defaults = {
        "input_path": "",
        "output_path": str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME),
        "backend": "browser-use",
        "display_mode": "无头",
        "thread_count": "2",
    }
    path = (config_path or gui_config_path()).resolve()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            return defaults
    except (OSError, ValueError, TypeError):
        return defaults

    result = dict(defaults)
    for key in ("input_path", "output_path"):
        value = payload.get(key)
        if isinstance(value, str):
            result[key] = value
    backend = payload.get("backend")
    if isinstance(backend, str) and backend in PROCESSING_BACKENDS:
        result["backend"] = backend
    display_mode = payload.get("display_mode")
    if isinstance(display_mode, str) and display_mode in DISPLAY_MODES:
        result["display_mode"] = display_mode
    try:
        thread_count = int(str(payload.get("thread_count", "2")).strip())
    except (TypeError, ValueError):
        thread_count = 2
    if thread_count >= 1:
        result["thread_count"] = str(thread_count)
    return result


def save_gui_config(config: Mapping[str, Any], config_path: Path | None = None) -> None:
    """原子保存 GUI 值；保存失败不阻断 GUI 或批处理。"""
    path = (config_path or gui_config_path()).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "input_path": str(config.get("input_path", "")),
        "output_path": str(config.get("output_path", "")),
        "backend": str(config.get("backend", "browser-use")),
        "display_mode": str(config.get("display_mode", "无头")),
        "thread_count": str(config.get("thread_count", "2")),
    }
    temporary = path.with_name(f".{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        temporary.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        os.replace(temporary, path)
    except OSError:
        pass
    finally:
        if temporary.exists():
            temporary.unlink()

FIRST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameFirstName"
LAST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameLastName"
BIRTH_MONTH_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthMonth"
BIRTH_DAY_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthDay"
BIRTH_YEAR_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthYear"
SSN_SELECTOR = "#fsa_Input_ForgotUsernameSsnInput"

RESULT_COLUMNS = (
    "Masked_Phone",
    "Masked_Email",
    "Recovery_Method",
    "Processing_Status",
    "Error",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
    "Result_Heading",
)


@dataclass(frozen=True)
class ProgressMetrics:
    """GUI 一次刷新所需的完整批次效率快照。"""

    total: int
    pending: int
    processing: int
    completed: int
    failed: int
    stopped: int
    terminal: int
    remaining: int
    percent: float
    recent_minute_count: int
    average_per_minute: float
    elapsed_seconds: float
    eta_seconds: float | None


class BatchProgressTracker:
    """根据数据库状态快照计算滚动一分钟速度、平均速度和 ETA。"""

    def __init__(self, clock: Callable[[], float] = time.monotonic) -> None:
        self._clock = clock
        self._counts = {
            "total": 0,
            "pending": 0,
            "processing": 0,
            "completed": 0,
            "failed": 0,
            "stopped": 0,
        }
        self._recent_events: deque[tuple[float, int]] = deque()
        self._started_at: float | None = None
        self._rate_started_at: float | None = None
        self._finished_at: float | None = None
        self._last_terminal: int | None = None
        self._processed_since_baseline = 0

    def start(self, now: float | None = None) -> None:
        current = self._clock() if now is None else float(now)
        for key in self._counts:
            self._counts[key] = 0
        self._recent_events.clear()
        self._started_at = current
        self._rate_started_at = None
        self._finished_at = None
        self._last_terminal = None
        self._processed_since_baseline = 0

    def finish(self, now: float | None = None) -> None:
        if self._started_at is None or self._finished_at is not None:
            return
        self._finished_at = self._clock() if now is None else float(now)

    def update(
        self, payload: Mapping[str, Any], now: float | None = None
    ) -> ProgressMetrics:
        current = self._clock() if now is None else float(now)
        if self._started_at is None:
            self.start(current)
        new_counts = {
            key: max(0, int(payload.get(key, 0))) for key in self._counts
        }
        terminal = (
            new_counts["completed"]
            + new_counts["failed"]
            + new_counts["stopped"]
        )
        # 多个 worker 会并发读取进度；较旧快照偶尔稍后到达，不能让 GUI 倒退。
        if self._last_terminal is not None and terminal < self._last_terminal:
            return self.snapshot(current)
        self._counts.update(new_counts)
        if self._last_terminal is None:
            # 首个快照可能已经包含导入格式错误，只作为速度基线而不计入吞吐量。
            self._last_terminal = terminal
            self._rate_started_at = current
        elif terminal > self._last_terminal:
            delta = terminal - self._last_terminal
            self._recent_events.append((current, delta))
            self._processed_since_baseline += delta
            self._last_terminal = terminal
        self._discard_old_events(current)
        return self.snapshot(current)

    def _discard_old_events(self, now: float) -> None:
        cutoff = now - 60.0
        while self._recent_events and self._recent_events[0][0] < cutoff:
            self._recent_events.popleft()

    def snapshot(self, now: float | None = None) -> ProgressMetrics:
        current = self._clock() if now is None else float(now)
        effective_now = self._finished_at if self._finished_at is not None else current
        self._discard_old_events(effective_now)
        total = self._counts["total"]
        terminal = (
            self._counts["completed"]
            + self._counts["failed"]
            + self._counts["stopped"]
        )
        terminal = min(total, terminal) if total else terminal
        remaining = max(0, total - terminal)
        percent = min(100.0, terminal * 100.0 / total) if total else 0.0
        recent_count = sum(delta for _event_time, delta in self._recent_events)
        elapsed = (
            max(0.0, effective_now - self._started_at)
            if self._started_at is not None
            else 0.0
        )
        rate_elapsed = (
            max(0.0, effective_now - self._rate_started_at)
            if self._rate_started_at is not None
            else 0.0
        )
        average_rate = (
            self._processed_since_baseline * 60.0 / rate_elapsed
            if self._processed_since_baseline and rate_elapsed > 0
            else 0.0
        )
        if remaining == 0 and total > 0:
            eta_seconds: float | None = 0.0
        elif average_rate > 0:
            eta_seconds = remaining * 60.0 / average_rate
        else:
            eta_seconds = None
        return ProgressMetrics(
            total=total,
            pending=self._counts["pending"],
            processing=self._counts["processing"],
            completed=self._counts["completed"],
            failed=self._counts["failed"],
            stopped=self._counts["stopped"],
            terminal=terminal,
            remaining=remaining,
            percent=percent,
            recent_minute_count=recent_count,
            average_per_minute=average_rate,
            elapsed_seconds=elapsed,
            eta_seconds=eta_seconds,
        )


def format_duration(seconds: float | None) -> str:
    """把 GUI 用秒数格式化为紧凑且稳定的中文时长。"""
    if seconds is None:
        return "--"
    total_seconds = max(0, int(float(seconds) + 0.999))
    days, remainder = divmod(total_seconds, 86400)
    hours, remainder = divmod(remainder, 3600)
    minutes, seconds_part = divmod(remainder, 60)
    clock_text = f"{hours:02d}:{minutes:02d}:{seconds_part:02d}"
    return f"{days}天 {clock_text}" if days else clock_text


@dataclass(frozen=True)
class AccountDetails:
    """StudentAid 账户找回页面所需字段。"""

    first_name: str
    last_name: str
    birth_month: str
    birth_day: str
    birth_year: str
    ssn: str
    original_ssn: str


@dataclass(frozen=True)
class RecoveryResult:
    """页面返回的可公开保存结果；联系方式保持页面原有脱敏形式。"""

    result_code: str
    heading: str
    masked_phone: str
    masked_email: str
    recovery_method: str


@dataclass(frozen=True)
class ImportedRecord:
    """一条已读取的输入资料；校验失败的行也会进入数据库。"""

    source_file: str
    source_sheet: str
    source_row: int
    original_fields: tuple[str, ...]
    details: AccountDetails | None
    address: str
    import_error: str = ""
    input_mapping: tuple[tuple[str, int], ...] = ()

    @property
    def record_key(self) -> str:
        if self.details is None:
            return ""
        return _input_row_key(
            self.details,
            self.original_fields,
            self.address,
            self.input_mapping,
        )


@dataclass(frozen=True)
class WorkItem:
    """SQLite、浏览器、累计输出和输入删行之间传递的任务。"""

    record_id: int
    details: AccountDetails
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    original_fields: tuple[str, ...] = ()
    address: str = ""
    input_mapping: tuple[tuple[str, int], ...] = ()

    @property
    def record_key(self) -> str:
        return _input_row_key(
            self.details,
            self.original_fields,
            self.address,
            self.input_mapping,
        )


class StopRequested(RuntimeError):
    """在网页处理步骤之间收到停止请求。"""


class PageSessionExpired(RuntimeError):
    """StudentAid 页面停留过久后返回 unknown error，需要重新打开页面。"""


class PageSubmissionStalled(RuntimeError):
    """Continue 已提交但站点长时间停在 Loading，需要重建浏览器会话。"""


class MissingRequiredField(ValueError):
    """输入源行缺少用户指定的 StudentAid 必填字段，应直接删除。"""


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _clean_error(exc: BaseException | str, limit: int = 1000) -> str:
    """生成适合写库和显示的单行错误，不包含输入资料。"""
    text = str(exc).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit] or type(exc).__name__


def _normalise_month(value: str) -> str:
    aliases = {
        "january": "01", "jan": "01", "february": "02", "feb": "02",
        "march": "03", "mar": "03", "april": "04", "apr": "04",
        "may": "05", "june": "06", "jun": "06", "july": "07",
        "jul": "07", "august": "08", "aug": "08", "september": "09",
        "sep": "09", "sept": "09", "october": "10", "oct": "10",
        "november": "11", "nov": "11", "december": "12", "dec": "12",
    }
    cleaned = value.strip().lower()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    month = aliases.get(cleaned) or (cleaned.zfill(2) if cleaned.isdigit() else "")
    if month not in {f"{number:02d}" for number in range(1, 13)}:
        raise ValueError("出生月份必须是 1-12、01-12 或英文月份名称")
    return month


def _normalise_day(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if not digits or not 1 <= int(digits) <= 31:
        raise ValueError("出生日期（日）必须是 1-31")
    return digits.zfill(2)


def _normalise_year(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) != 4 or not 1900 <= int(digits) <= date.today().year:
        raise ValueError("出生年份必须是四位数字，且不晚于当前年份")
    return digits


def _normalise_ssn(value: str) -> str:
    cleaned = value.strip()
    # 只移除 Excel 数字单元格常见的 .0 后缀；纯数字文本必须保留前导 0。
    if re.fullmatch(r"\d+\.0+", cleaned):
        cleaned = cleaned.split(".", 1)[0]
    digits = re.sub(r"\D", "", cleaned)
    # Excel 把首位 0 当数字格式丢掉时，允许 8 位数字恢复为 9 位。
    if len(digits) == 8:
        digits = digits.zfill(9)
    if len(digits) != 9:
        raise ValueError("Social Security Number 必须包含 9 位数字")
    return digits


def _validate_details(
    ssn_value: str,
    month_value: str,
    day_value: str,
    year_value: str,
    first_name: str,
    last_name: str,
) -> AccountDetails:
    ssn = _normalise_ssn(ssn_value)
    month = _normalise_month(month_value)
    day = _normalise_day(day_value)
    year = _normalise_year(year_value)
    first_name = first_name.strip()
    last_name = last_name.strip()

    if not first_name or len(first_name) > 35:
        raise ValueError("First Name 不能为空且长度不能超过 35 个字符")
    if not last_name or len(last_name) > 35:
        raise ValueError("Last Name 不能为空且长度不能超过 35 个字符")
    try:
        date(int(year), int(month), int(day))
    except ValueError as exc:
        raise ValueError("出生日期不是有效日期") from exc

    return AccountDetails(
        first_name=first_name,
        last_name=last_name,
        birth_month=month,
        birth_day=day,
        birth_year=year,
        ssn=ssn,
        original_ssn=ssn_value,
    )


def _parse_dob(value: str) -> tuple[str, str, str]:
    value = value.strip()
    compact_date = re.fullmatch(r"\d{8}", value)
    if compact_date:
        compact_formats: list[str] = []
        if 1900 <= int(value[:4]) <= date.today().year:
            compact_formats.append("%Y%m%d")
        if 1900 <= int(value[-4:]) <= date.today().year:
            compact_formats.append("%m%d%Y")
        for date_format in compact_formats:
            try:
                parsed = datetime.strptime(value, date_format).date()
                return f"{parsed.month:02d}", f"{parsed.day:02d}", f"{parsed.year:04d}"
            except ValueError:
                continue
    iso_prefix = re.match(r"^(\d{4}-\d{1,2}-\d{1,2})[T\s]", value)
    if iso_prefix:
        value = iso_prefix.group(1)
    short_year = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{3})", value)
    if short_year:
        month, day, year_tail = short_year.groups()
        candidates: list[date] = []
        for restored_year in (int(f"1{year_tail}"), int(f"{year_tail}0")):
            if not 1900 <= restored_year <= date.today().year:
                continue
            try:
                candidates.append(date(restored_year, int(month), int(day)))
            except ValueError:
                continue
        if len(candidates) == 1:
            parsed = candidates[0]
            return (
                f"{parsed.month:02d}",
                f"{parsed.day:02d}",
                f"{parsed.year:04d}",
            )
    for date_format in (
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m.%d.%Y", "%Y.%m.%d", "%Y年%m月%d日",
        "%m/%d/%y", "%m-%d-%y", "%B %d, %Y", "%b %d, %Y",
        "%B %d %Y", "%b %d %Y",
    ):
        try:
            parsed = datetime.strptime(value, date_format).date()
            return f"{parsed.month:02d}", f"{parsed.day:02d}", f"{parsed.year:04d}"
        except ValueError:
            continue
    raise ValueError(
        "DOB 格式无效，应为常见分隔日期、YYYYMMDD、ISO 日期或英文月份日期"
    )


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalise_header(value: str) -> str:
    """忽略分隔符和大小写，但保留中文表头供新格式识别。"""
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", value.strip().casefold())


HEADER_ALIASES: Mapping[str, set[str]] = {
    "result_ssn": {
        "resultssn", "matchedssn", "foundssn", "verifiedssn", "searchresultssn",
        "结果ssn", "查询结果ssn", "匹配ssn",
    },
    "result_dob": {
        "resultdob", "resultdateofbirth", "resultbirthdate", "resultbirthday",
        "matcheddob", "matchedbirthdate", "founddob", "verifieddob",
        "结果出生日期", "查询结果出生日期", "匹配出生日期", "结果生日",
    },
    "result_first_name": {
        "resultfirstname", "resultgivenname", "matchedfirstname", "foundfirstname",
        "verifiedfirstname", "结果名", "查询结果名", "匹配名",
    },
    "result_last_name": {
        "resultlastname", "resultsurname", "resultfamilyname", "matchedlastname",
        "foundlastname", "verifiedlastname", "结果姓", "查询结果姓", "匹配姓",
    },
    "result_full_name": {
        "resultfullname", "resultname", "matchedfullname", "foundfullname",
        "verifiedfullname", "结果姓名", "查询结果姓名", "匹配姓名",
    },
    "result_address": {
        "resultaddress", "matchedaddress", "foundaddress", "verifiedaddress",
        "结果地址", "查询结果地址", "匹配地址",
    },
    "ssn": {
        "ssn", "socialsecuritynumber", "socialsecurityno", "socialsecurity",
        "社会安全号", "社会保障号", "美国社会安全号",
    },
    "dob": {
        "dob", "dateofbirth", "birthdate", "knownbirthday", "出生日期",
    },
    "month": {"birthmonth", "dobmonth", "month", "mm"},
    "day": {"birthday", "dobday", "day", "dd"},
    "year": {"birthyear", "dobyear", "year", "yyyy"},
    "first_name": {"firstname", "givenname", "fname", "first", "名"},
    "last_name": {"lastname", "surname", "familyname", "lname", "last", "姓"},
    "address": {
        "address", "streetaddress", "mailingaddress", "homeaddress",
        "currentaddress", "地址", "现住址", "当前地址",
    },
    "full_name": {"fullname", "legalname", "personname", "姓名"},
    "preferred_dob": {"生日"},
}


def _semantic_header_kind(cleaned: str) -> str:
    """对未列出的表头按语义推断字段，兼容供应方任意前后缀。"""
    if not cleaned:
        return ""
    is_result = any(
        marker in cleaned
        for marker in (
            "result", "matched", "verified", "found", "searchresult",
            "结果", "查询结果", "匹配",
        )
    )
    has_ssn = "ssn" in cleaned or "socialsecurity" in cleaned or "社会安全" in cleaned
    has_dob = (
        "dob" in cleaned or "dateofbirth" in cleaned or "birthdate" in cleaned
        or "出生日期" in cleaned
    )
    has_first = "firstname" in cleaned or "givenname" in cleaned
    has_last = (
        "lastname" in cleaned or "surname" in cleaned or "familyname" in cleaned
    )
    has_full = "fullname" in cleaned or "legalname" in cleaned or "personname" in cleaned
    has_address = "address" in cleaned or "地址" in cleaned or "住址" in cleaned
    if has_ssn:
        return "result_ssn" if is_result else "ssn"
    if has_dob:
        return "result_dob" if is_result else "dob"
    if has_first:
        return "result_first_name" if is_result else "first_name"
    if has_last:
        return "result_last_name" if is_result else "last_name"
    if has_full:
        return "result_full_name" if is_result else "full_name"
    if has_address:
        return "result_address" if is_result else "address"
    return ""


def _header_mapping(row: Sequence[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        cleaned = _normalise_header(value)
        matched = False
        for canonical, aliases in HEADER_ALIASES.items():
            if cleaned in aliases and canonical not in mapping:
                mapping[canonical] = index
                matched = True
                break
        if not matched:
            canonical = _semantic_header_kind(cleaned)
            if canonical and canonical not in mapping:
                mapping[canonical] = index
    has_date = (
        "result_dob" in mapping
        or "preferred_dob" in mapping
        or "dob" in mapping
        or {"month", "day", "year"}.issubset(mapping)
    )
    has_name = (
        "result_full_name" in mapping
        or {"result_first_name", "result_last_name"}.issubset(mapping)
        or "full_name" in mapping
        or {"first_name", "last_name"}.issubset(mapping)
    )
    if ("result_ssn" in mapping or "ssn" in mapping) and has_date and has_name:
        return mapping
    return None


_NON_NAME_VALUES = {
    "male", "female", "unknown", "other", "yes", "no", "true", "false",
    "aries", "taurus", "gemini", "cancer", "leo", "virgo", "libra",
    "scorpio", "sagittarius", "capricorn", "aquarius", "pisces",
}

_US_STATE_ABBREVIATIONS = frozenset({
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
    "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
    "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY",
    "DC",
})


def _looks_like_headerless_ssn(value: str) -> bool:
    text = value.strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return len(re.sub(r"\D", "", text)) == 9 and not re.search(r"[A-Za-z@]", text)


def _looks_like_headerless_dob(value: str) -> bool:
    try:
        _parse_dob(value)
        return True
    except ValueError:
        return False


def _looks_like_headerless_name(value: str) -> bool:
    text = value.strip()
    if (
        not text
        or len(text) > 35
        or text.casefold() in _NON_NAME_VALUES
        or text in _US_STATE_ABBREVIATIONS
    ):
        return False
    if any(character.isdigit() for character in text) or "@" in text:
        return False
    return bool(re.fullmatch(r"[A-Za-zÀ-ÖØ-öø-ÿ'’ .-]+", text))


def _looks_like_headerless_address(value: str) -> bool:
    text = value.strip().casefold()
    if not text or "@" in text or not any(character.isdigit() for character in text):
        return False
    address_words = (
        " street", " st", " road", " rd", " avenue", " ave", " boulevard",
        " blvd", " drive", " dr", " lane", " ln", " court", " ct", " way",
        " place", " pl", " highway", " hwy", " circle", " cir", " terrace",
        " trail", " trl", " parkway", " pkwy", " apartment", " apt", " suite",
    )
    return text.count(",") >= 2 or any(word in f" {text}" for word in address_words)


def _column_match_ratio(
    rows: Sequence[Sequence[str]],
    column: int,
    predicate: Callable[[str], bool],
) -> float:
    values = [row[column].strip() for row in rows if column < len(row) and row[column].strip()]
    if not values:
        return 0.0
    return sum(1 for value in values if predicate(value)) / len(values)


def _infer_headerless_mapping(rows: Sequence[Sequence[str]]) -> dict[str, int] | None:
    """用多行统计推断无表头宽表；低置信度时继续交给原有位置解析。"""
    sample = [list(row) for row in rows[:200] if any(str(value).strip() for value in row)]
    if not sample:
        return None
    maximum_columns = max(len(row) for row in sample)
    ssn_scores = {
        index: _column_match_ratio(sample, index, _looks_like_headerless_ssn)
        for index in range(maximum_columns)
    }
    dob_scores = {
        index: _column_match_ratio(sample, index, _looks_like_headerless_dob)
        for index in range(maximum_columns)
    }
    ssn_index, ssn_score = max(ssn_scores.items(), key=lambda item: item[1])
    dob_candidates = {
        index: score for index, score in dob_scores.items() if index != ssn_index
    }
    if not dob_candidates:
        return None
    dob_index, dob_score = max(dob_candidates.items(), key=lambda item: item[1])
    if ssn_score < 0.80 or dob_score < 0.80:
        return None

    # 已有两种无表头紧凑 DOB 顺序继续交给原解析器，避免改变历史行为。
    if ssn_index == 0 and dob_index in {1, 3}:
        return None

    name_scores = {
        index: _column_match_ratio(sample, index, _looks_like_headerless_name)
        for index in range(maximum_columns)
        if index not in {ssn_index, dob_index}
    }
    between = [
        index for index in range(min(ssn_index, dob_index) + 1, max(ssn_index, dob_index))
        if name_scores.get(index, 0.0) >= 0.55
    ]
    if len(between) >= 2:
        name_candidates = between
    else:
        high_confidence = [
            index for index, score in sorted(name_scores.items()) if score >= 0.80
        ]
        contiguous_groups: list[list[int]] = []
        for index in high_confidence:
            if contiguous_groups and index == contiguous_groups[-1][-1] + 1:
                contiguous_groups[-1].append(index)
            else:
                contiguous_groups.append([index])
        paired_groups = [group for group in contiguous_groups if len(group) >= 2]
        if paired_groups:
            name_candidates = max(
                paired_groups,
                key=lambda group: (
                    len(group),
                    sum(name_scores[index] for index in group),
                    -group[0],
                ),
            )
        else:
            name_candidates = high_confidence
    if len(name_candidates) < 2:
        return None
    first_name_index = name_candidates[0]
    last_name_index = name_candidates[-1]
    if first_name_index == last_name_index:
        return None

    mapping = {
        "ssn": ssn_index,
        "dob": dob_index,
        "first_name": first_name_index,
        "last_name": last_name_index,
    }
    address_scores = {
        index: _column_match_ratio(sample, index, _looks_like_headerless_address)
        for index in range(maximum_columns)
        if index not in mapping.values()
    }
    if address_scores:
        address_index, address_score = max(address_scores.items(), key=lambda item: item[1])
        if address_score >= 0.55:
            mapping["address"] = address_index
    return mapping


def _resolve_input_mapping(
    rows: Sequence[Sequence[str]],
) -> tuple[dict[str, int] | None, bool]:
    """统一解析显式表头或无表头列映射；返回值同时标记首行是否为表头。"""
    if not rows:
        return None, False
    explicit_mapping = _header_mapping(rows[0])
    if explicit_mapping is not None:
        return explicit_mapping, True
    return _infer_headerless_mapping(rows), False


def _freeze_mapping(mapping: Mapping[str, int] | None) -> tuple[tuple[str, int], ...]:
    return tuple(sorted((str(key), int(index)) for key, index in (mapping or {}).items()))


def _mapping_dict(mapping: Sequence[tuple[str, int]]) -> dict[str, int]:
    return {str(key): int(index) for key, index in mapping}


def _split_full_name(value: str) -> tuple[str, str]:
    """将新格式 full_name 稳定拆为 First Name 和 Last Name。"""
    parts = value.strip().split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def _mapped_value(row: Sequence[str], mapping: Mapping[str, int], key: str) -> str:
    index = mapping.get(key)
    return row[index].strip() if index is not None and index < len(row) else ""


def _row_looks_like_split_date(row: Sequence[str]) -> bool:
    if len(row) < 6:
        return False
    try:
        if row[1].strip():
            _normalise_month(row[1])
        if row[2].strip():
            _normalise_day(row[2])
        if row[3].strip():
            _normalise_year(row[3])
        return True
    except ValueError:
        return False


def _row_looks_like_compact_date(row: Sequence[str]) -> bool:
    """识别两种无表头紧凑布局，不改变任意顺序的带表头输入。"""
    if len(row) < 4:
        return False
    for index in (1, 3):
        try:
            if row[index].strip():
                _parse_dob(row[index])
                return True
        except ValueError:
            continue
    return False


def _raise_for_missing_required_fields(
    ssn: str,
    month: str,
    day: str,
    year: str,
    first_name: str,
    last_name: str,
) -> None:
    fields = (
        ("SSN", ssn),
        ("Month", month),
        ("Day", day),
        ("Year", year),
        ("First Name", first_name),
        ("Last Name", last_name),
    )
    missing = [name for name, value in fields if not str(value).strip()]
    if missing:
        raise MissingRequiredField("缺少必填字段：" + ", ".join(missing))


def _parse_input_row(
    fields: Sequence[str],
    header: Mapping[str, int] | None,
) -> tuple[AccountDetails, str]:
    row = [str(value) for value in fields]
    if header is not None:
        # 聚合/背调结果表优先使用最终匹配身份；结果值为空时再回退原始字段。
        ssn = (
            _mapped_value(row, header, "result_ssn")
            or _mapped_value(row, header, "ssn")
        )
        fallback_full_name = _mapped_value(row, header, "full_name")
        if fallback_full_name:
            fallback_first_name, fallback_last_name = _split_full_name(
                fallback_full_name
            )
        else:
            fallback_first_name = _mapped_value(row, header, "first_name")
            fallback_last_name = _mapped_value(row, header, "last_name")
        result_full_name = _mapped_value(row, header, "result_full_name")
        if result_full_name:
            result_first_name, result_last_name = _split_full_name(result_full_name)
        else:
            result_first_name = _mapped_value(row, header, "result_first_name")
            result_last_name = _mapped_value(row, header, "result_last_name")
        first_name = result_first_name or fallback_first_name
        last_name = result_last_name or fallback_last_name
        address = (
            _mapped_value(row, header, "result_address")
            or _mapped_value(row, header, "address")
        )
        # MyLife 新格式以人工/最终整理的中文生日为准，空白时回退 known_birthday。
        dob = (
            _mapped_value(row, header, "result_dob")
            or _mapped_value(row, header, "preferred_dob")
            or _mapped_value(row, header, "dob")
        )
        if dob:
            month, day, year = _parse_dob(dob)
        elif (
            "result_dob" in header
            or "preferred_dob" in header
            or "dob" in header
        ):
            month = day = year = ""
        else:
            month = _mapped_value(row, header, "month")
            day = _mapped_value(row, header, "day")
            year = _mapped_value(row, header, "year")
    elif _row_looks_like_split_date(row):
        # 当前测试资料：SSN, 月, 日, 年, Last Name, First Name, Address。
        ssn, month, day, year = row[:4]
        last_name, first_name = row[4:6]
        address = ",".join(row[6:]).strip()
    else:
        if len(row) < 4:
            raise ValueError(
                "至少需要 4 列：SSN、DOB、First Name、Last Name"
            )
        ssn = row[0]
        # 同时兼容两种五列顺序：
        # 1) SSN, DOB, First Name, Last Name, Address
        # 2) SSN, First Name, Last Name, DOB, Address
        second_column_error: ValueError | None = None
        try:
            if not row[1].strip():
                raise ValueError("DOB 为空")
            month, day, year = _parse_dob(row[1])
            first_name, last_name = row[2:4]
        except ValueError as exc:
            second_column_error = exc
            try:
                if not row[3].strip():
                    raise ValueError("DOB 为空")
                month, day, year = _parse_dob(row[3])
                first_name, last_name = row[1:3]
            except ValueError:
                if not row[1].strip():
                    month = day = year = ""
                    first_name, last_name = row[2:4]
                else:
                    raise second_column_error
        address = ",".join(row[4:]).strip()

    first_name = first_name.strip()
    last_name = last_name.strip()
    _raise_for_missing_required_fields(
        ssn, month, day, year, first_name, last_name
    )

    return _validate_details(
        ssn, month, day, year, first_name, last_name
    ), address


def _records_from_rows(
    rows: Iterable[Sequence[Any]],
    input_path: Path,
    sheet_name: str = "",
) -> list[ImportedRecord]:
    materialised: list[tuple[int, list[str]]] = []
    for row_number, raw_row in enumerate(rows, start=1):
        row = [_cell_to_text(value) for value in raw_row]
        while row and row[-1] == "":
            row.pop()
        if row and any(value.strip() for value in row):
            materialised.append((row_number, row))

    if not materialised:
        return []

    mapping, has_header = _resolve_input_mapping(
        [row for _row_number, row in materialised]
    )
    data_rows = materialised[1:] if has_header else materialised
    imported: list[ImportedRecord] = []
    for row_number, row in data_rows:
        details: AccountDetails | None = None
        address = ""
        import_error = ""
        try:
            details, address = _parse_input_row(row, mapping)
        except Exception as exc:
            import_error = _clean_error(exc)
        imported.append(
            ImportedRecord(
                source_file=str(input_path.resolve()),
                source_sheet=sheet_name,
                source_row=row_number,
                original_fields=tuple(row),
                details=details,
                address=address,
                import_error=import_error,
                input_mapping=_freeze_mapping(mapping),
            )
        )
    return imported


def _rows_for_text_delimiter(text: str, delimiter: str) -> list[list[str]]:
    return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def _delimiter_layout_score(text: str, delimiter: str) -> tuple[float, int]:
    rows = [
        row for row in _rows_for_text_delimiter(text, delimiter)
        if row and any(str(value).strip() for value in row)
    ]
    if not rows:
        return 0.0, 0
    widths: dict[int, int] = {}
    for row in rows:
        widths[len(row)] = widths.get(len(row), 0) + 1
    modal_width, modal_count = max(
        widths.items(), key=lambda item: (item[1], item[0])
    )
    return modal_count / len(rows), modal_width


def _choose_text_delimiter(text: str) -> str:
    """优先保留 Sniffer 结果；稳定的四列以上 Tab 表可覆盖字段内逗号误判。"""
    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            return "\t"

    if "\t" in sample and delimiter != "\t":
        tab_consistency, tab_width = _delimiter_layout_score(text, "\t")
        detected_consistency, detected_width = _delimiter_layout_score(text, delimiter)
        if (
            tab_width >= 4
            and tab_consistency >= 0.80
            and tab_consistency >= detected_consistency
            and tab_width > detected_width
        ):
            return "\t"
    return delimiter


def _read_text_rows(input_path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = input_path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8") from last_error

    delimiter = _choose_text_delimiter(text)
    return _rows_for_text_delimiter(text, delimiter)


def load_input_records(input_path: Path) -> list[ImportedRecord]:
    """读取 CSV/SCV/TXT/XLSX，返回包括校验失败行在内的导入记录。"""
    input_path = input_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    suffix = input_path.suffix.casefold()
    records: list[ImportedRecord] = []
    if suffix in {".csv", ".scv", ".txt"}:
        records.extend(_records_from_rows(_read_text_rows(input_path), input_path))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                records.extend(
                    _records_from_rows(
                        worksheet.iter_rows(values_only=True),
                        input_path,
                        worksheet.title,
                    )
                )
        finally:
            workbook.close()
    else:
        raise ValueError("只支持 .csv、.scv、.txt 和 .xlsx 输入文件")

    if not records:
        raise ValueError("输入文件中没有可导入的资料行")
    return records


_SOURCE_FILE_LOCK = threading.RLock()


def _normalise_record_key(value: Any) -> str:
    """输出第一列/输入第一列比较键；SSN 忽略连字符、空格和 Excel .0。"""
    text = _cell_to_text(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    digits = re.sub(r"\D", "", text)
    if 8 <= len(digits) <= 9:
        return digits.zfill(9)
    return text.casefold()


def _normalise_row_field(value: Any) -> str:
    """整行去重字段：保留字段值、大小写、空白和列边界，只做类型文本化。"""
    return _cell_to_text(value)


def _canonical_input_columns(
    details: AccountDetails,
    original_fields: Sequence[Any],
    address: str = "",
    input_mapping: Sequence[tuple[str, int]] = (),
) -> list[str]:
    """原样保留全部输入列；仅无原始行的内部兼容对象使用五项兜底。"""
    original = [_cell_to_text(value) for value in original_fields]
    if original:
        return original
    return [
        details.original_ssn or details.ssn,
        _format_output_dob(details),
        details.first_name,
        details.last_name,
        address,
    ]


def _input_row_key(
    details: AccountDetails,
    original_fields: Sequence[Any],
    address: str = "",
    input_mapping: Sequence[tuple[str, int]] = (),
) -> str:
    columns = _canonical_input_columns(
        details, original_fields, address, input_mapping
    )
    return json.dumps(
        [_normalise_row_field(value) for value in columns],
        ensure_ascii=False,
        separators=(",", ":"),
    )


def _detect_text_format(path: Path) -> tuple[str, str, list[list[str]]]:
    raw = path.read_bytes()
    encoding = "utf-8-sig"
    text = ""
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8")
    delimiter = _choose_text_delimiter(text)
    rows = _rows_for_text_delimiter(text, delimiter)
    return encoding, delimiter, rows


def _is_cumulative_result_heading(value: Any) -> bool:
    """判断累计 CSV 单元格是否为程序允许落盘的明确结果状态。"""
    return _cell_to_text(value).strip() in KNOWN_RESULT_HEADINGS


def _trim_cumulative_alignment_padding(values: Sequence[Any]) -> list[str]:
    """移除程序为固定结果列增加的尾部空格；输入导入阶段本就会移除这些空格。"""
    trimmed = [_cell_to_text(value) for value in values]
    while trimmed and trimmed[-1] == "":
        trimmed.pop()
    return trimmed


def _split_cumulative_output_row(
    row: Sequence[Any],
) -> tuple[list[str], list[str], bool]:
    """拆分累计行并统一返回 Phone/Email/Method/Heading；第三项表示旧布局。"""
    values = [_cell_to_text(value) for value in row]
    if len(values) < CUMULATIVE_RESULT_FIELD_COUNT:
        return _trim_cumulative_alignment_padding(values), ["", "", "", ""], False
    input_columns = values[:-CUMULATIVE_RESULT_FIELD_COUNT]
    tail = values[-CUMULATIVE_RESULT_FIELD_COUNT:]
    if _is_cumulative_result_heading(tail[-1]):
        result_columns = tail
        legacy_layout = False
    else:
        heading, masked_phone, masked_email, recovery_method = tail
        result_columns = [masked_phone, masked_email, recovery_method, heading]
        legacy_layout = True
    return (
        _trim_cumulative_alignment_padding(input_columns),
        result_columns,
        legacy_layout,
    )


def read_output_row_keys(output_path: Path) -> set[str]:
    """读取累计输出的完整输入部分；固定宽度补齐格和末尾结果列不参与去重。"""
    output_path = output_path.resolve()
    if not output_path.is_file() or output_path.stat().st_size == 0:
        return set()
    _encoding, _delimiter, rows = _detect_text_format(output_path)
    keys: set[str] = set()
    for row in rows:
        input_columns, _result_columns, _legacy_layout = (
            _split_cumulative_output_row(row)
        )
        if not input_columns or not any(value.strip() for value in input_columns):
            continue
        keys.add(
            json.dumps(
                [_normalise_row_field(value) for value in input_columns],
                ensure_ascii=False,
                separators=(",", ":"),
            )
        )
    return keys


def read_output_first_column_keys(output_path: Path) -> set[str]:
    """历史 API 兼容名称；第二十六步起返回完整输入行键。"""
    return read_output_row_keys(output_path)


def filter_records_already_output(
    records: Sequence[ImportedRecord], existing_keys: set[str]
) -> tuple[list[ImportedRecord], set[str], int]:
    """不改源文件地过滤累计输出已有整行，供大型 XLSX 快速恢复使用。"""
    remaining: list[ImportedRecord] = []
    matched_keys: set[str] = set()
    skipped_count = 0
    for record in records:
        record_key = record.record_key
        if record_key and record_key in existing_keys:
            matched_keys.add(record_key)
            skipped_count += 1
        else:
            remaining.append(record)
    return remaining, matched_keys, skipped_count


def _atomic_write_text_rows(
    path: Path, rows: Sequence[Sequence[Any]], encoding: str, delimiter: str
) -> None:
    temporary = path.with_name(f".{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding=encoding, newline="") as stream:
            writer = csv.writer(stream, delimiter=delimiter, lineterminator="\n")
            writer.writerows(rows)
            stream.flush()
            os.fsync(stream.fileno())
        _atomic_replace_with_retry(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _atomic_replace_with_retry(source: Path, target: Path) -> None:
    """Windows 上外部扫描器短暂占用文件时重试原子替换。"""
    for attempt in range(1, 13):
        try:
            os.replace(source, target)
            return
        except PermissionError:
            if attempt >= 12:
                raise
            time.sleep(min(0.75, 0.1 * attempt))


def _copy_worksheet_row(worksheet: Any, source_row: int, target_row: int) -> None:
    """向前复制一行的值和常用格式；调用方保证目标行不晚于来源行。"""
    for column in range(1, worksheet.max_column + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        target._value = source._value
        target.data_type = source.data_type
        target._style = copy(source._style)
        target.comment = copy(source.comment) if source.comment is not None else None
        target._hyperlink = (
            copy(source.hyperlink) if source.hyperlink is not None else None
        )

    source_dimension = worksheet.row_dimensions[source_row]
    target_dimension = worksheet.row_dimensions[target_row]
    for attribute in (
        "height", "hidden", "outlineLevel", "collapsed", "thickTop", "thickBot",
    ):
        setattr(target_dimension, attribute, getattr(source_dimension, attribute))


def _delete_worksheet_rows_efficiently(
    worksheet: Any, rows_to_delete: Iterable[int]
) -> int:
    """单遍压缩 XLSX 行，避免对每个分散行执行一次全表搬移。"""
    original_max_row = int(worksheet.max_row or 0)
    deleted_rows = {
        int(row_number)
        for row_number in rows_to_delete
        if 1 <= int(row_number) <= original_max_row
    }
    if not deleted_rows:
        return 0

    # openpyxl 的 delete_rows 本身也不重写合并区域引用；遇到合并表格时按连续区间
    # 从底部删除，保持与旧实现最接近的行为并减少搬移次数。
    if worksheet.merged_cells.ranges:
        groups: list[tuple[int, int]] = []
        for row_number in sorted(deleted_rows):
            if groups and row_number == groups[-1][1] + 1:
                groups[-1] = (groups[-1][0], row_number)
            else:
                groups.append((row_number, row_number))
        for start, end in reversed(groups):
            worksheet.delete_rows(start, end - start + 1)
        return len(deleted_rows)

    target_row = 1
    for source_row in range(1, original_max_row + 1):
        if source_row in deleted_rows:
            continue
        if source_row != target_row:
            _copy_worksheet_row(worksheet, source_row, target_row)
        target_row += 1

    trailing_count = original_max_row - target_row + 1
    if trailing_count > 0:
        worksheet.delete_rows(target_row, trailing_count)
    for row_number in list(worksheet.row_dimensions):
        if int(row_number) >= target_row:
            del worksheet.row_dimensions[row_number]
    return len(deleted_rows)


_XLSX_ROW_XML_PATTERN = re.compile(
    br"<row\b[^>]*(?:/>|>.*?</row>)", re.DOTALL
)
_XLSX_ROW_NUMBER_PATTERN = re.compile(br'<row\b[^>]*\br="(\d+)"')


def _remove_xlsx_rows_from_xml(xml: bytes, rows_to_delete: set[int]) -> bytes:
    """直接压缩 worksheet XML；保留样式、共享字符串和其余 ZIP 部件。"""
    deleted = sorted({int(value) for value in rows_to_delete if int(value) >= 1})
    if not deleted:
        return xml
    deleted_set = set(deleted)
    output: list[bytes] = []
    cursor = 0
    old_max_row = 0
    for match in _XLSX_ROW_XML_PATTERN.finditer(xml):
        output.append(xml[cursor:match.start()])
        row_blob = match.group(0)
        number_match = _XLSX_ROW_NUMBER_PATTERN.search(row_blob)
        if number_match is None:
            output.append(row_blob)
            cursor = match.end()
            continue
        old_row = int(number_match.group(1))
        old_max_row = max(old_max_row, old_row)
        if old_row not in deleted_set:
            new_row = old_row - bisect_left(deleted, old_row)
            if new_row != old_row:
                old_bytes = str(old_row).encode("ascii")
                new_bytes = str(new_row).encode("ascii")
                row_blob = re.sub(
                    br'(<row\b[^>]*\br=")' + old_bytes + br'(")',
                    lambda part: part.group(1) + new_bytes + part.group(2),
                    row_blob,
                    count=1,
                )
                row_blob = re.sub(
                    br'(<c\b[^>]*\br="[A-Z]{1,3})' + old_bytes + br'(")',
                    lambda part: part.group(1) + new_bytes + part.group(2),
                    row_blob,
                )
            output.append(row_blob)
        cursor = match.end()
    output.append(xml[cursor:])
    rewritten = b"".join(output)
    new_max_row = max(1, old_max_row - bisect_left(deleted, old_max_row + 1))
    rewritten = re.sub(
        br'(<dimension\b[^>]*\bref="(?:[A-Z]{1,3}\d+:)?[A-Z]{1,3})\d+(")',
        lambda part: (
            part.group(1) + str(new_max_row).encode("ascii") + part.group(2)
        ),
        rewritten,
        count=1,
    )
    return rewritten


def _atomic_remove_xlsx_rows(
    input_path: Path, sheet_rows: Mapping[str, set[int]]
) -> None:
    """原子重写指定 worksheet XML，并在替换源文件前验证新 XLSX 可读取。"""
    requested = {
        str(sheet_path).replace("\\", "/"): {
            int(row_number) for row_number in row_numbers if int(row_number) >= 1
        }
        for sheet_path, row_numbers in sheet_rows.items()
        if row_numbers
    }
    if not requested:
        return
    temporary = input_path.with_name(
        f".{input_path.stem}.{os.getpid()}.{uuid.uuid4().hex}.tmp.xlsx"
    )
    try:
        with zipfile.ZipFile(input_path, "r") as source_archive:
            source_names = set(source_archive.namelist())
            missing = sorted(set(requested) - source_names)
            if missing:
                raise ValueError("XLSX 工作表部件不存在：" + ", ".join(missing))
            with zipfile.ZipFile(temporary, "w") as target_archive:
                for info in source_archive.infolist():
                    content = source_archive.read(info.filename)
                    rows = requested.get(info.filename)
                    if rows:
                        content = _remove_xlsx_rows_from_xml(content, rows)
                    target_archive.writestr(info, content)

        from openpyxl import load_workbook

        validation = load_workbook(temporary, read_only=True, data_only=True)
        validation.close()
        _atomic_replace_with_retry(temporary, input_path)
    finally:
        if temporary.exists():
            temporary.unlink()


def remove_input_rows_by_keys(input_path: Path, keys: set[str]) -> int:
    """按完整输入资料行原子删除；XLSX 保留样式和其他工作表。"""
    keys = {str(value) for value in keys if str(value)}
    if not keys:
        return 0
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            kept: list[list[str]] = []
            removed = 0
            meaningful_rows = [
                (index, row) for index, row in enumerate(rows)
                if row and any(str(value).strip() for value in row)
            ]
            mapping, has_header = _resolve_input_mapping(
                [row for _index, row in meaningful_rows]
            )
            header_index = meaningful_rows[0][0] if has_header and meaningful_rows else -1
            for index, row in enumerate(rows):
                if index == header_index:
                    kept.append(row)
                    continue
                try:
                    details, address = _parse_input_row(row, mapping)
                    key = _input_row_key(
                        details, row, address, _freeze_mapping(mapping)
                    )
                except Exception:
                    key = ""
                if key and key in keys:
                    removed += 1
                else:
                    kept.append(row)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path, read_only=True, data_only=True)
            removed = 0
            sheet_rows: dict[str, set[int]] = {}
            try:
                for worksheet in workbook.worksheets:
                    sample_rows: list[tuple[int, list[str]]] = []
                    for sample_row_number, raw_values in enumerate(
                        worksheet.iter_rows(
                            min_row=1,
                            max_row=min(int(worksheet.max_row or 0), 200),
                            values_only=True,
                        ),
                        start=1,
                    ):
                        sample_values = [_cell_to_text(value) for value in raw_values]
                        while sample_values and sample_values[-1] == "":
                            sample_values.pop()
                        if sample_values and any(value.strip() for value in sample_values):
                            sample_rows.append((sample_row_number, sample_values))
                    mapping, has_header = _resolve_input_mapping(
                        [values for _row_number, values in sample_rows]
                    )
                    header_row_number = (
                        sample_rows[0][0] if has_header and sample_rows else -1
                    )
                    rows_to_delete: set[int] = set()
                    for row_number, raw_values in enumerate(
                        worksheet.iter_rows(values_only=True), start=1
                    ):
                        if row_number == header_row_number:
                            continue
                        values = [_cell_to_text(value) for value in raw_values]
                        while values and values[-1] == "":
                            values.pop()
                        try:
                            details, address = _parse_input_row(values, mapping)
                            key = _input_row_key(
                                details, values, address, _freeze_mapping(mapping)
                            )
                        except Exception:
                            key = ""
                        if key and key in keys:
                            rows_to_delete.add(row_number)
                    if rows_to_delete:
                        sheet_rows[str(worksheet._worksheet_path)] = rows_to_delete
                        removed += len(rows_to_delete)
            finally:
                workbook.close()
            if removed:
                _atomic_remove_xlsx_rows(input_path, sheet_rows)
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def remove_duplicate_input_rows(input_path: Path) -> int:
    """批次内仅删除完全相同的额外行，保留每个完整行键的一条代表记录。"""
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            mapping = _header_mapping(rows[0]) if rows else None
            seen: set[str] = set()
            kept: list[list[str]] = []
            removed = 0
            for row_number, row in enumerate(rows, start=1):
                if row_number == 1 and mapping is not None:
                    kept.append(row)
                    continue
                try:
                    details, address = _parse_input_row(row, mapping)
                    key = _input_row_key(
                        details, row, address, _freeze_mapping(mapping)
                    )
                except Exception:
                    kept.append(row)
                    continue
                if key in seen:
                    removed += 1
                else:
                    seen.add(key)
                    kept.append(row)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path, read_only=True, data_only=True)
            removed = 0
            sheet_rows: dict[str, set[int]] = {}
            try:
                for worksheet in workbook.worksheets:
                    row_iterator = worksheet.iter_rows(values_only=True)
                    first_values = next(row_iterator, ())
                    header = [_cell_to_text(value) for value in first_values]
                    mapping = _header_mapping(header)
                    first_data_row = 2 if mapping is not None else 1
                    seen: set[str] = set()
                    rows_to_delete: set[int] = set()
                    rows = worksheet.iter_rows(values_only=True)
                    for row_number, raw_values in enumerate(rows, start=1):
                        if row_number < first_data_row:
                            continue
                        values = [_cell_to_text(value) for value in raw_values]
                        while values and values[-1] == "":
                            values.pop()
                        try:
                            details, address = _parse_input_row(values, mapping)
                            key = _input_row_key(
                                details, values, address, _freeze_mapping(mapping)
                            )
                        except Exception:
                            continue
                        if key in seen:
                            rows_to_delete.add(row_number)
                        else:
                            seen.add(key)
                    if rows_to_delete:
                        sheet_rows[str(worksheet._worksheet_path)] = rows_to_delete
                        removed += len(rows_to_delete)
            finally:
                workbook.close()
            if removed:
                _atomic_remove_xlsx_rows(input_path, sheet_rows)
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def remove_input_rows_by_locations(
    input_path: Path, records: Sequence[ImportedRecord]
) -> int:
    """按导入时的工作表/源行精确删除，兼容第一列 SSN 本身为空的记录。"""
    locations = {
        (record.source_sheet, int(record.source_row))
        for record in records
        if record.source_row > 0
    }
    if not locations:
        return 0
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            kept = [
                row
                for row_number, row in enumerate(rows, start=1)
                if ("", row_number) not in locations
            ]
            removed = len(rows) - len(kept)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path, read_only=True, data_only=True)
            removed = 0
            sheet_rows: dict[str, set[int]] = {}
            try:
                for worksheet in workbook.worksheets:
                    rows_to_delete = {
                        row_number
                        for sheet_name, row_number in locations
                        if sheet_name == worksheet.title
                        and 1 <= row_number <= int(worksheet.max_row or 0)
                    }
                    if rows_to_delete:
                        sheet_rows[str(worksheet._worksheet_path)] = rows_to_delete
                        removed += len(rows_to_delete)
            finally:
                workbook.close()
            if removed:
                _atomic_remove_xlsx_rows(input_path, sheet_rows)
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def _format_output_dob(details: AccountDetails) -> str:
    return f"{int(details.birth_month):02d}/{int(details.birth_day):02d}/{details.birth_year}"


def build_cumulative_output_row(
    item: WorkItem,
    result: RecoveryResult,
    input_width: int | None = None,
) -> list[str]:
    """原样透传输入列，补齐到固定宽度，并把结果状态放在唯一最后一列。"""
    input_columns = _canonical_input_columns(
        item.details, item.original_fields, item.address, item.input_mapping
    )
    target_width = max(len(input_columns), int(input_width or 0))
    padded_input = [
        *input_columns,
        *("" for _ in range(target_width - len(input_columns))),
    ]
    return [
        *padded_input,
        result.masked_phone,
        result.masked_email,
        result.recovery_method,
        result.heading,
    ]


def append_cumulative_result(
    output_path: Path, item: WorkItem, result: RecoveryResult
) -> bool:
    """实时追加累计结果；自动扩展固定输入宽度，完整资料行不重复追加。"""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    required_width = len(
        _canonical_input_columns(
            item.details, item.original_fields, item.address, item.input_mapping
        )
    )
    with _SOURCE_FILE_LOCK:
        layout = normalise_cumulative_output_layout(
            output_path, minimum_input_width=required_width
        )
        existing_keys = read_output_row_keys(output_path)
        if item.record_key and item.record_key in existing_keys:
            return False
        with output_path.open("a", encoding="utf-8-sig", newline="") as stream:
            csv.writer(
                stream, lineterminator="\n", quoting=csv.QUOTE_ALL
            ).writerow(
                build_cumulative_output_row(item, result, layout.input_width)
            )
            stream.flush()
            os.fsync(stream.fileno())
    return True


def append_cumulative_result_cached(
    output_path: Path,
    item: WorkItem,
    result: RecoveryResult,
    existing_keys: set[str],
    input_width: int | None = None,
) -> bool:
    """使用批次完整行键实时追加；调用方可传入已缓存的固定输入宽度。"""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    record_key = item.record_key
    required_width = len(
        _canonical_input_columns(
            item.details, item.original_fields, item.address, item.input_mapping
        )
    )
    with _SOURCE_FILE_LOCK:
        if record_key and record_key in existing_keys:
            return False
        target_width = max(required_width, int(input_width or 0))
        if input_width is None or required_width > int(input_width or 0):
            target_width = normalise_cumulative_output_layout(
                output_path, minimum_input_width=required_width
            ).input_width
        with output_path.open("a", encoding="utf-8-sig", newline="") as stream:
            csv.writer(
                stream, lineterminator="\n", quoting=csv.QUOTE_ALL
            ).writerow(build_cumulative_output_row(item, result, target_width))
            stream.flush()
            os.fsync(stream.fileno())
        if record_key:
            existing_keys.add(record_key)
    return True


@dataclass
class _PersistenceCommand:
    item: WorkItem
    result: RecoveryResult
    response: queue.Queue[tuple[bool, Any]]


_NO_PERSISTENCE_COMMAND = object()


class ResultPersistenceWriter:
    """串行写累计 CSV，并合并输入文件删行。

    浏览器 worker 只同步等待累计 CSV flush/fsync。CSV/TXT 沿用定时批量删行；XLSX
    在所有 worker 结束后一次性压缩，避免大表重写占住本线程而饿死后续输出命令。
    异常退出时启动同步仍会按完整行清理已经安全输出的完全相同资料。
    """

    def __init__(
        self,
        input_path: Path,
        output_path: Path,
        existing_keys: Iterable[str] = (),
        *,
        initial_pending_delete_keys: Iterable[str] = (),
        batch_size: int = INPUT_DELETE_BATCH_SIZE,
        flush_seconds: float = INPUT_DELETE_FLUSH_SECONDS,
        log_callback: Callable[[str], None] | None = None,
    ) -> None:
        self.input_path = input_path.resolve()
        self.output_path = output_path.resolve()
        self.existing_keys = {str(value) for value in existing_keys if str(value)}
        self._output_input_width = normalise_cumulative_output_layout(
            self.output_path
        ).input_width
        self.batch_size = max(1, int(batch_size))
        self.flush_seconds = max(0.05, float(flush_seconds))
        self.log_callback = log_callback
        self._defer_input_flush = self.input_path.suffix.casefold() == ".xlsx"
        self._commands: queue.Queue[_PersistenceCommand | None] = queue.Queue()
        self._pending_delete_keys: set[str] = {
            str(value) for value in initial_pending_delete_keys if str(value)
        }
        self._pending_lock = threading.Lock()
        self._thread = threading.Thread(
            target=self._run,
            name="studentaid-result-persistence",
            daemon=False,
        )
        self._started = False
        self._closed = False
        self._close_error: BaseException | None = None
        self._pending_since: float | None = (
            time.monotonic() if self._pending_delete_keys else None
        )

    @property
    def pending_delete_count(self) -> int:
        with self._pending_lock:
            return len(self._pending_delete_keys)

    def _log(self, message: str) -> None:
        if self.log_callback is None:
            return
        try:
            self.log_callback(message)
        except Exception:
            pass

    def start(self) -> None:
        if self._started:
            return
        self._started = True
        self._thread.start()

    def commit(self, item: WorkItem, result: RecoveryResult) -> bool:
        """累计 CSV 安全落盘后返回；输入删行在本线程继续批量执行。"""
        if not self._started or self._closed:
            raise RuntimeError("结果持久化线程未运行")
        response: queue.Queue[tuple[bool, Any]] = queue.Queue(maxsize=1)
        self._commands.put(_PersistenceCommand(item, result, response))
        try:
            ok, value = response.get(timeout=300)
        except queue.Empty as exc:
            raise RuntimeError("累计输出写入等待超时") from exc
        if ok:
            return bool(value)
        if isinstance(value, BaseException):
            raise value
        raise RuntimeError(str(value))

    def _try_flush_pending(self) -> None:
        with self._pending_lock:
            keys = set(self._pending_delete_keys)
        if not keys:
            return
        try:
            deleted = remove_input_rows_by_keys(self.input_path, keys)
        except Exception as exc:
            self._close_error = exc
            self._pending_since = time.monotonic()
            self._log(
                f"输入文件批量删除暂时失败，保留 {len(keys)} 个待同步结果："
                f"{_clean_error(exc)}"
            )
            return
        with self._pending_lock:
            self._pending_delete_keys.difference_update(keys)
            self._pending_since = (
                time.monotonic() if self._pending_delete_keys else None
            )
        self._close_error = None
        self._log(
            f"输入文件已批量删除 {deleted} 条匹配资料；"
            f"本次合并 {len(keys)} 个明确结果。"
        )

    def _run(self) -> None:
        closing = False
        while not closing:
            try:
                command = self._commands.get(timeout=0.25)
            except queue.Empty:
                command = _NO_PERSISTENCE_COMMAND
            if command is None:
                self._commands.task_done()
                closing = True
            elif command is not _NO_PERSISTENCE_COMMAND:
                try:
                    required_width = len(
                        _canonical_input_columns(
                            command.item.details,
                            command.item.original_fields,
                            command.item.address,
                            command.item.input_mapping,
                        )
                    )
                    if required_width > self._output_input_width:
                        layout = normalise_cumulative_output_layout(
                            self.output_path, minimum_input_width=required_width
                        )
                        self._output_input_width = layout.input_width
                        self._log(
                            "累计 CSV 已扩展输入区到 "
                            f"{self._output_input_width} 列；结果状态固定在最后一列。"
                        )
                    appended = append_cumulative_result_cached(
                        self.output_path,
                        command.item,
                        command.result,
                        self.existing_keys,
                        self._output_input_width,
                    )
                    record_key = command.item.record_key
                    if record_key:
                        with self._pending_lock:
                            if not self._pending_delete_keys:
                                self._pending_since = time.monotonic()
                            self._pending_delete_keys.add(record_key)
                    command.response.put((True, appended))
                except BaseException as exc:
                    command.response.put((False, exc))
                finally:
                    self._commands.task_done()

            pending_count = self.pending_delete_count
            flush_due = bool(
                pending_count
                and self._pending_since is not None
                and time.monotonic() - self._pending_since >= self.flush_seconds
            )
            if (
                not self._defer_input_flush
                and (
                    pending_count >= self.batch_size
                    or (pending_count and flush_due)
                )
            ):
                self._try_flush_pending()

        if self._defer_input_flush and self.pending_delete_count:
            self._log(
                f"累计输出已全部写入；正在一次性同步输入 XLSX 中的 "
                f"{self.pending_delete_count} 个完整行结果……"
            )
        for attempt in range(3):
            if not self.pending_delete_count:
                break
            self._try_flush_pending()
            if self.pending_delete_count and attempt < 2:
                time.sleep(0.2 * (attempt + 1))
        if self.pending_delete_count and self._close_error is None:
            self._close_error = RuntimeError(
                f"仍有 {self.pending_delete_count} 个输入结果等待删除"
            )

    def close(self) -> None:
        if not self._started or self._closed:
            return
        self._closed = True
        self._commands.put(None)
        self._thread.join(timeout=300)
        if self._thread.is_alive():
            raise RuntimeError("结果持久化线程关闭超时")
        if self._close_error is not None and self.pending_delete_count:
            raise RuntimeError(
                f"累计结果已保存，但输入文件仍有 {self.pending_delete_count} 个结果"
                "等待下次启动同步删除"
            ) from self._close_error


@dataclass(frozen=True)
class CumulativeOutputLayoutResult:
    changed: bool
    row_count: int
    input_width: int
    legacy_rows: int = 0
    padded_rows: int = 0
    empty_rows: int = 0


def normalise_cumulative_output_layout(
    output_path: Path, minimum_input_width: int = 0
) -> CumulativeOutputLayoutResult:
    """原子统一累计 CSV：固定输入宽度，Phone/Email/Method 在前，Heading 唯一末列。"""
    output_path = output_path.resolve()
    minimum_input_width = max(0, int(minimum_input_width))
    if not output_path.is_file() or output_path.stat().st_size == 0:
        return CumulativeOutputLayoutResult(False, 0, minimum_input_width)
    with _SOURCE_FILE_LOCK:
        raw = output_path.read_text(encoding="utf-8-sig")
        first_nonempty = next((line for line in raw.splitlines() if line.strip()), "")
        _encoding, delimiter, rows = _detect_text_format(output_path)
        parsed_rows: list[tuple[list[str], list[str]]] = []
        input_width = minimum_input_width
        legacy_rows = 0
        empty_rows = 0
        for row in rows:
            if not any(_cell_to_text(value).strip() for value in row):
                empty_rows += 1
                continue
            input_columns, result_columns, legacy_layout = (
                _split_cumulative_output_row(row)
            )
            parsed_rows.append((input_columns, result_columns))
            input_width = max(input_width, len(input_columns))
            legacy_rows += int(legacy_layout)
        normalised_rows = [
            [
                *input_columns,
                *("" for _ in range(input_width - len(input_columns))),
                *result_columns,
            ]
            for input_columns, result_columns in parsed_rows
        ]
        padded_rows = sum(
            len(input_columns) < input_width
            for input_columns, _result_columns in parsed_rows
        )
        needs_quote_all = bool(first_nonempty) and not first_nonempty.startswith('"')
        changed = bool(
            empty_rows
            or legacy_rows
            or delimiter != ","
            or needs_quote_all
            or rows != normalised_rows
        )
        if not changed:
            return CumulativeOutputLayoutResult(
                False, len(normalised_rows), input_width, 0, padded_rows, 0
            )
        temporary = output_path.with_name(
            f".{output_path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp"
        )
        try:
            with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.writer(
                    stream, delimiter=",", lineterminator="\n", quoting=csv.QUOTE_ALL
                )
                writer.writerows(normalised_rows)
                stream.flush()
                os.fsync(stream.fileno())
            _atomic_replace_with_retry(temporary, output_path)
        finally:
            if temporary.exists():
                temporary.unlink()
        return CumulativeOutputLayoutResult(
            True, len(normalised_rows), input_width, legacy_rows, padded_rows, empty_rows
        )


def ensure_cumulative_output_quote_all(output_path: Path) -> bool:
    """历史 API：同时完成全字段引号、空行清理和结果状态末列统一。"""
    return normalise_cumulative_output_layout(output_path).changed


def resolve_output_target(output_target: Path) -> tuple[Path, Path]:
    """兼容旧版目录参数；GUI 第十四步直接选择累计 CSV。"""
    output_target = output_target.resolve()
    if output_target.suffix.casefold() in {".csv", ".scv", ".txt"}:
        return output_target.parent, output_target
    return output_target, output_target / CUMULATIVE_OUTPUT_FILENAME


@dataclass(frozen=True)
class DatabaseClearResult:
    database_path: Path
    existed: bool
    batches_deleted: int = 0
    records_deleted: int = 0


def _current_windows_account() -> str:
    """返回 icacls 可识别的当前 Windows 账户名。"""
    domain = os.environ.get("USERDOMAIN", "").strip()
    username = os.environ.get("USERNAME", "").strip()
    if username:
        return f"{domain}\\{username}" if domain else username
    completed = subprocess.run(
        ["whoami.exe"],
        stdin=subprocess.DEVNULL,
        stdout=subprocess.PIPE,
        stderr=subprocess.DEVNULL,
        check=False,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    account = os.fsdecode(completed.stdout).strip()
    if completed.returncode or not account:
        raise PermissionError("无法识别当前 Windows 账户")
    return account


def _run_icacls_grant(path: Path, account: str, rights: str) -> bool:
    """使用参数数组调用 icacls；返回是否成功，不经过 shell 拼接路径。"""
    startupinfo = None
    if os.name == "nt" and hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0
    completed = subprocess.run(
        [
            "icacls.exe",
            str(path),
            "/grant:r",
            f"{account}:{rights}",
            "/C",
        ],
        stdin=subprocess.DEVNULL,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        startupinfo=startupinfo,
    )
    return completed.returncode == 0


def ensure_output_storage_writable(
    output_directory: Path, *existing_paths: Path
) -> None:
    """修复 Windows 输出目录继承 ACL，并补齐已存在输出文件的当前账户权限。"""
    output_directory = output_directory.resolve()
    output_directory.mkdir(parents=True, exist_ok=True)
    if os.name != "nt":
        return

    account = _current_windows_account()
    if not _run_icacls_grant(output_directory, account, "(OI)(CI)F"):
        raise PermissionError(f"无法修复输出目录写权限：{output_directory}")

    # 先改父目录的可继承 ACE；未保护 ACL 的既有文件会立即获得继承权限。
    # 再给明确输出文件补一个直接 ACE，兼容曾经关闭继承的历史文件。
    for path in existing_paths:
        path = path.resolve()
        if path.exists():
            _run_icacls_grant(path, account, "F")


def database_path_for_output_target(output_target: Path) -> Path:
    output_directory, _output_path = resolve_output_target(output_target)
    return output_directory / DATABASE_FILENAME


def clear_database_contents(database_path: Path) -> DatabaseClearResult:
    """清空批次数据库但保留表结构；不接触累计输出和输入资料。"""
    database_path = database_path.resolve()
    if not database_path.is_file():
        return DatabaseClearResult(database_path, False)

    ensure_output_storage_writable(
        database_path.parent,
        database_path,
        Path(f"{database_path}-wal"),
        Path(f"{database_path}-shm"),
    )
    connection = sqlite3.connect(database_path, timeout=10.0)
    try:
        connection.execute("PRAGMA busy_timeout=10000")
        connection.execute("PRAGMA foreign_keys=ON")
        tables = {
            str(row[0])
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        records_deleted = (
            int(connection.execute("SELECT COUNT(*) FROM records").fetchone()[0])
            if "records" in tables else 0
        )
        batches_deleted = (
            int(connection.execute("SELECT COUNT(*) FROM batches").fetchone()[0])
            if "batches" in tables else 0
        )
        connection.execute("BEGIN IMMEDIATE")
        if "records" in tables:
            connection.execute("DELETE FROM records")
        if "batches" in tables:
            connection.execute("DELETE FROM batches")
        if "sqlite_sequence" in tables:
            connection.execute(
                "DELETE FROM sqlite_sequence WHERE name IN ('records', 'batches')"
            )
        connection.commit()
        connection.execute("PRAGMA wal_checkpoint(TRUNCATE)")
        connection.execute("VACUUM")
        return DatabaseClearResult(
            database_path, True, batches_deleted, records_deleted
        )
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def _normalise_display_mode(display_mode: str) -> str:
    display_mode = display_mode.strip()
    if display_mode not in DISPLAY_MODES:
        raise ValueError(f"浏览器显示模式必须是：{', '.join(DISPLAY_MODES)}")
    return display_mode


def _is_headless(display_mode: str) -> bool:
    return _normalise_display_mode(display_mode) == "无头"


def _normal_chrome_user_agent(browser_version: str) -> str:
    match = re.search(r"\d+(?:\.\d+){0,3}", str(browser_version))
    version = match.group(0) if match else "120.0.0.0"
    parts = version.split(".")
    version = ".".join((parts + ["0", "0", "0", "0"])[:4])
    return (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        f"Chrome/{version} Safari/537.36"
    )


def _chrome_process_ids(command_line_token: str = "") -> set[int]:
    try:
        import psutil

        process_ids: set[int] = set()
        for process in psutil.process_iter(["pid", "name", "cmdline"]):
            try:
                if (process.info["name"] or "").casefold() != "chrome.exe":
                    continue
                command_line = " ".join(process.info["cmdline"] or [])
                if command_line_token and command_line_token not in command_line:
                    continue
                process_ids.add(int(process.info["pid"]))
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        return process_ids
    except Exception:
        return set()


def _hide_windows_for_processes(process_ids: set[int]) -> int:
    if os.name != "nt" or not process_ids:
        return 0
    try:
        import ctypes
        import ctypes.wintypes

        user32 = ctypes.windll.user32
        hidden = 0
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p
        )

        @callback_type
        def callback(window_handle: int, _parameter: int) -> bool:
            nonlocal hidden
            process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(window_handle, ctypes.byref(process_id))
            if process_id.value in process_ids and user32.IsWindowVisible(window_handle):
                user32.ShowWindow(window_handle, 0)
                hidden += 1
            return True

        user32.EnumWindows(callback, 0)
        return hidden
    except Exception:
        return 0


@dataclass
class BrowserLaunch:
    """浏览器启动结果；默认浏览器完全由当前工具拥有。"""

    browser: Browser
    external_browser: bool
    mode: str
    owns_browser: bool = False
    dedicated_profile: bool = False


class BrowserUseBrowserHost:
    """使用 browser-use BrowserProfile 启动独立 Chrome 和 CDP。

    browser-use CLI 的默认本机 daemon 是共享浏览器，不能安全供两个 worker 并发。
    本类复用 BrowserProfile 的完整 Chrome 参数，为每个 worker 创建独立进程、CDP
    端口和临时 user-data-dir；避免 BrowserSession 看门狗与 Playwright 双控制器在
    多线程结束时竞争标签页。表单由确定性 Playwright 定位器通过该 CDP 会话执行。
    """

    def __init__(self, worker_number: int, headless: bool = True) -> None:
        self.worker_number = worker_number
        self.headless = bool(headless)
        self._cdp_url = ""
        self._process: subprocess.Popen[Any] | None = None
        self._browser_pid: int | None = None
        self._profile_dir = Path(
            tempfile.mkdtemp(prefix=f"browser-use-user-data-dir-studentaid-{worker_number}-")
        )

    @property
    def cdp_url(self) -> str:
        return self._cdp_url

    @staticmethod
    def _find_free_port() -> int:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
            probe.bind(("127.0.0.1", 0))
            return int(probe.getsockname()[1])

    def _wait_until_ready(self, timeout: float) -> None:
        deadline = time.monotonic() + timeout
        last_error: BaseException | None = None
        while time.monotonic() < deadline:
            if self._process is not None and self._process.poll() is not None:
                raise RuntimeError(
                    f"browser-use Chrome 提前退出，退出码 {self._process.returncode}"
                )
            try:
                with urllib.request.urlopen(
                    f"{self._cdp_url}/json/version", timeout=1
                ) as response:
                    payload = json.load(response)
                if payload.get("webSocketDebuggerUrl"):
                    return
            except Exception as exc:
                last_error = exc
            time.sleep(0.1)
        raise RuntimeError("browser-use Chrome CDP 启动超时") from last_error

    def start(self, timeout: float = 60.0) -> None:
        try:
            try:
                from browser_use.browser.chrome import find_chrome_executable
                from browser_use.browser.profile import BrowserProfile
            except ImportError as exc:
                raise RuntimeError(
                    "未安装 browser-use，请双击一键启动 CMD，"
                    "或执行：pip install -r requirements.txt"
                ) from exc

            chrome_path = find_chrome_executable()
            if not chrome_path:
                raise RuntimeError("browser-use 未找到 Google Chrome 可执行文件")
            profile = BrowserProfile(
                executable_path=chrome_path,
                user_data_dir=str(self._profile_dir),
                headless=False,
                keep_alive=False,
                window_size=dict(SMALL_WINDOW_SIZE),
                viewport=dict(SMALL_VIEWPORT_SIZE),
                args=["--disable-blink-features=AutomationControlled"],
            )
            debug_port = self._find_free_port()
            launch_args = [*profile.get_args(), f"--remote-debugging-port={debug_port}"]
            self._cdp_url = f"http://127.0.0.1:{debug_port}"
            creationflags = (
                getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                if os.name == "nt"
                else 0
            )
            startupinfo = None
            if os.name == "nt" and self.headless:
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = 0
            self._process = subprocess.Popen(
                [str(chrome_path), *launch_args],
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=creationflags,
                startupinfo=startupinfo,
            )
            self._browser_pid = self._process.pid
            self._wait_until_ready(timeout)
        except Exception:
            try:
                self.close()
            except Exception:
                pass
            raise

    def _terminate_process_tree(self) -> None:
        if not self._browser_pid:
            return
        try:
            import psutil

            parent = psutil.Process(self._browser_pid)
            processes = parent.children(recursive=True) + [parent]
            for process in reversed(processes):
                try:
                    process.terminate()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            _gone, alive = psutil.wait_procs(processes, timeout=4)
            for process in alive:
                try:
                    process.kill()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except Exception:
            pass

    def _remove_profile_dir(self) -> None:
        last_error: OSError | None = None
        for _attempt in range(5):
            try:
                if self._profile_dir.exists():
                    shutil.rmtree(self._profile_dir)
                return
            except OSError as exc:
                last_error = exc
                time.sleep(0.2)
        if self._profile_dir.exists():
            raise RuntimeError(
                f"browser-use 临时缓存目录删除失败：{self._profile_dir}"
            ) from last_error

    def close(self) -> None:
        self._terminate_process_tree()
        if self._process is not None:
            try:
                self._process.wait(timeout=5)
            except Exception:
                try:
                    self._process.kill()
                    self._process.wait(timeout=5)
                except Exception:
                    pass
            self._process = None
        self._remove_profile_dir()


def launch_browser(playwright: Playwright, headless: bool = True) -> BrowserLaunch:
    """启动独立、隐藏 AutomationControlled 的真实 Chrome。

    实测旧版 ``playwright.chromium.launch(channel="chrome")`` 会令
    ``navigator.webdriver`` 为 true，StudentAid 接口随后长期停在 Loading。
    browser-use 使用的同名 Blink 开关可消除这个差异，同时仍由 Playwright
    管理独立 BrowserContext，便于彻底清除本批次浏览器数据。
    """
    try:
        browser = playwright.chromium.launch(
            channel="chrome",
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                f"--window-size={SMALL_WINDOW_SIZE['width']},{SMALL_WINDOW_SIZE['height']}",
                *(["--window-position=-32000,-32000"] if headless else []),
            ],
        )
    except Exception as exc:
        raise RuntimeError(
            "无法启动 Google Chrome。请双击一键启动 CMD 自动检查并安装 Chrome。"
        ) from exc
    return BrowserLaunch(
        browser,
        False,
        (
            "独立 Google Chrome 无头模式（AutomationControlled 已关闭）"
            if headless
            else "独立 Google Chrome 小窗口（900x720，AutomationControlled 已关闭）"
        ),
        owns_browser=True,
        dedicated_profile=True,
    )


def _check_stop(stop_event: threading.Event) -> None:
    if stop_event.is_set():
        raise StopRequested("用户已请求停止")


def step_1_open_retrieve_account_details(
    page: Page, stop_event: threading.Event
) -> None:
    last_error: BaseException | None = None
    for attempt in range(1, 4):
        _check_stop(stop_event)
        try:
            page.goto(
                RETRIEVE_ACCOUNT_DETAILS_URL,
                wait_until="commit",
                timeout=60_000,
            )
            _check_stop(stop_event)
            page.locator(FIRST_NAME_SELECTOR).wait_for(
                state="visible", timeout=60_000
            )
            page.wait_for_function(
                """
                () => {
                    const text = document.body?.innerText || "";
                    return !text.includes("Loading...")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameLastName")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameDateOfBirthMonth")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameSsnInput");
                }
                """,
                timeout=30_000,
            )
            return
        except StopRequested:
            raise
        except Exception as exc:
            last_error = exc
            if attempt >= 3:
                break
            page.wait_for_timeout(1_500 * attempt)
    raise RuntimeError(
        "StudentAid 页面连续 3 次无法打开；请先在 Chrome 中确认页面可访问，"
        "或切换 GUI 中的浏览器后端后重试。"
    ) from last_error


def _type_account_field(page: Page, selector: str, value: str) -> None:
    """用真实键盘事件输入并失焦，确保 StudentAid 前端状态完整更新。"""
    field = page.locator(selector)
    field.click(timeout=30_000)
    field.fill("")
    field.press_sequentially(value, delay=20)
    field.press("Tab")


def step_2_fill_account_details(
    page: Page, details: AccountDetails, stop_event: threading.Event
) -> None:
    _check_stop(stop_event)
    _type_account_field(page, FIRST_NAME_SELECTOR, details.first_name)
    _type_account_field(page, LAST_NAME_SELECTOR, details.last_name)
    page.locator(BIRTH_MONTH_SELECTOR).select_option(details.birth_month)
    page.locator(BIRTH_MONTH_SELECTOR).press("Tab")
    _type_account_field(page, BIRTH_DAY_SELECTOR, str(int(details.birth_day)))
    _type_account_field(page, BIRTH_YEAR_SELECTOR, details.birth_year)
    _type_account_field(page, SSN_SELECTOR, details.ssn)
    _check_stop(stop_event)
    try:
        page.wait_for_function(
            """
            () => {
                const selectors = [
                    "#fsa_Input_ForgotUsernameFirstName",
                    "#fsa_Input_ForgotUsernameLastName",
                    "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                    "#fsa_Input_ForgotUsernameDateOfBirthDay",
                    "#fsa_Input_ForgotUsernameDateOfBirthYear",
                    "#fsa_Input_ForgotUsernameSsnInput",
                ];
                const button = [...document.querySelectorAll("button")].find(
                    element => (element.innerText || "").trim() === "Continue"
                );
                return selectors.every(selector => {
                    const field = document.querySelector(selector);
                    return field && String(field.value || "").length > 0;
                }) && button && !button.disabled;
            }
            """,
            timeout=30_000,
        )
    except Exception as exc:
        raise PageSubmissionStalled(
            "资料已填写，但 Continue 在 30 秒内未进入可点击状态"
        ) from exc
    page.wait_for_timeout(200)


def _wait_for_submission_started(page: Page, timeout_ms: int) -> bool:
    try:
        page.wait_for_function(
            """
            () => {
                const text = document.body?.innerText || "";
                const button = [...document.querySelectorAll("button")].find(
                    element => (element.innerText || "").trim() === "Continue"
                );
                return text.includes("Loading...")
                    || text.includes("Account Not Found")
                    || text.includes("Recover my account with a photo ID")
                    || text.includes("Limit Reached: Try Again in 24 Hours")
                    || text.includes("Your Account Is Disabled")
                    || text.includes("Account Lookup Issue: Get Help")
                    || text.includes("We are unable to retrieve your log-in information. Access your account by recovering your account with a photo ID.")
                    || text.includes("Enter a valid Social Security number.")
                    || text.includes("An unknown error has occurred")
                    || location.pathname.endsWith("/username")
                    || (button && button.disabled);
            }
            """,
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def step_3_click_continue(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
) -> None:
    """点击真实 Continue，并在必要时按可验证顺序使用两级兜底。"""
    _check_stop(stop_event)
    continue_element = page.get_by_role("button", name="Continue", exact=True).first
    continue_element.wait_for(state="visible", timeout=30_000)
    if not continue_element.is_enabled():
        raise RuntimeError("Continue 按钮未启用，请检查字段输入状态")
    continue_element.scroll_into_view_if_needed(timeout=10_000)
    try:
        continue_element.click(timeout=10_000)
    except Exception:
        _report_stage(progress_callback, "Continue 标准点击未完成，正在使用 Enter 兜底")
    if _wait_for_submission_started(page, 5_000):
        return

    _check_stop(stop_event)
    _report_stage(progress_callback, "Continue 尚未触发，正在聚焦按钮并按 Enter")
    try:
        continue_element.focus()
        continue_element.press("Enter", timeout=5_000)
    except Exception:
        pass
    if _wait_for_submission_started(page, 5_000):
        return

    _check_stop(stop_event)
    _report_stage(progress_callback, "Continue 仍未触发，正在执行按钮 DOM click")
    try:
        continue_element.evaluate("button => button.click()")
    except Exception:
        pass
    if _wait_for_submission_started(page, 5_000):
        return

    raise PageSubmissionStalled("三种 Continue 点击方式均未进入提交或结果状态")


def _report_stage(
    progress_callback: Callable[[str], None] | None, message: str
) -> None:
    """发送不包含输入资料的阶段消息；UI 回调失败不能中断网页处理。"""
    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def step_4_judge_password_recovery(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
    *,
    timeout_ms: int = 60_000,
    poll_interval_ms: int = 125,
    heartbeat_seconds: float = 5.0,
    stalled_loading_seconds: float = 60.0,
) -> str:
    """等待明确页面结果；持续 Loading 会触发重建会话，而不是无限转圈。"""
    started = time.monotonic()
    deadline = started + max(1, timeout_ms) / 1000
    next_heartbeat = max(0.05, heartbeat_seconds)
    loading_started: float | None = None

    while True:
        _check_stop(stop_event)
        body_text = page.evaluate("document.body?.innerText || ''")
        normalised_body_text = re.sub(r"\s+", " ", body_text).strip()
        loading = "Loading..." in body_text
        now = time.monotonic()
        if loading:
            loading_started = loading_started or now
            if now - loading_started >= stalled_loading_seconds:
                raise PageSubmissionStalled(
                    f"点击 Continue 后 Loading 已持续 {int(now - loading_started)} 秒"
                )
        else:
            loading_started = None
            if "An unknown error has occurred" in body_text:
                raise PageSessionExpired(
                    "StudentAid 页面会话已失效，需要重新打开页面"
                )
            if "Something went wrong" in body_text:
                raise RuntimeError("StudentAid 返回页面错误，请稍后重试")
            if LIMIT_REACHED_HEADING in body_text:
                _report_stage(progress_callback, f"已识别结果：{LIMIT_REACHED_HEADING}")
                return "limit_reached"
            if ACCOUNT_DISABLED_HEADING in body_text:
                _report_stage(progress_callback, f"已识别结果：{ACCOUNT_DISABLED_HEADING}")
                return "account_disabled"
            if ACCOUNT_RECOVERY_IN_PROGRESS_HEADING in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{ACCOUNT_RECOVERY_IN_PROGRESS_HEADING}",
                )
                return "account_recovery_in_progress"
            if ACCOUNT_LOOKUP_ISSUE_HEADING in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{ACCOUNT_LOOKUP_ISSUE_HEADING}",
                )
                return "account_lookup_issue"
            if PHOTO_ID_RECOVERY_MESSAGE in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{PHOTO_ID_RECOVERY_MESSAGE}",
                )
                return "photo_id_recovery_required"
            if INVALID_SSN_MESSAGE in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{INVALID_SSN_MESSAGE}",
                )
                return "invalid_ssn"
            if ACCOUNT_NOT_FOUND_CREATE_HEADING in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{ACCOUNT_NOT_FOUND_CREATE_HEADING}",
                )
                return "account_not_found"
            if re.search(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                body_text,
                re.IGNORECASE | re.MULTILINE,
            ):
                _report_stage(progress_callback, "已识别结果：Account Not Found")
                return "account_not_found"
            if (
                "Retrieve Your Log-in Information" in body_text
                and "Recover my account with a photo ID" in body_text
            ):
                _report_stage(progress_callback, "已识别结果：Retrieve Your Log-in Information")
                return "can_recover"

        elapsed = now - started
        if now >= deadline:
            raise PageSubmissionStalled(
                f"等待 StudentAid 明确结果超时（{timeout_ms / 1000:g} 秒）"
            )
        if elapsed >= next_heartbeat:
            _report_stage(
                progress_callback,
                f"官方页面仍在处理中，已等待 {int(elapsed)} 秒",
            )
            while next_heartbeat <= elapsed:
                next_heartbeat += max(0.05, heartbeat_seconds)
        page.wait_for_timeout(
            min(poll_interval_ms, max(1, int((deadline - now) * 1000)))
        )


def _visible_text(page: Page, text: str) -> str:
    locator = page.get_by_text(text, exact=True).first
    return locator.inner_text().strip() if locator.is_visible() else ""


_MASK_CHARACTERS = "*•⦁●"
_MASK_CHARACTER_CLASS = re.escape(_MASK_CHARACTERS)
_MASKED_PHONE_RE = re.compile(
    rf"^\(\s*[{_MASK_CHARACTER_CLASS}]{{3}}\s*\)\s*"
    rf"[{_MASK_CHARACTER_CLASS}]{{3}}\s*\d{{4}}$"
)


def _normalise_contact_text(value: str) -> str:
    """保留掩码内容，只统一 DOM 中的不可见/重复空白。"""
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def _looks_like_masked_phone(value: str) -> bool:
    return bool(_MASKED_PHONE_RE.fullmatch(_normalise_contact_text(value)))


def _looks_like_masked_email(value: str) -> bool:
    value = _normalise_contact_text(value)
    if "@" not in value or not any(char in value for char in _MASK_CHARACTERS):
        return False
    local_part, separator, domain = value.partition("@")
    if not separator or not local_part or not domain or " " in value:
        return False
    # StudentAid 只遮挡本地部分；限定为一行邮箱格式，避免页脚说明文字被误记。
    return bool(
        re.fullmatch(
            rf"[A-Za-z0-9._%+\-{_MASK_CHARACTER_CLASS}]+@"
            r"[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            value,
        )
    )


def _visible_contact_texts(page: Page) -> list[str]:
    """只读取结果卡片的可见段落；精确类名失效时再退到可见 p。"""
    selectors = ("p.fsa-color-gray-60", "p.m-0.fsa-font-size-16", "p")
    for selector in selectors:
        locator = page.locator(selector)
        values: list[str] = []
        try:
            for index in range(locator.count()):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                value = _normalise_contact_text(item.inner_text())
                if value:
                    values.append(value)
        except Exception:
            values = []
        if any(
            _looks_like_masked_phone(value) or _looks_like_masked_email(value)
            for value in values
        ):
            return values
    return []


def collect_recovery_result(page: Page, recovery_status: str) -> RecoveryResult:
    if recovery_status == "limit_reached":
        heading = _visible_text(page, LIMIT_REACHED_HEADING) or LIMIT_REACHED_HEADING
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_disabled":
        heading = _visible_text(page, ACCOUNT_DISABLED_HEADING) or ACCOUNT_DISABLED_HEADING
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_recovery_in_progress":
        heading = (
            _visible_text(page, ACCOUNT_RECOVERY_IN_PROGRESS_HEADING)
            or ACCOUNT_RECOVERY_IN_PROGRESS_HEADING
        )
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_lookup_issue":
        heading = (
            _visible_text(page, ACCOUNT_LOOKUP_ISSUE_HEADING)
            or ACCOUNT_LOOKUP_ISSUE_HEADING
        )
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "photo_id_recovery_required":
        heading = (
            _visible_text(page, PHOTO_ID_RECOVERY_MESSAGE)
            or PHOTO_ID_RECOVERY_MESSAGE
        )
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "invalid_ssn":
        heading = _visible_text(page, INVALID_SSN_MESSAGE) or INVALID_SSN_MESSAGE
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_not_found":
        exact_heading = _visible_text(page, ACCOUNT_NOT_FOUND_CREATE_HEADING)
        if exact_heading:
            return RecoveryResult(recovery_status, exact_heading, "", "", "")
        locator = page.get_by_text(
            re.compile(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                re.IGNORECASE,
            )
        ).first
        heading = locator.inner_text().strip() if locator.is_visible() else "Account Not Found"
        return RecoveryResult(recovery_status, heading, "", "", "")

    heading = _visible_text(page, "Retrieve Your Log-in Information")
    masked_phone = ""
    masked_email = ""
    for value in _visible_contact_texts(page):
        if _looks_like_masked_phone(value) and not masked_phone:
            masked_phone = value
        elif _looks_like_masked_email(value) and not masked_email:
            masked_email = value
    recovery_method = _visible_text(page, "Recover my account with a photo ID")
    return RecoveryResult(
        recovery_status,
        heading,
        masked_phone,
        masked_email,
        recovery_method,
    )


HTTP_ERROR_RESULTS: Mapping[str, tuple[str, str]] = {
    "SSN_INVALID": ("invalid_ssn", INVALID_SSN_MESSAGE),
    "USER_NOT_FOUND": ("account_not_found", ACCOUNT_NOT_FOUND_CREATE_HEADING),
    "PII_MISMATCH": ("account_not_found", "Account Not Found"),
    "MULTIPLE_USERS_FOUND": ("account_lookup_issue", ACCOUNT_LOOKUP_ISSUE_HEADING),
    "DISABLED_BY_FSA": ("account_disabled", ACCOUNT_DISABLED_HEADING),
    "ACCT_RECOVERY_CASE_PENDING": (
        "account_recovery_in_progress", ACCOUNT_RECOVERY_IN_PROGRESS_HEADING,
    ),
    "LAST_NAME_DOB_SSN_COMBO_LOCKED": ("limit_reached", LIMIT_REACHED_HEADING),
    "SELF_SERVICE_UNAVAILABLE": (
        "photo_id_recovery_required", PHOTO_ID_RECOVERY_MESSAGE,
    ),
}
HTTP_MASK_CHARACTER = "\u2981"


def _mask_http_phone(value: str) -> str:
    """按官网当前展示形式仅保留手机尾四位。"""
    normalised = _normalise_contact_text(value)
    if _looks_like_masked_phone(normalised):
        return normalised
    digits = re.sub(r"\D", "", normalised)
    if len(digits) < 4:
        return ""
    return f"({HTTP_MASK_CHARACTER * 3}) {HTTP_MASK_CHARACTER * 3} {digits[-4:]}"


def _mask_http_email(value: str) -> str:
    """保留邮箱前两位和域名，其余本地部分使用官网掩码字符。"""
    normalised = _normalise_contact_text(value)
    if _looks_like_masked_email(normalised):
        return normalised
    local_part, separator, domain = normalised.partition("@")
    if not separator or not local_part or not domain or " " in normalised:
        return ""
    visible_length = min(2, len(local_part))
    hidden_length = max(1, len(local_part) - visible_length)
    return (
        f"{local_part[:visible_length]}"
        f"{HTTP_MASK_CHARACTER * hidden_length}@{domain}"
    )


def _http_contact_values(options: Any) -> tuple[str, str]:
    """兼容官网 selfServiceOptions 的对象、数组及类型/值包装形式。"""
    phones: list[str] = []
    emails: list[str] = []

    def add(kind: str, value: Any) -> None:
        if not isinstance(value, str):
            return
        cleaned = value.strip()
        if not cleaned:
            return
        if kind == "email" and "@" in cleaned:
            emails.append(cleaned)
        elif kind == "phone" and len(re.sub(r"\D", "", cleaned)) >= 4:
            phones.append(cleaned)

    def walk(node: Any) -> None:
        if isinstance(node, Mapping):
            type_text = " ".join(
                str(node.get(key, ""))
                for key in ("type", "contactType", "method", "channel", "optionType")
            ).casefold()
            scalar_values = [value for value in node.values() if isinstance(value, str)]
            if "email" in type_text:
                for value in scalar_values:
                    add("email", value)
            if "mobile" in type_text or "phone" in type_text or "sms" in type_text:
                for value in scalar_values:
                    add("phone", value)
            for key, value in node.items():
                normalised_key = re.sub(r"[^a-z]", "", str(key).casefold())
                if "email" in normalised_key:
                    add("email", value)
                if "phone" in normalised_key or "mobile" in normalised_key:
                    add("phone", value)
                if isinstance(value, (Mapping, list, tuple)):
                    walk(value)
        elif isinstance(node, (list, tuple)):
            for value in node:
                walk(value)

    walk(options)
    return (phones[0] if phones else "", emails[0] if emails else "")


def recovery_result_from_http_payload(payload: Any) -> RecoveryResult:
    """把官网 JSON 响应转换为既有 RecoveryResult，不接受未知明确状态。"""
    if not isinstance(payload, Mapping):
        raise RuntimeError("StudentAid HTTP 搜索返回了非 JSON 对象")
    raw_codes = payload.get("errorCodes", ())
    if isinstance(raw_codes, str):
        error_codes = [raw_codes]
    elif isinstance(raw_codes, Sequence):
        error_codes = [str(code) for code in raw_codes if str(code).strip()]
    else:
        error_codes = []
    if not error_codes and payload.get("errorCode"):
        error_codes = [str(payload["errorCode"])]
    if error_codes:
        for code in error_codes:
            if code in HTTP_ERROR_RESULTS:
                result_code, heading = HTTP_ERROR_RESULTS[code]
                return RecoveryResult(result_code, heading, "", "", "")
        raise RuntimeError("StudentAid HTTP 搜索返回了未识别的错误状态")

    options = payload.get("selfServiceOptions")
    if not isinstance(options, (Mapping, list, tuple)):
        raise RuntimeError("StudentAid HTTP 搜索没有返回明确结果或恢复选项")
    phone, email = _http_contact_values(options)
    return RecoveryResult(
        "can_recover",
        "Retrieve Your Log-in Information",
        _mask_http_phone(phone),
        _mask_http_email(email),
        "Recover my account with a photo ID",
    )


class HttpRecoverySession:
    """每个 worker 独立复用官网 HTTP 会话；不记录请求中的身份资料。"""

    def __init__(
        self,
        worker_number: int,
        session_factory: Callable[[], Any] | None = None,
    ) -> None:
        self.worker_number = worker_number
        self._session_factory = session_factory
        self._session: Any = None
        self._xsrf_token = ""
        self.browser_mode = f"HTTP 官方接口直连 (worker {worker_number})"

    @staticmethod
    def _base_headers(accept: str = "application/json") -> dict[str, str]:
        return {
            "Accept": accept,
            "Accept-Language": "en-US,en;q=0.9",
            "Origin": HTTP_ORIGIN,
            "Referer": RETRIEVE_ACCOUNT_DETAILS_URL,
            "User-Agent": HTTP_USER_AGENT,
        }

    @staticmethod
    def _ensure_success(response: Any, stage: str) -> None:
        status_code = int(getattr(response, "status_code", 0) or 0)
        if not 200 <= status_code < 300:
            raise RuntimeError(f"StudentAid HTTP 会话初始化失败（{stage}，HTTP {status_code}）")

    def _create_session(self) -> Any:
        if self._session_factory is not None:
            return self._session_factory()
        if requests is None:
            raise RuntimeError(
                "未安装 requests，请双击一键启动 CMD，或执行：pip install -r requirements.txt"
            )
        return requests.Session()

    def _close_session(self) -> None:
        session, self._session = self._session, None
        self._xsrf_token = ""
        if session is not None:
            try:
                session.close()
            except Exception:
                pass

    def _read_xsrf_cookie(self) -> str:
        cookies = getattr(self._session, "cookies", ())
        try:
            for cookie in cookies:
                if getattr(cookie, "name", "") == "XSRF-TOKEN":
                    return str(getattr(cookie, "value", "")).strip()
        except TypeError:
            pass
        getter = getattr(cookies, "get", None)
        if callable(getter):
            try:
                return str(getter("XSRF-TOKEN") or "").strip()
            except Exception:
                return ""
        return ""

    def _bootstrap(self) -> None:
        self._close_session()
        self._session = self._create_session()
        page_headers = self._base_headers("text/html,application/xhtml+xml")
        response = self._session.get(
            RETRIEVE_ACCOUNT_DETAILS_URL, headers=page_headers, timeout=HTTP_TIMEOUT,
        )
        self._ensure_success(response, "页面")
        response = self._session.get(
            HTTP_KEEP_SESSION_ALIVE_URL,
            headers=self._base_headers("*/*"), timeout=HTTP_TIMEOUT,
        )
        self._ensure_success(response, "keepSessionAlive")
        response = self._session.post(
            HTTP_SESSION_USER_URL, data="",
            headers=self._base_headers("application/json, text/plain, */*"),
            timeout=HTTP_TIMEOUT,
        )
        self._ensure_success(response, "sessionUser")
        response = self._session.get(
            HTTP_GATEWAY_SESSION_URL,
            headers=self._base_headers("application/json, text/plain, */*"),
            timeout=HTTP_TIMEOUT,
        )
        self._ensure_success(response, "gateway/session")
        self._xsrf_token = self._read_xsrf_cookie()
        if not self._xsrf_token:
            raise RuntimeError("StudentAid HTTP 会话没有返回 XSRF-TOKEN")

    def start(self) -> None:
        self._bootstrap()

    def _post_lookup(self, item: WorkItem) -> Any:
        details = item.details
        headers = self._base_headers("application/json, text/plain, */*")
        headers.update({
            "Content-Type": "application/json;charset=utf-8",
            "X-XSRF-TOKEN": self._xsrf_token,
        })
        body = {
            "lastName": details.last_name,
            "dob": (
                f"{int(details.birth_year):04d}-"
                f"{int(details.birth_month):02d}-{int(details.birth_day):02d}"
            ),
            "firstName": details.first_name,
            "ssn": details.ssn,
        }
        return self._session.post(
            HTTP_ACCOUNT_LOOKUP_URL, json=body, headers=headers, timeout=HTTP_TIMEOUT,
        )

    def process(
        self,
        item: WorkItem,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> RecoveryResult:
        _check_stop(stop_event)
        if self._session is None or not self._xsrf_token:
            _report_stage(progress_callback, "正在初始化 HTTP 搜索会话")
            self._bootstrap()
        for attempt in range(2):
            _check_stop(stop_event)
            _report_stage(progress_callback, "正在通过官方 HTTP 接口搜索")
            response = self._post_lookup(item)
            status_code = int(getattr(response, "status_code", 0) or 0)
            if status_code in {401, 403} and attempt == 0:
                _report_stage(progress_callback, "HTTP 会话已失效，正在自动重新初始化")
                self._bootstrap()
                continue
            if status_code == 429:
                raise RuntimeError("StudentAid HTTP 搜索触发限流（HTTP 429）")
            if status_code >= 500:
                raise RuntimeError(f"StudentAid HTTP 搜索服务端错误（HTTP {status_code}）")
            if not 200 <= status_code < 300:
                raise RuntimeError(f"StudentAid HTTP 搜索失败（HTTP {status_code}）")
            try:
                payload = response.json()
            except Exception as exc:
                raise RuntimeError("StudentAid HTTP 搜索返回了无效 JSON") from exc
            result = recovery_result_from_http_payload(payload)
            _report_stage(progress_callback, "官方 HTTP 接口已返回明确结果")
            return result
        raise RuntimeError("StudentAid HTTP 会话重新初始化后仍被拒绝")

    def prepare_for_next(
        self,
        _result_code: str,
        stop_event: threading.Event,
        _progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        _check_stop(stop_event)

    def recover_after_cleanup_error(self) -> None:
        self._close_session()

    def close(self) -> None:
        self._close_session()


class BrowserRecoverySession:
    """第十五步 worker 会话；每个线程拥有独立浏览器和缓存目录。"""

    def __init__(
        self,
        worker_number: int,
        backend: str = "playwright",
        display_mode: str = "无头",
    ) -> None:
        backend = backend.strip().casefold()
        if backend not in BROWSER_BACKENDS:
            raise ValueError(f"浏览器后端必须是：{', '.join(BROWSER_BACKENDS)}")
        self.worker_number = worker_number
        self.backend = backend
        self.display_mode = _normalise_display_mode(display_mode)
        self.headless = _is_headless(self.display_mode)
        self._playwright: Playwright | None = None
        self._browser: Browser | None = None
        self._browser_use_host: BrowserUseBrowserHost | None = None
        self._context: Any = None
        self._page: Page | None = None
        self._owns_context = False
        self._external_browser = False
        self._owns_browser = False
        self._dedicated_profile = False
        self.browser_mode = "未启动"

    def start(self) -> None:
        if sync_playwright is None:
            raise RuntimeError(
                "未安装 playwright，请双击一键启动 CMD，或执行：pip install -r requirements.txt"
            )
        self._playwright = sync_playwright().start()
        try:
            if self.backend == "browser-use":
                self._browser_use_host = BrowserUseBrowserHost(
                    self.worker_number, headless=self.headless
                )
                self._browser_use_host.start()
                try:
                    browser = self._playwright.chromium.connect_over_cdp(
                        self._browser_use_host.cdp_url, timeout=30_000
                    )
                except Exception as exc:
                    raise RuntimeError("Playwright 无法接管 browser-use 独立 CDP 会话") from exc
                launch = BrowserLaunch(
                    browser=browser,
                    external_browser=True,
                    mode=(
                        "browser-use 独立 Chrome "
                        f"{self.display_mode}模式 (worker {self.worker_number}, "
                        "AutomationControlled 已关闭)"
                    ),
                    owns_browser=False,
                    dedicated_profile=True,
                )
            else:
                launch = launch_browser(self._playwright, headless=self.headless)
            self._browser = launch.browser
            self._external_browser = launch.external_browser
            self._owns_browser = launch.owns_browser
            self._dedicated_profile = launch.dedicated_profile
            self.browser_mode = launch.mode
            browser_contexts = list(getattr(self._browser, "contexts", []))
            if self._external_browser:
                if not browser_contexts:
                    raise RuntimeError("Chrome CDP 没有可用的默认浏览器上下文")
                self._context = browser_contexts[0]
            else:
                self._context = self._browser.new_context(
                    viewport=dict(SMALL_VIEWPORT_SIZE), locale="en-US"
                )
                self._owns_context = True
            pages = list(getattr(self._context, "pages", []))
            self._page = pages[-1] if pages else self._context.new_page()
            self._configure_page(self._page)
            if self._page.evaluate("navigator.webdriver === true"):
                raise RuntimeError(
                    "检测到浏览器处于 Playwright 自动化启动模式；"
                    "该模式会让 StudentAid 在 Continue 后持续 Loading"
                )
            if self._dedicated_profile:
                # 上次进程若被强制关闭，先消除可能残留的 cookie、cache 和 storage。
                self._clear_browser_data_and_blank()
            self._hide_browser_window()
        except Exception:
            self.close()
            raise

    def _close_page(self) -> None:
        if self._page is not None:
            try:
                if not self._page.is_closed():
                    self._page.close()
            except Exception:
                pass
            self._page = None

    def _configure_page(self, page: Page) -> None:
        try:
            page.set_viewport_size(dict(SMALL_VIEWPORT_SIZE))
        except Exception:
            pass
        if not self.headless or self._browser is None or self._context is None:
            return
        user_agent = _normal_chrome_user_agent(self._browser.version)
        try:
            session = self._context.new_cdp_session(page)
            try:
                session.send(
                    "Emulation.setUserAgentOverride",
                    {
                        "userAgent": user_agent,
                        "acceptLanguage": "en-US,en;q=0.9",
                        "platform": "Win32",
                    },
                )
            finally:
                session.detach()
        except Exception as exc:
            raise RuntimeError("无法为无头 Chrome 设置正常浏览器标识") from exc

    def _hide_browser_window(self) -> None:
        if not self.headless or os.name != "nt":
            return
        owned_processes: set[int] = set()
        if (
            self._browser_use_host is not None
            and self._browser_use_host._browser_pid is not None
        ):
            try:
                import psutil

                root_process = psutil.Process(self._browser_use_host._browser_pid)
                owned_processes = {
                    root_process.pid,
                    *(child.pid for child in root_process.children(recursive=True)),
                }
            except Exception:
                owned_processes = set()
        else:
            owned_processes = _chrome_process_ids("--window-position=-32000,-32000")
        for _attempt in range(10):
            if _hide_windows_for_processes(owned_processes):
                return
            time.sleep(0.05)
            if self._browser_use_host is None:
                owned_processes = _chrome_process_ids(
                    "--window-position=-32000,-32000"
                )

    def _new_blank_page(self) -> Page:
        if self._context is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        self._close_page()
        self._page = self._context.new_page()
        self._configure_page(self._page)
        self._hide_browser_window()
        return self._page

    def _ensure_form_page(
        self,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None,
    ) -> Page:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        form_ready = False
        try:
            form_ready = bool(
                page.locator(FIRST_NAME_SELECTOR).is_visible(timeout=1_000)
                and "Loading..." not in page.locator("body").inner_text(timeout=1_000)
            )
        except Exception:
            form_ready = False
        if not form_ready:
            self._hide_browser_window()
            _report_stage(progress_callback, "正在打开 StudentAid 页面")
            step_1_open_retrieve_account_details(page, stop_event)
        return page

    def _clear_browser_data_and_blank(
        self,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        _report_stage(progress_callback, "正在清除 StudentAid 专用浏览器数据")
        origins = {"https://studentaid.gov"}
        try:
            discovered = page.evaluate(
                """
                () => [...new Set([
                    location.origin,
                    ...performance.getEntriesByType("resource").map(entry => {
                        try { return new URL(entry.name).origin; }
                        catch (_) { return ""; }
                    })
                ])]
                """
            )
            origins.update(
                origin for origin in discovered
                if isinstance(origin, str)
                and re.fullmatch(r"https://(?:[a-z0-9-]+\.)*studentaid\.gov", origin)
            )
        except Exception:
            pass
        try:
            session = self._context.new_cdp_session(page)
            try:
                session.send("Network.clearBrowserCache")
                if self._dedicated_profile:
                    session.send("Network.clearBrowserCookies")
                try:
                    session.send("ServiceWorker.stopAllWorkers")
                except Exception:
                    pass
                for origin in origins:
                    session.send(
                        "Storage.clearDataForOrigin",
                        {"origin": origin, "storageTypes": "all"},
                    )
            finally:
                try:
                    session.detach()
                except Exception:
                    pass
        except Exception:
            pass
        if self._dedicated_profile:
            try:
                self._context.clear_cookies()
            except Exception:
                pass
        self._new_blank_page()
        _report_stage(progress_callback, "浏览器数据已清除，已回到空白页")

    def process(
        self,
        item: WorkItem,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> RecoveryResult:
        if self._browser is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        last_error: BaseException | None = None
        for attempt in range(1, 4):
            try:
                page = self._ensure_form_page(stop_event, progress_callback)
                if attempt > 1:
                    _report_stage(
                        progress_callback,
                        f"已重建页面会话，正在重新填写资料（自动重试 {attempt - 1}/2）",
                    )
                else:
                    _report_stage(progress_callback, "页面已打开，正在填写资料")
                step_2_fill_account_details(page, item.details, stop_event)
                _report_stage(progress_callback, "资料已填写，正在点击 Continue")
                step_3_click_continue(page, stop_event, progress_callback)
                _report_stage(progress_callback, "已提交，正在等待官方结果")
                status = step_4_judge_password_recovery(
                    page, stop_event, progress_callback
                )
                return collect_recovery_result(page, status)
            except StopRequested:
                raise
            except (PageSessionExpired, PageSubmissionStalled) as exc:
                last_error = exc
                if attempt >= 3:
                    break
                _report_stage(
                    progress_callback,
                    "页面会话失效或提交持续转圈，正在自动清理并重建页面",
                )
                self._clear_browser_data_and_blank(progress_callback)
        raise RuntimeError(
            "StudentAid 页面清理并重填后仍未返回明确结果"
        ) from last_error

    def prepare_for_next(
        self,
        result_code: str,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        """必须在结果写入 SQLite/累计 CSV 后调用。"""
        _check_stop(stop_event)
        page = self._page
        if result_code in {
            "account_not_found",
            "limit_reached",
            "account_disabled",
            "account_lookup_issue",
            "account_recovery_in_progress",
            "photo_id_recovery_required",
            "invalid_ssn",
        }:
            self._clear_browser_data_and_blank(progress_callback)
            return
        if result_code != "can_recover":
            self._close_page()
            return
        if page is None or page.is_closed():
            raise RuntimeError("记录账户找回结果后页面已关闭，无法点击 Cancel")
        _report_stage(progress_callback, "结果已保存，正在点击 Cancel")
        cancel_matches = page.locator(
            "span", has_text=re.compile(r"^\s*Cancel\s*$", re.IGNORECASE)
        )
        cancel_span = None
        for index in range(cancel_matches.count()):
            candidate = cancel_matches.nth(index)
            if (
                candidate.is_visible()
                and candidate.inner_text().strip().casefold() == "cancel"
            ):
                cancel_span = candidate
                break
        if cancel_span is None:
            raise RuntimeError("找不到可见的 <span>Cancel</span>")

        original_url = page.url
        parent_control = cancel_span.locator(
            "xpath=ancestor::*[self::button or self::a or @role='button'][1]"
        )
        click_targets = [("Cancel 文本", cancel_span)]
        if parent_control.count() and parent_control.is_visible():
            click_targets.append(("Cancel 父控件", parent_control))

        cancel_triggered = False
        last_click_error: Exception | None = None
        for target_name, target in click_targets:
            try:
                target.scroll_into_view_if_needed(timeout=3_000)
                target.click(timeout=5_000)
                page.wait_for_function(
                    """
                    originalUrl => {
                        const visible = element => {
                            const style = getComputedStyle(element);
                            const box = element.getBoundingClientRect();
                            return style.visibility !== "hidden"
                                && style.display !== "none"
                                && box.width > 0 && box.height > 0;
                        };
                        const resultHeadingVisible = [...document.querySelectorAll("*")]
                            .some(element => element.textContent?.trim()
                                === "Retrieve Your Log-in Information" && visible(element));
                        const cancelVisible = [...document.querySelectorAll("span")]
                            .some(element => element.textContent?.trim().toLowerCase()
                                === "cancel" && visible(element));
                        const formVisible = (() => {
                            const element = document.querySelector(
                                "#fsa_Input_ForgotUsernameFirstName"
                            );
                            return Boolean(element && visible(element));
                        })();
                        return location.href !== originalUrl || formVisible
                            || !resultHeadingVisible || !cancelVisible;
                    }
                    """,
                    arg=original_url,
                    timeout=6_000,
                )
                cancel_triggered = True
                _report_stage(
                    progress_callback,
                    f"{target_name}已实际触发，结果页已离开",
                )
                break
            except Exception as exc:
                last_click_error = exc

        if not cancel_triggered:
            # Angular 模板偶尔拦截合成鼠标事件；原生 click 会在可见 span/父控件上
            # 冒泡到同一个 Angular 处理器。仍必须验证结果页确实发生变化。
            try:
                cancel_span.evaluate(
                    """
                    element => (element.closest("button,a,[role='button']") || element).click()
                    """
                )
                page.wait_for_function(
                    """
                    () => {
                        const visible = element => {
                            const style = getComputedStyle(element);
                            const box = element.getBoundingClientRect();
                            return style.visibility !== "hidden"
                                && style.display !== "none"
                                && box.width > 0 && box.height > 0;
                        };
                        return ![...document.querySelectorAll("span")].some(
                            element => element.textContent?.trim().toLowerCase()
                                === "cancel" && visible(element)
                        );
                    }
                    """,
                    timeout=6_000,
                )
                cancel_triggered = True
                _report_stage(progress_callback, "Cancel 原生事件已实际触发")
            except Exception as exc:
                last_click_error = exc

        if not cancel_triggered:
            raise RuntimeError("可见 Cancel 点击后结果页没有变化") from last_click_error

        try:
            page.locator(FIRST_NAME_SELECTOR).wait_for(state="visible", timeout=5_000)
        except Exception:
            _report_stage(
                progress_callback,
                "Cancel 已完成；正在重新打开账户找回空白表单",
            )
            step_1_open_retrieve_account_details(page, stop_event)
        page.wait_for_function(
            """
            () => {
                const ids = [
                    "#fsa_Input_ForgotUsernameFirstName",
                    "#fsa_Input_ForgotUsernameLastName",
                    "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                    "#fsa_Input_ForgotUsernameDateOfBirthDay",
                    "#fsa_Input_ForgotUsernameDateOfBirthYear",
                    "#fsa_Input_ForgotUsernameSsnInput",
                ];
                return ids.every(id => !(document.querySelector(id)?.value || ""));
            }
            """,
            timeout=15_000,
        )
        _report_stage(progress_callback, "Cancel 已确认完成，空白表单已就绪")

    def recover_after_cleanup_error(self) -> None:
        self._close_page()

    def close(self) -> None:
        if self.backend == "browser-use":
            # browser-use 使用一次性 user-data-dir。先结束精确 PID 树并删除整个
            # profile，比在失效 CDP 页面上逐项清理更彻底，也避免无头失败页令
            # Playwright disconnect/stop 长时间等待。
            self._page = None
            self._context = None
            self._owns_context = False
            self._browser = None
            self._external_browser = False
            self._owns_browser = False
            if self._browser_use_host is not None:
                try:
                    self._browser_use_host.close()
                finally:
                    self._browser_use_host = None
            if self._playwright is not None:
                try:
                    self._playwright.stop()
                except Exception:
                    pass
                self._playwright = None
            self._dedicated_profile = False
            return

        if self._context is not None and self._browser is not None:
            try:
                self._clear_browser_data_and_blank()
            except Exception:
                pass
        self._close_page()
        if self._context is not None and self._owns_context:
            try:
                self._context.close()
            except Exception:
                pass
        self._context = None
        self._owns_context = False
        if self._browser is not None and self._owns_browser:
            try:
                self._browser.close()
            except Exception:
                pass
        self._browser = None
        self._external_browser = False
        self._owns_browser = False
        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None
        self._dedicated_profile = False


@dataclass
class _DbCommand:
    action: str
    payload: dict[str, Any]
    response: queue.Queue[tuple[bool, Any]]


class DatabaseWriter:
    """SQLite 的唯一写入者；工作线程不得直接执行写 SQL。"""

    def __init__(self, database_path: Path) -> None:
        self.database_path = database_path.resolve()
        self._commands: queue.Queue[_DbCommand] = queue.Queue()
        self._ready = threading.Event()
        self._thread = threading.Thread(
            target=self._run,
            name="studentaid-db-writer",
            daemon=False,
        )
        self._startup_error: BaseException | None = None

    def start(self) -> None:
        ensure_output_storage_writable(
            self.database_path.parent,
            self.database_path,
            Path(f"{self.database_path}-wal"),
            Path(f"{self.database_path}-shm"),
        )
        self._thread.start()
        if not self._ready.wait(30):
            raise TimeoutError("SQLite 写入线程启动超时")
        if self._startup_error is not None:
            raise RuntimeError("SQLite 初始化失败") from self._startup_error

    def request(self, action: str, **payload: Any) -> Any:
        if not self._thread.is_alive():
            raise RuntimeError("SQLite 写入线程未运行")
        response: queue.Queue[tuple[bool, Any]] = queue.Queue(maxsize=1)
        self._commands.put(_DbCommand(action, payload, response))
        try:
            ok, value = response.get(timeout=300)
        except queue.Empty as exc:
            raise TimeoutError(f"SQLite 写入命令超时：{action}") from exc
        if ok:
            return value
        raise RuntimeError(f"SQLite 写入失败：{action}") from value

    def close(self) -> None:
        if self._thread.is_alive():
            try:
                self.request("shutdown")
            finally:
                self._thread.join(timeout=30)

    def _run(self) -> None:
        connection: sqlite3.Connection | None = None
        try:
            connection = sqlite3.connect(self.database_path, timeout=30)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA busy_timeout=30000")
            journal_mode = connection.execute("PRAGMA journal_mode=WAL").fetchone()[0]
            if str(journal_mode).casefold() != "wal":
                raise RuntimeError(f"无法启用 SQLite WAL，当前模式：{journal_mode}")
            connection.execute("PRAGMA synchronous=NORMAL")
            connection.execute("PRAGMA wal_autocheckpoint=1000")
            self._create_schema(connection)
            connection.commit()
        except BaseException as exc:
            self._startup_error = exc
            self._ready.set()
            if connection is not None:
                connection.close()
            return

        self._ready.set()
        assert connection is not None
        try:
            while True:
                command = self._commands.get()
                try:
                    value = self._handle(connection, command.action, command.payload)
                    connection.commit()
                    command.response.put((True, value))
                    if command.action == "shutdown":
                        break
                except BaseException as exc:
                    connection.rollback()
                    command.response.put((False, exc))
                finally:
                    self._commands.task_done()
        finally:
            connection.close()

    @staticmethod
    def _create_schema(connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS batches (
                batch_id TEXT PRIMARY KEY,
                input_path TEXT NOT NULL,
                output_directory TEXT NOT NULL,
                thread_count INTEGER NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                finished_at TEXT,
                export_path TEXT,
                total_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id TEXT NOT NULL REFERENCES batches(batch_id),
                source_file TEXT NOT NULL,
                source_sheet TEXT NOT NULL DEFAULT '',
                source_row INTEGER NOT NULL,
                original_fields_json TEXT NOT NULL,
                input_row_key TEXT NOT NULL DEFAULT '',
                ssn TEXT,
                birth_month TEXT,
                birth_day TEXT,
                birth_year TEXT,
                first_name TEXT,
                last_name TEXT,
                address TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL,
                attempt_count INTEGER NOT NULL DEFAULT 0,
                result_code TEXT NOT NULL DEFAULT '',
                result_heading TEXT NOT NULL DEFAULT '',
                masked_phone TEXT NOT NULL DEFAULT '',
                masked_email TEXT NOT NULL DEFAULT '',
                recovery_method TEXT NOT NULL DEFAULT '',
                error TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_records_batch_status
                ON records(batch_id, status, id);

            CREATE INDEX IF NOT EXISTS idx_records_batch_ssn_status
                ON records(batch_id, ssn, status);

            """
        )
        columns = {
            str(row[1]) for row in connection.execute("PRAGMA table_info(records)")
        }
        if "input_row_key" not in columns:
            connection.execute(
                "ALTER TABLE records ADD COLUMN input_row_key TEXT NOT NULL DEFAULT ''"
            )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_records_batch_row_key_status "
            "ON records(batch_id, input_row_key, status)"
        )

    @staticmethod
    def _handle(
        connection: sqlite3.Connection, action: str, payload: Mapping[str, Any]
    ) -> Any:
        if action == "create_batch":
            connection.execute(
                """
                INSERT INTO batches(
                    batch_id, input_path, output_directory, thread_count,
                    status, created_at
                ) VALUES (?, ?, ?, ?, 'importing', ?)
                """,
                (
                    payload["batch_id"], payload["input_path"],
                    payload["output_directory"], payload["thread_count"], _now_iso(),
                ),
            )
            return None

        if action == "insert_records":
            batch_id = str(payload["batch_id"])
            work_items: list[WorkItem] = []
            queued_keys: set[str] = set()
            records: Sequence[ImportedRecord] = payload["records"]
            for record in records:
                details = record.details
                status = "failed" if record.import_error else "pending"
                row_key = (
                    _input_row_key(
                        details, record.original_fields, record.address,
                        record.input_mapping,
                    )
                    if details is not None and not record.import_error
                    else ""
                )
                cursor = connection.execute(
                    """
                    INSERT INTO records(
                        batch_id, source_file, source_sheet, source_row,
                        original_fields_json, input_row_key, ssn,
                        birth_month, birth_day,
                        birth_year, first_name, last_name, address, status,
                        error, created_at, finished_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id, record.source_file, record.source_sheet,
                        record.source_row,
                        json.dumps(record.original_fields, ensure_ascii=False),
                        row_key,
                        details.ssn if details else None,
                        details.birth_month if details else None,
                        details.birth_day if details else None,
                        details.birth_year if details else None,
                        details.first_name if details else None,
                        details.last_name if details else None,
                        record.address, status, record.import_error, _now_iso(),
                        _now_iso() if record.import_error else None,
                    ),
                )
                if details is not None and not record.import_error:
                    item = WorkItem(
                        int(cursor.lastrowid), details, record.source_file,
                        record.source_sheet, record.source_row,
                        record.original_fields, record.address,
                        record.input_mapping,
                    )
                    if row_key not in queued_keys:
                        queued_keys.add(row_key)
                        work_items.append(item)
            connection.execute(
                """
                UPDATE batches SET status='running', total_count=? WHERE batch_id=?
                """,
                (len(records), batch_id),
            )
            return work_items

        if action == "mark_processing":
            connection.execute(
                """
                UPDATE records
                SET status='processing', started_at=?, attempt_count=attempt_count+1,
                    error=''
                WHERE id=? AND status='pending'
                """,
                (_now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_completed":
            result: RecoveryResult = payload["result"]
            connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE id=?
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["record_id"],
                ),
            )
            return None

        if action == "mark_completed_group":
            result = payload["result"]
            cursor = connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE batch_id=? AND input_row_key=?
                  AND status IN ('pending', 'processing')
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["batch_id"], payload["record_key"],
                ),
            )
            return int(cursor.rowcount)

        if action == "mark_retry_group":
            cursor = connection.execute(
                """
                UPDATE records
                SET status='pending', error=?, started_at=NULL, finished_at=NULL
                WHERE batch_id=? AND input_row_key=?
                  AND status IN ('pending', 'processing')
                """,
                (
                    payload.get("error", ""), payload["batch_id"],
                    payload["record_key"],
                ),
            )
            return int(cursor.rowcount)

        if action in {"mark_failed_group", "mark_stopped_group"}:
            status = "failed" if action == "mark_failed_group" else "stopped"
            allowed_statuses = (
                "('pending', 'processing', 'completed')"
                if action == "mark_failed_group"
                else "('pending', 'processing')"
            )
            cursor = connection.execute(
                f"""
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND input_row_key=? AND status IN {allowed_statuses}
                """,
                (
                    status, payload.get("error", ""), _now_iso(),
                    payload["batch_id"], payload["record_key"],
                ),
            )
            return int(cursor.rowcount)

        if action in {"mark_failed", "mark_stopped"}:
            status = "failed" if action == "mark_failed" else "stopped"
            connection.execute(
                "UPDATE records SET status=?, error=?, finished_at=? WHERE id=?",
                (status, payload.get("error", ""), _now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_remaining":
            status = str(payload["status"])
            connection.execute(
                """
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND status IN ('pending', 'processing')
                """,
                (
                    status, payload.get("error", ""), _now_iso(), payload["batch_id"],
                ),
            )
            return None

        if action == "status_counts":
            rows = connection.execute(
                """
                SELECT status, COUNT(*) AS count FROM records
                WHERE batch_id=? GROUP BY status
                """,
                (payload["batch_id"],),
            ).fetchall()
            return {str(row["status"]): int(row["count"]) for row in rows}

        if action == "finish_batch":
            connection.execute(
                """
                UPDATE batches SET status=?, finished_at=?, export_path=?
                WHERE batch_id=?
                """,
                (
                    payload["status"], _now_iso(), payload.get("export_path", ""),
                    payload["batch_id"],
                ),
            )
            return None

        if action == "checkpoint":
            return connection.execute("PRAGMA wal_checkpoint(PASSIVE)").fetchone()

        if action == "shutdown":
            return None

        raise ValueError(f"未知 SQLite 写入命令：{action}")


def export_batch_csv(
    database_path: Path, batch_id: str, output_directory: Path
) -> Path:
    """只读查询当前批次并原子生成带 BOM 的最终 CSV。"""
    output_directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_directory / f"studentaid_results_{timestamp}_{batch_id[:8]}.csv"
    temporary_path = output_path.with_name(f".{output_path.name}.{os.getpid()}.tmp")

    connection = sqlite3.connect(database_path, timeout=30)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            "SELECT * FROM records WHERE batch_id=? ORDER BY id",
            (batch_id,),
        ).fetchall()
    finally:
        connection.close()

    original_rows: list[list[str]] = []
    maximum_columns = 0
    for row in rows:
        fields = [str(value) for value in json.loads(row["original_fields_json"])]
        original_rows.append(fields)
        maximum_columns = max(maximum_columns, len(fields))

    try:
        with temporary_path.open("w", encoding="utf-8-sig", newline="") as output_file:
            writer = csv.writer(output_file, lineterminator="\n")
            writer.writerow(
                [*(f"Input_Column_{index}" for index in range(1, maximum_columns + 1)),
                 *RESULT_COLUMNS]
            )
            for row, original_fields in zip(rows, original_rows):
                padded = [*original_fields, *("" for _ in range(maximum_columns - len(original_fields)))]
                writer.writerow(
                    [
                        *padded,
                        row["masked_phone"], row["masked_email"],
                        row["recovery_method"], row["status"], row["error"],
                        Path(row["source_file"]).name, row["source_sheet"],
                        row["source_row"], row["result_heading"],
                    ]
                )
            output_file.flush()
            os.fsync(output_file.fileno())
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path


EventCallback = Callable[[str, dict[str, Any]], None]
SessionFactory = Callable[..., Any]


class BatchEngine:
    """第十二步并发处理：每个 worker 使用所选后端和显示模式。"""

    def __init__(
        self,
        event_callback: EventCallback | None = None,
        session_factory: SessionFactory = BrowserRecoverySession,
    ) -> None:
        self._callback = event_callback or (lambda _kind, _payload: None)
        self._session_factory = session_factory
        self._stop_event = threading.Event()
        self._state_lock = threading.Lock()
        self._running = False
        self._coordinator: threading.Thread | None = None

    @property
    def is_running(self) -> bool:
        with self._state_lock:
            return self._running

    def start(
        self,
        input_path: Path,
        output_target: Path,
        thread_count: int = 2,
        backend: str = "playwright",
        display_mode: str = "无头",
    ) -> None:
        input_path = input_path.resolve()
        output_directory, output_path = resolve_output_target(output_target)
        backend = backend.strip().casefold()
        display_mode = _normalise_display_mode(display_mode)
        if not input_path.is_file():
            raise FileNotFoundError(f"输入文件不存在：{input_path}")
        if input_path == output_path:
            raise ValueError("输入文件和累计输出文件不能是同一个文件")
        if thread_count < 1:
            raise ValueError("处理线程数必须是大于等于 1 的整数")
        if backend not in PROCESSING_BACKENDS:
            raise ValueError(f"搜索后端必须是：{', '.join(PROCESSING_BACKENDS)}")
        database_path = output_directory / DATABASE_FILENAME
        ensure_output_storage_writable(
            output_directory,
            output_path,
            database_path,
            Path(f"{database_path}-wal"),
            Path(f"{database_path}-shm"),
        )
        with self._state_lock:
            if self._running:
                raise RuntimeError("已有批次正在运行")
            self._running = True
        self._stop_event.clear()
        self._coordinator = threading.Thread(
            target=self._run,
            args=(input_path, output_path, thread_count, backend, display_mode),
            name="studentaid-coordinator",
            daemon=False,
        )
        self._coordinator.start()

    def stop(self) -> None:
        if self.is_running:
            self._stop_event.set()
            self._emit(
                "log",
                message="已收到停止请求；正在停止领取新任务、清除缓存并结束浏览器进程。",
            )

    def wait(self, timeout: float | None = None) -> bool:
        coordinator = self._coordinator
        if coordinator is not None:
            coordinator.join(timeout)
        return not self.is_running

    def _emit(self, kind: str, **payload: Any) -> None:
        try:
            self._callback(kind, payload)
        except Exception:
            pass

    def _publish_progress(
        self, writer: DatabaseWriter, batch_id: str, total: int
    ) -> dict[str, int]:
        counts = writer.request("status_counts", batch_id=batch_id)
        payload = {
            "total": total,
            "pending": int(counts.get("pending", 0)),
            "processing": int(counts.get("processing", 0)),
            "completed": int(counts.get("completed", 0)),
            "failed": int(counts.get("failed", 0)),
            "stopped": int(counts.get("stopped", 0)),
        }
        self._emit("progress", **payload)
        return payload

    def _new_session(
        self, worker_number: int, backend: str, display_mode: str
    ) -> Any:
        if self._session_factory is BrowserRecoverySession:
            if backend == "http":
                return HttpRecoverySession(worker_number)
            return self._session_factory(worker_number, backend, display_mode)
        return self._session_factory(worker_number)

    def _worker_loop(
        self,
        worker_number: int,
        backend: str,
        display_mode: str,
        tasks: queue.Queue[tuple[WorkItem, int] | None],
        writer: DatabaseWriter,
        batch_id: str,
        total: int,
        persistence: ResultPersistenceWriter,
    ) -> None:
        session = self._new_session(worker_number, backend, display_mode)
        try:
            session.start()
            self._emit(
                "log",
                message=(
                    f"处理线程 {worker_number} 已启动："
                    f"{getattr(session, 'browser_mode', '模式未知')}。"
                ),
            )
        except Exception as exc:
            self._emit(
                "log",
                message=f"处理线程 {worker_number} 启动失败：{_clean_error(exc)}",
            )
            try:
                session.close()
            except Exception:
                pass
            return

        try:
            while not self._stop_event.is_set():
                try:
                    queued_task = tasks.get(timeout=0.2)
                except queue.Empty:
                    continue
                if queued_task is None:
                    tasks.task_done()
                    break
                item, queue_attempt = queued_task
                if self._stop_event.is_set():
                    tasks.task_done()
                    break

                writer.request("mark_processing", record_id=item.record_id)
                self._publish_progress(writer, batch_id, total)
                progress = lambda stage, record_id=item.record_id: self._emit(
                    "log", message=f"记录 #{record_id}：{stage}"
                )
                try:
                    try:
                        result = session.process(item, self._stop_event, progress)
                    except StopRequested:
                        raise
                    except Exception as process_exc:
                        if queue_attempt >= MAX_QUEUE_ATTEMPTS:
                            raise
                        recover = getattr(
                            session, "recover_after_cleanup_error", None
                        )
                        if callable(recover):
                            try:
                                recover()
                            except Exception:
                                pass
                        retried = writer.request(
                            "mark_retry_group", batch_id=batch_id,
                            record_key=item.record_key,
                            error=_clean_error(process_exc),
                        )
                        tasks.put((item, queue_attempt + 1))
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：本轮未取得明确结果，"
                                f"当前整行 {retried} 条已回到队尾；稍后执行"
                                f"第 {queue_attempt + 1}/{MAX_QUEUE_ATTEMPTS} 轮。"
                            ),
                        )
                        continue
                    completed_group = writer.request(
                        "mark_completed_group", batch_id=batch_id,
                        record_key=item.record_key, result=result,
                    )
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：明确结果已实时写入 SQLite，"
                            f"完全相同整行完成 {completed_group} 条。"
                        ),
                    )
                    appended = persistence.commit(item, result)
                    if appended:
                        self._emit(
                            "log",
                            message=f"记录 #{item.record_id}：已实时追加到累计输出 CSV。",
                        )
                    else:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：累计输出中完整资料行已存在，"
                                "跳过完全重复追加。"
                            ),
                        )
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：明确结果已落盘；"
                            "输入删除已进入批量持久化队列，浏览器继续下一条。"
                        ),
                    )
                    try:
                        prepare = getattr(session, "prepare_for_next", None)
                        if callable(prepare):
                            prepare(result.result_code, self._stop_event, progress)
                    except StopRequested:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "收到停止请求，跳过下一页准备。"
                            ),
                        )
                    except Exception as cleanup_exc:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "页面清理失败，将为下一条重建页面："
                                f"{_clean_error(cleanup_exc)}"
                            ),
                        )
                        recover = getattr(session, "recover_after_cleanup_error", None)
                        if callable(recover):
                            recover()
                except StopRequested as exc:
                    writer.request(
                        "mark_stopped_group", batch_id=batch_id,
                        record_key=item.record_key,
                        error=_clean_error(exc),
                    )
                except Exception as exc:
                    writer.request(
                        "mark_failed_group", batch_id=batch_id,
                        record_key=item.record_key,
                        error=_clean_error(exc),
                    )
                    self._emit(
                        "log",
                        message=f"记录 #{item.record_id} 处理失败：{_clean_error(exc)}",
                    )
                    recover = getattr(session, "recover_after_cleanup_error", None)
                    if callable(recover):
                        try:
                            recover()
                        except Exception:
                            pass
                finally:
                    tasks.task_done()
                    self._publish_progress(writer, batch_id, total)
        finally:
            try:
                session.close()
                self._emit(
                    "log",
                    message=(
                        f"处理线程 {worker_number} 已清除浏览器缓存并结束浏览器进程。"
                    ),
                )
            except Exception as exc:
                self._emit(
                    "log",
                    message=f"处理线程 {worker_number} 关闭浏览器失败：{_clean_error(exc)}",
                )

    def _run(
        self,
        input_path: Path,
        output_path: Path,
        thread_count: int,
        backend: str,
        display_mode: str,
    ) -> None:
        batch_id = uuid.uuid4().hex
        output_directory = output_path.parent
        database_path = output_directory / DATABASE_FILENAME
        writer: DatabaseWriter | None = None
        persistence: ResultPersistenceWriter | None = None
        total = 0
        try:
            self._emit(
                "log",
                message=(
                    "正在读取累计输出完整资料行并导入输入文件；"
                    "大型 XLSX 在批次结束时一次性同步删除……"
                ),
            )
            layout = normalise_cumulative_output_layout(output_path)
            if layout.changed:
                self._emit(
                    "log",
                    message=(
                        f"累计 CSV 已统一 {layout.row_count} 行：输入区固定 "
                        f"{layout.input_width} 列，结果状态固定在最后第 "
                        f"{layout.input_width + CUMULATIVE_RESULT_FIELD_COUNT} 列；"
                        "旧布局、空行和双引号格式已原子迁移。"
                    ),
                )
            existing_keys = read_output_row_keys(output_path)
            defer_xlsx_rewrite = input_path.suffix.casefold() == ".xlsx"
            removed_before_start = (
                0
                if defer_xlsx_rewrite
                else remove_input_rows_by_keys(input_path, existing_keys)
            )
            if removed_before_start:
                self._emit(
                    "log",
                    message=f"输入文件已删除 {removed_before_start} 条输出中整行已存在的资料。",
                )
            try:
                records = load_input_records(input_path)
            except ValueError as exc:
                if "没有可导入的资料行" not in str(exc):
                    raise
                records = []
            matched_existing_keys: set[str] = set()
            if defer_xlsx_rewrite and existing_keys:
                records, matched_existing_keys, skipped_existing = (
                    filter_records_already_output(records, existing_keys)
                )
                if skipped_existing:
                    self._emit(
                        "log",
                        message=(
                            f"累计输出中已有 {skipped_existing} 条完整资料；"
                            "本次直接跳过搜索，并在批次结束时一次性从输入 XLSX 删除。"
                        ),
                    )
            missing_required = [
                record
                for record in records
                if record.import_error.startswith("缺少必填字段：")
            ]
            if missing_required:
                removed_missing = remove_input_rows_by_locations(
                    input_path, missing_required
                )
                self._emit(
                    "log",
                    message=(
                        f"输入文件已直接删除 {removed_missing} 条缺少必填字段的资料；"
                        "这些资料不写累计输出、不进入浏览器。"
                    ),
                )
                records = [
                    record for record in records if record not in missing_required
                ]
            total = len(records)
            if self._stop_event.is_set():
                raise StopRequested("导入阶段已停止")

            writer = DatabaseWriter(database_path)
            writer.start()
            writer.request(
                "create_batch",
                batch_id=batch_id,
                input_path=str(input_path),
                output_directory=str(output_directory),
                thread_count=thread_count,
            )
            work_items: list[WorkItem] = writer.request(
                "insert_records", batch_id=batch_id, records=records
            )
            persistence = ResultPersistenceWriter(
                input_path,
                output_path,
                existing_keys,
                initial_pending_delete_keys=matched_existing_keys,
                log_callback=lambda message: self._emit("log", message=message),
            )
            persistence.start()
            invalid_count = sum(bool(record.import_error) for record in records)
            valid_count = len(records) - invalid_count
            duplicate_count = valid_count - len(work_items)
            if duplicate_count:
                removed_duplicates = remove_duplicate_input_rows(
                    input_path,
                )
                self._emit(
                    "log",
                    message=(
                        f"输入文件已删除 {removed_duplicates} 条完全重复整行；"
                        "同一 SSN 但其他字段不同的资料全部保留。"
                    ),
                )
            self._emit(
                "log",
                message=(
                    f"输入剩余 {len(records)} 条；唯一浏览器任务 {len(work_items)} 个，"
                    f"完全重复整行 {duplicate_count} 条，格式错误 {invalid_count} 条。"
                    f"后端 {backend}，"
                    f"显示 {display_mode}，处理线程 {thread_count}。"
                ),
            )
            self._publish_progress(writer, batch_id, total)

            tasks: queue.Queue[tuple[WorkItem, int] | None] = queue.Queue()
            for item in work_items:
                tasks.put((item, 1))
            worker_count = min(thread_count, len(work_items))
            workers = [
                threading.Thread(
                    target=self._worker_loop,
                    args=(
                        number,
                        backend,
                        display_mode,
                        tasks,
                        writer,
                        batch_id,
                        total,
                        persistence,
                    ),
                    name=f"studentaid-worker-{number}",
                    daemon=False,
                )
                for number in range(1, worker_count + 1)
            ]
            for worker in workers:
                worker.start()
            while workers and not self._stop_event.is_set():
                with tasks.all_tasks_done:
                    unfinished_tasks = tasks.unfinished_tasks
                if unfinished_tasks == 0 or not any(
                    worker.is_alive() for worker in workers
                ):
                    break
                time.sleep(0.1)
            for _ in range(worker_count):
                tasks.put(None)
            for worker in workers:
                worker.join()
            persistence.close()
            persistence = None

            if self._stop_event.is_set():
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="stopped",
                    error="用户停止，任务尚未处理",
                )
                batch_status = "stopped"
            else:
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="failed",
                    error="任务未被浏览器处理",
                )
                batch_status = "completed"
            final_counts = self._publish_progress(writer, batch_id, total)
            writer.request(
                "finish_batch", batch_id=batch_id, status=batch_status,
                export_path=str(output_path),
            )
            writer.request("checkpoint")
            self._emit(
                "finished",
                batch_id=batch_id,
                database_path=str(database_path),
                export_path=str(output_path),
                status=batch_status,
                counts=final_counts,
            )
        except StopRequested as exc:
            self._emit("fatal", message=_clean_error(exc), stopped=True)
        except Exception as exc:
            self._emit(
                "fatal",
                message=_clean_error(exc),
                details=traceback.format_exc(limit=8),
                stopped=False,
            )
        finally:
            if persistence is not None:
                try:
                    persistence.close()
                except Exception as exc:
                    self._emit(
                        "log",
                        message=f"关闭结果持久化线程时出错：{_clean_error(exc)}",
                    )
            if writer is not None:
                try:
                    writer.close()
                except Exception as exc:
                    self._emit("log", message=f"关闭数据库时出错：{_clean_error(exc)}")
            with self._state_lock:
                self._running = False
            self._emit("idle", export_path=str(output_path))


class StudentAidApp:
    """配置可持久化、带实时效率看板的桌面 GUI。"""

    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("960x700")
        self.root.minsize(860, 620)

        saved_config = load_gui_config()
        self._config_path = gui_config_path()
        self._config_save_after_id: str | None = None
        self.input_var = tk.StringVar(value=saved_config["input_path"])
        self.output_var = tk.StringVar(value=saved_config["output_path"])
        self.backend_var = tk.StringVar(value=saved_config["backend"])
        self.display_mode_var = tk.StringVar(value=saved_config["display_mode"])
        self.thread_var = tk.StringVar(value=saved_config["thread_count"])
        self.status_var = tk.StringVar(value="就绪")
        self.progress_text_var = tk.StringVar(
            value="进度 0.0% | 已处理 0/0 | 剩余 0 | 待处理 0 | 处理中 0"
        )
        self.realtime_text_var = tk.StringVar(
            value=(
                "最近1分钟处理 0 条 | 平均 0.00 条/分钟 | 预计剩余 -- | "
                "已运行 00:00:00 | 明确结果 0 | 失败 0 | 停止 0"
            )
        )
        self._progress_tracker = BatchProgressTracker()
        self._last_metrics_render = 0.0
        self._ui_events: queue.Queue[tuple[str, dict[str, Any]]] = queue.Queue()
        self.engine = BatchEngine(self._backend_event)
        self._database_clear_active = False
        self._database_clear_thread: threading.Thread | None = None
        self._closing = False

        self._build_ui()
        for variable in (
            self.input_var,
            self.output_var,
            self.backend_var,
            self.display_mode_var,
            self.thread_var,
        ):
            variable.trace_add("write", self._schedule_config_save)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._poll_events)

    def _current_gui_config(self) -> dict[str, str]:
        return {
            "input_path": self.input_var.get().strip(),
            "output_path": self.output_var.get().strip(),
            "backend": self.backend_var.get().strip(),
            "display_mode": self.display_mode_var.get().strip(),
            "thread_count": self.thread_var.get().strip(),
        }

    def _save_gui_config(self) -> None:
        self._config_save_after_id = None
        save_gui_config(self._current_gui_config(), self._config_path)

    def _schedule_config_save(self, *_args: object) -> None:
        if self._config_save_after_id is not None:
            try:
                self.root.after_cancel(self._config_save_after_id)
            except Exception:
                pass
        self._config_save_after_id = self.root.after(300, self._save_gui_config)

    def _build_ui(self) -> None:
        ttk = self.ttk
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(9, weight=1)

        ttk.Label(main, text="导入文件：").grid(row=0, column=0, sticky="w", pady=6)
        self.input_entry = ttk.Entry(main, textvariable=self.input_var)
        self.input_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        self.input_button = ttk.Button(main, text="选择文件", command=self._choose_input)
        self.input_button.grid(row=0, column=2, pady=6)

        ttk.Label(main, text="累计输出 CSV：").grid(row=1, column=0, sticky="w", pady=6)
        self.output_entry = ttk.Entry(main, textvariable=self.output_var)
        self.output_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        self.output_button = ttk.Button(main, text="选择文件", command=self._choose_output)
        self.output_button.grid(row=1, column=2, pady=6)

        ttk.Label(main, text="搜索后端：").grid(row=2, column=0, sticky="w", pady=6)
        self.backend_combo = ttk.Combobox(
            main,
            textvariable=self.backend_var,
            values=PROCESSING_BACKENDS,
            width=18,
            state="readonly",
        )
        self.backend_combo.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="http=官方接口；browser-use/playwright=独立浏览器。")\
            .grid(row=2, column=1, sticky="w", padx=(180, 0), pady=6)

        ttk.Label(main, text="浏览器显示：").grid(row=3, column=0, sticky="w", pady=6)
        self.display_mode_combo = ttk.Combobox(
            main,
            textvariable=self.display_mode_var,
            values=DISPLAY_MODES,
            width=18,
            state="readonly",
        )
        self.display_mode_combo.grid(row=3, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="仅浏览器后端生效；HTTP 后端忽略此项。")\
            .grid(row=3, column=1, sticky="w", padx=(180, 0), pady=6)

        ttk.Label(main, text="处理线程数：").grid(row=4, column=0, sticky="w", pady=6)
        self.thread_spin = ttk.Entry(main, textvariable=self.thread_var, width=10)
        self.thread_spin.grid(row=4, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="任意正整数；每个线程使用独立 HTTP 会话或浏览器。")\
            .grid(row=4, column=1, sticky="w", padx=(100, 0), pady=6)

        button_bar = ttk.Frame(main)
        button_bar.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(10, 8))
        self.start_button = ttk.Button(button_bar, text="开始", command=self._start)
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(
            button_bar, text="停止", command=self._stop, state="disabled"
        )
        self.stop_button.pack(side="left", padx=8)
        self.clear_database_button = ttk.Button(
            button_bar, text="清空数据库", command=self._clear_database
        )
        self.clear_database_button.pack(side="left")
        ttk.Label(button_bar, textvariable=self.status_var).pack(side="right")

        self.progress = ttk.Progressbar(main, mode="determinate", maximum=1, value=0)
        self.progress.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(2, 4))
        ttk.Label(main, textvariable=self.progress_text_var).grid(
            row=7, column=0, columnspan=3, sticky="w", pady=(0, 2)
        )
        ttk.Label(main, textvariable=self.realtime_text_var).grid(
            row=8, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        log_frame = ttk.LabelFrame(main, text="运行日志（不显示 SSN 和原始资料）", padding=8)
        log_frame.grid(row=9, column=0, columnspan=3, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.log_text = self.tk.Text(log_frame, wrap="word", state="disabled")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

    def _choose_input(self) -> None:
        path = self.filedialog.askopenfilename(
            title="选择导入文件",
            filetypes=[
                ("支持的资料文件", "*.csv *.scv *.txt *.xlsx"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
                ("纯文本文件", "*.txt *.scv"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            self.input_var.set(path)
            current = self.output_var.get().strip()
            if not current or current == str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME):
                self.output_var.set(str(Path(path).parent / CUMULATIVE_OUTPUT_FILENAME))

    def _choose_output(self) -> None:
        current = Path(self.output_var.get().strip() or CUMULATIVE_OUTPUT_FILENAME)
        path = self.filedialog.asksaveasfilename(
            title="选择累计输出 CSV（已有文件会继续叠加）",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if path:
            self.output_var.set(path)

    def _clear_database(self) -> None:
        if self.engine.is_running:
            self.messagebox.showwarning(APP_TITLE, "请在当前批次结束后清空数据库。")
            return
        if getattr(self, "_database_clear_active", False):
            self.messagebox.showinfo(APP_TITLE, "数据库正在清空，请稍候。")
            return
        output_text = self.output_var.get().strip()
        if not output_text:
            self.messagebox.showerror(APP_TITLE, "请先选择累计输出 CSV。")
            return
        try:
            database_path = database_path_for_output_target(Path(output_text))
        except Exception as exc:
            self.messagebox.showerror(APP_TITLE, _clean_error(exc))
            return
        if not database_path.is_file():
            self.messagebox.showinfo(
                APP_TITLE, f"数据库当前不存在，无需清空。\n\n{database_path}"
            )
            return
        confirmed = self.messagebox.askyesno(
            APP_TITLE,
            (
                "确定清空当前数据库中的全部批次和记录吗？\n\n"
                f"数据库：{database_path}\n\n"
                "累计 CSV、输入文件和 GUI 配置不会被删除。"
            ),
        )
        if not confirmed:
            return

        self._database_clear_active = True
        self.status_var.set("正在清空数据库")
        self._set_running(False)
        self._append_log(f"正在后台清空数据库：{database_path}")
        self._database_clear_thread = threading.Thread(
            target=self._clear_database_worker,
            args=(database_path,),
            name="studentaid-database-clear",
            daemon=True,
        )
        try:
            self._database_clear_thread.start()
        except Exception as exc:
            self._database_clear_active = False
            self._database_clear_thread = None
            self._set_running(False)
            self.status_var.set("清空数据库失败")
            self.messagebox.showerror(APP_TITLE, f"清空数据库失败：{_clean_error(exc)}")
            return

    def _clear_database_worker(self, database_path: Path) -> None:
        """后台执行 SQLite 维护；只投递事件，不从工作线程调用 Tk。"""
        try:
            result = clear_database_contents(database_path)
        except Exception as exc:
            self._ui_events.put(
                ("database_clear_failed", {"message": _clean_error(exc)})
            )
        else:
            self._ui_events.put(("database_clear_finished", {"result": result}))

    def _finish_database_clear(self, result: DatabaseClearResult) -> None:
        self._database_clear_active = False
        self._database_clear_thread = None
        self._set_running(self.engine.is_running)
        message = (
            f"数据库已清空。\n\n批次：{result.batches_deleted} 条\n"
            f"记录：{result.records_deleted} 条\n\n{result.database_path}"
        )
        self.status_var.set("数据库已清空")
        self._append_log(
            f"数据库已清空：批次 {result.batches_deleted} 条，"
            f"记录 {result.records_deleted} 条。"
        )
        if not self._closing:
            self.messagebox.showinfo(APP_TITLE, message)

    def _fail_database_clear(self, message: str) -> None:
        self._database_clear_active = False
        self._database_clear_thread = None
        self._set_running(self.engine.is_running)
        self.status_var.set("清空数据库失败")
        self._append_log(f"清空数据库失败：{message}")
        if not self._closing:
            self.messagebox.showerror(APP_TITLE, f"清空数据库失败：{message}")

    def _set_running(self, running: bool) -> None:
        controls_blocked = running or bool(
            getattr(self, "_database_clear_active", False)
        )
        normal_or_disabled = "disabled" if controls_blocked else "normal"
        self.input_entry.configure(state=normal_or_disabled)
        self.output_entry.configure(state=normal_or_disabled)
        self.backend_combo.configure(
            state="disabled" if controls_blocked else "readonly"
        )
        self.display_mode_combo.configure(
            state="disabled" if controls_blocked else "readonly"
        )
        self.thread_spin.configure(state=normal_or_disabled)
        self.input_button.configure(state=normal_or_disabled)
        self.output_button.configure(state=normal_or_disabled)
        self.start_button.configure(state=normal_or_disabled)
        self.clear_database_button.configure(state=normal_or_disabled)
        self.stop_button.configure(state="normal" if running else "disabled")

    def _start(self) -> None:
        if getattr(self, "_database_clear_active", False):
            self.messagebox.showwarning(APP_TITLE, "数据库正在清空，请稍候再开始批次。")
            return
        try:
            input_path = Path(self.input_var.get().strip())
            output_target = Path(self.output_var.get().strip())
            thread_count = int(self.thread_var.get().strip())
            backend = self.backend_var.get().strip()
            display_mode = self.display_mode_var.get().strip()
            if thread_count < 1:
                raise ValueError("处理线程数必须是大于等于 1 的整数")
            self._save_gui_config()
            self.engine.start(
                input_path, output_target, thread_count, backend, display_mode
            )
        except Exception as exc:
            self.messagebox.showerror(APP_TITLE, _clean_error(exc))
            return
        self._progress_tracker.start()
        self._render_progress_metrics(force=True)
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self.status_var.set("运行中")
        self._set_running(True)
        self._append_log(
            f"批次已启动：后端 {self.backend_var.get()}，"
            f"显示 {self.display_mode_var.get()}，线程 {self.thread_var.get()}。"
        )

    def _stop(self) -> None:
        self.stop_button.configure(state="disabled")
        self.status_var.set("停止中")
        self.engine.stop()

    def _backend_event(self, kind: str, payload: dict[str, Any]) -> None:
        self._ui_events.put((kind, payload))

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _render_progress_metrics(self, force: bool = False) -> None:
        now = time.monotonic()
        if not force and now - self._last_metrics_render < 1.0:
            return
        self._last_metrics_render = now
        metrics = self._progress_tracker.snapshot(now)
        self.progress.configure(
            maximum=max(1, metrics.total), value=metrics.terminal
        )
        self.progress_text_var.set(
            f"进度 {metrics.percent:.1f}% | "
            f"已处理 {metrics.terminal}/{metrics.total} | "
            f"剩余 {metrics.remaining} | 待处理 {metrics.pending} | "
            f"处理中 {metrics.processing}"
        )
        self.realtime_text_var.set(
            f"最近1分钟处理 {metrics.recent_minute_count} 条 | "
            f"平均 {metrics.average_per_minute:.2f} 条/分钟 | "
            f"预计剩余 {format_duration(metrics.eta_seconds)} | "
            f"已运行 {format_duration(metrics.elapsed_seconds)} | "
            f"明确结果 {metrics.completed} | 失败 {metrics.failed} | "
            f"停止 {metrics.stopped}"
        )

    def _poll_events(self) -> None:
        try:
            while True:
                kind, payload = self._ui_events.get_nowait()
                if kind == "log":
                    self._append_log(str(payload.get("message", "")))
                elif kind == "progress":
                    self._progress_tracker.update(payload)
                    self._render_progress_metrics(force=True)
                elif kind == "finished":
                    self._progress_tracker.finish()
                    self._render_progress_metrics(force=True)
                    output_path = str(payload.get("export_path", ""))
                    database_path = str(payload.get("database_path", ""))
                    counts = payload.get("counts", {})
                    failed = int(counts.get("failed", 0)) if isinstance(counts, dict) else 0
                    if payload.get("status") == "stopped":
                        display_status = "已停止"
                    elif failed:
                        display_status = "完成（有失败）"
                    else:
                        display_status = "已完成"
                    self.status_var.set(display_status)
                    self._append_log(f"累计 CSV：{output_path}")
                    self._append_log(f"SQLite 数据库：{database_path}")
                    self.messagebox.showinfo(
                        APP_TITLE,
                        f"批次处理结束。\n\n累计 CSV：{output_path}\n数据库：{database_path}",
                    )
                elif kind == "fatal":
                    self._progress_tracker.finish()
                    self._render_progress_metrics(force=True)
                    message = str(payload.get("message", "未知错误"))
                    self.status_var.set("已停止" if payload.get("stopped") else "失败")
                    self._append_log(message)
                    if not payload.get("stopped"):
                        self.messagebox.showerror(APP_TITLE, message)
                elif kind == "database_clear_finished":
                    result = payload.get("result")
                    if isinstance(result, DatabaseClearResult):
                        self._finish_database_clear(result)
                    else:
                        self._fail_database_clear("清理线程没有返回有效结果")
                elif kind == "database_clear_failed":
                    self._fail_database_clear(
                        str(payload.get("message", "未知数据库错误"))
                    )
                elif kind == "idle":
                    self._progress_tracker.finish()
                    self._render_progress_metrics(force=True)
                    self._set_running(False)
                self._ui_events.task_done()
        except queue.Empty:
            pass

        if self.engine.is_running:
            self._render_progress_metrics()

        if (
            self._closing
            and not self.engine.is_running
            and not self._database_clear_active
        ):
            self.root.destroy()
            return
        self.root.after(100, self._poll_events)

    def _on_close(self) -> None:
        self._save_gui_config()
        if self._database_clear_active:
            if not self.messagebox.askyesno(
                APP_TITLE,
                "数据库正在清空。是否等待清理结束后自动退出？",
            ):
                return
            self._closing = True
            self.status_var.set("数据库清理结束后退出")
            self._set_running(False)
        elif self.engine.is_running:
            if not self.messagebox.askyesno(APP_TITLE, "任务仍在运行，是否停止并退出？"):
                return
            self._closing = True
            self._set_running(True)
            self.stop_button.configure(state="disabled")
            self.status_var.set("停止中")
            self.engine.stop()
        else:
            self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    StudentAidApp().run()


if __name__ == "__main__":
    main()
