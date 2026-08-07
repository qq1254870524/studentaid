import csv
import os
from pathlib import Path
import tempfile
import threading
import time
import unittest

from openpyxl import Workbook, load_workbook

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import ait5


class SuccessSession:
    prepared = []

    def __init__(self, worker_number: int) -> None:
        self.worker_number = worker_number
        self.browser_mode = "mock-dedicated-cdp"

    def start(self) -> None:
        pass

    def process(self, item, stop_event, progress_callback=None):
        if progress_callback:
            progress_callback("mock result")
        return ait5.RecoveryResult(
            result_code="account_not_found",
            heading="Account Not Found",
            masked_phone="",
            masked_email="",
            recovery_method="",
        )

    def prepare_for_next(self, result_code, stop_event, progress_callback=None):
        self.prepared.append(result_code)

    def close(self) -> None:
        pass


class RateLimitedSession(SuccessSession):
    def process(self, item, stop_event, progress_callback=None):
        raise ait5.SiteRateLimitReached("rate limited for test")


class StopAfterResultSession(SuccessSession):
    def process(self, item, stop_event, progress_callback=None):
        result = super().process(item, stop_event, progress_callback)
        stop_event.set()
        return result


class StudentAidStep10Tests(unittest.TestCase):
    def setUp(self):
        SuccessSession.prepared = []

    @staticmethod
    def _write_input(path: Path, keys=("111223333", "222334444")) -> None:
        rows = [
            [keys[0], "01", "02", "1990", "Example", "First", "Address 1"],
            [keys[1], "03", "04", "1991", "Example", "Second", "Address 2"],
        ]
        with path.open("w", encoding="utf-8", newline="") as stream:
            csv.writer(stream).writerows(rows)

    def _run_engine(self, input_path, output_path, session_factory=SuccessSession):
        events = []
        engine = ait5.BatchEngine(
            event_callback=lambda kind, payload: events.append((kind, payload)),
            session_factory=session_factory,
        )
        engine.start(input_path, output_path, 1)
        self.assertTrue(engine.wait(20), "engine did not finish")
        finished = [payload for kind, payload in events if kind == "finished"]
        self.assertEqual(len(finished), 1, events)
        return finished[0], events

    def test_cumulative_output_is_appended_and_input_rows_are_deleted(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)
            output_path.write_text(
                "111-22-3333,01/02/1990,First,Example,Address 1,Account Not Found,,,\n",
                encoding="utf-8-sig",
            )

            finished, _events = self._run_engine(input_path, output_path)

            self.assertEqual(finished["counts"]["completed"], 1)
            self.assertEqual(input_path.read_text(encoding="utf-8"), "")
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
            self.assertEqual([ait5._normalise_record_key(row[0]) for row in rows], [
                "111223333", "222334444"
            ])
            self.assertEqual(SuccessSession.prepared, ["account_not_found"])

            finished_again, _events = self._run_engine(input_path, output_path)
            self.assertEqual(finished_again["counts"]["completed"], 0)
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                self.assertEqual(len(list(csv.reader(stream))), 2)

    def test_rate_limit_keeps_current_and_remaining_input_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)

            finished, events = self._run_engine(
                input_path, output_path, RateLimitedSession
            )

            self.assertEqual(finished["status"], "rate_limited")
            self.assertEqual(finished["counts"]["failed"], 1)
            self.assertEqual(finished["counts"]["stopped"], 1)
            self.assertFalse(output_path.exists())
            self.assertEqual(len(ait5.load_input_records(input_path)), 2)
            messages = [p.get("message", "") for k, p in events if k == "log"]
            self.assertTrue(any("全部保留" in message for message in messages))

    def test_stop_after_result_does_not_overwrite_completed_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)

            finished, _events = self._run_engine(
                input_path, output_path, StopAfterResultSession
            )

            self.assertEqual(finished["status"], "stopped")
            self.assertEqual(finished["counts"]["completed"], 1)
            self.assertEqual(finished["counts"]["stopped"], 1)
            self.assertEqual(len(ait5.load_input_records(input_path)), 1)
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                self.assertEqual(len(list(csv.reader(stream))), 1)

    def test_xlsx_first_column_deletion_preserves_header_and_other_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["SSN", "DOB", "First Name", "Last Name", "Address"])
            sheet.append(["111-22-3333", "01/02/1990", "First", "Example", "A"])
            sheet.append(["222334444", "03/04/1991", "Second", "Example", "B"])
            workbook.save(path)
            workbook.close()

            removed = ait5.remove_input_rows_by_keys(path, {"111223333"})

            self.assertEqual(removed, 1)
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(rows[0][0], "SSN")
            self.assertEqual(rows[1][0], "222334444")

    @unittest.skipIf(ait5.sync_playwright is None, "playwright not installed")
    def test_limit_reached_is_detected_immediately(self):
        with ait5.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content("<h2>Limit Reached: Try Again in 24 Hours</h2>")
            started = time.monotonic()
            try:
                with self.assertRaises(ait5.SiteRateLimitReached):
                    ait5.step_4_judge_password_recovery(
                        page, threading.Event(), timeout_ms=5_000
                    )
            finally:
                browser.close()
            self.assertLess(time.monotonic() - started, 1)

    @unittest.skipIf(ait5.sync_playwright is None, "playwright not installed")
    def test_stalled_loading_requests_session_rebuild(self):
        with ait5.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content("<div>Loading...</div>")
            try:
                with self.assertRaises(ait5.PageSubmissionStalled):
                    ait5.step_4_judge_password_recovery(
                        page,
                        threading.Event(),
                        timeout_ms=1_000,
                        poll_interval_ms=10,
                        stalled_loading_seconds=0.05,
                    )
            finally:
                browser.close()

    @unittest.skipIf(ait5.sync_playwright is None, "playwright not installed")
    def test_retrieve_result_cancel_returns_to_empty_form(self):
        html = """
        <input id="fsa_Input_ForgotUsernameFirstName" value="First">
        <input id="fsa_Input_ForgotUsernameLastName" value="Last">
        <select id="fsa_Input_ForgotUsernameDateOfBirthMonth"><option value="01" selected>Jan</option></select>
        <input id="fsa_Input_ForgotUsernameDateOfBirthDay" value="2">
        <input id="fsa_Input_ForgotUsernameDateOfBirthYear" value="1990">
        <input id="fsa_Input_ForgotUsernameSsnInput" value="111223333">
        <button id="fsa_Button_ForgotUsernameCancel"><span>Cancel</span></button>
        <script>
        document.querySelector('button').onclick = () => {
          for (const x of document.querySelectorAll('input,select')) x.value = '';
        };
        </script>
        """
        with ait5.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 900, "height": 700})
            page.set_content(html)
            session = ait5.BrowserRecoverySession(1)
            session._browser = browser
            session._context = page.context
            session._page = page
            messages = []
            try:
                session.prepare_for_next(
                    "can_recover", threading.Event(), messages.append
                )
                self.assertEqual(page.locator(ait5.FIRST_NAME_SELECTOR).input_value(), "")
                self.assertTrue(any("Cancel 已完成" in value for value in messages))
            finally:
                browser.close()

    @unittest.skipIf(ait5.sync_playwright is None, "playwright not installed")
    def test_account_not_found_cleanup_clears_data_and_returns_blank(self):
        with ait5.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()
            context.add_cookies([
                {"name": "test", "value": "1", "url": "https://studentaid.gov"}
            ])
            session = ait5.BrowserRecoverySession(1)
            session._browser = browser
            session._context = context
            session._page = page
            session._dedicated_profile = True
            messages = []
            try:
                session.prepare_for_next(
                    "account_not_found", threading.Event(), messages.append
                )
                self.assertEqual(session._page.url, "about:blank")
                self.assertEqual(context.cookies(), [])
                self.assertTrue(any("数据已清除" in value for value in messages))
            finally:
                browser.close()

    @unittest.skipUnless(os.name == "nt", "dedicated Chrome integration is Windows-only")
    @unittest.skipIf(ait5.sync_playwright is None, "playwright not installed")
    def test_dedicated_chrome_cdp_has_no_webdriver_flag(self):
        old_cdp = os.environ.get("STUDENTAID_CDP_URL")
        os.environ.pop("STUDENTAID_CDP_URL", None)
        session = ait5.BrowserRecoverySession(1)
        try:
            session.start()
            self.assertFalse(session._page.evaluate("navigator.webdriver"))
            self.assertTrue(session._dedicated_profile)
            self.assertIn("AutomationControlled", session.browser_mode)
        finally:
            session.close()
            if old_cdp is None:
                os.environ.pop("STUDENTAID_CDP_URL", None)
            else:
                os.environ["STUDENTAID_CDP_URL"] = old_cdp


if __name__ == "__main__":
    unittest.main()
