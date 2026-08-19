from __future__ import annotations

import csv
import queue
import sqlite3
import sys
import tempfile
from pathlib import Path
import threading
import time
import unittest
from unittest import mock

sys.dont_write_bytecode = True
module_root = Path(__file__).resolve().parent
if not (module_root / "ait31.py").is_file():
    module_root = module_root.parent
sys.path.insert(0, str(module_root))

import ait32 as ait11


class GuiConfigPersistenceTests(unittest.TestCase):
    def test_save_and_load_restores_all_gui_values(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            config_path = Path(directory) / ait11.GUI_CONFIG_FILENAME
            expected = {
                "input_path": r"E:\studentaid\假资料测试2.xlsx",
                "output_path": r"E:\studentaid\StudentAid累计结果.csv",
                "backend": "playwright",
                "display_mode": "窗口",
                "thread_count": "12",
            }
            ait11.save_gui_config(expected, config_path)
            self.assertEqual(expected, ait11.load_gui_config(config_path))
            self.assertTrue(config_path.is_file())

    def test_invalid_or_corrupt_config_falls_back_without_crashing(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            config_path = Path(directory) / ait11.GUI_CONFIG_FILENAME
            config_path.write_text("{not-json", encoding="utf-8")
            loaded = ait11.load_gui_config(config_path)
            self.assertEqual("browser-use", loaded["backend"])
            self.assertEqual("无头", loaded["display_mode"])
            self.assertEqual("2", loaded["thread_count"])

            ait11.save_gui_config(
                {
                    "backend": "unknown",
                    "display_mode": "bad",
                    "thread_count": "0",
                },
                config_path,
            )
            loaded = ait11.load_gui_config(config_path)
            self.assertEqual("browser-use", loaded["backend"])
            self.assertEqual("无头", loaded["display_mode"])
            self.assertEqual("2", loaded["thread_count"])


class Step24RealtimeProgressMetricsTests(unittest.TestCase):
    def test_percentage_rolling_minute_rate_average_and_eta(self) -> None:
        tracker = ait11.BatchProgressTracker()
        tracker.start(now=100.0)
        baseline = tracker.update(
            {
                "total": 100,
                "pending": 100,
                "processing": 0,
                "completed": 0,
                "failed": 0,
                "stopped": 0,
            },
            now=100.0,
        )
        self.assertEqual(0, baseline.recent_minute_count)
        self.assertIsNone(baseline.eta_seconds)

        metrics = tracker.update(
            {
                "total": 100,
                "pending": 75,
                "processing": 5,
                "completed": 18,
                "failed": 2,
                "stopped": 0,
            },
            now=130.0,
        )
        self.assertEqual(20, metrics.terminal)
        self.assertEqual(80, metrics.remaining)
        self.assertEqual(20.0, metrics.percent)
        self.assertEqual(20, metrics.recent_minute_count)
        self.assertAlmostEqual(40.0, metrics.average_per_minute)
        self.assertAlmostEqual(120.0, metrics.eta_seconds or 0.0)

        after_window = tracker.snapshot(now=191.0)
        self.assertEqual(0, after_window.recent_minute_count)
        self.assertGreater(after_window.average_per_minute, 0)

    def test_initial_import_failures_are_not_counted_as_browser_speed(self) -> None:
        tracker = ait11.BatchProgressTracker()
        tracker.start(now=0.0)
        metrics = tracker.update(
            {
                "total": 10,
                "pending": 7,
                "processing": 0,
                "completed": 0,
                "failed": 3,
                "stopped": 0,
            },
            now=5.0,
        )
        self.assertEqual(30.0, metrics.percent)
        self.assertEqual(0, metrics.recent_minute_count)
        self.assertEqual(0.0, metrics.average_per_minute)
        self.assertIsNone(metrics.eta_seconds)

    def test_stale_event_does_not_move_processed_count_backwards(self) -> None:
        tracker = ait11.BatchProgressTracker()
        tracker.start(now=0.0)
        tracker.update(
            {"total": 10, "pending": 10, "processing": 0}, now=0.0
        )
        current = tracker.update(
            {"total": 10, "pending": 5, "processing": 1, "completed": 4},
            now=20.0,
        )
        stale = tracker.update(
            {"total": 10, "pending": 8, "processing": 1, "completed": 1},
            now=21.0,
        )
        self.assertEqual(current.terminal, stale.terminal)
        self.assertEqual(current.percent, stale.percent)

    def test_finish_freezes_elapsed_time_and_duration_format(self) -> None:
        tracker = ait11.BatchProgressTracker()
        tracker.start(now=10.0)
        tracker.update(
            {"total": 2, "pending": 2, "processing": 0}, now=10.0
        )
        tracker.update(
            {"total": 2, "pending": 0, "processing": 0, "completed": 2},
            now=70.0,
        )
        tracker.finish(now=80.0)
        frozen = tracker.snapshot(now=500.0)
        self.assertEqual(70.0, frozen.elapsed_seconds)
        self.assertEqual(0.0, frozen.eta_seconds)
        self.assertEqual("00:01:10", ait11.format_duration(frozen.elapsed_seconds))
        self.assertEqual("--", ait11.format_duration(None))
        self.assertEqual("1天 01:01:01", ait11.format_duration(90061))


def make_record(row: int, ssn: str) -> ait11.ImportedRecord:
    details = ait11.AccountDetails(
        first_name="FIRST",
        last_name="LAST",
        birth_month="01",
        birth_day="02",
        birth_year="1963",
        ssn=ssn,
        original_ssn=ssn,
    )
    return ait11.ImportedRecord(
        source_file="input.xlsx",
        source_sheet="Sheet1",
        source_row=row,
        original_fields=(ssn, "01", "02", "1963", "LAST", "FIRST", "ADDRESS"),
        details=details,
        address="ADDRESS",
    )


class Step16DuplicateQueueTests(unittest.TestCase):
    def setUp(self) -> None:
        self.connection = sqlite3.connect(":memory:")
        self.connection.row_factory = sqlite3.Row
        ait11.DatabaseWriter._create_schema(self.connection)
        ait11.DatabaseWriter._handle(
            self.connection,
            "create_batch",
            {
                "batch_id": "batch",
                "input_path": "input.xlsx",
                "output_directory": ".",
                "thread_count": 8,
            },
        )

    def tearDown(self) -> None:
        self.connection.close()

    def test_only_one_work_item_per_normalised_first_column(self) -> None:
        records = [
            make_record(1, "111223333"),
            make_record(2, "111223333"),
            make_record(3, "111223333"),
            make_record(4, "444556666"),
            ait11.ImportedRecord(
                source_file="input.xlsx",
                source_sheet="Sheet1",
                source_row=5,
                original_fields=("bad",),
                details=None,
                address="",
                import_error="Social Security Number 必须包含 9 位数字",
            ),
        ]
        work_items = ait11.DatabaseWriter._handle(
            self.connection,
            "insert_records",
            {"batch_id": "batch", "records": records},
        )

        self.assertEqual(2, len(work_items))
        self.assertEqual(["111223333", "444556666"], [x.details.ssn for x in work_items])
        counts = dict(
            self.connection.execute(
                "SELECT status, COUNT(*) FROM records GROUP BY status"
            ).fetchall()
        )
        self.assertEqual({"failed": 1, "pending": 4}, counts)

    def test_same_ssn_with_different_full_rows_creates_separate_tasks(self) -> None:
        first = make_record(1, "111223333")
        second = ait11.ImportedRecord(
            source_file=first.source_file,
            source_sheet=first.source_sheet,
            source_row=2,
            original_fields=(
                "111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS 2",
            ),
            details=first.details,
            address="ADDRESS 2",
        )
        work_items = ait11.DatabaseWriter._handle(
            self.connection,
            "insert_records",
            {"batch_id": "batch", "records": [first, second]},
        )
        self.assertEqual(2, len(work_items))
        self.assertEqual("111223333", work_items[0].details.ssn)
        self.assertEqual("111223333", work_items[1].details.ssn)
        self.assertNotEqual(work_items[0].record_key, work_items[1].record_key)

    def test_group_completion_and_failure_cover_all_duplicates(self) -> None:
        records = [
            make_record(1, "111223333"),
            make_record(2, "111223333"),
            make_record(3, "111223333"),
        ]
        work_items = ait11.DatabaseWriter._handle(
            self.connection,
            "insert_records",
            {"batch_id": "batch", "records": records},
        )
        ait11.DatabaseWriter._handle(
            self.connection,
            "mark_processing",
            {"record_id": work_items[0].record_id},
        )
        result = ait11.RecoveryResult(
            result_code="account_not_found",
            heading="Account Not Found",
            masked_phone="",
            masked_email="",
            recovery_method="",
        )
        completed = ait11.DatabaseWriter._handle(
            self.connection,
            "mark_completed_group",
            {
                "batch_id": "batch",
                "record_key": work_items[0].record_key,
                "result": result,
            },
        )
        self.assertEqual(3, completed)
        self.assertEqual(
            [("completed", 3)],
            [
                tuple(row)
                for row in self.connection.execute(
                    "SELECT status, COUNT(*) FROM records GROUP BY status"
                ).fetchall()
            ],
        )

        failed = ait11.DatabaseWriter._handle(
            self.connection,
            "mark_failed_group",
            {
                "batch_id": "batch",
                "record_key": work_items[0].record_key,
                "error": "output write failed",
            },
        )
        self.assertEqual(3, failed)
        self.assertEqual(
            [("failed", 3)],
            [
                tuple(row)
                for row in self.connection.execute(
                    "SELECT status, COUNT(*) FROM records GROUP BY status"
                ).fetchall()
            ],
        )

    def test_retry_returns_the_whole_duplicate_group_to_pending(self) -> None:
        records = [
            make_record(1, "111223333"),
            make_record(2, "111223333"),
            make_record(3, "111223333"),
        ]
        work_items = ait11.DatabaseWriter._handle(
            self.connection,
            "insert_records",
            {"batch_id": "batch", "records": records},
        )
        ait11.DatabaseWriter._handle(
            self.connection,
            "mark_processing",
            {"record_id": work_items[0].record_id},
        )

        retried = ait11.DatabaseWriter._handle(
            self.connection,
            "mark_retry_group",
            {
                "batch_id": "batch",
                "record_key": work_items[0].record_key,
                "error": "temporary page timeout",
            },
        )

        self.assertEqual(3, retried)
        self.assertEqual(
            [("pending", 3)],
            [
                tuple(row)
                for row in self.connection.execute(
                    "SELECT status, COUNT(*) FROM records GROUP BY status"
                ).fetchall()
            ],
        )
        representative = self.connection.execute(
            "SELECT attempt_count, started_at, finished_at, error "
            "FROM records WHERE id=?",
            (work_items[0].record_id,),
        ).fetchone()
        self.assertEqual(1, representative["attempt_count"])
        self.assertIsNone(representative["started_at"])
        self.assertIsNone(representative["finished_at"])
        self.assertEqual("temporary page timeout", representative["error"])


class Step16CumulativeCsvTests(unittest.TestCase):
    def test_quote_all_normalisation_also_removes_empty_rows(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            output = Path(directory) / "result.csv"
            output.write_text(
                '\n"key","01/02/1963","FIRST","LAST","ADDRESS",'
                '"Account Not Found","","",""\n\n',
                encoding="utf-8-sig",
            )

            self.assertTrue(ait11.ensure_cumulative_output_quote_all(output))
            text = output.read_text(encoding="utf-8-sig")
            rows = list(csv.reader(text.splitlines()))
            self.assertEqual(1, len(rows))
            self.assertEqual(9, len(rows[0]))
            self.assertTrue(text.startswith('"key","01/02/1963"'))
            self.assertFalse(ait11.ensure_cumulative_output_quote_all(output))


class Step35CumulativeResultLastColumnTests(unittest.TestCase):
    def test_mixed_legacy_widths_are_padded_and_heading_is_one_last_column(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            output = Path(directory) / "mixed.csv"
            short_input = ["short-1", "short-2"]
            wide_input = [f"wide-{index}" for index in range(1, 8)]
            legacy_rows = [
                [*short_input, "Account Not Found", "", "", ""],
                [
                    *wide_input,
                    "Retrieve Your Log-in Information",
                    "(⦁⦁⦁) ⦁⦁⦁ 1234",
                    "te⦁⦁@example.com",
                    "Recover my account with a photo ID",
                ],
            ]
            with output.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(legacy_rows)

            layout = ait11.normalise_cumulative_output_layout(output)
            self.assertTrue(layout.changed)
            self.assertEqual(2, layout.row_count)
            self.assertEqual(7, layout.input_width)
            self.assertEqual(2, layout.legacy_rows)
            rows = list(csv.reader(output.read_text(encoding="utf-8-sig").splitlines()))
            self.assertEqual({11}, {len(row) for row in rows})
            self.assertEqual("Account Not Found", rows[0][-1])
            self.assertEqual("Retrieve Your Log-in Information", rows[1][-1])
            self.assertEqual(["", "", ""], rows[0][-4:-1])
            self.assertEqual(
                [
                    "(⦁⦁⦁) ⦁⦁⦁ 1234",
                    "te⦁⦁@example.com",
                    "Recover my account with a photo ID",
                ],
                rows[1][-4:-1],
            )
            self.assertEqual(short_input, rows[0][:2])
            self.assertEqual(["" for _ in range(5)], rows[0][2:7])
            self.assertEqual(
                {
                    ait11.json.dumps(short_input, ensure_ascii=False, separators=(",", ":")),
                    ait11.json.dumps(wide_input, ensure_ascii=False, separators=(",", ":")),
                },
                ait11.read_output_row_keys(output),
            )
            self.assertFalse(ait11.normalise_cumulative_output_layout(output).changed)

    def test_wider_append_expands_existing_rows_and_keeps_heading_last(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            output = Path(directory) / "expand.csv"
            original_input = ["old-1", "old-2", "old-3", "old-4"]
            with output.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(
                    [*original_input, "Account Not Found", "", "", ""]
                )
            details = ait11.AccountDetails(
                first_name="FIRST", last_name="LAST", birth_month="01",
                birth_day="02", birth_year="1980", ssn="111223333",
                original_ssn="111223333",
            )
            fields = (
                "111223333", "01/02/1980", "FIRST", "LAST",
                "ADDRESS", "EXTRA-1", "EXTRA-2",
            )
            item = ait11.WorkItem(
                record_id=1, details=details, original_fields=fields, address="ADDRESS"
            )
            result = ait11.RecoveryResult(
                "can_recover", "Retrieve Your Log-in Information",
                "(⦁⦁⦁) ⦁⦁⦁ 5678", "fi⦁⦁@example.com",
                "Recover my account with a photo ID",
            )

            self.assertTrue(ait11.append_cumulative_result(output, item, result))
            rows = list(csv.reader(output.read_text(encoding="utf-8-sig").splitlines()))
            self.assertEqual({11}, {len(row) for row in rows})
            self.assertEqual(["Account Not Found", "Retrieve Your Log-in Information"], [row[-1] for row in rows])
            self.assertEqual(original_input, rows[0][:4])
            self.assertEqual(["", "", ""], rows[0][4:7])
            self.assertIn(item.record_key, ait11.read_output_row_keys(output))


def make_work_item(record_id: int, ssn: str) -> ait11.WorkItem:
    record = make_record(record_id, ssn)
    assert record.details is not None
    return ait11.WorkItem(
        record_id=record_id,
        details=record.details,
        source_file=record.source_file,
        source_sheet=record.source_sheet,
        source_row=record.source_row,
        original_fields=record.original_fields,
        address=record.address,
    )


class Step23PersistencePipelineTests(unittest.TestCase):
    def test_same_ssn_different_rows_are_both_appended_and_deleted(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            first = make_work_item(1, "111223333")
            second = ait11.WorkItem(
                record_id=2,
                details=first.details,
                source_file=first.source_file,
                source_sheet=first.source_sheet,
                source_row=2,
                original_fields=(
                    "111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS 2",
                ),
                address="ADDRESS 2",
            )
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(
                    [first.original_fields, second.original_fields]
                )
            result = ait11.RecoveryResult(
                "account_not_found", "Account Not Found", "", "", ""
            )
            writer = ait11.ResultPersistenceWriter(
                input_path, output_path, batch_size=2, flush_seconds=60
            )
            writer.start()
            self.assertTrue(writer.commit(first, result))
            self.assertTrue(writer.commit(second, result))
            writer.close()
            rows = list(
                csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual(2, len(rows))
            self.assertEqual("ADDRESS", rows[0][6])
            self.assertEqual("ADDRESS 2", rows[1][6])
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))

    def test_startup_sync_removes_only_exact_output_row(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            first = make_work_item(1, "111223333")
            second_fields = (
                "111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS 2",
            )
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows([first.original_fields, second_fields])
            result = ait11.RecoveryResult(
                "account_not_found", "Account Not Found", "", "", ""
            )
            ait11.append_cumulative_result(output_path, first, result)
            keys = ait11.read_output_row_keys(output_path)
            self.assertEqual(1, len(keys))
            self.assertEqual(1, ait11.remove_input_rows_by_keys(input_path, keys))
            remaining = list(
                csv.reader(input_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual([list(second_fields)], remaining)

    def test_batch_duplicate_cleanup_keeps_different_row_with_same_ssn(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "input.csv"
            rows = [
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS 2"],
            ]
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)
            records = ait11.load_input_records(path)
            self.assertEqual(1, ait11.remove_duplicate_input_rows(path))
            remaining = list(
                csv.reader(path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual([rows[0], rows[2]], remaining)

    def test_twenty_results_use_one_batched_input_rewrite(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            items = [
                make_work_item(index, f"{100000000 + index:09d}")
                for index in range(1, 21)
            ]
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(item.original_fields for item in items)
            result = ait11.RecoveryResult(
                result_code="account_not_found",
                heading="Account Not Found",
                masked_phone="",
                masked_email="",
                recovery_method="",
            )
            original_remove = ait11.remove_input_rows_by_keys
            delete_calls: list[set[str]] = []

            def tracked_remove(path, keys):
                delete_calls.append(set(keys))
                return original_remove(path, keys)

            ait11.remove_input_rows_by_keys = tracked_remove
            try:
                writer = ait11.ResultPersistenceWriter(
                    input_path,
                    output_path,
                    batch_size=20,
                    flush_seconds=60,
                )
                writer.start()
                self.assertTrue(all(writer.commit(item, result) for item in items))
                writer.close()
            finally:
                ait11.remove_input_rows_by_keys = original_remove

            self.assertEqual(1, len(delete_calls))
            self.assertEqual(20, len(delete_calls[0]))
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))
            rows = list(
                csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual(20, len(rows))

    def test_browser_commit_does_not_wait_for_slow_input_rewrite(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            item = make_work_item(1, "111223333")
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(item.original_fields)
            result = ait11.RecoveryResult(
                result_code="account_not_found",
                heading="Account Not Found",
                masked_phone="",
                masked_email="",
                recovery_method="",
            )
            original_remove = ait11.remove_input_rows_by_keys
            delete_started = threading.Event()
            allow_delete = threading.Event()

            def slow_remove(path, keys):
                delete_started.set()
                self.assertTrue(allow_delete.wait(3))
                return original_remove(path, keys)

            ait11.remove_input_rows_by_keys = slow_remove
            try:
                writer = ait11.ResultPersistenceWriter(
                    input_path,
                    output_path,
                    batch_size=1,
                    flush_seconds=60,
                )
                writer.start()
                started = time.perf_counter()
                self.assertTrue(writer.commit(item, result))
                elapsed = time.perf_counter() - started
                self.assertLess(elapsed, 0.5)
                self.assertTrue(delete_started.wait(1))
                allow_delete.set()
                writer.close()
            finally:
                allow_delete.set()
                ait11.remove_input_rows_by_keys = original_remove

            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))

    def test_duplicate_output_key_is_not_appended_twice_but_input_is_deleted(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            item = make_work_item(1, "111223333")
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(item.original_fields)
            result = ait11.RecoveryResult(
                result_code="account_not_found",
                heading="Account Not Found",
                masked_phone="",
                masked_email="",
                recovery_method="",
            )
            writer = ait11.ResultPersistenceWriter(
                input_path,
                output_path,
                batch_size=20,
                flush_seconds=60,
            )
            writer.start()
            self.assertTrue(writer.commit(item, result))
            self.assertFalse(writer.commit(item, result))
            writer.close()
            rows = list(
                csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual(1, len(rows))
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))

    def test_time_threshold_flushes_a_small_batch(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            item = make_work_item(1, "111223333")
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(item.original_fields)
            result = ait11.RecoveryResult(
                "account_not_found", "Account Not Found", "", "", ""
            )
            writer = ait11.ResultPersistenceWriter(
                input_path,
                output_path,
                batch_size=20,
                flush_seconds=0.1,
            )
            writer.start()
            try:
                self.assertTrue(writer.commit(item, result))
                deadline = time.monotonic() + 2
                while time.monotonic() < deadline:
                    try:
                        if not input_path.read_text(encoding="utf-8-sig"):
                            break
                    except PermissionError:
                        # E: 盘实时扫描器可能恰好撞上原子替换的极短窗口。
                        pass
                    time.sleep(0.02)
            finally:
                writer.close()
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))

    def test_output_remains_recoverable_when_input_rewrite_fails(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            item = make_work_item(1, "111223333")
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(item.original_fields)
            result = ait11.RecoveryResult(
                "account_not_found", "Account Not Found", "", "", ""
            )
            original_remove = ait11.remove_input_rows_by_keys

            def failed_remove(_path, _keys):
                raise PermissionError("input file is busy")

            ait11.remove_input_rows_by_keys = failed_remove
            try:
                writer = ait11.ResultPersistenceWriter(
                    input_path,
                    output_path,
                    batch_size=1,
                    flush_seconds=60,
                )
                writer.start()
                self.assertTrue(writer.commit(item, result))
                with self.assertRaisesRegex(RuntimeError, "等待下次启动同步删除"):
                    writer.close()
            finally:
                ait11.remove_input_rows_by_keys = original_remove

            output_keys = ait11.read_output_first_column_keys(output_path)
            self.assertIn(item.record_key, output_keys)
            self.assertTrue(input_path.read_text(encoding="utf-8-sig").strip())
            self.assertEqual(
                1,
                ait11.remove_input_rows_by_keys(input_path, output_keys),
            )
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))


class Step34LargeXlsxPersistenceTests(unittest.TestCase):
    def test_existing_output_rows_are_filtered_in_memory_by_full_row(self) -> None:
        first = make_record(1, "111223333")
        same_ssn_different_row = ait11.ImportedRecord(
            source_file=first.source_file,
            source_sheet=first.source_sheet,
            source_row=2,
            original_fields=(*first.original_fields[:-1], "ADDRESS 2"),
            details=first.details,
            address="ADDRESS 2",
        )
        remaining, matched_keys, skipped = ait11.filter_records_already_output(
            [first, same_ssn_different_row], {first.record_key}
        )
        self.assertEqual(1, skipped)
        self.assertEqual({first.record_key}, matched_keys)
        self.assertEqual([same_ssn_different_row], remaining)

    def test_xlsx_rewrite_is_deferred_until_all_csv_commits_finish(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.xlsx"
            output_path = root / "output.csv"
            workbook = Workbook()
            worksheet = workbook.active
            first = make_work_item(1, "111223333")
            second = make_work_item(2, "444556666")
            worksheet.append(list(first.original_fields))
            worksheet.append(list(second.original_fields))
            workbook.save(input_path)
            workbook.close()
            result = ait11.RecoveryResult(
                "account_not_found", "Account Not Found", "", "", ""
            )
            delete_started = threading.Event()
            allow_delete = threading.Event()
            original_remove = ait11.remove_input_rows_by_keys

            def slow_remove(_path, keys):
                delete_started.set()
                self.assertTrue(allow_delete.wait(timeout=5))
                return len(keys)

            ait11.remove_input_rows_by_keys = slow_remove
            writer = ait11.ResultPersistenceWriter(
                input_path,
                output_path,
                batch_size=1,
                flush_seconds=0.05,
            )
            writer.start()
            close_errors: list[BaseException] = []
            close_thread = None
            try:
                started_at = time.monotonic()
                self.assertTrue(writer.commit(first, result))
                self.assertTrue(writer.commit(second, result))
                self.assertLess(time.monotonic() - started_at, 2.0)
                time.sleep(0.15)
                self.assertFalse(delete_started.is_set())

                def close_writer():
                    try:
                        writer.close()
                    except BaseException as exc:
                        close_errors.append(exc)

                close_thread = threading.Thread(target=close_writer)
                close_thread.start()
                self.assertTrue(delete_started.wait(timeout=1))
                rows = list(
                    csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
                )
                self.assertEqual(2, len(rows))
            finally:
                allow_delete.set()
                if close_thread is not None:
                    close_thread.join(timeout=5)
                elif not writer._closed:
                    try:
                        writer.close()
                    except BaseException as exc:
                        close_errors.append(exc)
                ait11.remove_input_rows_by_keys = original_remove
            self.assertFalse(close_errors)
            self.assertTrue(close_thread is None or not close_thread.is_alive())

    def test_noncontiguous_xlsx_deletes_use_one_tail_compaction(self) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        workbook = Workbook()
        worksheet = workbook.active
        for number in range(1, 51):
            worksheet.append([number, f"row-{number}"])
        worksheet.cell(3, 2).fill = PatternFill(
            fill_type="solid", fgColor="FFFF0000"
        )
        original_delete_rows = worksheet.delete_rows
        delete_calls: list[tuple[int, int]] = []

        def tracked_delete_rows(index: int, amount: int = 1):
            delete_calls.append((index, amount))
            return original_delete_rows(index, amount)

        worksheet.delete_rows = tracked_delete_rows
        deleted = ait11._delete_worksheet_rows_efficiently(
            worksheet, {2, 25, 49}
        )
        remaining = [worksheet.cell(row, 1).value for row in range(1, 48)]
        self.assertEqual(3, deleted)
        self.assertEqual([value for value in range(1, 51) if value not in {2, 25, 49}], remaining)
        self.assertEqual(1, len(delete_calls))
        self.assertEqual("FFFF0000", worksheet.cell(2, 2).fill.fgColor.rgb)
        workbook.close()

    def test_zip_level_xlsx_delete_preserves_styles_and_other_sheets(self) -> None:
        import zipfile
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill

        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "styled.xlsx"
            workbook = Workbook()
            first_sheet = workbook.active
            first_sheet.title = "Data"
            for number in range(1, 11):
                first_sheet.append([number, f"row-{number}"])
            first_sheet.cell(3, 2).fill = PatternFill(
                fill_type="solid", fgColor="FFFF0000"
            )
            second_sheet = workbook.create_sheet("Keep")
            second_sheet.append(["unchanged", 123])
            workbook.save(path)
            workbook.close()
            with zipfile.ZipFile(path) as archive:
                original_styles = archive.read("xl/styles.xml")
                original_second_sheet = archive.read("xl/worksheets/sheet2.xml")

            ait11._atomic_remove_xlsx_rows(
                path, {"xl/worksheets/sheet1.xml": {2, 7}}
            )

            workbook = load_workbook(path, data_only=True)
            values = [
                workbook["Data"].cell(row, 1).value
                for row in range(1, workbook["Data"].max_row + 1)
            ]
            self.assertEqual([1, 3, 4, 5, 6, 8, 9, 10], values)
            self.assertEqual(
                "FFFF0000", workbook["Data"].cell(2, 2).fill.fgColor.rgb
            )
            self.assertEqual("unchanged", workbook["Keep"].cell(1, 1).value)
            workbook.close()
            with zipfile.ZipFile(path) as archive:
                self.assertEqual(original_styles, archive.read("xl/styles.xml"))
                self.assertEqual(
                    original_second_sheet,
                    archive.read("xl/worksheets/sheet2.xml"),
                )


class FakeSession:
    processed: list[str] = []
    lock = threading.Lock()

    def __init__(self, worker_number: int) -> None:
        self.worker_number = worker_number

    def start(self) -> None:
        return None

    def process(self, item, stop_event, progress):
        with self.lock:
            self.processed.append(item.details.ssn)
        return ait11.RecoveryResult(
            result_code="account_not_found",
            heading="Account Not Found",
            masked_phone="",
            masked_email="",
            recovery_method="",
        )

    def prepare_for_next(self, result_code, stop_event, progress) -> None:
        return None

    def recover_after_cleanup_error(self) -> None:
        return None

    def close(self) -> None:
        return None


class FlakySession(FakeSession):
    attempts: dict[str, int] = {}
    sequence: list[str] = []

    def process(self, item, stop_event, progress):
        with self.lock:
            attempts = self.attempts.get(item.details.ssn, 0) + 1
            self.attempts[item.details.ssn] = attempts
            self.sequence.append(item.details.ssn)
        if item.details.ssn == "111223333" and attempts == 1:
            raise RuntimeError("temporary page timeout")
        return ait11.RecoveryResult(
            result_code="account_not_found",
            heading="Account Not Found",
            masked_phone="",
            masked_email="",
            recovery_method="",
        )


class AlwaysFailSession(FakeSession):
    calls = 0

    def process(self, item, stop_event, progress):
        with self.lock:
            type(self).calls += 1
        raise RuntimeError("persistent page timeout")


class Step16BatchIntegrationTests(unittest.TestCase):
    def test_duplicate_rows_use_one_browser_task_and_finish_as_group(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            rows = [
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["444556666", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["444556666", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
            ]
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)
            output_path.write_text("\n", encoding="utf-8-sig")

            events: list[tuple[str, dict]] = []
            FakeSession.processed = []
            engine = ait11.BatchEngine(
                lambda kind, payload: events.append((kind, payload)),
                session_factory=FakeSession,
            )
            engine.start(
                input_path,
                output_path,
                thread_count=12,
                backend="playwright",
                display_mode="窗口",
            )
            self.assertTrue(engine.wait(10))
            self.assertFalse(any(kind == "fatal" for kind, _ in events))
            self.assertEqual(
                ["111223333", "444556666"],
                sorted(FakeSession.processed),
            )
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))

            output_rows = list(
                csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual(2, len(output_rows))
            self.assertTrue(all(len(row) == 11 for row in output_rows))

            connection = sqlite3.connect(root / ait11.DATABASE_FILENAME)
            try:
                status_counts = dict(
                    connection.execute(
                        "SELECT status, COUNT(*) FROM records GROUP BY status"
                    ).fetchall()
                )
            finally:
                connection.close()
            self.assertEqual({"completed": 5}, status_counts)
            connection = sqlite3.connect(root / ait11.DATABASE_FILENAME)
            try:
                configured_threads = connection.execute(
                    "SELECT thread_count FROM batches ORDER BY created_at DESC LIMIT 1"
                ).fetchone()[0]
            finally:
                connection.close()
            self.assertEqual(12, configured_threads)
            log_text = "\n".join(
                payload.get("message", "") for kind, payload in events if kind == "log"
            )
            self.assertIn("唯一浏览器任务 2 个", log_text)
            self.assertIn("完全重复整行 3 条", log_text)

    def test_transient_failure_is_retried_from_queue_tail(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            rows = [
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
                ["444556666", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"],
            ]
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)

            events: list[tuple[str, dict]] = []
            FlakySession.attempts = {}
            FlakySession.sequence = []
            engine = ait11.BatchEngine(
                lambda kind, payload: events.append((kind, payload)),
                session_factory=FlakySession,
            )
            engine.start(
                input_path,
                output_path,
                thread_count=1,
                backend="playwright",
                display_mode="窗口",
            )
            self.assertTrue(engine.wait(10))
            self.assertFalse(any(kind == "fatal" for kind, _ in events))
            self.assertEqual(
                {"111223333": 2, "444556666": 1}, FlakySession.attempts
            )
            self.assertEqual(
                ["111223333", "444556666", "111223333"],
                FlakySession.sequence,
            )
            self.assertEqual("", input_path.read_text(encoding="utf-8-sig"))
            output_rows = list(
                csv.reader(output_path.read_text(encoding="utf-8-sig").splitlines())
            )
            self.assertEqual(2, len(output_rows))
            self.assertTrue(all(len(row) == 11 for row in output_rows))

            connection = sqlite3.connect(root / ait11.DATABASE_FILENAME)
            try:
                status_counts = dict(
                    connection.execute(
                        "SELECT status, COUNT(*) FROM records GROUP BY status"
                    ).fetchall()
                )
                max_attempts = connection.execute(
                    "SELECT MAX(attempt_count) FROM records"
                ).fetchone()[0]
            finally:
                connection.close()
            self.assertEqual({"completed": 3}, status_counts)
            self.assertEqual(2, max_attempts)
            log_text = "\n".join(
                payload.get("message", "") for kind, payload in events if kind == "log"
            )
            self.assertIn("已回到队尾", log_text)
            self.assertIn("第 2/3 轮", log_text)

    def test_persistent_failure_stops_after_three_queue_rounds(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "output.csv"
            with input_path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerow(
                    ["111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS"]
                )

            events: list[tuple[str, dict]] = []
            AlwaysFailSession.calls = 0
            engine = ait11.BatchEngine(
                lambda kind, payload: events.append((kind, payload)),
                session_factory=AlwaysFailSession,
            )
            engine.start(
                input_path,
                output_path,
                thread_count=1,
                backend="playwright",
                display_mode="窗口",
            )
            self.assertTrue(engine.wait(10))
            self.assertFalse(any(kind == "fatal" for kind, _ in events))
            self.assertEqual(ait11.MAX_QUEUE_ATTEMPTS, AlwaysFailSession.calls)
            self.assertTrue(input_path.read_text(encoding="utf-8-sig").strip())
            self.assertFalse(output_path.exists())

            connection = sqlite3.connect(root / ait11.DATABASE_FILENAME)
            try:
                status, attempts, error = connection.execute(
                    "SELECT status, attempt_count, error FROM records"
                ).fetchone()
            finally:
                connection.close()
            self.assertEqual("failed", status)
            self.assertEqual(ait11.MAX_QUEUE_ATTEMPTS, attempts)
            self.assertEqual("persistent page timeout", error)


class _TextLocator:
    def __init__(self, text: str, visible: bool = True) -> None:
        self._text = text
        self._visible = visible

    @property
    def first(self):
        return self

    def is_visible(self) -> bool:
        return self._visible

    def inner_text(self) -> str:
        return self._text


class _ResultPage:
    def __init__(self, body_text: str) -> None:
        self.body_text = body_text

    def evaluate(self, _script: str) -> str:
        return self.body_text

    def wait_for_timeout(self, _milliseconds: int) -> None:
        return None

    def get_by_text(self, text, exact=False):
        if isinstance(text, str) and text in self.body_text:
            return _TextLocator(text)
        return _TextLocator("", visible=False)


class Step18ExplicitStatusTests(unittest.TestCase):
    def test_account_lookup_issue_is_an_explicit_result(self) -> None:
        page = _ResultPage(ait11.ACCOUNT_LOOKUP_ISSUE_HEADING)
        status = ait11.step_4_judge_password_recovery(
            page,
            threading.Event(),
            timeout_ms=100,
        )
        self.assertEqual("account_lookup_issue", status)
        result = ait11.collect_recovery_result(page, status)
        self.assertEqual(ait11.ACCOUNT_LOOKUP_ISSUE_HEADING, result.heading)
        self.assertEqual(("", "", ""), (
            result.masked_phone,
            result.masked_email,
            result.recovery_method,
        ))

    def test_photo_id_recovery_message_is_an_explicit_result(self) -> None:
        page = _ResultPage(
            ait11.PHOTO_ID_RECOVERY_MESSAGE.replace(
                "information. Access", "information.\n   Access"
            )
        )
        status = ait11.step_4_judge_password_recovery(
            page,
            threading.Event(),
            timeout_ms=100,
        )
        self.assertEqual("photo_id_recovery_required", status)
        result = ait11.collect_recovery_result(page, status)
        self.assertEqual(ait11.PHOTO_ID_RECOVERY_MESSAGE, result.heading)
        self.assertEqual(("", "", ""), (
            result.masked_phone,
            result.masked_email,
            result.recovery_method,
        ))

    def test_new_statuses_use_existing_terminal_cleanup_path(self) -> None:
        session = ait11.BrowserRecoverySession.__new__(ait11.BrowserRecoverySession)
        calls: list[str] = []
        session._page = None
        session._clear_browser_data_and_blank = lambda progress: calls.append("clear")
        session._close_page = lambda: calls.append("close")
        for status in (
            "account_lookup_issue",
            "photo_id_recovery_required",
            "invalid_ssn",
        ):
            session.prepare_for_next(status, threading.Event())
        self.assertEqual(["clear", "clear", "clear"], calls)

    def test_zero_threads_is_rejected_but_no_upper_limit_is_applied(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            input_path.write_text(
                "111223333,01,02,1963,LAST,FIRST,ADDRESS\n",
                encoding="utf-8-sig",
            )
            engine = ait11.BatchEngine(lambda *_: None, session_factory=FakeSession)
            with self.assertRaisesRegex(ValueError, "大于等于 1"):
                engine.start(
                    input_path,
                    root / "output.csv",
                    thread_count=0,
                    backend="playwright",
                    display_mode="窗口",
                )


class Step19DynamicColumnsAndSsnValidationTests(unittest.TestCase):
    def test_invalid_ssn_page_message_is_an_explicit_result(self) -> None:
        page = _ResultPage(ait11.INVALID_SSN_MESSAGE)
        status = ait11.step_4_judge_password_recovery(
            page,
            threading.Event(),
            timeout_ms=100,
        )
        self.assertEqual("invalid_ssn", status)
        result = ait11.collect_recovery_result(page, status)
        self.assertEqual(ait11.INVALID_SSN_MESSAGE, result.heading)
        self.assertEqual(("", "", ""), (
            result.masked_phone,
            result.masked_email,
            result.recovery_method,
        ))

    def test_columns_after_first_four_are_preserved_separately(self) -> None:
        fields = (
            "111223333",
            "06/29/1963",
            "FIRST",
            "LAST",
            "ADDRESS",
            "EMAIL",
            "PHONE",
        )
        details, address = ait11._parse_input_row(fields, None)
        item = ait11.WorkItem(
            record_id=1,
            details=details,
            original_fields=fields,
            address=address,
        )
        result = ait11.RecoveryResult(
            result_code="invalid_ssn",
            heading=ait11.INVALID_SSN_MESSAGE,
            masked_phone="",
            masked_email="",
            recovery_method="",
        )
        row = ait11.build_cumulative_output_row(item, result)
        self.assertEqual(list(fields), row[:7])
        self.assertEqual(
            ["", "", "", ait11.INVALID_SSN_MESSAGE],
            row[7:],
        )
        self.assertEqual(11, len(row))

    def test_four_column_input_needs_no_placeholder_column(self) -> None:
        fields = ("111223333", "1/2/1963", "FIRST", "LAST")
        details, address = ait11._parse_input_row(fields, None)
        item = ait11.WorkItem(
            record_id=1,
            details=details,
            original_fields=fields,
            address=address,
        )
        result = ait11.RecoveryResult("invalid_ssn", ait11.INVALID_SSN_MESSAGE, "", "", "")
        row = ait11.build_cumulative_output_row(item, result)
        self.assertEqual(list(fields), row[:-4])
        self.assertEqual("1/2/1963", row[1])
        self.assertEqual(8, len(row))

    def test_legacy_split_date_row_preserves_all_seven_input_columns(self) -> None:
        fields = ("111223333", "01", "02", "1963", "LAST", "FIRST", "ADDRESS")
        details, address = ait11._parse_input_row(fields, None)
        item = ait11.WorkItem(
            record_id=1,
            details=details,
            original_fields=fields,
            address=address,
        )
        result = ait11.RecoveryResult("invalid_ssn", ait11.INVALID_SSN_MESSAGE, "", "", "")
        row = ait11.build_cumulative_output_row(item, result)
        self.assertEqual(list(fields), row[:-4])
        self.assertEqual(11, len(row))

    def test_arbitrary_header_order_is_preserved_exactly(self) -> None:
        details = ait11.AccountDetails(
            first_name="FIRST",
            last_name="LAST",
            birth_month="01",
            birth_day="02",
            birth_year="1963",
            ssn="111223333",
            original_ssn="111223333",
        )
        item = ait11.WorkItem(
            record_id=1,
            details=details,
            original_fields=("FIRST", "111223333", "01/02/1963", "LAST", "ADDRESS"),
            address="ADDRESS",
        )
        result = ait11.RecoveryResult("invalid_ssn", ait11.INVALID_SSN_MESSAGE, "", "", "")
        row = ait11.build_cumulative_output_row(item, result)
        self.assertEqual(list(item.original_fields), row[:-4])
        self.assertEqual(9, len(row))


class Step25MyLifeFormatAdaptationTests(unittest.TestCase):
    header = (
        "primary_phone", "ssn", "known_birthday", "full_name", "age",
        "current_address", "city", "state", "zip_code", "email", "phone",
        "birth_year", "reference_age", "column_14", "column_15", "column_16",
        "生日", "性别", "星座", "备注原因",
    )

    def make_row(
        self,
        ssn: str = "487-78-6185",
        known_birthday: str = "1972-09-06",
        full_name: str = "Mary Dangerfield",
        final_birthday: str = "09/07/1972",
    ) -> tuple[str, ...]:
        return (
            "2817334560", ssn, known_birthday, full_name, "53",
            "1515 Rudel Dr #1108", "Tomball", "TX", "77375",
            "person@example.com", "(936) 253-4235", "1972", "53",
            "", "", "", final_birthday, "Female", "Virgo", "matched",
        )

    def test_header_maps_ssn_full_name_preferred_dob_and_address(self) -> None:
        mapping = ait11._header_mapping(self.header)
        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual(1, mapping["ssn"])
        self.assertEqual(2, mapping["dob"])
        self.assertEqual(3, mapping["full_name"])
        self.assertEqual(5, mapping["address"])
        self.assertEqual(16, mapping["preferred_dob"])

        details, address = ait11._parse_input_row(self.make_row(), mapping)
        self.assertEqual("487786185", details.ssn)
        self.assertEqual(("Mary", "Dangerfield"), (details.first_name, details.last_name))
        self.assertEqual(("09", "07", "1972"), (
            details.birth_month, details.birth_day, details.birth_year,
        ))
        self.assertEqual("1515 Rudel Dr #1108", address)

    def test_empty_preferred_dob_falls_back_to_known_birthday(self) -> None:
        mapping = ait11._header_mapping(self.header)
        assert mapping is not None
        details, _address = ait11._parse_input_row(
            self.make_row(final_birthday=""), mapping
        )
        self.assertEqual(("09", "06", "1972"), (
            details.birth_month, details.birth_day, details.birth_year,
        ))

    def test_output_preserves_every_original_column_then_appends_results(self) -> None:
        mapping = ait11._header_mapping(self.header)
        assert mapping is not None
        fields = self.make_row()
        details, address = ait11._parse_input_row(fields, mapping)
        item = ait11.WorkItem(
            record_id=1,
            details=details,
            original_fields=fields,
            address=address,
            input_mapping=ait11._freeze_mapping(mapping),
        )
        result = ait11.RecoveryResult(
            "account_not_found", "Account Not Found", "", "", "",
        )
        output = ait11.build_cumulative_output_row(item, result)
        self.assertEqual(item.record_key, ait11._input_row_key(details, fields, address, ait11._freeze_mapping(mapping)))
        self.assertEqual(list(fields), output[:-4])
        self.assertEqual(24, len(output))
        self.assertEqual(["", "", "", "Account Not Found"], output[-4:])

    def test_mylife_same_ssn_but_any_later_column_change_has_different_key(self) -> None:
        mapping = ait11._header_mapping(self.header)
        assert mapping is not None
        first_fields = self.make_row()
        second_values = list(first_fields)
        second_values[-1] = "different note"
        second_fields = tuple(second_values)
        first_details, first_address = ait11._parse_input_row(first_fields, mapping)
        second_details, second_address = ait11._parse_input_row(second_fields, mapping)
        frozen_mapping = ait11._freeze_mapping(mapping)
        self.assertNotEqual(
            ait11._input_row_key(
                first_details, first_fields, first_address, frozen_mapping
            ),
            ait11._input_row_key(
                second_details, second_fields, second_address, frozen_mapping
            ),
        )
        case_values = list(first_fields)
        case_values[-1] = "MATCHED"
        case_fields = tuple(case_values)
        case_details, case_address = ait11._parse_input_row(case_fields, mapping)
        self.assertNotEqual(
            ait11._input_row_key(
                first_details, first_fields, first_address, frozen_mapping
            ),
            ait11._input_row_key(
                case_details, case_fields, case_address, frozen_mapping
            ),
        )

    def test_delete_by_output_ssn_uses_mapped_second_column(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "mylife.csv"
            rows = [
                self.header,
                self.make_row(ssn="487-78-6185"),
                self.make_row(ssn="589-70-1425", full_name="Robin Rochford"),
            ]
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)
            record = ait11.load_input_records(path)[0]
            item = ait11.WorkItem(1, record.details, record.source_file, record.source_sheet, record.source_row, record.original_fields, record.address, record.input_mapping)
            removed = ait11.remove_input_rows_by_keys(path, {item.record_key})
            self.assertEqual(1, removed)
            remaining = list(csv.reader(path.read_text(encoding="utf-8-sig").splitlines()))
            self.assertEqual(2, len(remaining))
            self.assertEqual("ssn", remaining[0][1])
            self.assertEqual("589-70-1425", remaining[1][1])

    def test_mylife_format_imports_valid_and_missing_dob_rows(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "mylife.csv"
            rows = [
                self.header,
                self.make_row(ssn="487-78-6185"),
                self.make_row(
                    ssn="589-70-1425", full_name="Robin Rochford",
                    known_birthday="", final_birthday="",
                ),
            ]
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)
            records = ait11.load_input_records(path)
        valid = [record for record in records if not record.import_error]
        missing = [
            record for record in records
            if record.import_error.startswith("缺少必填字段：")
        ]
        self.assertEqual(2, len(records))
        self.assertEqual(1, len(valid))
        self.assertEqual(1, len(missing))
        self.assertTrue(all(dict(record.input_mapping).get("ssn") == 1 for record in records))


class Step20AccountRecoveryInProgressTests(unittest.TestCase):
    def test_account_recovery_in_progress_is_output_as_explicit_status(self) -> None:
        page = _ResultPage("Header\n  Account   Recovery In Progress  \nFooter")
        events: list[str] = []
        status = ait11.step_4_judge_password_recovery(
            page,
            threading.Event(),
            events.append,
            timeout_ms=100,
        )
        self.assertEqual("account_recovery_in_progress", status)
        result = ait11.collect_recovery_result(page, status)
        self.assertEqual(ait11.ACCOUNT_RECOVERY_IN_PROGRESS_HEADING, result.heading)
        self.assertEqual(("", "", ""), (
            result.masked_phone,
            result.masked_email,
            result.recovery_method,
        ))
        self.assertIn(
            f"已识别结果：{ait11.ACCOUNT_RECOVERY_IN_PROGRESS_HEADING}",
            events,
        )

    def test_account_recovery_in_progress_uses_terminal_cleanup(self) -> None:
        session = ait11.BrowserRecoverySession.__new__(ait11.BrowserRecoverySession)
        calls: list[str] = []
        session._page = None
        session._clear_browser_data_and_blank = lambda progress: calls.append("clear")
        session._close_page = lambda: calls.append("close")
        session.prepare_for_next(
            "account_recovery_in_progress",
            threading.Event(),
        )
        self.assertEqual(["clear"], calls)


class Step22AccountNotFoundStatusTests(unittest.TestCase):
    def test_create_new_account_heading_is_preserved_as_exact_result(self) -> None:
        page = _ResultPage(ait11.ACCOUNT_NOT_FOUND_CREATE_HEADING)
        events: list[str] = []
        status = ait11.step_4_judge_password_recovery(
            page,
            threading.Event(),
            events.append,
            timeout_ms=100,
        )
        self.assertEqual("account_not_found", status)
        result = ait11.collect_recovery_result(page, status)
        self.assertEqual(ait11.ACCOUNT_NOT_FOUND_CREATE_HEADING, result.heading)
        self.assertEqual(("", "", ""), (
            result.masked_phone,
            result.masked_email,
            result.recovery_method,
        ))
        self.assertIn(
            f"已识别结果：{ait11.ACCOUNT_NOT_FOUND_CREATE_HEADING}",
            events,
        )


class _HttpCookie:
    def __init__(self, name: str, value: str) -> None:
        self.name = name
        self.value = value


class _HttpResponse:
    def __init__(self, status_code: int = 200, payload=None) -> None:
        self.status_code = status_code
        self._payload = {} if payload is None else payload

    def json(self):
        return self._payload


class _HttpSession:
    def __init__(self, lookup_payload=None, lookup_status: int = 200) -> None:
        self.cookies: list[_HttpCookie] = []
        self.lookup_payload = lookup_payload or {
            "status": "ERROR", "errorCodes": ["USER_NOT_FOUND"],
        }
        self.lookup_status = lookup_status
        self.calls: list[tuple[str, str, dict]] = []
        self.closed = False

    def get(self, url: str, **kwargs):
        self.calls.append(("GET", url, kwargs))
        if url == ait11.HTTP_GATEWAY_SESSION_URL:
            self.cookies = [_HttpCookie("XSRF-TOKEN", "test-xsrf-token")]
        return _HttpResponse()

    def post(self, url: str, **kwargs):
        self.calls.append(("POST", url, kwargs))
        if url == ait11.HTTP_ACCOUNT_LOOKUP_URL:
            return _HttpResponse(self.lookup_status, self.lookup_payload)
        return _HttpResponse()

    def close(self) -> None:
        self.closed = True


class Step28HttpBackendTests(unittest.TestCase):
    def make_item(self) -> ait11.WorkItem:
        details = ait11.AccountDetails(
            first_name="Test", last_name="Person",
            birth_month="01", birth_day="02", birth_year="1980",
            ssn="987654320", original_ssn="987-65-4320",
        )
        return ait11.WorkItem(record_id=1, details=details)

    def test_http_is_added_without_changing_default_backend(self) -> None:
        self.assertEqual(("http", "browser-use", "playwright"), ait11.PROCESSING_BACKENDS)
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            config_path = Path(directory) / "missing.json"
            self.assertEqual("browser-use", ait11.load_gui_config(config_path)["backend"])
            config_path.write_text('{"backend":"http"}', encoding="utf-8")
            self.assertEqual("http", ait11.load_gui_config(config_path)["backend"])

    def test_http_session_bootstrap_and_lookup_request(self) -> None:
        fake = _HttpSession()
        session = ait11.HttpRecoverySession(3, session_factory=lambda: fake)
        session.start()
        result = session.process(self.make_item(), threading.Event())
        self.assertEqual("account_not_found", result.result_code)
        self.assertEqual(ait11.ACCOUNT_NOT_FOUND_CREATE_HEADING, result.heading)
        urls = [(method, url) for method, url, _kwargs in fake.calls]
        self.assertEqual([
            ("GET", ait11.RETRIEVE_ACCOUNT_DETAILS_URL),
            ("GET", ait11.HTTP_KEEP_SESSION_ALIVE_URL),
            ("POST", ait11.HTTP_SESSION_USER_URL),
            ("GET", ait11.HTTP_GATEWAY_SESSION_URL),
            ("POST", ait11.HTTP_ACCOUNT_LOOKUP_URL),
        ], urls)
        lookup_kwargs = fake.calls[-1][2]
        self.assertEqual("test-xsrf-token", lookup_kwargs["headers"]["X-XSRF-TOKEN"])
        self.assertEqual({
            "lastName": "Person", "dob": "1980-01-02",
            "firstName": "Test", "ssn": "987654320",
        }, lookup_kwargs["json"])

    def test_all_current_http_error_codes_map_to_existing_results(self) -> None:
        expected = {
            "SSN_INVALID": ("invalid_ssn", ait11.INVALID_SSN_MESSAGE),
            "USER_NOT_FOUND": ("account_not_found", ait11.ACCOUNT_NOT_FOUND_CREATE_HEADING),
            "PII_MISMATCH": ("account_not_found", "Account Not Found"),
            "MULTIPLE_USERS_FOUND": ("account_lookup_issue", ait11.ACCOUNT_LOOKUP_ISSUE_HEADING),
            "DISABLED_BY_FSA": ("account_disabled", ait11.ACCOUNT_DISABLED_HEADING),
            "ACCT_RECOVERY_CASE_PENDING": (
                "account_recovery_in_progress", ait11.ACCOUNT_RECOVERY_IN_PROGRESS_HEADING,
            ),
            "LAST_NAME_DOB_SSN_COMBO_LOCKED": ("limit_reached", ait11.LIMIT_REACHED_HEADING),
            "SELF_SERVICE_UNAVAILABLE": (
                "photo_id_recovery_required", ait11.PHOTO_ID_RECOVERY_MESSAGE,
            ),
        }
        for error_code, (result_code, heading) in expected.items():
            with self.subTest(error_code=error_code):
                result = ait11.recovery_result_from_http_payload({
                    "status": "ERROR", "errorCodes": [error_code],
                })
                self.assertEqual((result_code, heading), (result.result_code, result.heading))
                self.assertEqual(("", "", ""), (
                    result.masked_phone, result.masked_email, result.recovery_method,
                ))

    def test_success_contacts_are_masked_before_result_storage(self) -> None:
        result = ait11.recovery_result_from_http_payload({
            "status": "SUCCESS",
            "selfServiceOptions": [
                {"type": "MOBILE", "value": "+1 (202) 555-0183"},
                {"type": "EMAIL", "value": "alice@example.com"},
            ],
        })
        self.assertEqual("can_recover", result.result_code)
        self.assertEqual("(⦁⦁⦁) ⦁⦁⦁ 0183", result.masked_phone)
        self.assertEqual("al⦁⦁⦁@example.com", result.masked_email)
        self.assertEqual("Recover my account with a photo ID", result.recovery_method)

    def test_unknown_http_status_is_not_written_as_completed(self) -> None:
        with self.assertRaisesRegex(RuntimeError, "未识别"):
            ait11.recovery_result_from_http_payload({
                "status": "ERROR", "errorCodes": ["NEW_UNKNOWN_CODE"],
            })

    def test_engine_selects_http_session_only_for_http_backend(self) -> None:
        engine = ait11.BatchEngine()
        http_session = engine._new_session(1, "http", "无头")
        browser_session = engine._new_session(2, "playwright", "无头")
        self.assertIsInstance(http_session, ait11.HttpRecoverySession)
        self.assertIsInstance(browser_session, ait11.BrowserRecoverySession)


class Step29UniversalHeaderAdaptationTests(unittest.TestCase):
    result_header = [
        "query", "search_type", "status", "message", "created_at", "updated_at",
        "detail_index", "full_name", "former_names", "first_name", "middle_name",
        "last_name", "age", "address", "phone_numbers", "email_addresses",
        "primary_phone", "current_address", "property_info",
        "current_address_property_info", "past_addresses", "possible_relatives",
        "possible_associates", "search_attempt", "result_id", "result_first_name",
        "result_last_name", "result_middle_name", "result_zip_code", "result_address",
        "result_city", "result_state", "result_phone", "result_birth_date",
        "result_ssn", "status", "message",
    ]

    def make_result_row(self, **overrides: str) -> list[str]:
        values = {name: "" for name in self.result_header}
        values.update({
            "query": "input-key", "full_name": "Original Person",
            "first_name": "Original", "last_name": "Person",
            "address": "Original Address", "result_first_name": "Matched",
            "result_last_name": "Identity", "result_address": "Matched Address",
            "result_birth_date": "1980-01-02", "result_ssn": "487-78-6185",
        })
        values.update(overrides)
        return [values[name] for name in self.result_header]

    def test_result_table_header_is_detected_and_result_identity_has_priority(self) -> None:
        mapping = ait11._header_mapping(self.result_header)
        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual(34, mapping["result_ssn"])
        self.assertEqual(33, mapping["result_dob"])
        details, address = ait11._parse_input_row(self.make_result_row(), mapping)
        self.assertEqual("487786185", details.ssn)
        self.assertEqual(("Matched", "Identity"), (details.first_name, details.last_name))
        self.assertEqual(("01", "02", "1980"), (
            details.birth_month, details.birth_day, details.birth_year,
        ))
        self.assertEqual("Matched Address", address)

    def test_blank_result_names_fall_back_to_original_full_name(self) -> None:
        mapping = ait11._header_mapping(self.result_header)
        assert mapping is not None
        details, _address = ait11._parse_input_row(
            self.make_result_row(result_first_name="", result_last_name=""), mapping,
        )
        self.assertEqual(("Original", "Person"), (details.first_name, details.last_name))

    def test_unknown_vendor_prefixes_are_inferred_by_header_semantics(self) -> None:
        header = [
            "vendor_verified_social_security_number",
            "provider_matched_date_of_birth",
            "lookup_found_given_name",
            "lookup_found_family_name",
            "vendor_verified_address",
        ]
        mapping = ait11._header_mapping(header)
        self.assertEqual({
            "result_ssn": 0, "result_dob": 1, "result_first_name": 2,
            "result_last_name": 3, "result_address": 4,
        }, mapping)
        details, address = ait11._parse_input_row(
            ["487786185", "19800102", "Matched", "Identity", "Matched Address"],
            mapping,
        )
        self.assertEqual(("01", "02", "1980"), (
            details.birth_month, details.birth_day, details.birth_year,
        ))
        self.assertEqual("Matched Address", address)

    def test_chinese_semantic_headers_are_supported(self) -> None:
        header = ["美国社会安全号", "出生日期", "名", "姓", "当前地址"]
        mapping = ait11._header_mapping(header)
        self.assertIsNotNone(mapping)
        details, address = ait11._parse_input_row(
            ["487786185", "1980年01月02日", "First", "Last", "Address"], mapping,
        )
        self.assertEqual(("First", "Last", "Address"), (
            details.first_name, details.last_name, address,
        ))

    def test_common_compact_iso_and_dotted_dates_are_supported(self) -> None:
        expected = ("01", "02", "1980")
        for value in (
            "19800102", "01021980", "1980-01-02T00:00:00Z",
            "1980.01.02", "01.02.1980", "1980年01月02日",
        ):
            with self.subTest(value=value):
                self.assertEqual(expected, ait11._parse_dob(value))

    def test_live_811_workbook_imports_valid_rows_and_preserves_all_columns(self) -> None:
        path = Path(__file__).resolve().parent.parent / "存资料处" / "8.11 背调手机号 SSN.xlsx"
        if not path.is_file():
            path = Path(r"E:\studentaid\存资料处\8.11 背调手机号 SSN.xlsx")
        if not path.is_file():
            self.skipTest("现场 8.11 工作簿不存在")
        records = ait11.load_input_records(path)
        valid = [record for record in records if not record.import_error]
        self.assertEqual(736, len(records))
        self.assertEqual(465, len(valid))
        self.assertEqual({37}, {len(record.original_fields) for record in records})
        self.assertEqual(2, records[0].source_row)
        first = valid[0]
        item = ait11.WorkItem(
            record_id=1, details=first.details, source_file=first.source_file,
            source_sheet=first.source_sheet, source_row=first.source_row,
            original_fields=first.original_fields, address=first.address,
            input_mapping=first.input_mapping,
        )
        output = ait11.build_cumulative_output_row(
            item, ait11.RecoveryResult("account_not_found", "Account Not Found", "", "", ""),
        )
        self.assertEqual(list(first.original_fields), output[:37])
        self.assertEqual(41, len(output))


class Step30HeaderlessColumnInferenceTests(unittest.TestCase):
    def make_row(
        self,
        ssn: str = "487786185",
        first_name: str = "First",
        middle_name: str = "Middle",
        last_name: str = "Last",
        dob: str = "01/02/1980",
    ) -> list[str]:
        return [
            ssn, first_name, middle_name, last_name, dob,
            "masked@example.test", "2025550183", "123 Example St, Austin, TX 78701",
            "Female", "Capricorn",
        ]

    def test_headerless_wide_rows_infer_identity_columns_without_skipping_first_row(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "headerless.csv"
            rows = [
                self.make_row(),
                self.make_row("589701425", "Second", "", "Person", "1981-03-04"),
                self.make_row("376234235", "Third", "Q", "Identity", "19820506"),
            ]
            with path.open("w", encoding="utf-8-sig", newline="") as stream:
                csv.writer(stream).writerows(rows)
            records = ait11.load_input_records(path)
        self.assertEqual(3, len(records))
        self.assertTrue(all(not record.import_error for record in records))
        self.assertEqual({
            "address": 7, "dob": 4, "first_name": 1, "last_name": 3, "ssn": 0,
        }, dict(records[0].input_mapping))
        self.assertEqual(1, records[0].source_row)
        self.assertEqual(("First", "Last"), (
            records[0].details.first_name, records[0].details.last_name,
        ))

    def test_middle_name_is_not_used_as_last_name(self) -> None:
        mapping = ait11._infer_headerless_mapping([
            self.make_row(middle_name="A"), self.make_row(middle_name=""),
            self.make_row(middle_name="Beth"),
        ])
        self.assertIsNotNone(mapping)
        assert mapping is not None
        self.assertEqual(1, mapping["first_name"])
        self.assertEqual(3, mapping["last_name"])

    def test_existing_compact_headerless_orders_keep_legacy_parser(self) -> None:
        dob_second = [
            ["487786185", "01/02/1980", "First", "Last", "Address"],
            ["589701425", "03/04/1981", "Second", "Person", "Address"],
        ]
        dob_fourth = [
            ["487786185", "First", "Last", "01/02/1980", "Address"],
            ["589701425", "Second", "Person", "03/04/1981", "Address"],
        ]
        self.assertIsNone(ait11._infer_headerless_mapping(dob_second))
        self.assertIsNone(ait11._infer_headerless_mapping(dob_fourth))

    def test_live_817_workbook_all_rows_import_and_output_preserves_ten_columns(self) -> None:
        path = Path(__file__).resolve().parent.parent / "导入资料" / "8.17 测试.xlsx"
        if not path.is_file():
            path = Path(r"E:\studentaid\导入资料\8.17 测试.xlsx")
        if not path.is_file():
            self.skipTest("现场 8.17 工作簿不存在")
        records = ait11.load_input_records(path)
        self.assertEqual(936, len(records))
        self.assertEqual(936, sum(not record.import_error for record in records))
        self.assertEqual({10}, {len(record.original_fields) for record in records})
        self.assertEqual((1, 936), (records[0].source_row, records[-1].source_row))
        self.assertEqual({
            "address": 7, "dob": 4, "first_name": 1, "last_name": 3, "ssn": 0,
        }, dict(records[0].input_mapping))
        first = records[0]
        item = ait11.WorkItem(
            record_id=1, details=first.details, original_fields=first.original_fields,
            address=first.address, input_mapping=first.input_mapping,
        )
        output = ait11.build_cumulative_output_row(
            item, ait11.RecoveryResult("account_not_found", "Account Not Found", "", "", ""),
        )
        self.assertEqual(list(first.original_fields), output[:10])
        self.assertEqual(14, len(output))


class Step31FullRowDedupeAllPathsTests(unittest.TestCase):
    @staticmethod
    def make_row(email: str) -> list[str]:
        return [
            "487786185", "First", "Middle", "Last", "01/02/1980",
            email, "2025550183", "123 Example St, Austin, TX 78701",
            "Female", "Capricorn",
        ]

    def test_headerless_xlsx_removal_deletes_exact_rows_not_same_ssn_rows(self) -> None:
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "same_ssn_different_full_rows.xlsx"
            first_row = self.make_row("first@example.test")
            different_row = self.make_row("different@example.test")
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(first_row)
            worksheet.append(different_row)
            worksheet.append(first_row)
            workbook.save(path)
            workbook.close()

            records = ait11.load_input_records(path)
            keys = [
                ait11._input_row_key(
                    record.details, record.original_fields, record.address,
                    record.input_mapping,
                )
                for record in records
            ]
            self.assertEqual(keys[0], keys[2])
            self.assertNotEqual(keys[0], keys[1])
            self.assertEqual(2, ait11.remove_input_rows_by_keys(path, {keys[0]}))
            remaining = ait11.load_input_records(path)

        self.assertEqual(1, len(remaining))
        self.assertEqual(tuple(different_row), remaining[0].original_fields)
        self.assertEqual("487786185", remaining[0].details.ssn)

    def test_live_817_same_ssn_rows_stay_separate_when_full_rows_differ(self) -> None:
        path = Path(__file__).resolve().parent.parent / "导入资料" / "8.17 测试.xlsx"
        if not path.is_file():
            path = Path(r"E:\studentaid\导入资料\8.17 测试.xlsx")
        if not path.is_file():
            self.skipTest("现场 8.17 工作簿不存在")
        records = [
            record for record in ait11.load_input_records(path)
            if not record.import_error
        ]
        row_keys = {
            ait11._input_row_key(
                record.details, record.original_fields, record.address,
                record.input_mapping,
            )
            for record in records
        }
        unique_ssns = {record.details.ssn for record in records}
        self.assertEqual(936, len(records))
        self.assertEqual(936, len(row_keys))
        self.assertEqual(928, len(unique_ssns))


class Step37TabDelimitedHeaderlessIdentityTests(unittest.TestCase):
    @staticmethod
    def make_rows() -> list[list[str]]:
        return [
            [
                "2025550101", "1", "Alice", "Carter",
                "Age 46, Born March 1980",
                "alice@example.test|alice.alt@example.test",
                "123 Example St", "Phoenix", "AZ", "85001",
                "123456789", "03/25/1980",
            ],
            [
                "2025550102", "1", "Brian", "Miller",
                "Age 51, Born July 1975",
                "brian@example.test|brian.alt@example.test|third@example.test",
                "456 Sample Ave", "Tucson", "AZ", "85701",
                "234567890", "07/30/1975",
            ],
            [
                "2025550103", "1", "Chloe", "Wilson",
                "Age 59, Born March 1967", "chloe@example.test",
                "789 Fixture Rd", "Little Rock", "AR", "72201",
                "345678901", "03/22/1967",
            ],
            [
                "2025550104", "1", "Daniel", "Anderson",
                "Age 61, Born January 1965", "daniel@example.test",
                "321 Test Dr", "Flagstaff", "AZ", "86001",
                "456789012", "01/17/1965",
            ],
        ]

    def write_tab_file(self, path: Path) -> list[list[str]]:
        rows = self.make_rows()
        with path.open("w", encoding="utf-8", newline="") as stream:
            csv.writer(stream, delimiter="\t").writerows(rows)
        return rows

    def test_tab_wins_over_one_comma_inside_each_age_field(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "tab_with_age_commas.txt"
            expected = self.write_tab_file(path)
            encoding, delimiter, detected = ait11._detect_text_format(path)
            loaded = ait11._read_text_rows(path)
        self.assertIn(encoding, {"utf-8-sig", "utf-8"})
        self.assertEqual("\t", delimiter)
        self.assertEqual(expected, detected)
        self.assertEqual(expected, loaded)
        self.assertEqual({12}, {len(row) for row in loaded})

    def test_twelve_column_identity_mapping_uses_adjacent_names_not_city_or_state(self) -> None:
        rows = self.make_rows()
        mapping = ait11._infer_headerless_mapping(rows)
        self.assertEqual(
            {"address": 6, "dob": 11, "first_name": 2, "last_name": 3, "ssn": 10},
            mapping,
        )
        self.assertFalse(ait11._looks_like_headerless_name("AZ"))
        self.assertTrue(ait11._looks_like_headerless_name("Alice"))

    def test_all_rows_import_and_all_original_columns_are_preserved(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "twelve_columns.txt"
            expected = self.write_tab_file(path)
            records = ait11.load_input_records(path)
        self.assertEqual(4, len(records))
        self.assertTrue(all(not record.import_error for record in records))
        self.assertEqual({12}, {len(record.original_fields) for record in records})
        self.assertEqual(expected[0], list(records[0].original_fields))
        self.assertEqual(("Alice", "Carter"), (
            records[0].details.first_name, records[0].details.last_name,
        ))
        self.assertEqual("123456789", records[0].details.ssn)
        self.assertEqual(("03", "25", "1980"), (
            records[0].details.birth_month, records[0].details.birth_day,
            records[0].details.birth_year,
        ))
        self.assertEqual("123 Example St", records[0].address)

    def test_exact_row_removal_reuses_tab_detection_and_preserves_other_rows(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            path = Path(directory) / "remove_one.txt"
            expected = self.write_tab_file(path)
            records = ait11.load_input_records(path)
            key = ait11._input_row_key(
                records[1].details, records[1].original_fields,
                records[1].address, records[1].input_mapping,
            )
            self.assertEqual(1, ait11.remove_input_rows_by_keys(path, {key}))
            encoding, delimiter, remaining = ait11._detect_text_format(path)
        self.assertIn(encoding, {"utf-8-sig", "utf-8"})
        self.assertEqual("\t", delimiter)
        self.assertEqual([expected[0], expected[2], expected[3]], remaining)


class _GuiValue:
    def __init__(self, value="") -> None:
        self.value = value

    def get(self):
        return self.value

    def set(self, value) -> None:
        self.value = value


class _GuiControl:
    def __init__(self) -> None:
        self.state = ""

    def configure(self, **kwargs) -> None:
        if "state" in kwargs:
            self.state = kwargs["state"]


class _GuiMessages:
    def __init__(self, confirm: bool = True) -> None:
        self.confirm = confirm
        self.calls: list[tuple[str, str]] = []

    def askyesno(self, _title: str, message: str) -> bool:
        self.calls.append(("askyesno", message))
        return self.confirm

    def showinfo(self, _title: str, message: str) -> None:
        self.calls.append(("showinfo", message))

    def showwarning(self, _title: str, message: str) -> None:
        self.calls.append(("showwarning", message))

    def showerror(self, _title: str, message: str) -> None:
        self.calls.append(("showerror", message))


class Step33GuiAsyncClearDatabaseTests(unittest.TestCase):
    @staticmethod
    def create_database(path: Path) -> None:
        connection = sqlite3.connect(path)
        try:
            connection.executescript(
                """
                CREATE TABLE batches(batch_id TEXT PRIMARY KEY);
                CREATE TABLE records(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    batch_id TEXT NOT NULL
                );
                CREATE TABLE settings(name TEXT PRIMARY KEY, value TEXT NOT NULL);
                INSERT INTO batches(batch_id) VALUES ('batch-1'), ('batch-2');
                INSERT INTO records(batch_id) VALUES ('batch-1'), ('batch-1'), ('batch-2');
                INSERT INTO settings(name, value) VALUES ('keep', 'yes');
                """
            )
            connection.commit()
        finally:
            connection.close()

    def test_clear_database_removes_batches_records_and_keeps_schema(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            database_path = Path(directory) / ait11.DATABASE_FILENAME
            self.create_database(database_path)
            result = ait11.clear_database_contents(database_path)
            connection = sqlite3.connect(database_path)
            try:
                counts = (
                    connection.execute("SELECT COUNT(*) FROM batches").fetchone()[0],
                    connection.execute("SELECT COUNT(*) FROM records").fetchone()[0],
                    connection.execute("SELECT COUNT(*) FROM settings").fetchone()[0],
                )
                sequence_count = connection.execute(
                    "SELECT COUNT(*) FROM sqlite_sequence WHERE name='records'"
                ).fetchone()[0]
            finally:
                connection.close()
        self.assertTrue(result.existed)
        self.assertEqual((2, 3), (result.batches_deleted, result.records_deleted))
        self.assertEqual((0, 0, 1), counts)
        self.assertEqual(0, sequence_count)

    def test_missing_database_is_reported_without_creating_a_file(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            database_path = Path(directory) / ait11.DATABASE_FILENAME
            result = ait11.clear_database_contents(database_path)
            self.assertFalse(result.existed)
            self.assertFalse(database_path.exists())

    @staticmethod
    def create_gui_app(output_path: Path):
        app = ait11.StudentAidApp.__new__(ait11.StudentAidApp)
        app.engine = type("Engine", (), {"is_running": False})()
        app.output_var = _GuiValue(str(output_path))
        app.status_var = _GuiValue("就绪")
        app.messagebox = _GuiMessages(confirm=True)
        app._database_clear_active = False
        app._database_clear_thread = None
        app._closing = False
        app._ui_events = queue.Queue()
        app.test_logs = []
        app._append_log = app.test_logs.append
        for name in (
            "input_entry", "output_entry", "backend_combo", "display_mode_combo",
            "thread_spin", "input_button", "output_button", "start_button",
            "stop_button", "clear_database_button",
        ):
            setattr(app, name, _GuiControl())
        return app

    def test_gui_clear_database_runs_in_worker_and_does_not_touch_csv(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            folder = Path(directory)
            output_path = folder / ait11.CUMULATIVE_OUTPUT_FILENAME
            output_path.write_text("original-output\n", encoding="utf-8")
            database_path = folder / ait11.DATABASE_FILENAME
            self.create_database(database_path)
            app = self.create_gui_app(output_path)
            app._clear_database()
            thread = app._database_clear_thread
            self.assertIsNotNone(thread)
            thread.join(timeout=10)
            self.assertFalse(thread.is_alive())
            kind, payload = app._ui_events.get_nowait()
            self.assertEqual("database_clear_finished", kind)
            self.assertTrue(app._database_clear_active)
            self.assertEqual("正在清空数据库", app.status_var.get())
            self.assertFalse(
                any(kind == "showinfo" for kind, _message in app.messagebox.calls)
            )
            app._finish_database_clear(payload["result"])
            connection = sqlite3.connect(database_path)
            try:
                remaining = connection.execute("SELECT COUNT(*) FROM records").fetchone()[0]
            finally:
                connection.close()
            output_text = output_path.read_text(encoding="utf-8")
        self.assertEqual(0, remaining)
        self.assertEqual("original-output\n", output_text)
        self.assertEqual("数据库已清空", app.status_var.get())
        self.assertTrue(any(kind == "askyesno" for kind, _message in app.messagebox.calls))
        self.assertTrue(app.test_logs)
        self.assertFalse(app._database_clear_active)
        self.assertEqual("normal", app.clear_database_button.state)

    def test_slow_database_clear_returns_to_gui_immediately(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            folder = Path(directory)
            output_path = folder / ait11.CUMULATIVE_OUTPUT_FILENAME
            output_path.write_text("keep\n", encoding="utf-8")
            database_path = folder / ait11.DATABASE_FILENAME
            self.create_database(database_path)
            app = self.create_gui_app(output_path)
            entered = threading.Event()
            release = threading.Event()
            original_clear = ait11.clear_database_contents

            def slow_clear(path: Path):
                entered.set()
                release.wait(timeout=5)
                return ait11.DatabaseClearResult(path.resolve(), True, 2, 3)

            ait11.clear_database_contents = slow_clear
            try:
                started_at = time.monotonic()
                app._clear_database()
                elapsed = time.monotonic() - started_at
                self.assertTrue(entered.wait(timeout=1))
                self.assertLess(elapsed, 1.0)
                self.assertTrue(app._database_clear_active)
                self.assertTrue(app._database_clear_thread.is_alive())
                self.assertEqual("disabled", app.start_button.state)
                self.assertEqual("disabled", app.clear_database_button.state)
                self.assertEqual(
                    ["askyesno"], [kind for kind, _message in app.messagebox.calls]
                )
            finally:
                release.set()
                if app._database_clear_thread is not None:
                    app._database_clear_thread.join(timeout=5)
                ait11.clear_database_contents = original_clear

            kind, payload = app._ui_events.get_nowait()
            self.assertEqual("database_clear_finished", kind)
            app._finish_database_clear(payload["result"])
            self.assertEqual("数据库已清空", app.status_var.get())
            self.assertEqual("normal", app.start_button.state)

    def test_running_state_disables_button_and_method_has_second_guard(self) -> None:
        app = ait11.StudentAidApp.__new__(ait11.StudentAidApp)
        app._database_clear_active = False
        for name in (
            "input_entry", "output_entry", "backend_combo", "display_mode_combo",
            "thread_spin", "input_button", "output_button", "start_button",
            "stop_button", "clear_database_button",
        ):
            setattr(app, name, _GuiControl())
        app._set_running(True)
        self.assertEqual("disabled", app.clear_database_button.state)
        app._set_running(False)
        self.assertEqual("normal", app.clear_database_button.state)

        app._database_clear_active = True
        app._set_running(False)
        self.assertEqual("disabled", app.start_button.state)
        self.assertEqual("disabled", app.clear_database_button.state)

        app.engine = type("Engine", (), {"is_running": True})()
        app._database_clear_active = False
        app.messagebox = _GuiMessages()
        app.output_var = _GuiValue("unused.csv")
        app._clear_database()
        self.assertEqual("showwarning", app.messagebox.calls[0][0])
        self.assertFalse(any(kind == "askyesno" for kind, _message in app.messagebox.calls))


class Step36WindowsOutputAclTests(unittest.TestCase):
    @unittest.skipUnless(ait11.os.name == "nt", "Windows ACL only")
    def test_output_directory_and_existing_files_receive_current_account_grants(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            output_directory = Path(directory) / "export"
            output_directory.mkdir()
            output_path = output_directory / "result.csv"
            output_path.write_text("row\n", encoding="utf-8")
            missing_sidecar = output_directory / "studentaid.sqlite3-wal"
            completed = type("Completed", (), {"returncode": 0, "stdout": b""})()
            with mock.patch.object(
                ait11, "_current_windows_account", return_value=r"HOST\USER"
            ), mock.patch.object(ait11.subprocess, "run", return_value=completed) as run:
                ait11.ensure_output_storage_writable(
                    output_directory, output_path, missing_sidecar
                )

            commands = [call.args[0] for call in run.call_args_list]
            self.assertEqual(2, len(commands))
            self.assertEqual("icacls.exe", commands[0][0])
            self.assertEqual(str(output_directory.resolve()), commands[0][1])
            self.assertIn(r"HOST\USER:(OI)(CI)F", commands[0])
            self.assertEqual(str(output_path.resolve()), commands[1][1])
            self.assertIn(r"HOST\USER:F", commands[1])
            self.assertFalse(any(str(missing_sidecar) in command for command in commands))

    @unittest.skipUnless(ait11.os.name == "nt", "Windows ACL only")
    def test_directory_acl_failure_stops_before_sqlite_open(self) -> None:
        with tempfile.TemporaryDirectory(dir=Path(__file__).parent) as directory:
            output_directory = Path(directory) / "export"
            completed = type("Completed", (), {"returncode": 5, "stdout": b"denied"})()
            with mock.patch.object(
                ait11, "_current_windows_account", return_value=r"HOST\USER"
            ), mock.patch.object(ait11.subprocess, "run", return_value=completed):
                with self.assertRaisesRegex(PermissionError, "输出目录写权限"):
                    ait11.ensure_output_storage_writable(output_directory)


if __name__ == "__main__":
    unittest.main(verbosity=2)





