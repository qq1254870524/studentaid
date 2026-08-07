import csv
import os
from pathlib import Path
import tempfile
import threading
import time
import unittest

from openpyxl import Workbook, load_workbook

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
import ait6


class SuccessSession:
    prepared = []
    closed_workers = []
    active = 0
    max_active = 0
    activity_lock = threading.Lock()

    def __init__(self, worker_number: int) -> None:
        self.worker_number = worker_number
        self.browser_mode = "mock-dedicated-cdp"

    def start(self) -> None:
        pass

    def process(self, item, stop_event, progress_callback=None):
        with self.activity_lock:
            type(self).active += 1
            type(self).max_active = max(type(self).max_active, type(self).active)
        if progress_callback:
            progress_callback("mock result")
        try:
            time.sleep(0.1)
            return ait6.RecoveryResult(
                result_code="account_not_found",
                heading="Account Not Found",
                masked_phone="",
                masked_email="",
                recovery_method="",
            )
        finally:
            with self.activity_lock:
                type(self).active -= 1

    def prepare_for_next(self, result_code, stop_event, progress_callback=None):
        self.prepared.append(result_code)

    def close(self) -> None:
        self.closed_workers.append(self.worker_number)


class LimitResultSession(SuccessSession):
    def process(self, item, stop_event, progress_callback=None):
        return ait6.RecoveryResult(
            result_code="limit_reached",
            heading=ait6.LIMIT_REACHED_HEADING,
            masked_phone="",
            masked_email="",
            recovery_method="",
        )


class StopAfterResultSession(SuccessSession):
    def process(self, item, stop_event, progress_callback=None):
        result = super().process(item, stop_event, progress_callback)
        stop_event.set()
        return result


class StudentAidStep11Tests(unittest.TestCase):
    def setUp(self):
        SuccessSession.prepared = []
        SuccessSession.closed_workers = []
        SuccessSession.active = 0
        SuccessSession.max_active = 0

    @staticmethod
    def _write_input(path: Path, keys=("111223333", "222334444")) -> None:
        rows = [
            [keys[0], "01", "02", "1990", "Example", "First", "Address 1"],
            [keys[1], "03", "04", "1991", "Example", "Second", "Address 2"],
        ]
        with path.open("w", encoding="utf-8", newline="") as stream:
            csv.writer(stream).writerows(rows)

    def _run_engine(
        self,
        input_path,
        output_path,
        session_factory=SuccessSession,
        thread_count=2,
        backend="playwright",
    ):
        events = []
        engine = ait6.BatchEngine(
            event_callback=lambda kind, payload: events.append((kind, payload)),
            session_factory=session_factory,
        )
        engine.start(input_path, output_path, thread_count, backend)
        self.assertTrue(engine.wait(20), "engine did not finish")
        finished = [payload for kind, payload in events if kind == "finished"]
        self.assertEqual(len(finished), 1, events)
        return finished[0], events

    def test_cumulative_output_is_appended_and_input_rows_are_deleted(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)
            output_path.write_text(
                "111-22-3333,01/02/1990,First,Example,Address 1,Account Not Found,,,\n",
                encoding="utf-8-sig",
            )

            finished, _events = self._run_engine(input_path, output_path)

            self.assertEqual(finished["counts"]["completed"], 1)
            self.assertEqual(input_path.read_text(encoding="utf-8"), "")
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
            self.assertEqual([ait6._normalise_record_key(row[0]) for row in rows], [
                "111223333", "222334444"
            ])
            self.assertEqual(SuccessSession.prepared, ["account_not_found"])

            finished_again, _events = self._run_engine(input_path, output_path)
            self.assertEqual(finished_again["counts"]["completed"], 0)
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                self.assertEqual(len(list(csv.reader(stream))), 2)

    def test_limit_reached_is_normal_output_and_deletes_input_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)

            finished, events = self._run_engine(
                input_path, output_path, LimitResultSession
            )

            self.assertEqual(finished["status"], "completed")
            self.assertEqual(finished["counts"]["completed"], 2)
            self.assertEqual(finished["counts"]["failed"], 0)
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.reader(stream))
            self.assertEqual([row[5] for row in rows], [
                ait6.LIMIT_REACHED_HEADING,
                ait6.LIMIT_REACHED_HEADING,
            ])
            self.assertEqual(input_path.read_text(encoding="utf-8"), "")
            messages = [p.get("message", "") for k, p in events if k == "log"]
            self.assertFalse(any("全部保留" in message for message in messages))

    def test_two_workers_process_concurrently_and_both_close(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)

            finished, _events = self._run_engine(input_path, output_path)

            self.assertEqual(finished["counts"]["completed"], 2)
            self.assertGreaterEqual(SuccessSession.max_active, 2)
            self.assertEqual(sorted(SuccessSession.closed_workers), [1, 2])

    def test_reference_output_preserves_original_five_column_dob(self):
        details = ait6._validate_details(
            "014646401", "09", "07", "1980", "Joel", "Beauregard"
        )
        item = ait6.WorkItem(
            record_id=1,
            details=details,
            original_fields=(
                "014646401",
                "September 07, 1980",
                "Joel",
                "Beauregard",
                "1408 TEASLEY LN 1321,DENTON, TX 76205",
            ),
            address="1408 TEASLEY LN 1321,DENTON, TX 76205",
        )
        result = ait6.RecoveryResult(
            "limit_reached", ait6.LIMIT_REACHED_HEADING, "", "", ""
        )

        row = ait6.build_cumulative_output_row(item, result)

        self.assertEqual(row[1], "September 07, 1980")
        self.assertEqual(len(row), 9)
        self.assertEqual(row[5], ait6.LIMIT_REACHED_HEADING)

    def test_english_month_dob_from_fake_workbook_is_valid(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "english-dob.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append([
                "014646401",
                "September 07, 1980",
                "Joel",
                "Beauregard",
                "Example Address",
            ])
            workbook.save(path)
            workbook.close()

            records = ait6.load_input_records(path)

        self.assertEqual(len(records), 1)
        self.assertIsNotNone(records[0].details)
        self.assertFalse(records[0].import_error)

    def test_backend_selection_validation(self):
        self.assertEqual(ait6.BROWSER_BACKENDS, ("browser-use", "playwright"))
        with self.assertRaises(ValueError):
            ait6.BrowserRecoverySession(1, "unsupported")

    def test_stop_after_result_does_not_overwrite_completed_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            input_path = root / "input.csv"
            output_path = root / "cumulative.csv"
            self._write_input(input_path)

            finished, _events = self._run_engine(
                input_path, output_path, StopAfterResultSession, thread_count=1
            )

            self.assertEqual(finished["status"], "stopped")
            self.assertEqual(finished["counts"]["completed"], 1)
            self.assertEqual(finished["counts"]["stopped"], 1)
            self.assertEqual(len(ait6.load_input_records(input_path)), 1)
            with output_path.open(encoding="utf-8-sig", newline="") as stream:
                self.assertEqual(len(list(csv.reader(stream))), 1)

    def test_xlsx_first_column_deletion_preserves_header_and_other_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "input.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["SSN", "DOB", "First Name", "Last Name", "Address"])
            sheet.append(["111-22-3333", "01/02/1990", "First", "Example", "A"])
            sheet.append(["222334444", "03/04/1991", "Second", "Example", "B"])
            workbook.save(path)
            workbook.close()

            removed = ait6.remove_input_rows_by_keys(path, {"111223333"})

            self.assertEqual(removed, 1)
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                rows = list(workbook.active.iter_rows(values_only=True))
            finally:
                workbook.close()
            self.assertEqual(rows[0][0], "SSN")
            self.assertEqual(rows[1][0], "222334444")

    @unittest.skipIf(ait6.sync_playwright is None, "playwright not installed")
    def test_limit_reached_is_detected_immediately(self):
        with ait6.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content("<h2>Limit Reached: Try Again in 24 Hours</h2>")
            started = time.monotonic()
            try:
                status = ait6.step_4_judge_password_recovery(
                    page, threading.Event(), timeout_ms=5_000
                )
                result = ait6.collect_recovery_result(page, status)
            finally:
                browser.close()
            self.assertEqual(status, "limit_reached")
            self.assertEqual(result.heading, ait6.LIMIT_REACHED_HEADING)
            self.assertLess(time.monotonic() - started, 1)

    @unittest.skipIf(ait6.sync_playwright is None, "playwright not installed")
    def test_stalled_loading_requests_session_rebuild(self):
        with ait6.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_content("<div>Loading...</div>")
            try:
                with self.assertRaises(ait6.PageSubmissionStalled):
                    ait6.step_4_judge_password_recovery(
                        page,
                        threading.Event(),
                        timeout_ms=1_000,
                        poll_interval_ms=10,
                        stalled_loading_seconds=0.05,
                    )
            finally:
                browser.close()

    @unittest.skipIf(ait6.sync_playwright is None, "playwright not installed")
    def test_retrieve_result_cancel_returns_to_empty_form(self):
        html = """
        <input id="fsa_Input_ForgotUsernameFirstName" value="First">
        <input id="fsa_Input_ForgotUsernameLastName" value="Last">
        <select id="fsa_Input_ForgotUsernameDateOfBirthMonth"><option value="01" selected>Jan</option></select>
        <input id="fsa_Input_ForgotUsernameDateOfBirthDay" value="2">
        <input id="fsa_Input_ForgotUsernameDateOfBirthYear" value="1990">
        <input id="fsa_Input_ForgotUsernameSsnInput" value="111223333">
        <button id="fsa_Button_ForgotUsernameCancel"><span>Cancel</span></button>
        <script>
        document.querySelector('button').onclick = () => {
          for (const x of document.querySelectorAll('input,select')) x.value = '';
        };
        </script>
        """
        with ait6.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            page = browser.new_page(viewport={"width": 900, "height": 700})
            page.set_content(html)
            session = ait6.BrowserRecoverySession(1)
            session._browser = browser
            session._context = page.context
            session._page = page
            messages = []
            try:
                session.prepare_for_next(
                    "can_recover", threading.Event(), messages.append
                )
                self.assertEqual(page.locator(ait6.FIRST_NAME_SELECTOR).input_value(), "")
                self.assertTrue(any("Cancel 已完成" in value for value in messages))
            finally:
                browser.close()

    @unittest.skipIf(ait6.sync_playwright is None, "playwright not installed")
    def test_account_not_found_cleanup_clears_data_and_returns_blank(self):
        with ait6.sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()
            context.add_cookies([
                {"name": "test", "value": "1", "url": "https://studentaid.gov"}
            ])
            session = ait6.BrowserRecoverySession(1)
            session._browser = browser
            session._context = context
            session._page = page
            session._dedicated_profile = True
            messages = []
            try:
                session.prepare_for_next(
                    "account_not_found", threading.Event(), messages.append
                )
                self.assertEqual(session._page.url, "about:blank")
                self.assertEqual(context.cookies(), [])
                self.assertTrue(any("数据已清除" in value for value in messages))
            finally:
                browser.close()

    @unittest.skipUnless(os.name == "nt", "dedicated Chrome integration is Windows-only")
    @unittest.skipIf(ait6.sync_playwright is None, "playwright not installed")
    def test_dedicated_chrome_cdp_has_no_webdriver_flag(self):
        session = ait6.BrowserRecoverySession(1, "playwright")
        try:
            session.start()
            self.assertFalse(session._page.evaluate("navigator.webdriver"))
            self.assertTrue(session._dedicated_profile)
            self.assertIn("AutomationControlled", session.browser_mode)
        finally:
            session.close()


if __name__ == "__main__":
    unittest.main()
