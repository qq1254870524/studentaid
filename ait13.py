"""
StudentAid 浏览器自动化工具 v13。

2026-08-08 第十八步新增明确状态与任意线程版更新：
1. 点击 Continue 后若出现 ``Account Lookup Issue: Get Help``，将该完整文案作为
   正常明确结果写入累计 CSV，随后沿用终态清理流程处理下一条。
2. 点击 Continue 后若出现 ``We are unable to retrieve your log-in information. Access
   your account by recovering your account with a photo ID.``，将该完整文案作为正常
   明确结果写入累计 CSV，不再误入普通 Retrieve 联系方式提取或等待超时。
3. GUI 线程输入改为可直接填写任意正整数，批处理层同步移除 8 线程硬上限；默认值、
   每线程独立浏览器、停止清缓存、结果落盘和输入删除等既有逻辑保持不变。

2026-08-08 第十七步瞬时失败队尾重试版更新：
1. 浏览器任务完成三次页面会话重建后仍未取得明确结果时，不再立即作为本批终态
   失败；任务会回到队尾，先让其余资料继续处理，避免单条慢响应持续占用线程。
2. 每个唯一任务最多执行三轮队列尝试，每轮仍保留原有三次页面清理、重开、重填；
   下一轮会在当前队列后方执行，让官网瞬时故障有充分恢复时间。
3. SQLite 新增同组 retry 状态回退：代表任务与同第一列重复行一起回到 pending，
   attempt_count 继续累计；最终明确结果仍只追加一行并一次删除全部匹配输入行。
4. 协调器改为等待动态队列真正清空后再发送线程结束信号，保证运行中追加到队尾的
   重试任务不会落在结束信号之后而被遗漏；停止按钮仍可立即结束并保留未完成输入。

2026-08-08 第十六步大批量去重稳定版更新：
1. 同一批输入按规范化后的第一列只创建一个浏览器任务；重复资料不再被 8 个线程
   同时领取，避免同一 SSN 重复打开页面、重复点击 Continue 和重复触发官网限流。
2. 唯一任务取得明确结果后，SQLite 会把同一批次、同一第一列的全部重复记录一起
   标记完成；累计 CSV 仍只写一行，输入文件仍一次删除全部匹配行。
3. 唯一任务失败或停止时，同组重复记录一起进入相同终态并保留在输入文件，后续
   批次可以整体重试，不会出现没有队列任务却长期停在 pending 的情况。
4. 累计 CSV 启动规范化现在同时删除全空物理行；即使文件已经是全字段双引号格式，
   也会清除文件开头或中间的空白行，保持固定 9 列数据连续。

2026-08-08 第十五步批量输入兼容版更新：
1. 兼容 ``MM/DD/YYY`` 或 ``MM-DD-YYY`` 的三位年份：仅当补前导 ``1`` 后落在
   1900 到当前年份且能组成真实日期时才接受，避免把其他错误日期静默改写。
2. First Name、Last Name、Month、Day、Year、SSN 任意必填项为空时，直接从输入
   文件删除该源行，不写累计输出，也不启动浏览器处理该行。
3. 输入处理发生在校验阶段；累计输出生日仍严格为单列 ``MM/DD/YYYY``，
   其他无法可靠恢复的格式错误继续保留在输入文件，不会被误删。

2026-08-08 第十四步生日格式修复版更新：
1. 累计输出第 2 列生日强制统一为 ``MM/DD/YYYY``，月份和日期始终补足两位。
2. 不再保留五列输入中的英文月份、短日期或其他原始显示格式；输入可以继续使用
   支持的多种生日写法，但输出只采用校验后的 month/day/year 生成单列标准日期。
3. 七列拆分生日输入仍按月、日、年读取，累计 CSV 不会把生日拆成三列；每行继续
   固定为 5 个资料列加 4 个结果列，共 9 列。
4. 现场已有累计 CSV 的 17 条生日已全部是单列 ``MM/DD/YYYY``，无需重写现有结果。

2026-08-08 第十三步联系方式修复版更新：
1. 修复找回结果页手机号漏记：StudentAid 实际掩码字符为 ``⦁``（U+2981），
   旧版只识别 ``*`` 和 ``•``，导致 ``(⦁⦁⦁) ⦁⦁⦁ 8139`` 未写入累计 CSV。
2. 联系方式改为优先读取结果卡片中可见的 ``p.fsa-color-gray-60``，手机号和
   邮箱分别按掩码格式判断；页面有值就原样记录，没有就严格保持空白。
3. 兜底扫描也只限定在可见 ``p`` 元素，避免把页脚邮箱或页面其他文字误当成
   找回邮箱；兼容 ``⦁``、``•``、``●`` 和 ``*`` 掩码字符。
4. Retrieve 的 Cancel 改为直接点击可见 ``<span>Cancel</span>``，并验证结果页
   实际离开；旧版仅发出按钮坐标点击，现场存在没有点上的情况。点击成功后复用同一个
   worker、浏览器、页面会话和空白表单继续填下一条，不清缓存、不重启 Chrome。
5. Account Not Found 保存后清除全部浏览器数据、回到 about:blank，下一条才重新打开
   指定网址；这条清理规则不用于 Retrieve。
6. 保持累计输出、明确结果后删除输入行，以及启动时按累计 CSV 第一列清理输入行。

2026-08-08 第十二步无头稳定版更新：
1. GUI 新增“浏览器显示”下拉框，选项严格为“窗口”和“无头”；默认“无头”。
2. 窗口模式显示 900x720 的小 Chrome 窗口，页面 viewport 固定为 880x650；
   无头模式完全不创建可见浏览器窗口。
3. browser-use / playwright 与窗口 / 无头可自由组合；每个 worker 继续拥有独立
   Chrome 进程、CDP/Context 和缓存目录，不复用用户已有浏览器。
4. 批处理日志记录本次后端、显示模式和线程数；运行中禁止修改两个下拉选项。
5. 停止或结束时两种显示模式都执行相同的 cookie/cache/storage/service worker
   清理，并结束本批次浏览器；browser-use 临时 profile 同步删除。
6. 新增四组合启动参数、GUI 默认值和浏览器生命周期回归；使用假资料测试4.xlsx
   的隔离副本对四种组合逐一现场验证。
7. 修复小窗口下 Continue 坐标与页面缩放不一致导致鼠标落点未触发：优先使用
   Playwright Locator 的真实按钮点击，未进入提交状态时依次用 Enter 和 DOM click 兜底，
   每种方式都必须确认 Loading、按钮禁用、结果文案或目标路径后才进入结果等待。
8. 实测 Chromium 原生 headless 会被 StudentAid 以 HTTP/2 协议错误拒绝；“无头”改用
   Windows 隐藏的普通 Chrome 网络栈，启动时只隐藏本批次精确 PID 的窗口，用户看不到
   浏览器且不会影响已有 Chrome。导航以 commit 返回，随后仍严格等待表单字段和 Loading。
9. browser-use 结束顺序改为先结束精确 PID 树、删除一次性 profile，再断开 CDP，避免
   失败页面令关闭阶段卡住；四组合 2 线程现场矩阵 8/8 完成且残留 Chrome 为 0。

2026-08-08 第十一步稳定版更新：
1. GUI 新增 browser-use / playwright 后端下拉选项；两种模式都为每个 worker
   创建独立浏览器进程和独立临时配置目录，默认使用 2 个处理线程。
2. browser-use 模式直接采用 BrowserProfile 的 Chrome 参数、隔离目录和 CDP
   启动策略，再执行现有确定性表单流程，避免默认共享 daemon 在并发时互相抢焦点。
3. Limit Reached: Try Again in 24 Hours 改为正常明确结果：实时追加累计 CSV，
   最终状态原样写入第 6 列，随后删除对应输入行并继续处理下一条。
4. 停止或批次结束时，每个 worker 都清除 cookie、cache、storage、service worker，
   回到空白页并结束本脚本创建的浏览器进程；browser-use 临时配置目录同步删除。
5. 恢复并发任务队列和单独 SQLite writer，累计输出与输入删行继续使用进程内锁和
   原子文件替换，两个浏览器线程不会同时破坏 CSV/XLSX。
6. 键盘逐字输入延迟由 60ms 调整为 20ms，字段完成后的稳定等待由 500ms 调整为
   200ms；仍保留真实键盘事件、失焦和 Continue 启用校验。
7. 累计 CSV 的 DOB 对 5 列无表头输入保留原始显示格式，与实际输出参考一致。
8. 根据假资料和实际输出参考补齐 ``September 07, 1980`` 这类英文月份 DOB
   导入格式；此类有效资料不再作为格式错误留在输入文件。

2026-08-08 第十步稳定版更新：
1. Account Not Found 明确结果实时保存后，清除 StudentAid 专用浏览器全部站点数据，
   回到 about:blank；处理下一条时重新打开找回页面。
2. Retrieve Your Log-in Information 明确结果实时保存后，点击可见 Cancel，
   等待返回空白输入表单，再填写下一条资料。
3. 采用严格顺序处理，避免多个共享 CDP 页面同时清理 cookie/session 导致提交互相干扰。
4. 累计结果实时追加到用户选择的 CSV，不覆盖旧数据；程序启动时读取输出第一列，
   输入第一列已存在于输出的资料会直接从输入文件原子删除。
5. 每条明确结果按“先写 SQLite、再追加累计 CSV、最后删除输入行”的顺序落盘，
   中断重启时可依靠输出第一列去重并继续。
6. 定位 Continue 转圈根因：Playwright 直接启动 Chrome 时 navigator.webdriver=true；
   新版复用 browser-use 的 AutomationControlled 启动参数，实测启动后该标志为 false。
7. Continue 改用可见按钮坐标鼠标点击；持续 Loading 自动清理、重开、重填并重试。
8. 识别官方 Limit Reached: Try Again in 24 Hours，立即停止批次并保留当前及后续输入。
9. browser-use 实测 Account Not Found、全数据清理、about:blank 重开及官方九次限制；
   Retrieve 结果后的 Cancel 使用真实按钮结构并由 Playwright DOM 回归覆盖。
10. 新增统一一键安装启动 CMD；依赖存在就跳过，缺少时才安装。

2026-08-07 第九步源码校验更新：
1. 使用 browser-use 独立核验 StudentAid 真实页面字段、Continue 按钮和提交状态时间线。
2. 删除“姓名输入框消失即判定可找回”的宽泛条件；只有明确结果标志才允许完成，
   防止页面跳转或短暂重绘被误判为成功。
3. 页面结果等待改为可停止的短轮询；保留 Loading 等待，并准确识别站点错误。
4. CDP 模式复用 Chrome 默认浏览器上下文并新开页面，保持与 browser-use 实测一致；
   输入改用真实键盘事件、字段失焦和提交前稳定等待，修复过快 fill 导致站点返回 unknown error。
5. 检测到因页面停留过久产生的 unknown error 时自动新开页面、重新填表并提交一次。
6. 每条记录实时输出“打开页面、填写资料、已提交、等待结果、已识别结果”等安全阶段日志；
   长时间等待时每 5 秒输出心跳，最终状态实时写 SQLite 并刷新 GUI，不记录原始资料。
7. Playwright 仍是正式运行根基，browser-use 仅用于独立校验。

2026-08-07 第八步实测更新：
1. 保留第七步的 GUI、SQLite WAL、并发处理和 CSV 导出能力。
2. 新增 Chrome CDP 连接模式：优先复用已通过站点网络校验的本机 Chrome，
   解决新启动 Playwright 浏览器访问 StudentAid 时的 HTTP/2 协议错误。
3. 支持 STUDENTAID_CDP_URL 显式指定 CDP 地址；未设置时自动探测
   http://127.0.0.1:9223，设置为 off/none/disabled 可关闭自动探测。
4. 新增页面导航重试、共享浏览器安全断开和浏览器模式日志。
5. 修复输入页标题被误判为找回成功的问题；现在会等待 Loading 结束，
   并把站点未知错误准确记录为失败。
6. 修复纯文本 CSV 中全数字 SSN 被 float 转换吞掉前导 0 的问题。
7. 第八步使用测试资料完成 GUI、浏览器、SQLite 和导出链路实测。

2026-08-07 第七步更新（原 ait2.py 日期记录受本机时钟影响）：
1. 新增 Tkinter GUI，可选择 CSV/SCV/TXT/XLSX 输入文件和结果目录。
2. 导入资料先进入 SQLite；数据库固定启用 WAL。
3. 使用多个独立 Playwright 处理线程和一个专用 SQLite 写入线程。
4. 支持开始、停止、线程数设置和实时进度；停止后不再领取新任务。
5. 每个批次完成或停止后自动导出 UTF-8 BOM CSV。
6. 兼容无表头的 ``SSN,月,日,年,姓,名,地址``、旧版
   ``SSN,DOB,First Name,Last Name,Address``，以及带常见英文表头的资料。

旧版 ait.py 保持不变。本文件不把 SSN 或整条原始资料打印到日志。
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
import json
import os
from pathlib import Path
import queue
import re
import shutil
import socket
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
from typing import Any, Callable, Iterable, Mapping, Sequence
import urllib.request
import uuid

sys.dont_write_bytecode = True

try:
    from playwright.sync_api import Browser, Page, Playwright, sync_playwright
except ImportError:  # GUI 仍可启动，并在开始处理时给出明确错误。
    Browser = Page = Playwright = Any  # type: ignore[assignment,misc]
    sync_playwright = None


APP_TITLE = "StudentAid 批量处理工具 - 第十八步新增状态与任意线程版"
DATABASE_FILENAME = "studentaid.sqlite3"
CUMULATIVE_OUTPUT_FILENAME = "StudentAid累计结果.csv"
LIMIT_REACHED_HEADING = "Limit Reached: Try Again in 24 Hours"
ACCOUNT_DISABLED_HEADING = "Your Account Is Disabled"
ACCOUNT_LOOKUP_ISSUE_HEADING = "Account Lookup Issue: Get Help"
PHOTO_ID_RECOVERY_MESSAGE = (
    "We are unable to retrieve your log-in information. "
    "Access your account by recovering your account with a photo ID."
)
BROWSER_BACKENDS = ("browser-use", "playwright")
DISPLAY_MODES = ("窗口", "无头")
MAX_QUEUE_ATTEMPTS = 3
SMALL_WINDOW_SIZE = {"width": 900, "height": 720}
SMALL_VIEWPORT_SIZE = {"width": 880, "height": 650}
RETRIEVE_ACCOUNT_DETAILS_URL = (
    "https://studentaid.gov/fsa-id/sign-in/retrieve-account-details"
)

FIRST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameFirstName"
LAST_NAME_SELECTOR = "#fsa_Input_ForgotUsernameLastName"
BIRTH_MONTH_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthMonth"
BIRTH_DAY_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthDay"
BIRTH_YEAR_SELECTOR = "#fsa_Input_ForgotUsernameDateOfBirthYear"
SSN_SELECTOR = "#fsa_Input_ForgotUsernameSsnInput"

RESULT_COLUMNS = (
    "Result_Heading",
    "Masked_Phone",
    "Masked_Email",
    "Recovery_Method",
    "Processing_Status",
    "Error",
    "Source_File",
    "Source_Sheet",
    "Source_Row",
)


@dataclass(frozen=True)
class AccountDetails:
    """StudentAid 账户找回页面所需字段。"""

    first_name: str
    last_name: str
    birth_month: str
    birth_day: str
    birth_year: str
    ssn: str
    original_ssn: str


@dataclass(frozen=True)
class RecoveryResult:
    """页面返回的可公开保存结果；联系方式保持页面原有脱敏形式。"""

    result_code: str
    heading: str
    masked_phone: str
    masked_email: str
    recovery_method: str


@dataclass(frozen=True)
class ImportedRecord:
    """一条已读取的输入资料；校验失败的行也会进入数据库。"""

    source_file: str
    source_sheet: str
    source_row: int
    original_fields: tuple[str, ...]
    details: AccountDetails | None
    address: str
    import_error: str = ""


@dataclass(frozen=True)
class WorkItem:
    """SQLite、浏览器、累计输出和输入删行之间传递的任务。"""

    record_id: int
    details: AccountDetails
    source_file: str = ""
    source_sheet: str = ""
    source_row: int = 0
    original_fields: tuple[str, ...] = ()
    address: str = ""

    @property
    def record_key(self) -> str:
        first_value = self.original_fields[0] if self.original_fields else self.details.original_ssn
        return _normalise_record_key(first_value)


class StopRequested(RuntimeError):
    """在网页处理步骤之间收到停止请求。"""


class PageSessionExpired(RuntimeError):
    """StudentAid 页面停留过久后返回 unknown error，需要重新打开页面。"""


class PageSubmissionStalled(RuntimeError):
    """Continue 已提交但站点长时间停在 Loading，需要重建浏览器会话。"""


class MissingRequiredField(ValueError):
    """输入源行缺少用户指定的 StudentAid 必填字段，应直接删除。"""


def _now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def _clean_error(exc: BaseException | str, limit: int = 1000) -> str:
    """生成适合写库和显示的单行错误，不包含输入资料。"""
    text = str(exc).replace("\r", " ").replace("\n", " ").strip()
    return text[:limit] or type(exc).__name__


def _normalise_month(value: str) -> str:
    aliases = {
        "january": "01", "jan": "01", "february": "02", "feb": "02",
        "march": "03", "mar": "03", "april": "04", "apr": "04",
        "may": "05", "june": "06", "jun": "06", "july": "07",
        "jul": "07", "august": "08", "aug": "08", "september": "09",
        "sep": "09", "sept": "09", "october": "10", "oct": "10",
        "november": "11", "nov": "11", "december": "12", "dec": "12",
    }
    cleaned = value.strip().lower()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    month = aliases.get(cleaned) or (cleaned.zfill(2) if cleaned.isdigit() else "")
    if month not in {f"{number:02d}" for number in range(1, 13)}:
        raise ValueError("出生月份必须是 1-12、01-12 或英文月份名称")
    return month


def _normalise_day(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if not digits or not 1 <= int(digits) <= 31:
        raise ValueError("出生日期（日）必须是 1-31")
    return digits.zfill(2)


def _normalise_year(value: str) -> str:
    cleaned = value.strip()
    if re.fullmatch(r"\d+(?:\.0+)?", cleaned):
        cleaned = str(int(float(cleaned)))
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) != 4 or not 1900 <= int(digits) <= date.today().year:
        raise ValueError("出生年份必须是四位数字，且不晚于当前年份")
    return digits


def _normalise_ssn(value: str) -> str:
    cleaned = value.strip()
    # 只移除 Excel 数字单元格常见的 .0 后缀；纯数字文本必须保留前导 0。
    if re.fullmatch(r"\d+\.0+", cleaned):
        cleaned = cleaned.split(".", 1)[0]
    digits = re.sub(r"\D", "", cleaned)
    # Excel 把首位 0 当数字格式丢掉时，允许 8 位数字恢复为 9 位。
    if len(digits) == 8:
        digits = digits.zfill(9)
    if len(digits) != 9:
        raise ValueError("Social Security Number 必须包含 9 位数字")
    return digits


def _validate_details(
    ssn_value: str,
    month_value: str,
    day_value: str,
    year_value: str,
    first_name: str,
    last_name: str,
) -> AccountDetails:
    ssn = _normalise_ssn(ssn_value)
    month = _normalise_month(month_value)
    day = _normalise_day(day_value)
    year = _normalise_year(year_value)
    first_name = first_name.strip()
    last_name = last_name.strip()

    if not first_name or len(first_name) > 35:
        raise ValueError("First Name 不能为空且长度不能超过 35 个字符")
    if not last_name or len(last_name) > 35:
        raise ValueError("Last Name 不能为空且长度不能超过 35 个字符")
    try:
        date(int(year), int(month), int(day))
    except ValueError as exc:
        raise ValueError("出生日期不是有效日期") from exc

    return AccountDetails(
        first_name=first_name,
        last_name=last_name,
        birth_month=month,
        birth_day=day,
        birth_year=year,
        ssn=ssn,
        original_ssn=ssn_value,
    )


def _parse_dob(value: str) -> tuple[str, str, str]:
    value = value.strip()
    short_year = re.fullmatch(r"(\d{1,2})[/-](\d{1,2})[/-](\d{3})", value)
    if short_year:
        month, day, year_tail = short_year.groups()
        candidates: list[date] = []
        for restored_year in (int(f"1{year_tail}"), int(f"{year_tail}0")):
            if not 1900 <= restored_year <= date.today().year:
                continue
            try:
                candidates.append(date(restored_year, int(month), int(day)))
            except ValueError:
                continue
        if len(candidates) == 1:
            parsed = candidates[0]
            return (
                f"{parsed.month:02d}",
                f"{parsed.day:02d}",
                f"{parsed.year:04d}",
            )
    for date_format in (
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%y", "%m-%d-%y", "%B %d, %Y", "%b %d, %Y",
        "%B %d %Y", "%b %d %Y",
    ):
        try:
            parsed = datetime.strptime(value, date_format).date()
            return f"{parsed.month:02d}", f"{parsed.day:02d}", f"{parsed.year:04d}"
        except ValueError:
            continue
    raise ValueError(
        "DOB 格式无效，应为 MM/DD/YYYY、MM-DD-YYYY、YYYY-MM-DD "
        "或英文月份日期（如 September 07, 1980）"
    )


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, date):
        return value.strftime("%m/%d/%Y")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalise_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.strip().casefold())


HEADER_ALIASES: Mapping[str, set[str]] = {
    "ssn": {"ssn", "socialsecuritynumber", "socialsecurityno", "socialsecurity"},
    "dob": {"dob", "dateofbirth", "birthdate"},
    "month": {"birthmonth", "dobmonth", "month", "mm"},
    "day": {"birthday", "dobday", "day", "dd"},
    "year": {"birthyear", "dobyear", "year", "yyyy"},
    "first_name": {"firstname", "givenname", "fname", "first"},
    "last_name": {"lastname", "surname", "familyname", "lname", "last"},
    "address": {"address", "streetaddress", "mailingaddress", "homeaddress"},
}


def _header_mapping(row: Sequence[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(row):
        cleaned = _normalise_header(value)
        for canonical, aliases in HEADER_ALIASES.items():
            if cleaned in aliases and canonical not in mapping:
                mapping[canonical] = index
                break
    has_date = "dob" in mapping or {"month", "day", "year"}.issubset(mapping)
    if "ssn" in mapping and has_date and {"first_name", "last_name"}.issubset(mapping):
        return mapping
    return None


def _mapped_value(row: Sequence[str], mapping: Mapping[str, int], key: str) -> str:
    index = mapping.get(key)
    return row[index].strip() if index is not None and index < len(row) else ""


def _row_looks_like_split_date(row: Sequence[str]) -> bool:
    if len(row) < 6:
        return False
    try:
        if row[1].strip():
            _normalise_month(row[1])
        if row[2].strip():
            _normalise_day(row[2])
        if row[3].strip():
            _normalise_year(row[3])
        return True
    except ValueError:
        return False


def _raise_for_missing_required_fields(
    ssn: str,
    month: str,
    day: str,
    year: str,
    first_name: str,
    last_name: str,
) -> None:
    fields = (
        ("SSN", ssn),
        ("Month", month),
        ("Day", day),
        ("Year", year),
        ("First Name", first_name),
        ("Last Name", last_name),
    )
    missing = [name for name, value in fields if not str(value).strip()]
    if missing:
        raise MissingRequiredField("缺少必填字段：" + ", ".join(missing))


def _parse_input_row(
    fields: Sequence[str],
    header: Mapping[str, int] | None,
) -> tuple[AccountDetails, str]:
    row = [str(value) for value in fields]
    if header is not None:
        ssn = _mapped_value(row, header, "ssn")
        first_name = _mapped_value(row, header, "first_name")
        last_name = _mapped_value(row, header, "last_name")
        address = _mapped_value(row, header, "address")
        dob = _mapped_value(row, header, "dob")
        if dob:
            month, day, year = _parse_dob(dob)
        else:
            month = _mapped_value(row, header, "month")
            day = _mapped_value(row, header, "day")
            year = _mapped_value(row, header, "year")
    elif _row_looks_like_split_date(row):
        # 当前测试资料：SSN, 月, 日, 年, Last Name, First Name, Address。
        ssn, month, day, year = row[:4]
        last_name, first_name = row[4:6]
        address = ",".join(row[6:]).strip()
    else:
        if len(row) < 4:
            raise ValueError(
                "至少需要 4 列：SSN、DOB、First Name、Last Name"
            )
        ssn = row[0]
        # 同时兼容两种五列顺序：
        # 1) SSN, DOB, First Name, Last Name, Address
        # 2) SSN, First Name, Last Name, DOB, Address
        second_column_error: ValueError | None = None
        try:
            if not row[1].strip():
                raise ValueError("DOB 为空")
            month, day, year = _parse_dob(row[1])
            first_name, last_name = row[2:4]
        except ValueError as exc:
            second_column_error = exc
            try:
                if not row[3].strip():
                    raise ValueError("DOB 为空")
                month, day, year = _parse_dob(row[3])
                first_name, last_name = row[1:3]
            except ValueError:
                if not row[1].strip():
                    month = day = year = ""
                    first_name, last_name = row[2:4]
                else:
                    raise second_column_error
        address = ",".join(row[4:]).strip()

    first_name = first_name.strip()
    last_name = last_name.strip()
    _raise_for_missing_required_fields(
        ssn, month, day, year, first_name, last_name
    )

    return _validate_details(
        ssn, month, day, year, first_name, last_name
    ), address


def _records_from_rows(
    rows: Iterable[Sequence[Any]],
    input_path: Path,
    sheet_name: str = "",
) -> list[ImportedRecord]:
    materialised: list[tuple[int, list[str]]] = []
    for row_number, raw_row in enumerate(rows, start=1):
        row = [_cell_to_text(value) for value in raw_row]
        while row and row[-1] == "":
            row.pop()
        if row and any(value.strip() for value in row):
            materialised.append((row_number, row))

    if not materialised:
        return []

    mapping = _header_mapping(materialised[0][1])
    data_rows = materialised[1:] if mapping is not None else materialised
    imported: list[ImportedRecord] = []
    for row_number, row in data_rows:
        details: AccountDetails | None = None
        address = ""
        import_error = ""
        try:
            details, address = _parse_input_row(row, mapping)
        except Exception as exc:
            import_error = _clean_error(exc)
        imported.append(
            ImportedRecord(
                source_file=str(input_path.resolve()),
                source_sheet=sheet_name,
                source_row=row_number,
                original_fields=tuple(row),
                details=details,
                address=address,
                import_error=import_error,
            )
        )
    return imported


def _read_text_rows(input_path: Path) -> list[list[str]]:
    last_error: UnicodeDecodeError | None = None
    text = ""
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = input_path.read_text(encoding=encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8") from last_error

    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            delimiter = "\t"
    return [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def load_input_records(input_path: Path) -> list[ImportedRecord]:
    """读取 CSV/SCV/TXT/XLSX，返回包括校验失败行在内的导入记录。"""
    input_path = input_path.resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"输入文件不存在：{input_path}")

    suffix = input_path.suffix.casefold()
    records: list[ImportedRecord] = []
    if suffix in {".csv", ".scv", ".txt"}:
        records.extend(_records_from_rows(_read_text_rows(input_path), input_path))
    elif suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from exc
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                records.extend(
                    _records_from_rows(
                        worksheet.iter_rows(values_only=True),
                        input_path,
                        worksheet.title,
                    )
                )
        finally:
            workbook.close()
    else:
        raise ValueError("只支持 .csv、.scv、.txt 和 .xlsx 输入文件")

    if not records:
        raise ValueError("输入文件中没有可导入的资料行")
    return records


_SOURCE_FILE_LOCK = threading.RLock()


def _normalise_record_key(value: Any) -> str:
    """输出第一列/输入第一列比较键；SSN 忽略连字符、空格和 Excel .0。"""
    text = _cell_to_text(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    digits = re.sub(r"\D", "", text)
    if 8 <= len(digits) <= 9:
        return digits.zfill(9)
    return text.casefold()


def _detect_text_format(path: Path) -> tuple[str, str, list[list[str]]]:
    raw = path.read_bytes()
    encoding = "utf-8-sig"
    text = ""
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            text = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("文本文件编码无法识别，请另存为 UTF-8")
    sample = text[:8192]
    delimiter = ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
    except csv.Error:
        if "\t" in sample:
            delimiter = "\t"
    rows = [list(row) for row in csv.reader(text.splitlines(), delimiter=delimiter)]
    return encoding, delimiter, rows


def read_output_first_column_keys(output_path: Path) -> set[str]:
    """读取累计输出第一列；空文件或不存在时返回空集合。"""
    output_path = output_path.resolve()
    if not output_path.is_file() or output_path.stat().st_size == 0:
        return set()
    _encoding, _delimiter, rows = _detect_text_format(output_path)
    return {
        key for row in rows if row and (key := _normalise_record_key(row[0]))
        and key not in {"ssn", "socialsecuritynumber", "inputcolumn1"}
    }


def _atomic_write_text_rows(
    path: Path, rows: Sequence[Sequence[Any]], encoding: str, delimiter: str
) -> None:
    temporary = path.with_name(f".{path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding=encoding, newline="") as stream:
            writer = csv.writer(stream, delimiter=delimiter, lineterminator="\n")
            writer.writerows(rows)
            stream.flush()
            os.fsync(stream.fileno())
        _atomic_replace_with_retry(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def _atomic_replace_with_retry(source: Path, target: Path) -> None:
    """Windows 上外部扫描器短暂占用文件时重试原子替换。"""
    for attempt in range(1, 13):
        try:
            os.replace(source, target)
            return
        except PermissionError:
            if attempt >= 12:
                raise
            time.sleep(min(0.75, 0.1 * attempt))


def remove_input_rows_by_keys(input_path: Path, keys: set[str]) -> int:
    """按第一列原子删除输入行；XLSX 保留样式和其他工作表。"""
    keys = {key for value in keys if (key := _normalise_record_key(value))}
    if not keys:
        return 0
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            kept: list[list[str]] = []
            removed = 0
            mapping = _header_mapping(rows[0]) if rows else None
            for index, row in enumerate(rows):
                if index == 0 and mapping is not None:
                    kept.append(row)
                    continue
                key = _normalise_record_key(row[0]) if row else ""
                if key and key in keys:
                    removed += 1
                else:
                    kept.append(row)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path)
            removed = 0
            temporary = input_path.with_name(
                f".{input_path.stem}.{os.getpid()}.{uuid.uuid4().hex}.tmp.xlsx"
            )
            try:
                for worksheet in workbook.worksheets:
                    header_values = [_cell_to_text(cell.value) for cell in worksheet[1]]
                    mapping = _header_mapping(header_values)
                    first_data_row = 2 if mapping is not None else 1
                    rows_to_delete = []
                    for row_number in range(first_data_row, worksheet.max_row + 1):
                        key = _normalise_record_key(worksheet.cell(row_number, 1).value)
                        if key and key in keys:
                            rows_to_delete.append(row_number)
                    for row_number in reversed(rows_to_delete):
                        worksheet.delete_rows(row_number, 1)
                    removed += len(rows_to_delete)
                if removed:
                    workbook.save(temporary)
                    workbook.close()
                    _atomic_replace_with_retry(temporary, input_path)
                else:
                    workbook.close()
            finally:
                try:
                    workbook.close()
                except Exception:
                    pass
                if temporary.exists():
                    temporary.unlink()
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def remove_input_rows_by_locations(
    input_path: Path, records: Sequence[ImportedRecord]
) -> int:
    """按导入时的工作表/源行精确删除，兼容第一列 SSN 本身为空的记录。"""
    locations = {
        (record.source_sheet, int(record.source_row))
        for record in records
        if record.source_row > 0
    }
    if not locations:
        return 0
    input_path = input_path.resolve()
    with _SOURCE_FILE_LOCK:
        suffix = input_path.suffix.casefold()
        if suffix in {".csv", ".scv", ".txt"}:
            encoding, delimiter, rows = _detect_text_format(input_path)
            kept = [
                row
                for row_number, row in enumerate(rows, start=1)
                if ("", row_number) not in locations
            ]
            removed = len(rows) - len(kept)
            if removed:
                _atomic_write_text_rows(input_path, kept, encoding, delimiter)
            return removed

        if suffix == ".xlsx":
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("修改 XLSX 需要安装 openpyxl") from exc
            workbook = load_workbook(input_path)
            temporary = input_path.with_name(
                f".{input_path.stem}.{os.getpid()}.{uuid.uuid4().hex}.tmp.xlsx"
            )
            removed = 0
            try:
                for worksheet in workbook.worksheets:
                    rows_to_delete = sorted(
                        {
                            row_number
                            for sheet_name, row_number in locations
                            if sheet_name == worksheet.title
                            and 1 <= row_number <= worksheet.max_row
                        },
                        reverse=True,
                    )
                    for row_number in rows_to_delete:
                        worksheet.delete_rows(row_number, 1)
                    removed += len(rows_to_delete)
                if removed:
                    workbook.save(temporary)
                    workbook.close()
                    _atomic_replace_with_retry(temporary, input_path)
                else:
                    workbook.close()
            finally:
                try:
                    workbook.close()
                except Exception:
                    pass
                if temporary.exists():
                    temporary.unlink()
            return removed
        raise ValueError("只支持修改 .csv、.scv、.txt 和 .xlsx 输入文件")


def _format_output_dob(details: AccountDetails) -> str:
    return f"{int(details.birth_month):02d}/{int(details.birth_day):02d}/{details.birth_year}"


def build_cumulative_output_row(item: WorkItem, result: RecoveryResult) -> list[str]:
    """生成固定 9 列结果；第 2 列生日始终为 MM/DD/YYYY。"""
    first_column = (
        item.original_fields[0] if item.original_fields else item.details.original_ssn
    )
    return [
        first_column,
        _format_output_dob(item.details),
        item.details.first_name,
        item.details.last_name,
        item.address,
        result.heading,
        result.masked_phone,
        result.masked_email,
        result.recovery_method,
    ]


def append_cumulative_result(
    output_path: Path, item: WorkItem, result: RecoveryResult
) -> bool:
    """实时追加累计结果；第一列已存在时不重复追加。"""
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with _SOURCE_FILE_LOCK:
        existing_keys = read_output_first_column_keys(output_path)
        if item.record_key and item.record_key in existing_keys:
            return False
        with output_path.open("a", encoding="utf-8-sig", newline="") as stream:
            csv.writer(
                stream, lineterminator="\n", quoting=csv.QUOTE_ALL
            ).writerow(
                build_cumulative_output_row(item, result)
            )
            stream.flush()
            os.fsync(stream.fileno())
    return True


def ensure_cumulative_output_quote_all(output_path: Path) -> bool:
    """把累计 CSV 规范为非空、全字段引号，避免 LibreOffice 错误分列。"""
    output_path = output_path.resolve()
    if not output_path.is_file() or output_path.stat().st_size == 0:
        return False
    raw = output_path.read_text(encoding="utf-8-sig")
    first_nonempty = next((line for line in raw.splitlines() if line.strip()), "")
    _encoding, delimiter, rows = _detect_text_format(output_path)
    nonempty_rows = [
        row
        for row in rows
        if any(_cell_to_text(value).strip() for value in row)
    ]
    has_empty_rows = len(nonempty_rows) != len(rows)
    needs_quote_all = bool(first_nonempty) and not first_nonempty.startswith('"')
    if not has_empty_rows and not needs_quote_all:
        return False
    temporary = output_path.with_name(
        f".{output_path.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp"
    )
    try:
        with temporary.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.writer(
                stream, delimiter=delimiter, lineterminator="\n", quoting=csv.QUOTE_ALL
            )
            writer.writerows(nonempty_rows)
            stream.flush()
            os.fsync(stream.fileno())
        _atomic_replace_with_retry(temporary, output_path)
    finally:
        if temporary.exists():
            temporary.unlink()
    return True


def resolve_output_target(output_target: Path) -> tuple[Path, Path]:
    """兼容旧版目录参数；GUI 第十四步直接选择累计 CSV。"""
    output_target = output_target.resolve()
    if output_target.suffix.casefold() in {".csv", ".scv", ".txt"}:
        return output_target.parent, output_target
    return output_target, output_target / CUMULATIVE_OUTPUT_FILENAME


def _normalise_display_mode(display_mode: str) -> str:
    display_mode = display_mode.strip()
    if display_mode not in DISPLAY_MODES:
        raise ValueError(f"浏览器显示模式必须是：{', '.join(DISPLAY_MODES)}")
    return display_mode


def _is_headless(display_mode: str) -> bool:
    return _normalise_display_mode(display_mode) == "无头"


def _normal_chrome_user_agent(browser_version: str) -> str:
    match = re.search(r"\d+(?:\.\d+){0,3}", str(browser_version))
    version = match.group(0) if match else "120.0.0.0"
    parts = version.split(".")
    version = ".".join((parts + ["0", "0", "0", "0"])[:4])
    return (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        f"Chrome/{version} Safari/537.36"
    )


def _chrome_process_ids(command_line_token: str = "") -> set[int]:
    try:
        import psutil

        process_ids: set[int] = set()
        for process in psutil.process_iter(["pid", "name", "cmdline"]):
            try:
                if (process.info["name"] or "").casefold() != "chrome.exe":
                    continue
                command_line = " ".join(process.info["cmdline"] or [])
                if command_line_token and command_line_token not in command_line:
                    continue
                process_ids.add(int(process.info["pid"]))
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        return process_ids
    except Exception:
        return set()


def _hide_windows_for_processes(process_ids: set[int]) -> int:
    if os.name != "nt" or not process_ids:
        return 0
    try:
        import ctypes
        import ctypes.wintypes

        user32 = ctypes.windll.user32
        hidden = 0
        callback_type = ctypes.WINFUNCTYPE(
            ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p
        )

        @callback_type
        def callback(window_handle: int, _parameter: int) -> bool:
            nonlocal hidden
            process_id = ctypes.c_ulong()
            user32.GetWindowThreadProcessId(window_handle, ctypes.byref(process_id))
            if process_id.value in process_ids and user32.IsWindowVisible(window_handle):
                user32.ShowWindow(window_handle, 0)
                hidden += 1
            return True

        user32.EnumWindows(callback, 0)
        return hidden
    except Exception:
        return 0


@dataclass
class BrowserLaunch:
    """浏览器启动结果；默认浏览器完全由当前工具拥有。"""

    browser: Browser
    external_browser: bool
    mode: str
    owns_browser: bool = False
    dedicated_profile: bool = False


class BrowserUseBrowserHost:
    """使用 browser-use BrowserProfile 启动独立 Chrome 和 CDP。

    browser-use CLI 的默认本机 daemon 是共享浏览器，不能安全供两个 worker 并发。
    本类复用 BrowserProfile 的完整 Chrome 参数，为每个 worker 创建独立进程、CDP
    端口和临时 user-data-dir；避免 BrowserSession 看门狗与 Playwright 双控制器在
    多线程结束时竞争标签页。表单由确定性 Playwright 定位器通过该 CDP 会话执行。
    """

    def __init__(self, worker_number: int, headless: bool = True) -> None:
        self.worker_number = worker_number
        self.headless = bool(headless)
        self._cdp_url = ""
        self._process: subprocess.Popen[Any] | None = None
        self._browser_pid: int | None = None
        self._profile_dir = Path(
            tempfile.mkdtemp(prefix=f"browser-use-user-data-dir-studentaid-{worker_number}-")
        )

    @property
    def cdp_url(self) -> str:
        return self._cdp_url

    @staticmethod
    def _find_free_port() -> int:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as probe:
            probe.bind(("127.0.0.1", 0))
            return int(probe.getsockname()[1])

    def _wait_until_ready(self, timeout: float) -> None:
        deadline = time.monotonic() + timeout
        last_error: BaseException | None = None
        while time.monotonic() < deadline:
            if self._process is not None and self._process.poll() is not None:
                raise RuntimeError(
                    f"browser-use Chrome 提前退出，退出码 {self._process.returncode}"
                )
            try:
                with urllib.request.urlopen(
                    f"{self._cdp_url}/json/version", timeout=1
                ) as response:
                    payload = json.load(response)
                if payload.get("webSocketDebuggerUrl"):
                    return
            except Exception as exc:
                last_error = exc
            time.sleep(0.1)
        raise RuntimeError("browser-use Chrome CDP 启动超时") from last_error

    def start(self, timeout: float = 60.0) -> None:
        try:
            try:
                from browser_use.browser.chrome import find_chrome_executable
                from browser_use.browser.profile import BrowserProfile
            except ImportError as exc:
                raise RuntimeError(
                    "未安装 browser-use，请双击一键启动 CMD，"
                    "或执行：pip install -r requirements.txt"
                ) from exc

            chrome_path = find_chrome_executable()
            if not chrome_path:
                raise RuntimeError("browser-use 未找到 Google Chrome 可执行文件")
            profile = BrowserProfile(
                executable_path=chrome_path,
                user_data_dir=str(self._profile_dir),
                headless=False,
                keep_alive=False,
                window_size=dict(SMALL_WINDOW_SIZE),
                viewport=dict(SMALL_VIEWPORT_SIZE),
                args=["--disable-blink-features=AutomationControlled"],
            )
            debug_port = self._find_free_port()
            launch_args = [*profile.get_args(), f"--remote-debugging-port={debug_port}"]
            self._cdp_url = f"http://127.0.0.1:{debug_port}"
            creationflags = (
                getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
                if os.name == "nt"
                else 0
            )
            startupinfo = None
            if os.name == "nt" and self.headless:
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow = 0
            self._process = subprocess.Popen(
                [str(chrome_path), *launch_args],
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=creationflags,
                startupinfo=startupinfo,
            )
            self._browser_pid = self._process.pid
            self._wait_until_ready(timeout)
        except Exception:
            try:
                self.close()
            except Exception:
                pass
            raise

    def _terminate_process_tree(self) -> None:
        if not self._browser_pid:
            return
        try:
            import psutil

            parent = psutil.Process(self._browser_pid)
            processes = parent.children(recursive=True) + [parent]
            for process in reversed(processes):
                try:
                    process.terminate()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            _gone, alive = psutil.wait_procs(processes, timeout=4)
            for process in alive:
                try:
                    process.kill()
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
        except Exception:
            pass

    def _remove_profile_dir(self) -> None:
        last_error: OSError | None = None
        for _attempt in range(5):
            try:
                if self._profile_dir.exists():
                    shutil.rmtree(self._profile_dir)
                return
            except OSError as exc:
                last_error = exc
                time.sleep(0.2)
        if self._profile_dir.exists():
            raise RuntimeError(
                f"browser-use 临时缓存目录删除失败：{self._profile_dir}"
            ) from last_error

    def close(self) -> None:
        self._terminate_process_tree()
        if self._process is not None:
            try:
                self._process.wait(timeout=5)
            except Exception:
                try:
                    self._process.kill()
                    self._process.wait(timeout=5)
                except Exception:
                    pass
            self._process = None
        self._remove_profile_dir()


def launch_browser(playwright: Playwright, headless: bool = True) -> BrowserLaunch:
    """启动独立、隐藏 AutomationControlled 的真实 Chrome。

    实测旧版 ``playwright.chromium.launch(channel="chrome")`` 会令
    ``navigator.webdriver`` 为 true，StudentAid 接口随后长期停在 Loading。
    browser-use 使用的同名 Blink 开关可消除这个差异，同时仍由 Playwright
    管理独立 BrowserContext，便于彻底清除本批次浏览器数据。
    """
    try:
        browser = playwright.chromium.launch(
            channel="chrome",
            headless=False,
            args=[
                "--disable-blink-features=AutomationControlled",
                f"--window-size={SMALL_WINDOW_SIZE['width']},{SMALL_WINDOW_SIZE['height']}",
                *(["--window-position=-32000,-32000"] if headless else []),
            ],
        )
    except Exception as exc:
        raise RuntimeError(
            "无法启动 Google Chrome。请双击一键启动 CMD 自动检查并安装 Chrome。"
        ) from exc
    return BrowserLaunch(
        browser,
        False,
        (
            "独立 Google Chrome 无头模式（AutomationControlled 已关闭）"
            if headless
            else "独立 Google Chrome 小窗口（900x720，AutomationControlled 已关闭）"
        ),
        owns_browser=True,
        dedicated_profile=True,
    )


def _check_stop(stop_event: threading.Event) -> None:
    if stop_event.is_set():
        raise StopRequested("用户已请求停止")


def step_1_open_retrieve_account_details(
    page: Page, stop_event: threading.Event
) -> None:
    last_error: BaseException | None = None
    for attempt in range(1, 4):
        _check_stop(stop_event)
        try:
            page.goto(
                RETRIEVE_ACCOUNT_DETAILS_URL,
                wait_until="commit",
                timeout=60_000,
            )
            _check_stop(stop_event)
            page.locator(FIRST_NAME_SELECTOR).wait_for(
                state="visible", timeout=60_000
            )
            page.wait_for_function(
                """
                () => {
                    const text = document.body?.innerText || "";
                    return !text.includes("Loading...")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameLastName")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameDateOfBirthMonth")
                        && !!document.querySelector("#fsa_Input_ForgotUsernameSsnInput");
                }
                """,
                timeout=30_000,
            )
            return
        except StopRequested:
            raise
        except Exception as exc:
            last_error = exc
            if attempt >= 3:
                break
            page.wait_for_timeout(1_500 * attempt)
    raise RuntimeError(
        "StudentAid 页面连续 3 次无法打开；请先在 Chrome 中确认页面可访问，"
        "或切换 GUI 中的浏览器后端后重试。"
    ) from last_error


def _type_account_field(page: Page, selector: str, value: str) -> None:
    """用真实键盘事件输入并失焦，确保 StudentAid 前端状态完整更新。"""
    field = page.locator(selector)
    field.click(timeout=30_000)
    field.fill("")
    field.press_sequentially(value, delay=20)
    field.press("Tab")


def step_2_fill_account_details(
    page: Page, details: AccountDetails, stop_event: threading.Event
) -> None:
    _check_stop(stop_event)
    _type_account_field(page, FIRST_NAME_SELECTOR, details.first_name)
    _type_account_field(page, LAST_NAME_SELECTOR, details.last_name)
    page.locator(BIRTH_MONTH_SELECTOR).select_option(details.birth_month)
    page.locator(BIRTH_MONTH_SELECTOR).press("Tab")
    _type_account_field(page, BIRTH_DAY_SELECTOR, str(int(details.birth_day)))
    _type_account_field(page, BIRTH_YEAR_SELECTOR, details.birth_year)
    _type_account_field(page, SSN_SELECTOR, details.ssn)
    _check_stop(stop_event)
    try:
        page.wait_for_function(
            """
            () => {
                const selectors = [
                    "#fsa_Input_ForgotUsernameFirstName",
                    "#fsa_Input_ForgotUsernameLastName",
                    "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                    "#fsa_Input_ForgotUsernameDateOfBirthDay",
                    "#fsa_Input_ForgotUsernameDateOfBirthYear",
                    "#fsa_Input_ForgotUsernameSsnInput",
                ];
                const button = [...document.querySelectorAll("button")].find(
                    element => (element.innerText || "").trim() === "Continue"
                );
                return selectors.every(selector => {
                    const field = document.querySelector(selector);
                    return field && String(field.value || "").length > 0;
                }) && button && !button.disabled;
            }
            """,
            timeout=30_000,
        )
    except Exception as exc:
        raise PageSubmissionStalled(
            "资料已填写，但 Continue 在 30 秒内未进入可点击状态"
        ) from exc
    page.wait_for_timeout(200)


def _wait_for_submission_started(page: Page, timeout_ms: int) -> bool:
    try:
        page.wait_for_function(
            """
            () => {
                const text = document.body?.innerText || "";
                const button = [...document.querySelectorAll("button")].find(
                    element => (element.innerText || "").trim() === "Continue"
                );
                return text.includes("Loading...")
                    || text.includes("Account Not Found")
                    || text.includes("Recover my account with a photo ID")
                    || text.includes("Limit Reached: Try Again in 24 Hours")
                    || text.includes("Your Account Is Disabled")
                    || text.includes("Account Lookup Issue: Get Help")
                    || text.includes("We are unable to retrieve your log-in information. Access your account by recovering your account with a photo ID.")
                    || text.includes("An unknown error has occurred")
                    || location.pathname.endsWith("/username")
                    || (button && button.disabled);
            }
            """,
            timeout=timeout_ms,
        )
        return True
    except Exception:
        return False


def step_3_click_continue(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
) -> None:
    """点击真实 Continue，并在必要时按可验证顺序使用两级兜底。"""
    _check_stop(stop_event)
    continue_element = page.get_by_role("button", name="Continue", exact=True).first
    continue_element.wait_for(state="visible", timeout=30_000)
    if not continue_element.is_enabled():
        raise RuntimeError("Continue 按钮未启用，请检查字段输入状态")
    continue_element.scroll_into_view_if_needed(timeout=10_000)
    try:
        continue_element.click(timeout=10_000)
    except Exception:
        _report_stage(progress_callback, "Continue 标准点击未完成，正在使用 Enter 兜底")
    if _wait_for_submission_started(page, 5_000):
        return

    _check_stop(stop_event)
    _report_stage(progress_callback, "Continue 尚未触发，正在聚焦按钮并按 Enter")
    try:
        continue_element.focus()
        continue_element.press("Enter", timeout=5_000)
    except Exception:
        pass
    if _wait_for_submission_started(page, 5_000):
        return

    _check_stop(stop_event)
    _report_stage(progress_callback, "Continue 仍未触发，正在执行按钮 DOM click")
    try:
        continue_element.evaluate("button => button.click()")
    except Exception:
        pass
    if _wait_for_submission_started(page, 5_000):
        return

    raise PageSubmissionStalled("三种 Continue 点击方式均未进入提交或结果状态")


def _report_stage(
    progress_callback: Callable[[str], None] | None, message: str
) -> None:
    """发送不包含输入资料的阶段消息；UI 回调失败不能中断网页处理。"""
    if progress_callback is None:
        return
    try:
        progress_callback(message)
    except Exception:
        pass


def step_4_judge_password_recovery(
    page: Page,
    stop_event: threading.Event,
    progress_callback: Callable[[str], None] | None = None,
    *,
    timeout_ms: int = 60_000,
    poll_interval_ms: int = 125,
    heartbeat_seconds: float = 5.0,
    stalled_loading_seconds: float = 60.0,
) -> str:
    """等待明确页面结果；持续 Loading 会触发重建会话，而不是无限转圈。"""
    started = time.monotonic()
    deadline = started + max(1, timeout_ms) / 1000
    next_heartbeat = max(0.05, heartbeat_seconds)
    loading_started: float | None = None

    while True:
        _check_stop(stop_event)
        body_text = page.evaluate("document.body?.innerText || ''")
        normalised_body_text = re.sub(r"\s+", " ", body_text).strip()
        loading = "Loading..." in body_text
        now = time.monotonic()
        if loading:
            loading_started = loading_started or now
            if now - loading_started >= stalled_loading_seconds:
                raise PageSubmissionStalled(
                    f"点击 Continue 后 Loading 已持续 {int(now - loading_started)} 秒"
                )
        else:
            loading_started = None
            if "An unknown error has occurred" in body_text:
                raise PageSessionExpired(
                    "StudentAid 页面会话已失效，需要重新打开页面"
                )
            if "Something went wrong" in body_text:
                raise RuntimeError("StudentAid 返回页面错误，请稍后重试")
            if LIMIT_REACHED_HEADING in body_text:
                _report_stage(progress_callback, f"已识别结果：{LIMIT_REACHED_HEADING}")
                return "limit_reached"
            if ACCOUNT_DISABLED_HEADING in body_text:
                _report_stage(progress_callback, f"已识别结果：{ACCOUNT_DISABLED_HEADING}")
                return "account_disabled"
            if ACCOUNT_LOOKUP_ISSUE_HEADING in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{ACCOUNT_LOOKUP_ISSUE_HEADING}",
                )
                return "account_lookup_issue"
            if PHOTO_ID_RECOVERY_MESSAGE in normalised_body_text:
                _report_stage(
                    progress_callback,
                    f"已识别结果：{PHOTO_ID_RECOVERY_MESSAGE}",
                )
                return "photo_id_recovery_required"
            if re.search(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                body_text,
                re.IGNORECASE | re.MULTILINE,
            ):
                _report_stage(progress_callback, "已识别结果：Account Not Found")
                return "account_not_found"
            if (
                "Retrieve Your Log-in Information" in body_text
                and "Recover my account with a photo ID" in body_text
            ):
                _report_stage(progress_callback, "已识别结果：Retrieve Your Log-in Information")
                return "can_recover"

        elapsed = now - started
        if now >= deadline:
            raise PageSubmissionStalled(
                f"等待 StudentAid 明确结果超时（{timeout_ms / 1000:g} 秒）"
            )
        if elapsed >= next_heartbeat:
            _report_stage(
                progress_callback,
                f"官方页面仍在处理中，已等待 {int(elapsed)} 秒",
            )
            while next_heartbeat <= elapsed:
                next_heartbeat += max(0.05, heartbeat_seconds)
        page.wait_for_timeout(
            min(poll_interval_ms, max(1, int((deadline - now) * 1000)))
        )


def _visible_text(page: Page, text: str) -> str:
    locator = page.get_by_text(text, exact=True).first
    return locator.inner_text().strip() if locator.is_visible() else ""


_MASK_CHARACTERS = "*•⦁●"
_MASK_CHARACTER_CLASS = re.escape(_MASK_CHARACTERS)
_MASKED_PHONE_RE = re.compile(
    rf"^\(\s*[{_MASK_CHARACTER_CLASS}]{{3}}\s*\)\s*"
    rf"[{_MASK_CHARACTER_CLASS}]{{3}}\s*\d{{4}}$"
)


def _normalise_contact_text(value: str) -> str:
    """保留掩码内容，只统一 DOM 中的不可见/重复空白。"""
    return re.sub(r"\s+", " ", value.replace("\xa0", " ")).strip()


def _looks_like_masked_phone(value: str) -> bool:
    return bool(_MASKED_PHONE_RE.fullmatch(_normalise_contact_text(value)))


def _looks_like_masked_email(value: str) -> bool:
    value = _normalise_contact_text(value)
    if "@" not in value or not any(char in value for char in _MASK_CHARACTERS):
        return False
    local_part, separator, domain = value.partition("@")
    if not separator or not local_part or not domain or " " in value:
        return False
    # StudentAid 只遮挡本地部分；限定为一行邮箱格式，避免页脚说明文字被误记。
    return bool(
        re.fullmatch(
            rf"[A-Za-z0-9._%+\-{_MASK_CHARACTER_CLASS}]+@"
            r"[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            value,
        )
    )


def _visible_contact_texts(page: Page) -> list[str]:
    """只读取结果卡片的可见段落；精确类名失效时再退到可见 p。"""
    selectors = ("p.fsa-color-gray-60", "p.m-0.fsa-font-size-16", "p")
    for selector in selectors:
        locator = page.locator(selector)
        values: list[str] = []
        try:
            for index in range(locator.count()):
                item = locator.nth(index)
                if not item.is_visible():
                    continue
                value = _normalise_contact_text(item.inner_text())
                if value:
                    values.append(value)
        except Exception:
            values = []
        if any(
            _looks_like_masked_phone(value) or _looks_like_masked_email(value)
            for value in values
        ):
            return values
    return []


def collect_recovery_result(page: Page, recovery_status: str) -> RecoveryResult:
    if recovery_status == "limit_reached":
        heading = _visible_text(page, LIMIT_REACHED_HEADING) or LIMIT_REACHED_HEADING
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_disabled":
        heading = _visible_text(page, ACCOUNT_DISABLED_HEADING) or ACCOUNT_DISABLED_HEADING
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_lookup_issue":
        heading = (
            _visible_text(page, ACCOUNT_LOOKUP_ISSUE_HEADING)
            or ACCOUNT_LOOKUP_ISSUE_HEADING
        )
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "photo_id_recovery_required":
        heading = (
            _visible_text(page, PHOTO_ID_RECOVERY_MESSAGE)
            or PHOTO_ID_RECOVERY_MESSAGE
        )
        return RecoveryResult(recovery_status, heading, "", "", "")

    if recovery_status == "account_not_found":
        locator = page.get_by_text(
            re.compile(
                r"^\s*Account Not Found(?:\s*:\s*Create a New Account)?\s*$",
                re.IGNORECASE,
            )
        ).first
        heading = locator.inner_text().strip() if locator.is_visible() else "Account Not Found"
        return RecoveryResult(recovery_status, heading, "", "", "")

    heading = _visible_text(page, "Retrieve Your Log-in Information")
    masked_phone = ""
    masked_email = ""
    for value in _visible_contact_texts(page):
        if _looks_like_masked_phone(value) and not masked_phone:
            masked_phone = value
        elif _looks_like_masked_email(value) and not masked_email:
            masked_email = value
    recovery_method = _visible_text(page, "Recover my account with a photo ID")
    return RecoveryResult(
        recovery_status,
        heading,
        masked_phone,
        masked_email,
        recovery_method,
    )


class BrowserRecoverySession:
    """第十五步 worker 会话；每个线程拥有独立浏览器和缓存目录。"""

    def __init__(
        self,
        worker_number: int,
        backend: str = "playwright",
        display_mode: str = "无头",
    ) -> None:
        backend = backend.strip().casefold()
        if backend not in BROWSER_BACKENDS:
            raise ValueError(f"浏览器后端必须是：{', '.join(BROWSER_BACKENDS)}")
        self.worker_number = worker_number
        self.backend = backend
        self.display_mode = _normalise_display_mode(display_mode)
        self.headless = _is_headless(self.display_mode)
        self._playwright: Playwright | None = None
        self._browser: Browser | None = None
        self._browser_use_host: BrowserUseBrowserHost | None = None
        self._context: Any = None
        self._page: Page | None = None
        self._owns_context = False
        self._external_browser = False
        self._owns_browser = False
        self._dedicated_profile = False
        self.browser_mode = "未启动"

    def start(self) -> None:
        if sync_playwright is None:
            raise RuntimeError(
                "未安装 playwright，请双击一键启动 CMD，或执行：pip install -r requirements.txt"
            )
        self._playwright = sync_playwright().start()
        try:
            if self.backend == "browser-use":
                self._browser_use_host = BrowserUseBrowserHost(
                    self.worker_number, headless=self.headless
                )
                self._browser_use_host.start()
                try:
                    browser = self._playwright.chromium.connect_over_cdp(
                        self._browser_use_host.cdp_url, timeout=30_000
                    )
                except Exception as exc:
                    raise RuntimeError("Playwright 无法接管 browser-use 独立 CDP 会话") from exc
                launch = BrowserLaunch(
                    browser=browser,
                    external_browser=True,
                    mode=(
                        "browser-use 独立 Chrome "
                        f"{self.display_mode}模式 (worker {self.worker_number}, "
                        "AutomationControlled 已关闭)"
                    ),
                    owns_browser=False,
                    dedicated_profile=True,
                )
            else:
                launch = launch_browser(self._playwright, headless=self.headless)
            self._browser = launch.browser
            self._external_browser = launch.external_browser
            self._owns_browser = launch.owns_browser
            self._dedicated_profile = launch.dedicated_profile
            self.browser_mode = launch.mode
            browser_contexts = list(getattr(self._browser, "contexts", []))
            if self._external_browser:
                if not browser_contexts:
                    raise RuntimeError("Chrome CDP 没有可用的默认浏览器上下文")
                self._context = browser_contexts[0]
            else:
                self._context = self._browser.new_context(
                    viewport=dict(SMALL_VIEWPORT_SIZE), locale="en-US"
                )
                self._owns_context = True
            pages = list(getattr(self._context, "pages", []))
            self._page = pages[-1] if pages else self._context.new_page()
            self._configure_page(self._page)
            if self._page.evaluate("navigator.webdriver === true"):
                raise RuntimeError(
                    "检测到浏览器处于 Playwright 自动化启动模式；"
                    "该模式会让 StudentAid 在 Continue 后持续 Loading"
                )
            if self._dedicated_profile:
                # 上次进程若被强制关闭，先消除可能残留的 cookie、cache 和 storage。
                self._clear_browser_data_and_blank()
            self._hide_browser_window()
        except Exception:
            self.close()
            raise

    def _close_page(self) -> None:
        if self._page is not None:
            try:
                if not self._page.is_closed():
                    self._page.close()
            except Exception:
                pass
            self._page = None

    def _configure_page(self, page: Page) -> None:
        try:
            page.set_viewport_size(dict(SMALL_VIEWPORT_SIZE))
        except Exception:
            pass
        if not self.headless or self._browser is None or self._context is None:
            return
        user_agent = _normal_chrome_user_agent(self._browser.version)
        try:
            session = self._context.new_cdp_session(page)
            try:
                session.send(
                    "Emulation.setUserAgentOverride",
                    {
                        "userAgent": user_agent,
                        "acceptLanguage": "en-US,en;q=0.9",
                        "platform": "Win32",
                    },
                )
            finally:
                session.detach()
        except Exception as exc:
            raise RuntimeError("无法为无头 Chrome 设置正常浏览器标识") from exc

    def _hide_browser_window(self) -> None:
        if not self.headless or os.name != "nt":
            return
        owned_processes: set[int] = set()
        if (
            self._browser_use_host is not None
            and self._browser_use_host._browser_pid is not None
        ):
            try:
                import psutil

                root_process = psutil.Process(self._browser_use_host._browser_pid)
                owned_processes = {
                    root_process.pid,
                    *(child.pid for child in root_process.children(recursive=True)),
                }
            except Exception:
                owned_processes = set()
        else:
            owned_processes = _chrome_process_ids("--window-position=-32000,-32000")
        for _attempt in range(10):
            if _hide_windows_for_processes(owned_processes):
                return
            time.sleep(0.05)
            if self._browser_use_host is None:
                owned_processes = _chrome_process_ids(
                    "--window-position=-32000,-32000"
                )

    def _new_blank_page(self) -> Page:
        if self._context is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        self._close_page()
        self._page = self._context.new_page()
        self._configure_page(self._page)
        self._hide_browser_window()
        return self._page

    def _ensure_form_page(
        self,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None,
    ) -> Page:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        form_ready = False
        try:
            form_ready = bool(
                page.locator(FIRST_NAME_SELECTOR).is_visible(timeout=1_000)
                and "Loading..." not in page.locator("body").inner_text(timeout=1_000)
            )
        except Exception:
            form_ready = False
        if not form_ready:
            self._hide_browser_window()
            _report_stage(progress_callback, "正在打开 StudentAid 页面")
            step_1_open_retrieve_account_details(page, stop_event)
        return page

    def _clear_browser_data_and_blank(
        self,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        page = self._page
        if page is None or page.is_closed():
            page = self._new_blank_page()
        _report_stage(progress_callback, "正在清除 StudentAid 专用浏览器数据")
        origins = {"https://studentaid.gov"}
        try:
            discovered = page.evaluate(
                """
                () => [...new Set([
                    location.origin,
                    ...performance.getEntriesByType("resource").map(entry => {
                        try { return new URL(entry.name).origin; }
                        catch (_) { return ""; }
                    })
                ])]
                """
            )
            origins.update(
                origin for origin in discovered
                if isinstance(origin, str)
                and re.fullmatch(r"https://(?:[a-z0-9-]+\.)*studentaid\.gov", origin)
            )
        except Exception:
            pass
        try:
            session = self._context.new_cdp_session(page)
            try:
                session.send("Network.clearBrowserCache")
                if self._dedicated_profile:
                    session.send("Network.clearBrowserCookies")
                try:
                    session.send("ServiceWorker.stopAllWorkers")
                except Exception:
                    pass
                for origin in origins:
                    session.send(
                        "Storage.clearDataForOrigin",
                        {"origin": origin, "storageTypes": "all"},
                    )
            finally:
                try:
                    session.detach()
                except Exception:
                    pass
        except Exception:
            pass
        if self._dedicated_profile:
            try:
                self._context.clear_cookies()
            except Exception:
                pass
        self._new_blank_page()
        _report_stage(progress_callback, "浏览器数据已清除，已回到空白页")

    def process(
        self,
        item: WorkItem,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> RecoveryResult:
        if self._browser is None:
            raise RuntimeError("浏览器处理会话尚未启动")
        last_error: BaseException | None = None
        for attempt in range(1, 4):
            try:
                page = self._ensure_form_page(stop_event, progress_callback)
                if attempt > 1:
                    _report_stage(
                        progress_callback,
                        f"已重建页面会话，正在重新填写资料（自动重试 {attempt - 1}/2）",
                    )
                else:
                    _report_stage(progress_callback, "页面已打开，正在填写资料")
                step_2_fill_account_details(page, item.details, stop_event)
                _report_stage(progress_callback, "资料已填写，正在点击 Continue")
                step_3_click_continue(page, stop_event, progress_callback)
                _report_stage(progress_callback, "已提交，正在等待官方结果")
                status = step_4_judge_password_recovery(
                    page, stop_event, progress_callback
                )
                return collect_recovery_result(page, status)
            except StopRequested:
                raise
            except (PageSessionExpired, PageSubmissionStalled) as exc:
                last_error = exc
                if attempt >= 3:
                    break
                _report_stage(
                    progress_callback,
                    "页面会话失效或提交持续转圈，正在自动清理并重建页面",
                )
                self._clear_browser_data_and_blank(progress_callback)
        raise RuntimeError(
            "StudentAid 页面清理并重填后仍未返回明确结果"
        ) from last_error

    def prepare_for_next(
        self,
        result_code: str,
        stop_event: threading.Event,
        progress_callback: Callable[[str], None] | None = None,
    ) -> None:
        """必须在结果写入 SQLite/累计 CSV 后调用。"""
        _check_stop(stop_event)
        page = self._page
        if result_code in {
            "account_not_found",
            "limit_reached",
            "account_disabled",
            "account_lookup_issue",
            "photo_id_recovery_required",
        }:
            self._clear_browser_data_and_blank(progress_callback)
            return
        if result_code != "can_recover":
            self._close_page()
            return
        if page is None or page.is_closed():
            raise RuntimeError("记录账户找回结果后页面已关闭，无法点击 Cancel")
        _report_stage(progress_callback, "结果已保存，正在点击 Cancel")
        cancel_matches = page.locator(
            "span", has_text=re.compile(r"^\s*Cancel\s*$", re.IGNORECASE)
        )
        cancel_span = None
        for index in range(cancel_matches.count()):
            candidate = cancel_matches.nth(index)
            if (
                candidate.is_visible()
                and candidate.inner_text().strip().casefold() == "cancel"
            ):
                cancel_span = candidate
                break
        if cancel_span is None:
            raise RuntimeError("找不到可见的 <span>Cancel</span>")

        original_url = page.url
        parent_control = cancel_span.locator(
            "xpath=ancestor::*[self::button or self::a or @role='button'][1]"
        )
        click_targets = [("Cancel 文本", cancel_span)]
        if parent_control.count() and parent_control.is_visible():
            click_targets.append(("Cancel 父控件", parent_control))

        cancel_triggered = False
        last_click_error: Exception | None = None
        for target_name, target in click_targets:
            try:
                target.scroll_into_view_if_needed(timeout=3_000)
                target.click(timeout=5_000)
                page.wait_for_function(
                    """
                    originalUrl => {
                        const visible = element => {
                            const style = getComputedStyle(element);
                            const box = element.getBoundingClientRect();
                            return style.visibility !== "hidden"
                                && style.display !== "none"
                                && box.width > 0 && box.height > 0;
                        };
                        const resultHeadingVisible = [...document.querySelectorAll("*")]
                            .some(element => element.textContent?.trim()
                                === "Retrieve Your Log-in Information" && visible(element));
                        const cancelVisible = [...document.querySelectorAll("span")]
                            .some(element => element.textContent?.trim().toLowerCase()
                                === "cancel" && visible(element));
                        const formVisible = (() => {
                            const element = document.querySelector(
                                "#fsa_Input_ForgotUsernameFirstName"
                            );
                            return Boolean(element && visible(element));
                        })();
                        return location.href !== originalUrl || formVisible
                            || !resultHeadingVisible || !cancelVisible;
                    }
                    """,
                    arg=original_url,
                    timeout=6_000,
                )
                cancel_triggered = True
                _report_stage(
                    progress_callback,
                    f"{target_name}已实际触发，结果页已离开",
                )
                break
            except Exception as exc:
                last_click_error = exc

        if not cancel_triggered:
            # Angular 模板偶尔拦截合成鼠标事件；原生 click 会在可见 span/父控件上
            # 冒泡到同一个 Angular 处理器。仍必须验证结果页确实发生变化。
            try:
                cancel_span.evaluate(
                    """
                    element => (element.closest("button,a,[role='button']") || element).click()
                    """
                )
                page.wait_for_function(
                    """
                    () => {
                        const visible = element => {
                            const style = getComputedStyle(element);
                            const box = element.getBoundingClientRect();
                            return style.visibility !== "hidden"
                                && style.display !== "none"
                                && box.width > 0 && box.height > 0;
                        };
                        return ![...document.querySelectorAll("span")].some(
                            element => element.textContent?.trim().toLowerCase()
                                === "cancel" && visible(element)
                        );
                    }
                    """,
                    timeout=6_000,
                )
                cancel_triggered = True
                _report_stage(progress_callback, "Cancel 原生事件已实际触发")
            except Exception as exc:
                last_click_error = exc

        if not cancel_triggered:
            raise RuntimeError("可见 Cancel 点击后结果页没有变化") from last_click_error

        try:
            page.locator(FIRST_NAME_SELECTOR).wait_for(state="visible", timeout=5_000)
        except Exception:
            _report_stage(
                progress_callback,
                "Cancel 已完成；正在重新打开账户找回空白表单",
            )
            step_1_open_retrieve_account_details(page, stop_event)
        page.wait_for_function(
            """
            () => {
                const ids = [
                    "#fsa_Input_ForgotUsernameFirstName",
                    "#fsa_Input_ForgotUsernameLastName",
                    "#fsa_Input_ForgotUsernameDateOfBirthMonth",
                    "#fsa_Input_ForgotUsernameDateOfBirthDay",
                    "#fsa_Input_ForgotUsernameDateOfBirthYear",
                    "#fsa_Input_ForgotUsernameSsnInput",
                ];
                return ids.every(id => !(document.querySelector(id)?.value || ""));
            }
            """,
            timeout=15_000,
        )
        _report_stage(progress_callback, "Cancel 已确认完成，空白表单已就绪")

    def recover_after_cleanup_error(self) -> None:
        self._close_page()

    def close(self) -> None:
        if self.backend == "browser-use":
            # browser-use 使用一次性 user-data-dir。先结束精确 PID 树并删除整个
            # profile，比在失效 CDP 页面上逐项清理更彻底，也避免无头失败页令
            # Playwright disconnect/stop 长时间等待。
            self._page = None
            self._context = None
            self._owns_context = False
            self._browser = None
            self._external_browser = False
            self._owns_browser = False
            if self._browser_use_host is not None:
                try:
                    self._browser_use_host.close()
                finally:
                    self._browser_use_host = None
            if self._playwright is not None:
                try:
                    self._playwright.stop()
                except Exception:
                    pass
                self._playwright = None
            self._dedicated_profile = False
            return

        if self._context is not None and self._browser is not None:
            try:
                self._clear_browser_data_and_blank()
            except Exception:
                pass
        self._close_page()
        if self._context is not None and self._owns_context:
            try:
                self._context.close()
            except Exception:
                pass
        self._context = None
        self._owns_context = False
        if self._browser is not None and self._owns_browser:
            try:
                self._browser.close()
            except Exception:
                pass
        self._browser = None
        self._external_browser = False
        self._owns_browser = False
        if self._playwright is not None:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None
        self._dedicated_profile = False


@dataclass
class _DbCommand:
    action: str
    payload: dict[str, Any]
    response: queue.Queue[tuple[bool, Any]]


class DatabaseWriter:
    """SQLite 的唯一写入者；工作线程不得直接执行写 SQL。"""

    def __init__(self, database_path: Path) -> None:
        self.database_path = database_path.resolve()
        self._commands: queue.Queue[_DbCommand] = queue.Queue()
        self._ready = threading.Event()
        self._thread = threading.Thread(
            target=self._run,
            name="studentaid-db-writer",
            daemon=False,
        )
        self._startup_error: BaseException | None = None

    def start(self) -> None:
        self.database_path.parent.mkdir(parents=True, exist_ok=True)
        self._thread.start()
        if not self._ready.wait(30):
            raise TimeoutError("SQLite 写入线程启动超时")
        if self._startup_error is not None:
            raise RuntimeError("SQLite 初始化失败") from self._startup_error

    def request(self, action: str, **payload: Any) -> Any:
        if not self._thread.is_alive():
            raise RuntimeError("SQLite 写入线程未运行")
        response: queue.Queue[tuple[bool, Any]] = queue.Queue(maxsize=1)
        self._commands.put(_DbCommand(action, payload, response))
        try:
            ok, value = response.get(timeout=300)
        except queue.Empty as exc:
            raise TimeoutError(f"SQLite 写入命令超时：{action}") from exc
        if ok:
            return value
        raise RuntimeError(f"SQLite 写入失败：{action}") from value

    def close(self) -> None:
        if self._thread.is_alive():
            try:
                self.request("shutdown")
            finally:
                self._thread.join(timeout=30)

    def _run(self) -> None:
        connection: sqlite3.Connection | None = None
        try:
            connection = sqlite3.connect(self.database_path, timeout=30)
            connection.row_factory = sqlite3.Row
            connection.execute("PRAGMA busy_timeout=30000")
            journal_mode = connection.execute("PRAGMA journal_mode=WAL").fetchone()[0]
            if str(journal_mode).casefold() != "wal":
                raise RuntimeError(f"无法启用 SQLite WAL，当前模式：{journal_mode}")
            connection.execute("PRAGMA synchronous=NORMAL")
            connection.execute("PRAGMA wal_autocheckpoint=1000")
            self._create_schema(connection)
            connection.commit()
        except BaseException as exc:
            self._startup_error = exc
            self._ready.set()
            if connection is not None:
                connection.close()
            return

        self._ready.set()
        assert connection is not None
        try:
            while True:
                command = self._commands.get()
                try:
                    value = self._handle(connection, command.action, command.payload)
                    connection.commit()
                    command.response.put((True, value))
                    if command.action == "shutdown":
                        break
                except BaseException as exc:
                    connection.rollback()
                    command.response.put((False, exc))
                finally:
                    self._commands.task_done()
        finally:
            connection.close()

    @staticmethod
    def _create_schema(connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS batches (
                batch_id TEXT PRIMARY KEY,
                input_path TEXT NOT NULL,
                output_directory TEXT NOT NULL,
                thread_count INTEGER NOT NULL,
                status TEXT NOT NULL,
                created_at TEXT NOT NULL,
                finished_at TEXT,
                export_path TEXT,
                total_count INTEGER NOT NULL DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id TEXT NOT NULL REFERENCES batches(batch_id),
                source_file TEXT NOT NULL,
                source_sheet TEXT NOT NULL DEFAULT '',
                source_row INTEGER NOT NULL,
                original_fields_json TEXT NOT NULL,
                ssn TEXT,
                birth_month TEXT,
                birth_day TEXT,
                birth_year TEXT,
                first_name TEXT,
                last_name TEXT,
                address TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL,
                attempt_count INTEGER NOT NULL DEFAULT 0,
                result_code TEXT NOT NULL DEFAULT '',
                result_heading TEXT NOT NULL DEFAULT '',
                masked_phone TEXT NOT NULL DEFAULT '',
                masked_email TEXT NOT NULL DEFAULT '',
                recovery_method TEXT NOT NULL DEFAULT '',
                error TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL,
                started_at TEXT,
                finished_at TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_records_batch_status
                ON records(batch_id, status, id);

            CREATE INDEX IF NOT EXISTS idx_records_batch_ssn_status
                ON records(batch_id, ssn, status);
            """
        )

    @staticmethod
    def _handle(
        connection: sqlite3.Connection, action: str, payload: Mapping[str, Any]
    ) -> Any:
        if action == "create_batch":
            connection.execute(
                """
                INSERT INTO batches(
                    batch_id, input_path, output_directory, thread_count,
                    status, created_at
                ) VALUES (?, ?, ?, ?, 'importing', ?)
                """,
                (
                    payload["batch_id"], payload["input_path"],
                    payload["output_directory"], payload["thread_count"], _now_iso(),
                ),
            )
            return None

        if action == "insert_records":
            batch_id = str(payload["batch_id"])
            work_items: list[WorkItem] = []
            queued_keys: set[str] = set()
            records: Sequence[ImportedRecord] = payload["records"]
            for record in records:
                details = record.details
                status = "failed" if record.import_error else "pending"
                cursor = connection.execute(
                    """
                    INSERT INTO records(
                        batch_id, source_file, source_sheet, source_row,
                        original_fields_json, ssn, birth_month, birth_day,
                        birth_year, first_name, last_name, address, status,
                        error, created_at, finished_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id, record.source_file, record.source_sheet,
                        record.source_row,
                        json.dumps(record.original_fields, ensure_ascii=False),
                        details.ssn if details else None,
                        details.birth_month if details else None,
                        details.birth_day if details else None,
                        details.birth_year if details else None,
                        details.first_name if details else None,
                        details.last_name if details else None,
                        record.address, status, record.import_error, _now_iso(),
                        _now_iso() if record.import_error else None,
                    ),
                )
                if (
                    details is not None
                    and not record.import_error
                    and details.ssn not in queued_keys
                ):
                    queued_keys.add(details.ssn)
                    work_items.append(
                        WorkItem(
                            int(cursor.lastrowid), details, record.source_file,
                            record.source_sheet, record.source_row,
                            record.original_fields, record.address,
                        )
                    )
            connection.execute(
                """
                UPDATE batches SET status='running', total_count=? WHERE batch_id=?
                """,
                (len(records), batch_id),
            )
            return work_items

        if action == "mark_processing":
            connection.execute(
                """
                UPDATE records
                SET status='processing', started_at=?, attempt_count=attempt_count+1,
                    error=''
                WHERE id=? AND status='pending'
                """,
                (_now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_completed":
            result: RecoveryResult = payload["result"]
            connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE id=?
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["record_id"],
                ),
            )
            return None

        if action == "mark_completed_group":
            result = payload["result"]
            cursor = connection.execute(
                """
                UPDATE records
                SET status='completed', result_code=?, result_heading=?,
                    masked_phone=?, masked_email=?, recovery_method=?, error='',
                    finished_at=?
                WHERE batch_id=? AND ssn=?
                  AND status IN ('pending', 'processing')
                """,
                (
                    result.result_code, result.heading, result.masked_phone,
                    result.masked_email, result.recovery_method, _now_iso(),
                    payload["batch_id"], payload["ssn"],
                ),
            )
            return int(cursor.rowcount)

        if action == "mark_retry_group":
            cursor = connection.execute(
                """
                UPDATE records
                SET status='pending', error=?, started_at=NULL, finished_at=NULL
                WHERE batch_id=? AND ssn=?
                  AND status IN ('pending', 'processing')
                """,
                (
                    payload.get("error", ""), payload["batch_id"], payload["ssn"],
                ),
            )
            return int(cursor.rowcount)

        if action in {"mark_failed_group", "mark_stopped_group"}:
            status = "failed" if action == "mark_failed_group" else "stopped"
            allowed_statuses = (
                "('pending', 'processing', 'completed')"
                if action == "mark_failed_group"
                else "('pending', 'processing')"
            )
            cursor = connection.execute(
                f"""
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND ssn=? AND status IN {allowed_statuses}
                """,
                (
                    status, payload.get("error", ""), _now_iso(),
                    payload["batch_id"], payload["ssn"],
                ),
            )
            return int(cursor.rowcount)

        if action in {"mark_failed", "mark_stopped"}:
            status = "failed" if action == "mark_failed" else "stopped"
            connection.execute(
                "UPDATE records SET status=?, error=?, finished_at=? WHERE id=?",
                (status, payload.get("error", ""), _now_iso(), payload["record_id"]),
            )
            return None

        if action == "mark_remaining":
            status = str(payload["status"])
            connection.execute(
                """
                UPDATE records SET status=?, error=?, finished_at=?
                WHERE batch_id=? AND status IN ('pending', 'processing')
                """,
                (
                    status, payload.get("error", ""), _now_iso(), payload["batch_id"],
                ),
            )
            return None

        if action == "status_counts":
            rows = connection.execute(
                """
                SELECT status, COUNT(*) AS count FROM records
                WHERE batch_id=? GROUP BY status
                """,
                (payload["batch_id"],),
            ).fetchall()
            return {str(row["status"]): int(row["count"]) for row in rows}

        if action == "finish_batch":
            connection.execute(
                """
                UPDATE batches SET status=?, finished_at=?, export_path=?
                WHERE batch_id=?
                """,
                (
                    payload["status"], _now_iso(), payload.get("export_path", ""),
                    payload["batch_id"],
                ),
            )
            return None

        if action == "checkpoint":
            return connection.execute("PRAGMA wal_checkpoint(PASSIVE)").fetchone()

        if action == "shutdown":
            return None

        raise ValueError(f"未知 SQLite 写入命令：{action}")


def export_batch_csv(
    database_path: Path, batch_id: str, output_directory: Path
) -> Path:
    """只读查询当前批次并原子生成带 BOM 的最终 CSV。"""
    output_directory.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_directory / f"studentaid_results_{timestamp}_{batch_id[:8]}.csv"
    temporary_path = output_path.with_name(f".{output_path.name}.{os.getpid()}.tmp")

    connection = sqlite3.connect(database_path, timeout=30)
    connection.row_factory = sqlite3.Row
    try:
        rows = connection.execute(
            "SELECT * FROM records WHERE batch_id=? ORDER BY id",
            (batch_id,),
        ).fetchall()
    finally:
        connection.close()

    original_rows: list[list[str]] = []
    maximum_columns = 0
    for row in rows:
        fields = [str(value) for value in json.loads(row["original_fields_json"])]
        original_rows.append(fields)
        maximum_columns = max(maximum_columns, len(fields))

    try:
        with temporary_path.open("w", encoding="utf-8-sig", newline="") as output_file:
            writer = csv.writer(output_file, lineterminator="\n")
            writer.writerow(
                [*(f"Input_Column_{index}" for index in range(1, maximum_columns + 1)),
                 *RESULT_COLUMNS]
            )
            for row, original_fields in zip(rows, original_rows):
                padded = [*original_fields, *("" for _ in range(maximum_columns - len(original_fields)))]
                writer.writerow(
                    [
                        *padded,
                        row["result_heading"], row["masked_phone"],
                        row["masked_email"], row["recovery_method"], row["status"],
                        row["error"], Path(row["source_file"]).name,
                        row["source_sheet"], row["source_row"],
                    ]
                )
            output_file.flush()
            os.fsync(output_file.fileno())
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path


EventCallback = Callable[[str, dict[str, Any]], None]
SessionFactory = Callable[..., Any]


class BatchEngine:
    """第十二步并发处理：每个 worker 使用所选后端和显示模式。"""

    def __init__(
        self,
        event_callback: EventCallback | None = None,
        session_factory: SessionFactory = BrowserRecoverySession,
    ) -> None:
        self._callback = event_callback or (lambda _kind, _payload: None)
        self._session_factory = session_factory
        self._stop_event = threading.Event()
        self._state_lock = threading.Lock()
        self._running = False
        self._coordinator: threading.Thread | None = None

    @property
    def is_running(self) -> bool:
        with self._state_lock:
            return self._running

    def start(
        self,
        input_path: Path,
        output_target: Path,
        thread_count: int = 2,
        backend: str = "playwright",
        display_mode: str = "无头",
    ) -> None:
        input_path = input_path.resolve()
        output_directory, output_path = resolve_output_target(output_target)
        backend = backend.strip().casefold()
        display_mode = _normalise_display_mode(display_mode)
        if not input_path.is_file():
            raise FileNotFoundError(f"输入文件不存在：{input_path}")
        if input_path == output_path:
            raise ValueError("输入文件和累计输出文件不能是同一个文件")
        if thread_count < 1:
            raise ValueError("处理线程数必须是大于等于 1 的整数")
        if backend not in BROWSER_BACKENDS:
            raise ValueError(f"浏览器后端必须是：{', '.join(BROWSER_BACKENDS)}")
        output_directory.mkdir(parents=True, exist_ok=True)
        with self._state_lock:
            if self._running:
                raise RuntimeError("已有批次正在运行")
            self._running = True
        self._stop_event.clear()
        self._coordinator = threading.Thread(
            target=self._run,
            args=(input_path, output_path, thread_count, backend, display_mode),
            name="studentaid-coordinator",
            daemon=False,
        )
        self._coordinator.start()

    def stop(self) -> None:
        if self.is_running:
            self._stop_event.set()
            self._emit(
                "log",
                message="已收到停止请求；正在停止领取新任务、清除缓存并结束浏览器进程。",
            )

    def wait(self, timeout: float | None = None) -> bool:
        coordinator = self._coordinator
        if coordinator is not None:
            coordinator.join(timeout)
        return not self.is_running

    def _emit(self, kind: str, **payload: Any) -> None:
        try:
            self._callback(kind, payload)
        except Exception:
            pass

    def _publish_progress(
        self, writer: DatabaseWriter, batch_id: str, total: int
    ) -> dict[str, int]:
        counts = writer.request("status_counts", batch_id=batch_id)
        payload = {
            "total": total,
            "pending": int(counts.get("pending", 0)),
            "processing": int(counts.get("processing", 0)),
            "completed": int(counts.get("completed", 0)),
            "failed": int(counts.get("failed", 0)),
            "stopped": int(counts.get("stopped", 0)),
        }
        self._emit("progress", **payload)
        return payload

    def _new_session(
        self, worker_number: int, backend: str, display_mode: str
    ) -> Any:
        if self._session_factory is BrowserRecoverySession:
            return self._session_factory(worker_number, backend, display_mode)
        return self._session_factory(worker_number)

    def _worker_loop(
        self,
        worker_number: int,
        backend: str,
        display_mode: str,
        tasks: queue.Queue[tuple[WorkItem, int] | None],
        writer: DatabaseWriter,
        batch_id: str,
        total: int,
        input_path: Path,
        output_path: Path,
    ) -> None:
        session = self._new_session(worker_number, backend, display_mode)
        try:
            session.start()
            self._emit(
                "log",
                message=(
                    f"处理线程 {worker_number} 已启动："
                    f"{getattr(session, 'browser_mode', '模式未知')}。"
                ),
            )
        except Exception as exc:
            self._emit(
                "log",
                message=f"处理线程 {worker_number} 启动失败：{_clean_error(exc)}",
            )
            try:
                session.close()
            except Exception:
                pass
            return

        try:
            while not self._stop_event.is_set():
                try:
                    queued_task = tasks.get(timeout=0.2)
                except queue.Empty:
                    continue
                if queued_task is None:
                    tasks.task_done()
                    break
                item, queue_attempt = queued_task
                if self._stop_event.is_set():
                    tasks.task_done()
                    break

                writer.request("mark_processing", record_id=item.record_id)
                self._publish_progress(writer, batch_id, total)
                progress = lambda stage, record_id=item.record_id: self._emit(
                    "log", message=f"记录 #{record_id}：{stage}"
                )
                try:
                    try:
                        result = session.process(item, self._stop_event, progress)
                    except StopRequested:
                        raise
                    except Exception as process_exc:
                        if queue_attempt >= MAX_QUEUE_ATTEMPTS:
                            raise
                        recover = getattr(
                            session, "recover_after_cleanup_error", None
                        )
                        if callable(recover):
                            try:
                                recover()
                            except Exception:
                                pass
                        retry_group = writer.request(
                            "mark_retry_group", batch_id=batch_id,
                            ssn=item.details.ssn, error=_clean_error(process_exc),
                        )
                        tasks.put((item, queue_attempt + 1))
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：本轮未取得明确结果，"
                                f"同组 {retry_group} 条已回到队尾；稍后执行"
                                f"第 {queue_attempt + 1}/{MAX_QUEUE_ATTEMPTS} 轮。"
                            ),
                        )
                        continue
                    completed_group = writer.request(
                        "mark_completed_group", batch_id=batch_id,
                        ssn=item.details.ssn, result=result,
                    )
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：明确结果已实时写入 SQLite，"
                            f"同组完成 {completed_group} 条。"
                        ),
                    )
                    appended = append_cumulative_result(output_path, item, result)
                    if appended:
                        self._emit(
                            "log",
                            message=f"记录 #{item.record_id}：已实时追加到累计输出 CSV。",
                        )
                    else:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：累计输出第一列已存在，"
                                "跳过重复追加。"
                            ),
                        )
                    deleted = remove_input_rows_by_keys(input_path, {item.record_key})
                    self._emit(
                        "log",
                        message=(
                            f"记录 #{item.record_id}：明确结果已落盘，"
                            f"输入文件已删除 {deleted} 条匹配资料。"
                        ),
                    )
                    try:
                        prepare = getattr(session, "prepare_for_next", None)
                        if callable(prepare):
                            prepare(result.result_code, self._stop_event, progress)
                    except StopRequested:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "收到停止请求，跳过下一页准备。"
                            ),
                        )
                    except Exception as cleanup_exc:
                        self._emit(
                            "log",
                            message=(
                                f"记录 #{item.record_id}：结果已安全保存；"
                                "页面清理失败，将为下一条重建页面："
                                f"{_clean_error(cleanup_exc)}"
                            ),
                        )
                        recover = getattr(session, "recover_after_cleanup_error", None)
                        if callable(recover):
                            recover()
                except StopRequested as exc:
                    writer.request(
                        "mark_stopped_group", batch_id=batch_id,
                        ssn=item.details.ssn, error=_clean_error(exc),
                    )
                except Exception as exc:
                    writer.request(
                        "mark_failed_group", batch_id=batch_id,
                        ssn=item.details.ssn, error=_clean_error(exc),
                    )
                    self._emit(
                        "log",
                        message=f"记录 #{item.record_id} 处理失败：{_clean_error(exc)}",
                    )
                    recover = getattr(session, "recover_after_cleanup_error", None)
                    if callable(recover):
                        try:
                            recover()
                        except Exception:
                            pass
                finally:
                    tasks.task_done()
                    self._publish_progress(writer, batch_id, total)
        finally:
            try:
                session.close()
                self._emit(
                    "log",
                    message=(
                        f"处理线程 {worker_number} 已清除浏览器缓存并结束浏览器进程。"
                    ),
                )
            except Exception as exc:
                self._emit(
                    "log",
                    message=f"处理线程 {worker_number} 关闭浏览器失败：{_clean_error(exc)}",
                )

    def _run(
        self,
        input_path: Path,
        output_path: Path,
        thread_count: int,
        backend: str,
        display_mode: str,
    ) -> None:
        batch_id = uuid.uuid4().hex
        output_directory = output_path.parent
        database_path = output_directory / DATABASE_FILENAME
        writer: DatabaseWriter | None = None
        total = 0
        try:
            self._emit("log", message="正在读取累计输出第一列并同步输入文件……")
            if ensure_cumulative_output_quote_all(output_path):
                self._emit(
                    "log",
                    message=(
                        "累计 CSV 已迁移为全字段双引号格式；"
                        "LibreOffice 打开时生日保持单列 MM/DD/YYYY。"
                    ),
                )
            existing_keys = read_output_first_column_keys(output_path)
            removed_before_start = remove_input_rows_by_keys(input_path, existing_keys)
            if removed_before_start:
                self._emit(
                    "log",
                    message=f"输入文件已删除 {removed_before_start} 条输出中已存在的资料。",
                )
            try:
                records = load_input_records(input_path)
            except ValueError as exc:
                if "没有可导入的资料行" not in str(exc):
                    raise
                records = []
            missing_required = [
                record
                for record in records
                if record.import_error.startswith("缺少必填字段：")
            ]
            if missing_required:
                removed_missing = remove_input_rows_by_locations(
                    input_path, missing_required
                )
                self._emit(
                    "log",
                    message=(
                        f"输入文件已直接删除 {removed_missing} 条缺少必填字段的资料；"
                        "这些资料不写累计输出、不进入浏览器。"
                    ),
                )
                records = [
                    record for record in records if record not in missing_required
                ]
            total = len(records)
            if self._stop_event.is_set():
                raise StopRequested("导入阶段已停止")

            writer = DatabaseWriter(database_path)
            writer.start()
            writer.request(
                "create_batch",
                batch_id=batch_id,
                input_path=str(input_path),
                output_directory=str(output_directory),
                thread_count=thread_count,
            )
            work_items: list[WorkItem] = writer.request(
                "insert_records", batch_id=batch_id, records=records
            )
            invalid_count = sum(bool(record.import_error) for record in records)
            valid_count = len(records) - invalid_count
            duplicate_count = valid_count - len(work_items)
            self._emit(
                "log",
                message=(
                    f"输入剩余 {len(records)} 条；唯一浏览器任务 {len(work_items)} 个，"
                    f"同第一列重复 {duplicate_count} 条，格式错误 {invalid_count} 条。"
                    f"后端 {backend}，"
                    f"显示 {display_mode}，处理线程 {thread_count}。"
                ),
            )
            self._publish_progress(writer, batch_id, total)

            tasks: queue.Queue[tuple[WorkItem, int] | None] = queue.Queue()
            for item in work_items:
                tasks.put((item, 1))
            worker_count = min(thread_count, len(work_items))
            workers = [
                threading.Thread(
                    target=self._worker_loop,
                    args=(
                        number,
                        backend,
                        display_mode,
                        tasks,
                        writer,
                        batch_id,
                        total,
                        input_path,
                        output_path,
                    ),
                    name=f"studentaid-worker-{number}",
                    daemon=False,
                )
                for number in range(1, worker_count + 1)
            ]
            for worker in workers:
                worker.start()
            while workers and not self._stop_event.is_set():
                with tasks.all_tasks_done:
                    unfinished_tasks = tasks.unfinished_tasks
                if unfinished_tasks == 0 or not any(
                    worker.is_alive() for worker in workers
                ):
                    break
                time.sleep(0.1)
            for _ in range(worker_count):
                tasks.put(None)
            for worker in workers:
                worker.join()

            if self._stop_event.is_set():
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="stopped",
                    error="用户停止，任务尚未处理",
                )
                batch_status = "stopped"
            else:
                writer.request(
                    "mark_remaining", batch_id=batch_id, status="failed",
                    error="任务未被浏览器处理",
                )
                batch_status = "completed"
            final_counts = self._publish_progress(writer, batch_id, total)
            writer.request(
                "finish_batch", batch_id=batch_id, status=batch_status,
                export_path=str(output_path),
            )
            writer.request("checkpoint")
            self._emit(
                "finished",
                batch_id=batch_id,
                database_path=str(database_path),
                export_path=str(output_path),
                status=batch_status,
                counts=final_counts,
            )
        except StopRequested as exc:
            self._emit("fatal", message=_clean_error(exc), stopped=True)
        except Exception as exc:
            self._emit(
                "fatal",
                message=_clean_error(exc),
                details=traceback.format_exc(limit=8),
                stopped=False,
            )
        finally:
            if writer is not None:
                try:
                    writer.close()
                except Exception as exc:
                    self._emit("log", message=f"关闭数据库时出错：{_clean_error(exc)}")
            with self._state_lock:
                self._running = False
            self._emit("idle", export_path=str(output_path))


class StudentAidApp:
    """第十五步批量输入兼容版桌面 GUI。"""

    def __init__(self) -> None:
        import tkinter as tk
        from tkinter import filedialog, messagebox, ttk

        self.tk = tk
        self.ttk = ttk
        self.filedialog = filedialog
        self.messagebox = messagebox
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("920x650")
        self.root.minsize(820, 570)

        self.input_var = tk.StringVar()
        self.output_var = tk.StringVar(value=str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME))
        self.backend_var = tk.StringVar(value="browser-use")
        self.display_mode_var = tk.StringVar(value="无头")
        self.thread_var = tk.StringVar(value="2")
        self.status_var = tk.StringVar(value="就绪")
        self.progress_text_var = tk.StringVar(value="总数 0 | 完成 0 | 失败 0 | 停止 0")
        self._ui_events: queue.Queue[tuple[str, dict[str, Any]]] = queue.Queue()
        self.engine = BatchEngine(self._backend_event)
        self._closing = False

        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.after(100, self._poll_events)

    def _build_ui(self) -> None:
        ttk = self.ttk
        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)
        main.columnconfigure(1, weight=1)
        main.rowconfigure(8, weight=1)

        ttk.Label(main, text="导入文件：").grid(row=0, column=0, sticky="w", pady=6)
        self.input_entry = ttk.Entry(main, textvariable=self.input_var)
        self.input_entry.grid(row=0, column=1, sticky="ew", padx=8, pady=6)
        self.input_button = ttk.Button(main, text="选择文件", command=self._choose_input)
        self.input_button.grid(row=0, column=2, pady=6)

        ttk.Label(main, text="累计输出 CSV：").grid(row=1, column=0, sticky="w", pady=6)
        self.output_entry = ttk.Entry(main, textvariable=self.output_var)
        self.output_entry.grid(row=1, column=1, sticky="ew", padx=8, pady=6)
        self.output_button = ttk.Button(main, text="选择文件", command=self._choose_output)
        self.output_button.grid(row=1, column=2, pady=6)

        ttk.Label(main, text="浏览器后端：").grid(row=2, column=0, sticky="w", pady=6)
        self.backend_combo = ttk.Combobox(
            main,
            textvariable=self.backend_var,
            values=BROWSER_BACKENDS,
            width=18,
            state="readonly",
        )
        self.backend_combo.grid(row=2, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="browser-use 和 playwright 都使用每线程独立浏览器。")\
            .grid(row=2, column=1, sticky="w", padx=(180, 0), pady=6)

        ttk.Label(main, text="浏览器显示：").grid(row=3, column=0, sticky="w", pady=6)
        self.display_mode_combo = ttk.Combobox(
            main,
            textvariable=self.display_mode_var,
            values=DISPLAY_MODES,
            width=18,
            state="readonly",
        )
        self.display_mode_combo.grid(row=3, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="窗口=900x720 小窗口；无头=不显示浏览器窗口。")\
            .grid(row=3, column=1, sticky="w", padx=(180, 0), pady=6)

        ttk.Label(main, text="处理线程数：").grid(row=4, column=0, sticky="w", pady=6)
        self.thread_spin = ttk.Entry(main, textvariable=self.thread_var, width=10)
        self.thread_spin.grid(row=4, column=1, sticky="w", padx=8, pady=6)
        ttk.Label(main, text="任意正整数；每个线程独立缓存、页面和浏览器进程。")\
            .grid(row=4, column=1, sticky="w", padx=(100, 0), pady=6)

        button_bar = ttk.Frame(main)
        button_bar.grid(row=5, column=0, columnspan=3, sticky="ew", pady=(10, 8))
        self.start_button = ttk.Button(button_bar, text="开始", command=self._start)
        self.start_button.pack(side="left")
        self.stop_button = ttk.Button(
            button_bar, text="停止", command=self._stop, state="disabled"
        )
        self.stop_button.pack(side="left", padx=8)
        ttk.Label(button_bar, textvariable=self.status_var).pack(side="right")

        self.progress = ttk.Progressbar(main, mode="determinate", maximum=1, value=0)
        self.progress.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(2, 4))
        ttk.Label(main, textvariable=self.progress_text_var).grid(
            row=7, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        log_frame = ttk.LabelFrame(main, text="运行日志（不显示 SSN 和原始资料）", padding=8)
        log_frame.grid(row=8, column=0, columnspan=3, sticky="nsew")
        log_frame.rowconfigure(0, weight=1)
        log_frame.columnconfigure(0, weight=1)
        self.log_text = self.tk.Text(log_frame, wrap="word", state="disabled")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

    def _choose_input(self) -> None:
        path = self.filedialog.askopenfilename(
            title="选择导入文件",
            filetypes=[
                ("支持的资料文件", "*.csv *.scv *.txt *.xlsx"),
                ("CSV 文件", "*.csv"),
                ("Excel 文件", "*.xlsx"),
                ("纯文本文件", "*.txt *.scv"),
                ("所有文件", "*.*"),
            ],
        )
        if path:
            self.input_var.set(path)
            current = self.output_var.get().strip()
            if not current or current == str(Path.cwd() / CUMULATIVE_OUTPUT_FILENAME):
                self.output_var.set(str(Path(path).parent / CUMULATIVE_OUTPUT_FILENAME))

    def _choose_output(self) -> None:
        current = Path(self.output_var.get().strip() or CUMULATIVE_OUTPUT_FILENAME)
        path = self.filedialog.asksaveasfilename(
            title="选择累计输出 CSV（已有文件会继续叠加）",
            initialdir=str(current.parent),
            initialfile=current.name,
            defaultextension=".csv",
            filetypes=[("CSV 文件", "*.csv"), ("所有文件", "*.*")],
        )
        if path:
            self.output_var.set(path)

    def _set_running(self, running: bool) -> None:
        normal_or_disabled = "disabled" if running else "normal"
        self.input_entry.configure(state=normal_or_disabled)
        self.output_entry.configure(state=normal_or_disabled)
        self.backend_combo.configure(state="disabled" if running else "readonly")
        self.display_mode_combo.configure(state="disabled" if running else "readonly")
        self.thread_spin.configure(state=normal_or_disabled)
        self.input_button.configure(state=normal_or_disabled)
        self.output_button.configure(state=normal_or_disabled)
        self.start_button.configure(state=normal_or_disabled)
        self.stop_button.configure(state="normal" if running else "disabled")

    def _start(self) -> None:
        try:
            input_path = Path(self.input_var.get().strip())
            output_target = Path(self.output_var.get().strip())
            thread_count = int(self.thread_var.get().strip())
            backend = self.backend_var.get().strip()
            display_mode = self.display_mode_var.get().strip()
            self.engine.start(
                input_path, output_target, thread_count, backend, display_mode
            )
        except Exception as exc:
            self.messagebox.showerror(APP_TITLE, _clean_error(exc))
            return
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self.status_var.set("运行中")
        self._set_running(True)
        self._append_log(
            f"批次已启动：后端 {self.backend_var.get()}，"
            f"显示 {self.display_mode_var.get()}，线程 {self.thread_var.get()}。"
        )

    def _stop(self) -> None:
        self.stop_button.configure(state="disabled")
        self.status_var.set("停止中")
        self.engine.stop()

    def _backend_event(self, kind: str, payload: dict[str, Any]) -> None:
        self._ui_events.put((kind, payload))

    def _append_log(self, message: str) -> None:
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _poll_events(self) -> None:
        try:
            while True:
                kind, payload = self._ui_events.get_nowait()
                if kind == "log":
                    self._append_log(str(payload.get("message", "")))
                elif kind == "progress":
                    total = int(payload.get("total", 0))
                    completed = int(payload.get("completed", 0))
                    failed = int(payload.get("failed", 0))
                    stopped = int(payload.get("stopped", 0))
                    processing = int(payload.get("processing", 0))
                    terminal = completed + failed + stopped
                    self.progress.configure(maximum=max(1, total), value=terminal)
                    self.progress_text_var.set(
                        f"总数 {total} | 完成 {completed} | 失败 {failed} | "
                        f"停止 {stopped} | 处理中 {processing}"
                    )
                elif kind == "finished":
                    output_path = str(payload.get("export_path", ""))
                    database_path = str(payload.get("database_path", ""))
                    counts = payload.get("counts", {})
                    failed = int(counts.get("failed", 0)) if isinstance(counts, dict) else 0
                    if payload.get("status") == "stopped":
                        display_status = "已停止"
                    elif failed:
                        display_status = "完成（有失败）"
                    else:
                        display_status = "已完成"
                    self.status_var.set(display_status)
                    self._append_log(f"累计 CSV：{output_path}")
                    self._append_log(f"SQLite 数据库：{database_path}")
                    self.messagebox.showinfo(
                        APP_TITLE,
                        f"批次处理结束。\n\n累计 CSV：{output_path}\n数据库：{database_path}",
                    )
                elif kind == "fatal":
                    message = str(payload.get("message", "未知错误"))
                    self.status_var.set("已停止" if payload.get("stopped") else "失败")
                    self._append_log(message)
                    if not payload.get("stopped"):
                        self.messagebox.showerror(APP_TITLE, message)
                elif kind == "idle":
                    self._set_running(False)
                self._ui_events.task_done()
        except queue.Empty:
            pass

        if self._closing and not self.engine.is_running:
            self.root.destroy()
            return
        self.root.after(100, self._poll_events)

    def _on_close(self) -> None:
        if self.engine.is_running:
            if not self.messagebox.askyesno(APP_TITLE, "任务仍在运行，是否停止并退出？"):
                return
            self._closing = True
            self._set_running(True)
            self.stop_button.configure(state="disabled")
            self.status_var.set("停止中")
            self.engine.stop()
        else:
            self.root.destroy()

    def run(self) -> None:
        self.root.mainloop()


def main() -> None:
    StudentAidApp().run()


if __name__ == "__main__":
    main()
